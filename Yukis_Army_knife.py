from tkinter import *
from tkinter import ttk
import pyperclip as clip
from tkinter import messagebox
import requests
import webbrowser
import my_icon
import socket
import re
import time
import os
from tkinterdnd2 import *
import random
from yt_dlp import YoutubeDL
from PIL import Image, ImageTk,ExifTags,ImageOps,ImageDraw,ImageFont
import qrcode
import cv2
import numpy as np
from moviepy.editor import concatenate_audioclips
import sys
from moviepy.editor import concatenate_videoclips
from moviepy.editor import VideoFileClip
import keyboard as key
import pyautogui
import threading
import pypdf
import speedtest
import io
import base64
import svgwrite
import shutil
import win32gui
import win32con
from pynput import keyboard
import winotify
from tkinter import filedialog
import win32process
import psutil
import subprocess
import win32api
import win32ui
import urllib.parse
from moviepy.editor import AudioFileClip
from  randomuser import *
from datetime import datetime,timedelta
import time
from gtts import gTTS
from playsound import playsound
from urllib.parse import urlparse
import hashlib
from tkinter import colorchooser
import win32_setctime
from PIL import ImageEnhance
from tkinter import scrolledtext
import jaconv
from moviepy.video.fx.speedx import speedx
import mistune
from moviepy.editor import ImageClip
import zipfile
import speech_recognition as sr
from moviepy.video.fx.resize import resize
from moviepy.video.fx.blackwhite import blackwhite
from moviepy.video.fx.lum_contrast import lum_contrast
from moviepy.video.fx.margin import margin
from googletrans import Translator,LANGUAGES
from bs4 import BeautifulSoup
from moviepy.video.fx.mirror_x import mirror_x
from moviepy.video.fx.mirror_y import mirror_y
import ctypes
import math
import string
import pykakasi
from ttkthemes import *
from moviepy.video.fx.all import fadein
from moviepy.video.fx.all import fadeout
import Crypto
from Crypto.Cipher import AES
from youtube_transcript_api import YouTubeTranscriptApi
from PIL import ImageGrab
from urllib.parse import urljoin,unquote
from database import *
from mutagen.easyid3 import EasyID3
from mutagen.id3 import ID3
import base64
import pystray
import win32com.client
import openpyxl
import csv
import unicodedata
from moviepy.editor import ImageSequenceClip
import sv_ttk
import winsound
import win32clipboard
from jeraconv import jeraconv
import datauri
from functools import partial
import getpass
import tkinter.simpledialog as simpledialog
from chardet import detect
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import mdx_truly_sane_lists
import markdown
from tkinterweb import HtmlFrame
import pillow_heif
import pyotp
from moviepy.video.fx.rotate import rotate
from moviepy.audio.fx.all import audio_normalize
from moviepy.video.fx.loop import loop
from moviepy.video.fx.invert_colors import invert_colors
from moviepy.audio.fx.audio_fadein import audio_fadein
from moviepy.audio.fx.audio_fadeout import audio_fadeout
import yaml
from Crypto.PublicKey import RSA
from Crypto.Cipher import AES, PKCS1_OAEP
from Crypto.Hash import SHA256
import diff_match_patch
from magika import Magika
from pygments import highlight
from pygments.lexers import get_lexer_by_name
from pygments.formatters import ImageFormatter
from pygments.styles import STYLE_MAP
from pygments.lexers import get_all_lexers

version="6.6"

class MyListBox(Listbox):
    def __init__(self, master, inlist, **kwargs):
        super().__init__(master, **kwargs)
        self.inlist = inlist
        for i in inlist:
            self.insert(END, i)

    def insert_inlist(self, inlist):
        self.inlist = inlist
        self.delete(0, END)
        for i in inlist:
            self.insert(END, i)

    def pack1(self, **kwargs):
        super().pack(**kwargs)


class ScrolledList(MyListBox):
    def __init__(self, master, **kwargs):
        self.frame=ttk.Frame(master)
        super().__init__(self.frame, **kwargs)
        self.scrollbar = ttk.Scrollbar(self.frame, orient=VERTICAL, command=self.yview)
        self.configure(yscrollcommand=self.scrollbar.set)
        self.pack1(side=LEFT, fill=BOTH, expand=True)
        self.scrollbar.pack(side=LEFT, fill=Y)

    def pack(self, **kwargs):
        self.frame.pack(**kwargs)

    def grid(self, **kwargs):
        self.frame.grid(**kwargs)

    def place(self, **kwargs):
        self.frame.place(**kwargs)


class RightClickMenu(Menu):
    def __init__(self, master,func,tree_rclick=False,list_rclick=False, **kwargs): # func = [[label,command],[label,command]...]
        def showMenu(e):
            if tree_rclick:
                x, y = e.x, e.y
                item = master.identify_row(y)
                if item:
                    master.selection_set(item)
                    master.focus(item)
                else:
                    # アイテムがない場合は選択を解除する
                    master.selection_remove(master.selection())
            elif list_rclick:
                x, y = e.x, e.y
                item = master.nearest(y)
                #if item:
                #print(item)
                master.select_clear(0, END)
                master.selection_set(item)

            self.post(e.x_root, e.y_root)

        kwargs['tearoff'] = 0
        Menu.__init__(self, master, **kwargs)
        master.bind("<Button-3>", showMenu)

        for i in range(len(func)):
            self.add_command(label=func[i][0], command=func[i][1])


class SortTreeview(ttk.Treeview):
    # columns = [[A,B,C],[D,E,F]...]
    # arrRows = list(string)
    # arrColWidth = list(num)
    # arrColAlignment = list(e,w,center)
    # arrSortType = list(name,num,date,multidecimal,numcomma)

    def __init__(self,master,arrRows,arrColWidth=[],arrColAlignment=[],arrSortType=[],**kwargs):
        kwargs['show'] ="headings"
        self.kwargs=kwargs
        self.master=master
        super().__init__(self.master,**kwargs)
        self.arrlbHeader = kwargs["columns"]
        self.arrRows = arrRows

        if arrColWidth==[]:
            self.width=[100 for i in range(len(self.arrlbHeader))]
        else:
            self.width=arrColWidth

        if arrColAlignment==[]:
            self.alignment=["center" for i in range(len(self.arrlbHeader))]
        else:
            self.alignment=arrColAlignment

        if arrSortType==[]:
            self.sorttype=["name" for i in range(len(self.arrlbHeader))]
        else:
            self.sorttype=arrSortType
        for iCount in range(len(self.arrlbHeader)):
            strHdr = self.arrlbHeader[iCount]
            self.heading(strHdr, text=strHdr.title(), sort_by=self.sorttype[iCount])
            self.column(self.arrlbHeader[iCount], width=self.width[iCount], stretch=True, anchor=self.alignment[iCount])
        for iCount in range(len(self.arrRows)):
            self.insert("", "end", values=self.arrRows[iCount])

    def heading(self, column, sort_by=None, **kwargs):
        if sort_by and not hasattr(kwargs, 'command'):
            func = getattr(self, f"_sort_by_{sort_by}", None)
            if func:
                kwargs['command'] = partial(func, column, False)
        return super().heading(column, **kwargs)

    def _sort(self, column, reverse, data_type, callback):
        l = [(self.set(k, column), k) for k in self.get_children('')]
        l.sort(key=lambda t: data_type(t[0]), reverse=reverse)
        for index, (_, k) in enumerate(l):
            self.move(k, '', index)
        self.heading(column, command=partial(callback, column, not reverse))

    def _sort_by_num(self, column, reverse):
        self._sort(column, reverse, int, self._sort_by_num)

    def _sort_by_name(self, column, reverse):
        self._sort(column, reverse, str, self._sort_by_name)

    def _sort_by_date(self, column, reverse):
        def _str_to_datetime(string):
            return datetime.datetime.strptime(string, "%Y-%m-%d")
        self._sort(column, reverse, _str_to_datetime, self._sort_by_date)

    def _sort_by_multidecimal(self, column, reverse):
        def _multidecimal_to_str(string):
            arrString = string.split(".")
            strNum = ""
            for iValue in arrString:
                strValue = f"{int(iValue):02}"
                strNum = "".join([strNum, str(strValue)])
            strNum = "".join([strNum, "0000000"])
            return int(strNum[:8])
        self._sort(column, reverse, _multidecimal_to_str, self._sort_by_multidecimal)

    def _sort_by_numcomma(self, column, reverse):
        def _numcomma_to_num(string):
            return int(string.replace(",", ""))
        self._sort(column, reverse, _numcomma_to_num, self._sort_by_numcomma)

    def pack1(self, **kwargs):
        super().pack(**kwargs)

    def configure_arr(self, arrRows=[[]]):
        self.arrRows = arrRows
        for item in self.get_children():
            self.delete(item)

        for iCount in range(len(self.arrlbHeader)):
            strHdr = self.arrlbHeader[iCount]
            self.heading(strHdr, text=strHdr.title(), sort_by=self.sorttype[iCount])
            self.column(self.arrlbHeader[iCount], width=self.width[iCount], stretch=True, anchor=self.alignment[iCount])
        for iCount in range(len(self.arrRows)):
            self.insert("", "end", values=self.arrRows[iCount])

class ScrolledTree(SortTreeview):
    def __init__(self, master, **kwargs):
        self.frame=ttk.Frame(master)
        super().__init__(self.frame, **kwargs)
        self.scrollbar = ttk.Scrollbar(self.frame, orient=VERTICAL, command=self.yview)
        self.configure(yscrollcommand=self.scrollbar.set)
        self.pack1(side=LEFT, fill=BOTH, expand=True)
        self.scrollbar.pack(side=LEFT, fill=Y)

    def pack(self, **kwargs):
        self.frame.pack(**kwargs)

    def grid(self, **kwargs):
        self.frame.grid(**kwargs)

    def place(self, **kwargs):
        self.frame.place(**kwargs)

class Free_window(Toplevel):

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Toplevelウィンドウにフォーカスが当たったときに親ウィンドウを非表示にする
        if hide_mainwindow==1:
            self.bind("<FocusIn>", hide_parent)

        # Toplevelウィンドウからフォーカスが外れたときに親ウィンドウを再表示する
        #self.bind("<FocusOut>", show_parent)

    def move_window(self,wiget):
        self.isMouseDown = False

        def mouseDown(event):
            self.isMouseDown = True
            self.origin = (event.x, event.y)

        def mouseRelease(event):
            self.isMouseDown = False

        def mouseMove(event):
            if self.isMouseDown:
                x = self.winfo_x() + (event.x - self.origin[0])
                y = self.winfo_y() + (event.y - self.origin[1])
                self.geometry("+%s+%s" % (x, y))

        wiget.bind("<Button>", mouseDown)
        wiget.bind("<ButtonRelease>", mouseRelease)
        wiget.bind("<Motion>", mouseMove)

class ScrolledText(scrolledtext.ScrolledText):
    def __init__(self, *args, **kwargs):
        kwargs['undo'] = True
        kwargs["font"]=("BIZ UDPGothic",)
        #kwargs["bg"]=rgbToHex((40, 40, 40))
        #kwargs["fg"]=rgbToHex((230, 230, 230))
        #kwargs["insertbackground"]=rgbToHex((255, 255, 255))
        super().__init__(*args, **kwargs)

class Tk(ThemedTk, TkinterDnD.DnDWrapper):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.TkdndVersion = TkinterDnD._require(self)

class Searchbox(ttk.Combobox):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        val_list=kwargs["values"]
        list_button=val_list.copy()
        val_list = StringVar()
        self.configure(textvariable=val_list)

        def handle_keypress(event):
            if event.keysym == 'Return':
                self.event_generate('<Down>')
            elif event.keysym == 'Delete':
                self.delete(0,END)

        def on_combobox_key_release(event):
            current_text = val_list.get()
            filtered_list = [item for item in list_button if current_text.lower() in item.lower()]
            self["values"] = filtered_list

        self.bind('<KeyRelease>', on_combobox_key_release)
        self.bind('<KeyPress>', handle_keypress)


# メイン画面
def set_frame1(page_x):
    global frame1,button45,buttonY,button143,combobox_frame,button_function_list,dist_button
    frame1 = ttk.Frame(root)
    root.protocol("WM_DELETE_WINDOW", frame_delete)
    random.seed(None)
    buttonY=Button(root)
    execute()
    men=Menu(root)
    root.config(menu=men)
    men1=Menu(men,relief=GROOVE,tearoff=0)
    men2=Menu(men,relief=GROOVE,tearoff=0)
    men3=Menu(men,relief=GROOVE,tearoff=0)
    men.add_cascade(label='メニュー', menu=men1)
    men.add_cascade(label='テーマ変更', menu=men2)
    men.add_cascade(label='設定', menu=men3)
    men3.add_command(label='スタートアップ登録', command=startup_reg)
    men1.add_command(label='バージョン確認', command=lambda: messagebox.showinfo("バージョン",f"Yuki's Army knife ver{version}"))
    men1.add_command(label='ホームページ', command=home_page)
    men1.add_command(label='新規ver確認',command=check_new_ver)
    #men1.add_command(label='ミニモード',command=mini_win)
    men2.add_command(label='デフォルト', command=lambda:set_theme(0))
    men2.add_command(label='モダン', command=lambda:set_theme(1))
    men2.add_command(label='ダーク', command=lambda:set_theme(2))
    men2.add_command(label='デジタル', command=lambda:set_theme(3))
    men2.add_command(label='ナチュラル', command=lambda:set_theme(5))
    men2.add_command(label='システマチック', command=lambda:set_theme(6))
    men2.add_command(label='ケラミック', command=lambda:set_theme(7))
    men2.add_command(label='木材風イエロー', command=lambda:set_theme(8))
    men2.add_command(label='プラスチック', command=lambda:set_theme(9))
    #men2.add_command(label='ラディエント', command=lambda:set_theme(10))
    men2.add_command(label='XP風', command=lambda:set_theme(11))
    men2.add_command(label='砂風', command=lambda:set_theme(12))
    men2.add_command(label='クラッシック', command=lambda:set_theme(13))
    men2.add_command(label='モダンダーク', command=lambda:set_theme(14))
    men2.add_command(label='モダンライト', command=lambda:set_theme(15))
    men3.add_command(label='タスクアイコンに登録', command=button_task_icon)
    men1.add_command(label='再起動', command=restart)
    men3.add_command(label='詳細設定', command=advanced_setting)
    men3.add_command(label='設定フォルダ', command=setting_folder)

    notebook = ttk.Notebook(frame1)
    frame_text=ttk.Frame(notebook,name="text")
    frame_video=ttk.Frame(notebook,name="video")
    frame_image=ttk.Frame(notebook,name="image")
    frame_other=ttk.Frame(notebook,name="other")
    frame_text1=ttk.Frame(frame_text,name="text1",relief=RIDGE,borderwidth=3,padding=7)
    frame_text2=ttk.Frame(frame_text,name="text2",relief=RIDGE,borderwidth=3,padding=7)
    frame_text3=ttk.Frame(frame_text,name="text3",relief=RIDGE,borderwidth=3,padding=7)
    frame_video1=ttk.Frame(frame_video,name="video1",relief=RIDGE,borderwidth=3,padding=7)
    frame_video2=ttk.Frame(frame_video,name="video2",relief=RIDGE,borderwidth=3,padding=7)
    frame_video3=ttk.Frame(frame_video,name="video3",relief=RIDGE,borderwidth=3,padding=7)
    frame_image1=ttk.Frame(frame_image,name="image1",relief=RIDGE,borderwidth=3,padding=7)
    frame_image2=ttk.Frame(frame_image,name="image2",relief=RIDGE,borderwidth=3,padding=7)
    frame_image3=ttk.Frame(frame_image,name="image3",relief=RIDGE,borderwidth=3,padding=7)
    frame_other1=ttk.Frame(frame_other,name="other1",relief=RIDGE,borderwidth=3,padding=7)
    frame_other2=ttk.Frame(frame_other,name="other2",relief=RIDGE,borderwidth=3,padding=7)
    frame_other3=ttk.Frame(frame_other,name="other3",relief=RIDGE,borderwidth=3,padding=7)
    frame_config=ttk.Frame(frame1)
    notebook.add(frame_text, text="テキスト・情報")
    notebook.add(frame_video, text="動画・音声")
    notebook.add(frame_image, text="画像・PDF")
    notebook.add(frame_other, text="システム・小物")
    buttonX=ttk.Checkbutton(
        frame_config,
        text=main_checkbox_name,
        onvalue=1,
        offvalue=0,
        variable=window_front,
        command=execute
        )
    #buttonZ=Button(
    #    frame_config,
    #    text="ミニモード",
    #    command=mini_win,
    #    bg="gray",
    #    fg="white"
    #)
    button1 = ttk.Button(
        frame_text2,
        width=13,
        text='文字カウント',
        command=count
    )
    button2 = ttk.Button(
        frame_text1,
        width=13,
        text='空白削除',
        command=blank
    )
    button3 = ttk.Button(
        frame_text1,
        width=13,
        text='置換',
        command=tikan
    )
    button4=ttk.Button(
        frame_text2,
        width=13,
        text="ネット検索",
        command=serch
    )
    button5=ttk.Button(
        frame_text3,
        width=13,
        text="抽選",
        command=choose
    )
    button6=ttk.Button(
        frame_other1,
        width=13,
        text="IPアドレス",
        command=IP
    )
    button7=ttk.Button(
        frame_other3,
        width=13,
        text="タイマー",
        command=timer
    )
    button8=ttk.Button(
        frame_other2,
        width=13,
        text="フォルダ作成",
        command=folder
    )
    button9=ttk.Button(
        frame_other2,
        width=13,
        text="パス取得",
        command=local_pass
    )
    button10=ttk.Button(
        frame_other3,
        width=13,
        text="カウンター",
        command=count_down
    )
    button11=ttk.Button(
        frame_video3,
        width=13,
        text="動画DL",
        command=video_dwonload
    )
    button12=ttk.Button(
        frame_image2,
        width=13,
        text="画像リサイズ",
        command=image_resize
    )
    button13=ttk.Button(
        frame_image3,
        width=13,
        text="画像変換",
        command=image_change
    )
    button14=ttk.Button(
        frame_image3,
        width=13,
        text="QR作成",
        command=qr_generate
    )
    button15=ttk.Button(
        frame_image3,
        width=13,
        text="QR読み取り",
        command=qr_reader
    )
    button16=ttk.Button(
        frame_image2,
        width=13,
        text="自動顔モザイク",
        command=face_mosaic
    )
    button17=ttk.Button(
        frame_image2,
        width=13,
        text="画像を回転",
        command=img_rotate
    )
    button18=ttk.Button(
        frame_image2,
        width=13,
        text="グレースケール",
        command=img_gray
    )
    button19=ttk.Button(
        frame_video1,
        width=13,
        text="動画カット",
        command=video_cut
    )
    button20=ttk.Button(
        frame_image2,
        width=13,
        text="トリミング",
        command=img_triming
    )
    button21=ttk.Button(
        frame_image3,
        width=13,
        text="画像圧縮",
        command=img_press
    )
    button22=ttk.Button(
        frame_other3,
        width=13,
        text="自動クリック",
        command=mouse_click
    )
    button23=ttk.Button(
        frame_video2,
        width=13,
        text="動画→音声",
        command=video_to_voice
    )
    button24=ttk.Button(
        frame_image1,
        width=13,
        text="PDF分解・抽出",
        command=pdf_split
    )
    button25=ttk.Button(
        frame_image1,
        width=13,
        text="PDF結合",
        command=pdf_merge
    )
    button26=ttk.Button(
        frame_other3,
        width=13,
        text="赤シート",
        command=red_sheet
    )
    button27=ttk.Button(
        frame_other1,
        width=13,
        text="スピードテスト",
        command=speed_check
    )
    button28=ttk.Button(
        frame_image3,
        width=13,
        text="EXIF",
        command=exif
    )
    button29=ttk.Button(
        frame_other3,
        width=13,
        text="付箋作成",
        command=memo
    )
    button30=ttk.Button(
        frame_video2,
        width=13,
        text="動画→GIF",
        command=video_to_gif
    )
    button31=ttk.Button(
        frame_image3,
        width=13,
        text="SVG変換",
        command=svg_con
    )
    button32=ttk.Button(
        frame_image3,
        width=13,
        text="透過PNG作成",
        command=img_inv
    )
    button33=ttk.Button(
        frame_other1,
        width=13,
        text="窓を最前面",
        command=front_window
    )
    button34=ttk.Button(
        frame_other3,
        width=13,
        text="画像付箋",
        command=screen_memo
    )
    button35=ttk.Button(
        frame_text3,
        width=13,
        text="短縮URL",
        command=short_url
    )
    button36=ttk.Button(
        frame_other1,
        width=13,
        text="窓のexe取得",
        command=window_path
    )
    button37=ttk.Button(
        frame_image3,
        width=13,
        text="アイコン抽出",
        command=icon_image
    )
    button38=ttk.Button(
        frame_other1,
        width=13,
        text="スリープ防止",
        command=sleep_stop
        )
    button39=ttk.Button(
        frame_text2,
        width=13,
        text="テキスト取得",
        command=txt_get
    )
    button40=ttk.Button(
        frame_video1,
        width=13,
        text="音声カット",
        command=voice_cut
    )
    button41=ttk.Button(
        frame_text3,
        width=13,
        text="ダミー人間",
        command=dumy_human
    )
    button42=ttk.Button(
        frame_other3,
        width=13,
        text="簡易時計",
        command=clock
    )
    button43=ttk.Button(
        frame_text3,
        width=13,
        text="進数変換",
        command=num_change
    )
    button44=ttk.Button(
        frame_text2,
        width=13,
        text="合体ペースト",
        command=sum_paste
    )
    button45=ttk.Button(
        frame_text2,
        width=13,
        text="連続コピペ",
        command=copy_paste
    )
    button46=ttk.Button(
        frame_other1,
        width=13,
        text="長さ測定",
        command=length_ruler
    )
    button47=ttk.Button(
        frame_video3,
        width=13,
        text="文章読み上げ",
        command=voice_read
    )
    button48=ttk.Button(
        frame_text3,
        width=13,
        text="パスワード生成",
        command=password
    )
    button49=ttk.Button(
        frame_other3,
        width=13,
        text="カラーパレット",
        command=color_bullet
    )
    button50=ttk.Button(
        frame_other2,
        width=13,
        text="更新日時編集",
        command=file_day
    )
    button51=ttk.Button(
        frame_video2,
        width=13,
        text="mp3↔wav",
        command=mp3_wav
    )
    button52=ttk.Button(
        frame_other1,
        width=13,
        text="システム使用率",
        command=pc_info
    )
    button53=ttk.Button(
        frame_video3,
        width=13,
        text="GIF作成",
        command=gif_create
    )
    button54=ttk.Button(
        frame_image2,
        width=13,
        text="画像余白追加",
        command=img_blank
    )
    button55=ttk.Button(
        frame_image2,
        width=13,
        text="画像調整",
        command=image_set
    )
    button56=ttk.Button(
        frame_image2,
        width=13,
        text="ガンマ補正",
        command=img_gamma
    )
    button57=ttk.Button(
        frame_image2,
        width=13,
        text="モザイク画像",
        command=area_mozaic
    )
    button58=ttk.Button(
        frame_text1,
        width=13,
        text="文字変換",
        command=translate
    )
    button59=ttk.Button(
        frame_video1,
        width=13,
        text="音量変更",
        command=video_volume
    )
    button60=ttk.Button(
        frame_video1,
        width=13,
        text="速度変更",
        command=video_speed
    )
    button61=ttk.Button(
        frame_video3,
        width=13,
        text="動画スクショ",
        command=video_shot
    )
    button62=ttk.Button(
        frame_video3,
        width=13,
        text="動画情報",
        command=video_info
    )
    button63=ttk.Button(
        frame_text3,
        width=13,
        text="md→html",
        command=md2html
    )
    button64=ttk.Button(
        frame_video1,
        width=13,
        text="動画結合",
        command=video_concat
    )
    button65=ttk.Button(
        frame_video2,
        width=13,
        text="音声→動画",
        command=voice2video
    )
    button66=ttk.Button(
        frame_other2,
        width=13,
        text="zip作成",
        command=compress
    )
    button67=ttk.Button(
        frame_other2,
        width=13,
        text="zip解凍",
        command=decompress
    )
    button68=ttk.Button(
        frame_image3,
        width=13,
        text="画像情報",
        command=image_info
    )
    button69=ttk.Button(
        frame_other1,
        width=13,
        text="窓透明化",
        command=window_toumei
    )
    button70=ttk.Button(
        frame_text3,
        width=13,
        text="URL変換",
        command=url_change
    )
    button71=ttk.Button(
        frame_video3,
        width=13,
        text="文字起こし",
        command=voice2read
    )
    button72=ttk.Button(
        frame_video1,
        width=13,
        text="動画リサイズ",
        command=video_size
    )
    button73=ttk.Button(
        frame_video1,
        width=13,
        text="動画白黒化",
        command=video_BW
    )
    button74=ttk.Button(
        frame_video1,
        width=13,
        text="動画色調整",
        command=video_light
    )
    button75=ttk.Button(
        frame_video1,
        width=13,
        text="動画余白",
        command=video_frame
    )
    #button76=ttk.Button(
    #    frame_text3,
    #    width=13,
    #    text="色コード変換",
    #    command=color_code
    #)
    button76=ttk.Button(
        frame_other3,
        width=13,
        text="マスコット化",
        command=desc_mascot
    )
    button77=ttk.Button(
        frame_text3,
        width=13,
        text="翻訳",
        command=translate_text
    )
    button78=ttk.Button(
        frame_text2,
        width=13,
        text="パスワード確認",
        command=pass_check
    )
    button79=ttk.Button(
        frame_video2,
        width=13,
        text="動画→映像",
        command=voice_delete
    )
    button80=ttk.Button(
        frame_text3,
        width=13,
        text="サイトタイトル",
        command=site_title
    )
    button81=ttk.Button(
        frame_other1,
        width=13,
        text="通信監視",
        command=network_monitor
        )
    button82=ttk.Button(
        frame_text1,
        width=13,
        text="改行調整",
        command=n_del
    )
    button83=ttk.Button(
        frame_image2,
        width=13,
        text="画像反転",
        command=image_mirror
    )
    button84=ttk.Button(
        frame_video1,
        width=13,
        text="動画反転",
        command=video_mirror
    )
    button85=ttk.Button(
        frame_text1,
        width=13,
        text="プレーン貼付",
        command=plain_paste
    )
    button86=ttk.Button(
        frame_other1,
        width=13,
        text="電源操作",
        command=power_control
    )
    button87=ttk.Button(
        frame_text1,
        width=13,
        text="ソート",
        command=sort_list
    )
    button89=ttk.Button(
        frame_image1,
        width=13,
        text="PDFパス解除",
        command=pdf_pass
    )
    button88=ttk.Button(
        frame_other3,
        width=13,
        text="BMI計算",
        command=bmi
    )
    button90=ttk.Button(
        frame_other1,
        width=13,
        text="通信先表示",
        command=net_info
    )
    button91=ttk.Button(
        frame_video3,
        width=13,
        text="GIF分解",
        command=gif_split
    )
    button92=ttk.Button(
        frame_image2,
        width=13,
        text="台形補正",
        command=trapezoid
    )
    button93=ttk.Button(
        frame_image2,
        width=13,
        text="書類影削除",
        command=shadow_delete
    )
    button94=ttk.Button(
        frame_video2,
        width=13,
        text="動画変換",
        command=video_convert
    )
    button95=ttk.Button(
        frame_video1,
        width=13,
        text="FPS変更",
        command=video_fps
    )
    button96=ttk.Button(
        frame_video1,
        width=13,
        text="動画音声変更",
        command=video_bgm
    )
    button97=ttk.Button(
        frame_text1,
        width=13,
        text="重複削除",
        command=duplication
    )
    button98=ttk.Button(
        frame_video3,
        width=13,
        text="サムネ取得",
        command=video_thumbnail
    )
    button99=ttk.Button(
        frame_text3,
        width=13,
        text="住所英語化",
        command=address_english
    )
    button100=ttk.Button(
        frame_other1,
        width=13,
        text="入力一時停止",
        command=input_stop
    )
    button101=ttk.Button(
        frame_text1,
        width=13,
        text="連番リスト",
        command=renban_list
    )
    button102=ttk.Button(
        frame_video3,
        width=13,
        text="WEBカメラ",
        command=web_camera
    )
    button103=ttk.Button(
        frame_video1,
        width=13,
        text="動画フェード",
        command=vidwo_fade
    )
    button104=ttk.Button(
        frame_other2,
        width=13,
        text="ファイル暗号化",
        command=aes_file
    )
    button105=ttk.Button(
        frame_video3,
        width=13,
        text="動画字幕DL",
        command=video_str
    )
    button106=ttk.Button(
        frame_other2,
        width=13,
        text="完全削除",
        command=delete2file
    )
    button107=ttk.Button(
        frame_other1,
        width=13,
        text="窓中心配置",
        command=window_center
    )
    button108=ttk.Button(
        frame_image3,
        width=13,
        text="画像結合",
        command=image_join
    )
    button109=ttk.Button(
        frame_image3,
        width=13,
        text="コピー画像取得",
        command=get_copyimage
    )
    button110=ttk.Button(
        frame_text2,
        width=13,
        text="テキスト検索",
        command=text_search
    )
    button111=ttk.Button(
        frame_image3,
        width=13,
        text="サイト画像DL",
        command=site_image
    )
    button112=ttk.Button(
        frame_other3,
        width=13,
        text="ダイス",
        command=dice
    )
    button113=ttk.Button(
        frame_text3,
        width=13,
        text="ダミー文章",
        command=dummy_txt
    )
    button114=ttk.Button(
        frame_image3,
        width=13,
        text="ダミー画像",
        command=dummy_img
    )
    button115=ttk.Button(
        frame_text3,
        width=13,
        text="UNIXタイム",
        command=unix_time
    )
    button116=ttk.Button(
        frame_video3,
        width=13,
        text="web動画情報",
        command=web_video_info
    )
    button117=ttk.Button(
        frame_text1,
        width=13,
        text="空行削除",
        command=empty_line
    )
    button118=ttk.Button(
        frame_video1,
        width=13,
        text="MP3タグ編集",
        command=mp3_tag
    )
    button119=ttk.Button(
        frame_image1,
        width=13,
        text="PDF文章抽出",
        command=pdf_text
    )
    button120=ttk.Button(
        frame_other2,
        width=13,
        text="ファイルハッシュ",
        command=file_hash
    )
    button121=ttk.Button(
        frame_text3,
        width=13,
        text="文字列ハッシュ",
        command=text_hash
    )
    button122=ttk.Button(
        frame_image3,
        width=13,
        text="画像→base64",
        command=image_base64
    )
    button123=ttk.Button(
        frame_text3,
        width=13,
        text="URLショトカ",
        command=url_shotcut
    )
    button124=ttk.Button(
        frame_other1,
        width=13,
        text="Explore再起動",
        command=explore_restart
    )
    button125=ttk.Button(
        frame_other1,
        width=13,
        text="神モード",
        command=godmode
    )
    button126=ttk.Button(
        frame_text1,
        width=13,
        text="筆記体変換",
        command=cursive
    )
    button127=ttk.Button(
        frame_other3,
        width=13,
        text="割り勘計算",
        command=warikan
    )
    button128=ttk.Button(
        frame_other3,
        width=13,
        text="カラーピッカー",
        command=color_picker
    )
    button129=ttk.Button(
        frame_other3,
        width=13,
        text="拡大鏡",
        command=magnifier
    )
    button130=ttk.Button(
        frame_text3,
        width=13,
        text="Excel↔CSV",
        command=excel_csv
    )
    button131=ttk.Button(
        frame_text3,
        width=13,
        text="JSON↔CSV",
        command=json_csv
    )
    button132=ttk.Button(
        frame_text1,
        width=13,
        text="UTF正規化",
        command=unicode_normalize
    )
    button133=ttk.Button(
        frame_image3,
        width=13,
        text="office画像抽出",
        command=office_image
    )
    button134=ttk.Button(
        frame_video3,
        width=13,
        text="GIF撮影",
        command=gif_video
    )
    button135=ttk.Button(
        frame_video3,
        width=13,
        text="メトローム",
        command=metrome
    )
    button136=ttk.Button(
        frame_text3,
        width=13,
        text="OGP取得",
        command=ogp
    )
    button137=ttk.Button(
        frame_image3,
        width=13,
        text="画像直接コピー",
        command=copy_image
    )
    button138=ttk.Button(
        frame_text3,
        width=13,
        text="クラウドDL",
        command=cloud_download
    )
    button139=ttk.Button(
        frame_video1,
        width=13,
        text="音声結合",
        command=voice_concat
    )
    button140=ttk.Button(
        frame_text3,
        width=13,
        text="JSON整形",
        command=json_format
    )
    button141=ttk.Button(
        frame_other1,
        width=13,
        text="環境変数",
        command=env_word
    )
    button142=ttk.Button(
        frame_text3,
        width=13,
        text="西暦↔和暦",
        command=seiwa
    )
    button143=ttk.Button(
        frame_other1,
        width=13,
        text="壁紙スラショ",
        command=wallpaper
    )
    button144=ttk.Button(
        frame_image1,
        width=13,
        text="PDF画像抽出",
        command=pdf_image
    )
    button145=ttk.Button(
        frame_image1,
        width=13,
        text="PDF圧縮",
        command=pdf_press
    )
    button146=ttk.Button(
        frame_image1,
        width=13,
        text="PDF回転",
        command=pdf_rotate
    )
    button147=ttk.Button(
        frame_other1,
        width=13,
        text="窓強制終了",
        command=window_kill
    )
    button148=ttk.Button(
        frame_other3,
        width=13,
        text="日数計算",
        command=time_calc
    )
    button149=ttk.Button(
        frame_other3,
        width=13,
        text="黄金比計算",
        command=golden_ratio
    )
    button150=ttk.Button(
        frame_video3,
        width=13,
        text="電子音ピアノ",
        command=piano
    )
    button151=ttk.Button(
        frame_other3,
        width=13,
        text="為替変換",
        command=exchange
    )
    button152=ttk.Button(
        frame_other3,
        width=13,
        text="得点計算",
        command=score_calc
    )
    button153=ttk.Button(
        frame_video3,
        width=13,
        text="動画経過画像",
        command=video_time_image
    )
    button154=ttk.Button(
        frame_image1,
        width=13,
        text="PDF挿入",
        command=pdf_insert
    )
    button155=ttk.Button(
        frame_other3,
        width=13,
        text="単語帳",
        command=word_book
    )
    button156=ttk.Button(
        frame_text3,
        width=13,
        text="カラー情報",
        command=color_name
    )
    button157=ttk.Button(
        frame_other3,
        width=13,
        text="自動操作",
        command=auto_oparation
    )
    button158=ttk.Button(
        frame_text2,
        width=13,
        text="強制ペースト",
        command=compulsion_paste
    )
    button159=ttk.Button(
        frame_other3,
        width=13,
        text="窓をトレイ",
        command=window_tray
    )
    button160=ttk.Button(
        frame_text3,
        width=13,
        text="文字コード変更",
        command=code_change
    )
    button161=ttk.Button(
        frame_text2,
        width=13,
        text="文字コード判別",
        command=code_check
    )
    button162=ttk.Button(
        frame_other3,
        width=13,
        text="マウス強調",
        command=mouse_highlight
    )
    button163=ttk.Button(
        frame_image3,
        width=13,
        text="画像→Excel",
        command=image_excel
    )
    button164=ttk.Button(
        frame_text2,
        width=13,
        text="MD編集",
        command=markdown_edit
    )
    button165=ttk.Button(
        frame_image3,
        width=13,
        text="HEIC変換",
        command=heic_convert
    )
    button166=ttk.Button(
        frame_other3,
        width=13,
        text="ワンタイムパス",
        command=totp_generator
    )
    button167=ttk.Button(
        frame_video1,
        width=13,
        text="等間隔分割",
        command=video_equally
    )
    button168=ttk.Button(
        frame_video1,
        width=13,
        text="動画音声分割",
        command=video_audio_split
    )
    button169=ttk.Button(
        frame_image3,
        text="数式画像",
        width=13,
        command=formula_image
    )
    button170=ttk.Button(
        frame_video1,
        width=13,
        text="動画回転",
        command=video_rotate
    )
    button171=ttk.Button(
        frame_video1,
        width=13,
        text="音量正規化",
        command=video_normalize
    )
    button172=ttk.Button(
        frame_video1,
        width=13,
        text="動画音声ループ",
        command=video_loop
    )
    button173=ttk.Button(
        frame_video1,
        width=13,
        text="動画色反転",
        command=video_invert
    )
    button174=ttk.Button(
        frame_video1,
        width=13,
        text="音声フェード",
        command=audio_fade
    )
    button175=ttk.Button(
        frame_text3,
        width=13,
        text="JSON↔YAML",
        command=json_yaml
    )
    button176=ttk.Button(
        frame_image3,
        width=13,
        text="重畳表示",
        command=image_superimposed
    )
    button177=ttk.Button(
        frame_image1,
        width=13,
        text="PDF全文検索",
        command=pdf_search
    )
    button178=ttk.Button(
        frame_video3,
        width=13,
        text="QRカメラ取得",
        command=qr_camera
    )
    button179=ttk.Button(
        frame_image3,
        width=13,
        text="文字画像生成",
        command=text_image
    )
    button180=ttk.Button(
        frame_text3,
        width=13,
        text="サムネmd化",
        command=thumbnail_md
    )
    button181=ttk.Button(
        frame_text2,
        width=13,
        text="定型文",
        command=text_template
    )
    button182=ttk.Button(
        frame_text3,
        width=13,
        text="命名規則",
        command=naming_rule
    )
    button183=ttk.Button(
        frame_text3,
        width=13,
        text="csv↔md表変換",
        command=csv_md
    )
    button184=ttk.Button(
        frame_text1,
        width=13,
        text="コピー書換",
        command=copy_replace
    )
    button185=ttk.Button(
        frame_other2,
        width=13,
        text="RSA暗号化",
        command=rsa_encrypt
    )
    button186=ttk.Button(
        frame_text3,
        width=13,
        text="テキストdiff",
        command=text_diff
    )
    button187=ttk.Button(
        frame_image2,
        width=13,
        text="画像丸切取",
        command=image_round
    )
    button188=ttk.Button(
        frame_other2,
        width=13,
        text="ファイル判別",
        command=file_type
    )
    button189=ttk.Button(
        frame_image3,
        width=13,
        text="WEBP圧縮",
        command=webp_compress
    )
    button190=ttk.Button(
        frame_other2,
        width=13,
        text="ファイル削除",
        command=file_delete
    )
    button191=ttk.Button(
        frame_other1,
        width=13,
        text="プロセス優先度",
        command=process_priority
    )
    button192=ttk.Button(
        frame_image3,
        width=13,
        text="コード画像化",
        command=code_image
    )
    button193=ttk.Button(
        frame_image3,
        width=13,
        text="画像比較",
        command=image_compare
    )

    kinou=193

    label_text=Label(frame_text,text="テキスト・情報",bg="green",fg="white")
    label_image=Label(frame_image,text="画像・PDF",bg="#800000",fg="white")
    label_video=Label(frame_video,text="動画・音声",bg="blue",fg="white")
    label_other=Label(frame_other,text="ファイル・小物ツール",bg="purple",fg="white")
    label_text1=Label(frame_text1,text="テキスト編集",fg="green")
    label_text2=Label(frame_text2,text="その他",fg="green")
    label_text3=Label(frame_text3,text="情報変換・生成",fg="green")
    label_video1=Label(frame_video1,text="編集",fg="blue")
    label_video2=Label(frame_video2,text="変換",fg="blue")
    label_video3=Label(frame_video3,text="その他",fg="blue")
    label_image1=Label(frame_image2,text="画像編集",fg="#800000")
    label_image2=Label(frame_image3,text="画像その他",fg="#800000")
    label_image3=Label(frame_image1,text="PDF編集",fg="#800000")
    label_other1=Label(frame_other3,text="小物ツール",fg="purple")
    label_other2=Label(frame_other2,text="フォルダ・ファイル",fg="purple")
    label_other3=Label(frame_other1,text="システム",fg="purple")

    def on_combobox_key_release(event):
        current_text = combobox_var.get()
        # 部分一致検索
        filtered_list = [item for item in list_button if current_text.lower() in item.lower()]
        # リスト更新
        combobox_frame['values'] = filtered_list

    def combo_selected(event):
        name_button=combobox_frame.get()
        dist_button[name_button]()

    combobox_var = StringVar()
    combobox_frame = ttk.Combobox(frame_config, textvariable=combobox_var)

    frame1.pack()

    frame_config.pack(side=TOP)
    #s = ttk.Style(root)
    #s.configure("TNotebook", tabposition='n')
    notebook.pack(side=TOP,expand=True, fill='both')

    buttonX.grid(row=0,column=0)
    #buttonZ.grid(row=0,column=0)
    combobox_frame.grid(row=0,column=1,columnspan=2)

    #frame_text.pack(side=TOP)
    label_text.pack(side=TOP)
    frame_text1.pack(side=TOP)
    frame_text3.pack(side=TOP)
    frame_text2.pack(side=TOP)
    label_text1.grid(row=0,column=0)
    label_text3.grid(row=0,column=0)
    label_text2.grid(row=0,column=0)

    #frame_image.pack(side=TOP)
    label_image.pack(side=TOP)
    frame_image1.pack(side=TOP)
    frame_image2.pack(side=TOP)
    frame_image3.pack(side=TOP)
    label_image1.grid(row=0,column=0)
    label_image2.grid(row=0,column=0)
    label_image3.grid(row=0,column=0)

    #frame_video.pack(side=TOP)
    label_video.pack(side=TOP)
    frame_video1.pack(side=TOP)
    frame_video2.pack(side=TOP)
    frame_video3.pack(side=TOP)
    label_video1.grid(row=0,column=0)
    label_video2.grid(row=0,column=0)
    label_video3.grid(row=0,column=0)

    #frame_other.pack(side=TOP)
    label_other.pack(side=TOP)
    frame_other1.pack(side=TOP)
    frame_other2.pack(side=TOP)
    frame_other3.pack(side=TOP)
    label_other1.grid(row=0,column=0)
    label_other2.grid(row=0,column=0)
    label_other3.grid(row=0,column=0)

    notebook.select(page_x)
    combobox_frame.focus_set()

    def on_focus_in(event):
        if frame_config.winfo_exists():
            combobox_frame.focus_set()
    def handle_keypress(event):
        if event.keysym == 'Return':
            show_dropdown(event)
        elif event.keysym == 'Delete':
            combobox_frame.delete(0,END)
    def show_dropdown(event):
        combobox_frame.event_generate('<Down>')

    root.bind('<FocusIn>', on_focus_in)
    combobox_frame.bind('<KeyPress>', handle_keypress)

    dist_button={}
    list_button=[]
    frame_textN1=1
    frame_textN2=1
    frame_textN3=1
    frame_videoN1=1
    frame_videoN2=1
    frame_videoN3=1
    frame_otherN1=1
    frame_otherN2=1
    frame_otherN3=1
    frame_imageN1=1
    frame_imageN2=1
    frame_imageN3=1
    width_range=main_frame_width
    for d in range(1,kinou+1):
        if f"button{str(d)}" in globals():
            dist_button[globals()["button"+str(d)]["text"]]=globals()["button"+str(d)].invoke
            list_button.append(globals()["button"+str(d)]["text"])
            if check_frame(globals()["button"+str(d)])=="text1":
                globals()["button"+str(d)].grid(row=frame_textN1//width_range,column=frame_textN1%width_range)
                frame_textN1+=1
            elif check_frame(globals()["button"+str(d)])=="text3":
                globals()["button"+str(d)].grid(row=frame_textN3//width_range,column=frame_textN3%width_range)
                frame_textN3+=1
            elif check_frame(globals()["button"+str(d)])=="text2":
                globals()["button"+str(d)].grid(row=frame_textN2//width_range,column=frame_textN2%width_range)
                frame_textN2+=1
            elif check_frame(globals()["button"+str(d)])=="video1":
                globals()["button"+str(d)].grid(row=frame_videoN1//width_range,column=frame_videoN1%width_range)
                frame_videoN1+=1
            elif check_frame(globals()["button"+str(d)])=="video2":
                globals()["button"+str(d)].grid(row=frame_videoN2//width_range,column=frame_videoN2%width_range)
                frame_videoN2+=1
            elif check_frame(globals()["button"+str(d)])=="video3":
                globals()["button"+str(d)].grid(row=frame_videoN3//width_range,column=frame_videoN3%width_range)
                frame_videoN3+=1
            elif check_frame(globals()["button"+str(d)])=="image1":
                globals()["button"+str(d)].grid(row=frame_imageN1//width_range,column=frame_imageN1%width_range)
                frame_imageN1+=1
            elif check_frame(globals()["button"+str(d)])=="image2":
                globals()["button"+str(d)].grid(row=frame_imageN2//width_range,column=frame_imageN2%width_range)
                frame_imageN2+=1
            elif check_frame(globals()["button"+str(d)])=="image3":
                globals()["button"+str(d)].grid(row=frame_imageN3//width_range,column=frame_imageN3%width_range)
                frame_imageN3+=1
            elif check_frame(globals()["button"+str(d)])=="other1":
                globals()["button"+str(d)].grid(row=frame_otherN1//width_range,column=frame_otherN1%width_range)
                frame_otherN1+=1
            elif check_frame(globals()["button"+str(d)])=="other2":
                globals()["button"+str(d)].grid(row=frame_otherN2//width_range,column=frame_otherN2%width_range)
                frame_otherN2+=1
            elif check_frame(globals()["button"+str(d)])=="other3":
                globals()["button"+str(d)].grid(row=frame_otherN3//width_range,column=frame_otherN3%width_range)
                frame_otherN3+=1
        else:
            dist_button[locals()["button"+str(d)]["text"]]=locals()["button"+str(d)].invoke
            list_button.append(locals()["button"+str(d)]["text"])
            if check_frame(locals()["button"+str(d)])=="text1":
                locals()["button"+str(d)].grid(row=frame_textN1//width_range,column=frame_textN1%width_range)
                frame_textN1+=1
            elif check_frame(locals()["button"+str(d)])=="text3":
                locals()["button"+str(d)].grid(row=frame_textN3//width_range,column=frame_textN3%width_range)
                frame_textN3+=1
            elif check_frame(locals()["button"+str(d)])=="text2":
                locals()["button"+str(d)].grid(row=frame_textN2//width_range,column=frame_textN2%width_range)
                frame_textN2+=1
            elif check_frame(locals()["button"+str(d)])=="video1":
                locals()["button"+str(d)].grid(row=frame_videoN1//width_range,column=frame_videoN1%width_range)
                frame_videoN1+=1
            elif check_frame(locals()["button"+str(d)])=="video2":
                locals()["button"+str(d)].grid(row=frame_videoN2//width_range,column=frame_videoN2%width_range)
                frame_videoN2+=1
            elif check_frame(locals()["button"+str(d)])=="video3":
                locals()["button"+str(d)].grid(row=frame_videoN3//width_range,column=frame_videoN3%width_range)
                frame_videoN3+=1
            elif check_frame(locals()["button"+str(d)])=="image1":
                locals()["button"+str(d)].grid(row=frame_imageN1//width_range,column=frame_imageN1%width_range)
                frame_imageN1+=1
            elif check_frame(locals()["button"+str(d)])=="image2":
                locals()["button"+str(d)].grid(row=frame_imageN2//width_range,column=frame_imageN2%width_range)
                frame_imageN2+=1
            elif check_frame(locals()["button"+str(d)])=="image3":
                locals()["button"+str(d)].grid(row=frame_imageN3//width_range,column=frame_imageN3%width_range)
                frame_imageN3+=1
            elif check_frame(locals()["button"+str(d)])=="other1":
                locals()["button"+str(d)].grid(row=frame_otherN1//width_range,column=frame_otherN1%width_range)
                frame_otherN1+=1
            elif check_frame(locals()["button"+str(d)])=="other2":
                locals()["button"+str(d)].grid(row=frame_otherN2//width_range,column=frame_otherN2%width_range)
                frame_otherN2+=1
            elif check_frame(locals()["button"+str(d)])=="other3":
                locals()["button"+str(d)].grid(row=frame_otherN3//width_range,column=frame_otherN3%width_range)
                frame_otherN3+=1

    combobox_frame['values'] = list_button
    button_function_list=list_button.copy()
    combobox_frame.bind('<KeyRelease>', on_combobox_key_release)
    combobox_frame.bind("<<ComboboxSelected>>",combo_selected)

    def on_mouse_wheel(event):
        try:
            # タブインデックスを取得
            current_tab = notebook.index(notebook.select())
            # ホイールの方向に応じてタブを移動
            if event.delta > 0 or event.num == 3:  # ホイール上/左ボタン（Linux）
                notebook.select(current_tab - 1 if current_tab > 0 else 0)
            elif event.delta < 0 or event.num == 4:  # ホイール下/右ボタン（Linux）
                notebook.select(current_tab + 1 if current_tab < notebook.index("end") - 1 else notebook.index("end") - 1)
        except:
            pass

    combobox_frame.unbind("<MouseWheel>")
    root.bind("<MouseWheel>", on_mouse_wheel)



# 共通機能
def hide_parent(event=None):
    root.iconify()  # 親ウィンドウを非表示にする

def show_parent(event=None):
    root.deiconify()  # 親ウィンドウを再表示する



# OpenCV保存
def cv2_write(img,path):
    ext=os.path.splitext(path)[1]
    _,buf = cv2.imencode(ext, img)
    buf.tofile(path)

# OpenCV読込
def cv2_save(path):
    img = cv2.imdecode(
    np.fromfile(path, dtype=np.uint8),cv2.IMREAD_UNCHANGED)
    return img

# Pillow → OpenCV
def pil2cv(image):
    new_image = np.array(image, dtype=np.uint8)
    if new_image.ndim == 2:  # モノクロ
        pass
    elif new_image.shape[2] == 3:  # カラー
        new_image = cv2.cvtColor(new_image, cv2.COLOR_RGB2BGR)
    elif new_image.shape[2] == 4:  # 透過
        new_image = cv2.cvtColor(new_image, cv2.COLOR_RGBA2BGRA)
    return new_image

# OpenCV → Pillow
def cv2pil(image):
    new_image = image.copy()
    if new_image.ndim == 2:  # モノクロ
        pass
    elif new_image.shape[2] == 3:  # カラー
        new_image = cv2.cvtColor(new_image, cv2.COLOR_BGR2RGB)
    elif new_image.shape[2] == 4:  # 透過
        new_image = cv2.cvtColor(new_image, cv2.COLOR_BGRA2RGBA)
    new_image = Image.fromarray(new_image)
    return new_image


def repaired_position():
    root.update_idletasks()
    ws=root.winfo_screenwidth()
    hs=root.winfo_screenheight()
    ws1=root.winfo_width()
    hs1=root.winfo_height()
    x=(ws/2)-(ws1/2)
    y=(hs/2)-(hs1/2)
    root.geometry('+%d+%d'%(x,y))

def main_frame_delete():
    root.unbind("<MouseWheel>")
    frame1.destroy()

def advanced_setting():
    def adv_setting_update():
        adv_setting["width_num"]=int(entry1.get())
        adv_setting["front_screen"]=var2.get()
        adv_setting["multi_window"]=var3.get()
        adv_setting["auto_update_chack"]=var4.get()
        adv_setting["show_center"]=var5.get()
        adv_setting["intermediate_screen"]=var6.get()
        adv_setting["hide_mainwindow"]=var7.get()

        with open(os.getcwd()+"/config/adv_setting.json", "w") as file:
            json.dump(adv_setting,file)
        messagebox.showinfo(title="保存", message="保存しました。\n再起動されます")
        restart()

    root1=Free_window()
    with open(os.getcwd()+"/config/adv_setting.json", "r") as file:
        adv_setting =json.load(file)
    if "width_num" not in adv_setting:
        adv_setting["width_num"]=4
    if "front_screen" not in adv_setting:
        adv_setting["front_screen"]=1
    if "multi_window" not in adv_setting:
        adv_setting["multi_window"]=0
    if "auto_update_chack" not in adv_setting:
        adv_setting["auto_update_chack"]=0
    if "show_center" not in adv_setting:
        adv_setting["show_center"]=0
    if "intermediate_screen" not in adv_setting:
        adv_setting["intermediate_screen"]=1
    if "hide_mainwindow" not in adv_setting:
        adv_setting["hide_mainwindow"]=0


    root1.title("高度な設定")
    root1.attributes("-topmost", True)

    label1=Label(root1,text="横幅数：")
    entry1=Spinbox(root1,width=10,from_=1,to=10,increment=1)
    entry1.delete(0,"end")
    entry1.insert(END,adv_setting["width_num"])
    label1.grid(row=0,column=0)
    entry1.grid(row=0,column=1)

    var2=IntVar()
    var2.set(adv_setting["front_screen"])
    label2=Label(root1,text="常時最前面に固定：")
    check2=Checkbutton(root1,variable=var2,onvalue=1,offvalue=0)
    label2.grid(row=1,column=0)
    check2.grid(row=1,column=1)

    var3=IntVar()
    var3.set(adv_setting["multi_window"])
    label3=Label(root1,text="複数起動を許可するか：")
    check3=Checkbutton(root1,variable=var3,onvalue=1,offvalue=0)
    label3.grid(row=2,column=0)
    check3.grid(row=2,column=1)

    var4=IntVar()
    var4.set(adv_setting["auto_update_chack"])
    label4=Label(root1,text="起動時に自動的に更新チェックをするか：")
    check4=Checkbutton(root1,variable=var4,onvalue=1,offvalue=0)
    label4.grid(row=3,column=0)
    check4.grid(row=3,column=1)

    var5=IntVar()
    var5.set(adv_setting["show_center"])
    label5=Label(root1,text="表示時に中央に表示するか：")
    check5=Checkbutton(root1,variable=var5,onvalue=1,offvalue=0)
    label5.grid(row=4,column=0)
    check5.grid(row=4,column=1)

    var6=IntVar()
    var6.set(adv_setting["intermediate_screen"])
    label6=Label(root1,text="一部機能で中間画面を表示するか：")
    check6=Checkbutton(root1,variable=var6,onvalue=0,offvalue=1)
    label6.grid(row=5,column=0)
    check6.grid(row=5,column=1)

    var7=IntVar()
    var7.set(adv_setting["hide_mainwindow"])
    label7=Label(root1,text="サブ画面アクティブ時にメイン画面を隠す：")
    check7=Checkbutton(root1,variable=var7,onvalue=1,offvalue=0)
    label7.grid(row=6,column=0)
    check7.grid(row=6,column=1)


    button=ttk.Button(root1,text="保存",command=adv_setting_update)
    button.grid(row=7,column=0,columnspan=2)



def rgbToHex(rgb):
    return "#%02x%02x%02x" % rgb

def home_page():
    root1=Free_window()
    root1.title("ホームページ一覧")
    root1.attributes("-topmost", True)
    ws=root1.winfo_screenwidth()
    hs=root1.winfo_screenheight()
    x=(ws/2)-250
    y=(hs/2)-300
    root1.geometry('+%d+%d'%(x,y))
    button1=ttk.Button(root1,text="配布ページ(最新VerDL・更新情報)",width=40,command=lambda:webbrowser.open("https://yukisakura.notion.site/Yuki-s-Army-knife-8b11161af892436a9dd9cfa64724dec9"))
    button2=ttk.Button(root1,text="Github(ソースコード・歴代VerDL)",width=40,command=lambda:webbrowser.open("https://github.com/yukisakura001/Yukis_Army_knife"))
    button3=ttk.Button(root1,text="Twitter(更新情報)",width=40,command=lambda:webbrowser.open("https://twitter.com/yukisakura001"))

    button1.grid(row=0,column=1)
    button2.grid(row=1,column=1)
    button3.grid(row=2,column=1)



def restart():
    subprocess.Popen(["Yukis_Army_knife.exe"])
    stop_tkinter()

def button_task_icon():

    def click_close():

        messagebox.showinfo(title="終了", message="再起動したあとに\nタスクトレイに表示されます。")
        root1.destroy()
        restart()

    def list_delete():
        nonlocal task_setting
        try:
            select_num = listbox.curselection()
            func_delete_name=listbox.get(select_num[0])
            func_delete_num=button_function_list.index(func_delete_name)+1
        except:
            return
        res=messagebox.askquestion(title="削除確認", message="この機能をタスクトレイから\n削除しますか？")
        if res != "yes":
            return
        task_setting=task_setting.replace(str(func_delete_num)+"\n","")
        with open(os.getcwd()+"/config/task_button.txt", "w") as file:
            file.write(task_setting)
        var=select_num_get()
        listbox.config(listvariable=var)
        listbox.update()

    def swap_lines(text, line1, line2):
        lines = text.split('\n')

        try:
            lines[line1], lines[line2] = lines[line2], lines[line1]
        except IndexError:
            print("指定した行が存在しません。")

        new_text = '\n'.join(lines)

        return new_text


    def select_num_get():
        select_num=[]
        for i in task_setting.split("\n")[:-1]:
            select_num.append(button_function_list[int(i)-1])
        return StringVar(value=select_num)

    def set_task_function(func_name):
        nonlocal task_setting
        try:
            func_num=button_function_list.index(func_name)+1
            task_setting=task_setting+str(func_num)+"\n"
            with open(os.getcwd()+"/config/task_button.txt", "w") as file:
                file.write(task_setting)
            var=select_num_get()
            listbox.config(listvariable=var)
            listbox.update()
        except:
            return

    def showMenu(event):
        try:
            x, y = event.x, event.y
            index = listbox.nearest(y)
            if index != -1:
                listbox.selection_clear(0, END)  # 一旦すべての選択を解除
                listbox.selection_set(index)  # 指定されたアイテムを選択
            pmenu.post(event.x_root, event.y_root)

        except:
            return

    def move_up():
        nonlocal task_setting
        # 指定した行とその上の行を入れ替える
        try:
            select_num = listbox.curselection()
        except:
            return
        if select_num[0]==0:
            return
        task_setting=swap_lines(task_setting,select_num[0],select_num[0]-1)
        with open(os.getcwd()+"/config/task_button.txt", "w") as file:
            file.write(task_setting)
        var=select_num_get()
        listbox.config(listvariable=var)
        listbox.update()

    def move_down():
        nonlocal task_setting
        # 指定した行とその下の行を入れ替える
        try:
            select_num = listbox.curselection()
        except:
            return
        if select_num[0]==len(task_setting.split("\n"))-2:
            return
        task_setting=swap_lines(task_setting,select_num[0],select_num[0]+1)
        with open(os.getcwd()+"/config/task_button.txt", "w") as file:
            file.write(task_setting)
        var=select_num_get()
        listbox.config(listvariable=var)
        listbox.update()

    if not os.path.exists(os.getcwd()+"/config/task_button.txt"):
        with open(os.getcwd()+"/config/task_button.txt", "w") as file:
            file.write("")
    with open(os.getcwd()+"/config/task_button.txt", "r") as file:
        task_setting = file.read()
    root1=Free_window()
    root1.title("トレイを設定")
    root1.attributes("-topmost", True)

    var=select_num_get()

    label1=Label(root1,text="タスクトレイに表示する機能を登録してください。")
    searchbox=Searchbox(root1,width=30,values=button_function_list)
    button1=ttk.Button(root1,text="登録",command=lambda:set_task_function(searchbox.get()))
    listbox=Listbox(root1,width=30,height=10,listvariable=var)

    listbox.bind("<Button-3>", showMenu)
    root1.protocol("WM_DELETE_WINDOW", click_close)
    pmenu = Menu(root1, tearoff=0)
    pmenu.add_command(label="削除", command=list_delete)
    pmenu.add_command(label="上に移動",command=move_up)
    pmenu.add_command(label="下に移動",command=move_down)

    label1.pack()
    searchbox.pack()
    button1.pack()
    listbox.pack(fill="both",expand=True)


def listener_window():
    global main_show_shrotcut
    ctrl=0
    shift=0
    def on_press(key):
        nonlocal ctrl,shift
        if main_show_shrotcut==1:
            return
        if key == keyboard.Key.space and ctrl==1 and shift==1:
            if root.state() == "normal":
                if show_center==1:
                    repaired_position()
                frame_delete()
            else:
                if show_center==1:
                    repaired_position()
                root.deiconify()
                combobox_frame.focus_set()

        if key == keyboard.Key.ctrl_l:
            ctrl=1
        if key == keyboard.Key.shift:
            shift=1

    def on_release(key):
        nonlocal ctrl,shift
        if key == keyboard.Key.ctrl_l:
            ctrl=0
        if key == keyboard.Key.shift:
            shift=0

    with keyboard.Listener(on_press=on_press,on_release=on_release) as listener_def:
        listener_def.join()

def check_new_ver():
    url = 'https://www.dropbox.com/scl/fi/5y15joc9artguntvecp7e/version.txt?rlkey=pps35nwsvdzsytn3l2hvrxyit&dl=1'
    response = requests.get(url)
    if float(version) < float(response.text.rstrip()):
        res=messagebox.askquestion(title="バージョン確認", message=f"新規バージョン{response.text.rstrip()}があります。ダウンロードしますか？")
        if res == "yes":
            #webbrowser.open("https://www.dropbox.com/scl/fo/frsya7l4zh14z30fpuk0n/h?rlkey=2jncv45l8jhoe7ezls9iqn51v&dl=0")
            # ユーザー名を取得
            user_name = getpass.getuser()
            # フォルダパスを組み立てる
            folder_path = fr'C:\Users\{user_name}\AppData\Local\Programs\Yukis Army knife'
            current_directory = os.getcwd()
            if current_directory == folder_path:
                def update_soft():
                    urllib.request.urlretrieve(url, save_path)
                    subprocess.Popen(save_path)
                    stop_tkinter()

                # ダウンロードするファイルのURLを指定
                url = "https://www.dropbox.com/scl/fi/del2oxu9w5aq4d4b2aul7/Yukis_Army_knife_Installer.exe?rlkey=zs2w4wjb9kai0m6axku5v46ui&dl=1"
                # ファイルを保存するフォルダとファイル名を指定
                make_folder("temp1")
                save_path = os.getcwd()+"/temp1/Yukis_Army_knife.exe"
                # URLからファイルをダウンロードして指定したフォルダに保存
                messagebox.showinfo("情報","ダウンロード中です。")
                t=threading.Thread(target=update_soft,daemon=True)
                t.start()

            else:
                def update_soft():
                    urllib.request.urlretrieve(url, save_path)
                    messagebox.showinfo("情報","ダウンロードが完了しました。")

                url="https://www.dropbox.com/scl/fi/zxji4x5epgddma25ew48q/Yukis_Army_knife_portable.zip?rlkey=xoqsbfn12hvjbuqv5iq3k8a4w&dl=1"
                save_path = filedialog.asksaveasfilename(initialdir = os.getcwd(),title = "保存場所を選択",filetypes = [("zipファイル","*.zip")],initialfile="Yukis_Army_knife.zip")
                t=threading.Thread(target=update_soft,daemon=True)
                t.start()
                messagebox.showinfo("情報","ダウンロード中です。")


    else:
        messagebox.showinfo("情報","最新バージョンです。")

def startup_reg():
    def create_shortcut(target_path, shortcut_path):
        shell = win32com.client.Dispatch("WScript.Shell")
        shortcut = shell.CreateShortCut(shortcut_path)
        shortcut.Targetpath = target_path
        shortcut.WorkingDirectory = os.path.dirname(target_path)
        shortcut.save()

    target_file = os.getcwd()+"/Yukis_Army_knife.exe"
    shortcut_file = os.path.expandvars("%APPDATA%\\Microsoft\\Windows\\Start Menu\\Programs\\Startup")+"/Yukis_Army_knife.lnk"

    create_shortcut(target_file, shortcut_file)
    messagebox.showinfo("情報","スタートアップに登録しました。")

def esc_key_pressed(event):
    if event.keysym == "Escape":
        buttonY.invoke()

def check_frame(button):
    parent = button.winfo_parent()
    return parent.split(".")[-1]

def set_theme(theme):
    if theme==0:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("0")
            s.theme_use("vista")
    elif theme==1:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("modern")
            s.theme_use("adapta")
    elif theme==2:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("dark")
            s.theme_use("black")
    elif theme==3:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("digital")
        s.theme_use("aquativo")
    elif theme==5:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("natural")
        s.theme_use("clearlooks")
    elif theme==6:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("systemati")
        s.theme_use("elegance")
    elif theme==7:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("keramik")
        s.theme_use("keramik")
    elif theme==8:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("wood-like")
        s.theme_use("kroc")
    elif theme==9:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("plastik")
        s.theme_use("plastik")
    elif theme==10:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("radiance")
        s.theme_use("radiance")
    elif theme==11:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("XPblue")
        s.theme_use("winxpblue")
    elif theme==12:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("sand")
        s.theme_use("scidsand")
    elif theme==13:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("classic")
        s.theme_use("clam")
    elif theme==14:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("modern_dark")
        sv_ttk.set_theme("dark")
    elif theme==15:
        with open(os.getcwd()+"/config/style.txt", "w") as file:
            file.write("modern_light")
        sv_ttk.set_theme("light")

    s.configure("TNotebook", tabposition='n')
    if frame1.winfo_exists():
        main_frame_delete()
    else:
        frame.destroy()
    set_frame1(0)


def file_mult(x):
    #input_str = x
    #input_str=input_str+" "
    #result = re.findall(r"(.+?)\.([a-zA-Z0-9]+)\s", input_str)
    #output_list = [f"{item[0]}.{item[1]}" for item in result]
    output_list = re.split(r"\s(?=[A-Za-z]+:)", x)
    return output_list

def main_frame(page_n):
    global frame,buttonY
    frame.destroy()
    set_frame1(page_n)

def frame_delete():
    if buttonY["state"] =="normal":
        buttonY.invoke()
    combobox_frame.delete(0,END)
    root.withdraw()

def stop_tkinter():
    global buttonY
    root.withdraw()
    if os.path.exists(os.getcwd()+"/temp1"):
        shutil.rmtree(os.getcwd()+"/temp1", ignore_errors=True)
    if buttonY["state"] =="normal":
        buttonY.invoke()
    root.destroy()

def execute():
    global window_front,main_show_shrotcut

    #def on_focus_out1(event):
    #    if event.widget == root:
    #        frame_delete()
    #    else:
    #    # ルートウィンドウ以外のウィジェットによるフォーカスアウトは無視
    #        pass
#
    #if window_front.get() == 1:
    #    #root.attributes("-topmost", False)
    #    root.bind("<FocusOut>", on_focus_out1)
    #elif window_front.get() ==0:
    #    #root.attributes("-topmost", True)
    #    root.unbind("<FocusOut>")
    if window_front.get() == 1:
        main_show_shrotcut=1
    elif window_front.get() ==0:
        main_show_shrotcut=0

def make_folder(x):
    if not os.path.exists(x):
        os.makedirs(x)

def delete_folder(x):
    if os.path.exists(x):
        shutil.rmtree(x)

def taskarea():
    global icon,task_menu

    toast = winotify.Notification(
            title="Yukis Army knifeを起動しました",
            msg="タスクトレイにアイコンが表示されています",
            app_id="Yukis_Army_knife",
        )
    toast.show()

    def action_func(icon,item):
        root.deiconify()
        if buttonY["state"] =="normal":
            buttonY.invoke()
        if show_center==1:
            repaired_position()
        dist_button[str(item)]()

    if not os.path.exists(os.getcwd()+"/config/task_button.txt"):
        with open(os.getcwd()+"/config/task_button.txt", "w") as file:
            file.write("")

    img=my_icon.get_img()
    image = Image.open(img)
    func_list=[]
    with open(os.getcwd()+"/config/task_button.txt", "r") as file:
        task_setting = file.read()
    for num,i in enumerate(task_setting.split("\n")[:-1]):
        #print(button_function_list[int(i) - 1])
        item_name = button_function_list[int(i) - 1]
        func_list.append(pystray.MenuItem(item_name, action=action_func))
    func_list.append(pystray.MenuItem('ーーーーーーーーーーー',action=None,enabled=False))
    func_list.append(pystray.MenuItem(text="ウィンドウ表示", action=left_click_action, default=True))
    func_list.append(pystray.MenuItem(text="再起動", action=restart))
    func_list.append(pystray.MenuItem(text="ソフトを終了", action=stop_tkinter))


    task_menu=pystray.Menu(*func_list)

    icon = pystray.Icon(
        name="Yukis_Army_knife",
        icon=image,
        title="Yukis_Army_knife",
        menu=task_menu
    )
    icon.run()

def left_click_action(icon,item):
    global root
    if root.state() == "normal":
        frame_delete()
    else:
        root.deiconify()
        if show_center==1:
            root.geometry('+%d+%d'%(x,y))
        combobox_frame.focus_set()


def setting_folder():
    subprocess.Popen(["explorer", r"config"], shell=True)


# 機能の関数

def count():
    global frame,buttonY
    def count1():
        try:
            x=clip.paste().replace("\r","") # 全文字
            y=x.replace('\n', '').replace("\r","").replace("\r\n","") # 改行なし
            z=y.replace(" ","").replace("　","").strip() # 空白なし
            s=[line for line in clip.paste().splitlines() if line.strip()]
            byte=len(x.encode("utf-8"))
            return messagebox.showinfo("文字数カウント",f"全文字：{len(x)}\n改行なし：{len(y)}\n改行・空白なし：{len(z)}\n行数（空行なし）：{len(s)}\nバイト数：{byte}")
        except:
            return messagebox.showerror("エラー","うまくクリップボードを取得できませんでした。")

    if intermediate_screen==0:
        main_frame_delete()
        frame = ttk.Frame(root, padding=16)
        button=ttk.Button(frame,text='文字カウント',command=count1,padding=[80,10,80,10])
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        button.grid(row=1,column=0,columnspan=2)
    elif intermediate_screen==1:
        count1()


def blank():
    global frame,buttonY
    def blank1():
        try:
            text_data=clip.paste()
            x=text_data.replace(" ","").replace("　","").strip()
            clip.copy(x)
            messagebox.showinfo("終了","終了しました")
        except:
            messagebox.showerror("エラー","うまくクリップボードを取得できませんでした。")

    if intermediate_screen==0:
        main_frame_delete()
        frame = ttk.Frame(root, padding=16)
        button=ttk.Button(frame,text='空白削除',command=blank1,padding=[80,10,80,10])
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")

        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        button.grid(row=1,column=0,columnspan=2)
    elif intermediate_screen==1:
        blank1()


def tikan():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=16)
    frame.pack()

    def clip_replace():
        # 入力された正規表現パターンと置換文字列を取得
        pattern = pattern_entry.get()
        replacement = replacement_entry.get()
        text =clip.paste()

        try:
            # 正規表現を使用してテキストを置換
            replaced_text = re.sub(pattern, replacement, text)
            text_entry.delete("1.0", END)
            text_entry.insert("1.0", replaced_text)
        except re.error as e:
            messagebox.showerror("正規表現エラー", e)

    def replace_text():
        pattern = pattern_entry.get()
        replacement = replacement_entry.get()
        text = text_entry.get("1.0",END)

        try:
            replaced_text = re.sub(pattern, replacement, text)
            text_entry.delete("1.0", END)
            text_entry.insert("1.0", replaced_text)
        except re.error as e:
            messagebox.showerror("正規表現エラー", e)

    pattern_label = ttk.Label(frame, text="置換される文字列:")

    pattern_entry = ttk.Entry(frame, width=50)

    replacement_label = ttk.Label(frame, text="置換する文字列:")

    replacement_entry = ttk.Entry(frame, width=50)

    text_label = ttk.Label(frame, text="置換対象テキスト:")

    text_entry =ScrolledText(frame, height=10, width=50,undo=True)

    frameY=ttk.Frame(frame)
    replace_button = ttk.Button(frameY, text="テキストを置換", command=replace_text)
    replace_clip_button = ttk.Button(frameY, text="コピーの置換", command=clip_replace)

    frameX=ttk.Frame(frame)
    buttonX=ttk.Checkbutton(frameX,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frameX,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    button_X=ttk.Button(frame,text="コピー",command=lambda:clip.copy(text_entry.get("1.0",END)))

    frameX.pack()

    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)

    pattern_label.pack()
    pattern_entry.pack()
    replacement_label.pack()
    replacement_entry.pack()
    frameY.pack()
    replace_button.grid(row=0,column=1)
    replace_clip_button.grid(row=0,column=0)
    text_label.pack()
    text_entry.pack()
    button_X.pack()

def serch():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)
    def serch1(x):
        if var.get()==1:
            query=clip.paste()
        elif var.get()==0:
            query=entry.get()
        if x==0:
            req= requests.get('https://www.google.co.jp/search', params={'q': query})
            webbrowser.open(urllib.parse.unquote(req.url))
        elif x==1:
            req= requests.get('https://www.google.co.jp/search', params={'q': query,'tbm':'isch'})
            webbrowser.open(urllib.parse.unquote(req.url))
        elif x==2:
            req= requests.get('https://ja.wikipedia.org/wiki/', params={'search': query})
            webbrowser.open(urllib.parse.unquote(req.url))
        elif x==3:
            query_encoded = urllib.parse.quote(query)
            url = f'https://scholar.google.com/scholar?q={query_encoded}'
            req = requests.get(url)
            webbrowser.open(urllib.parse.unquote(req.url))
        elif x==4:
            url = f'https://you.com/search?q={query}&fromSearchBar=true&tbm=youchat&'
            webbrowser.open(url)
        elif x==5:
            url=f"https://www.youtube.com/results?search_query={query}"
            webbrowser.open(url)

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    button1=ttk.Button(frame,text='Google',command=lambda:serch1(0),width=10)
    button2=ttk.Button(frame,text='Google画像',command=lambda:serch1(1),width=10)
    button3=ttk.Button(frame,text="wikipedia",command=lambda:serch1(2),width=10)
    button4=ttk.Button(frame,text="論文検索",command=lambda:serch1(3),width=10)
    button5=ttk.Button(frame,text="AI検索",command=lambda:serch1(4),width=10)
    button6=ttk.Button(frame,text="YouTube",command=lambda:serch1(5),width=10)
    var=IntVar()
    var.set(1)
    radio1=ttk.Radiobutton(frame,text="テキストボックスから検索",variable=var,value=0)
    radio2=ttk.Radiobutton(frame,text="クリップボードから検索",variable=var,value=1)
    entry=ttk.Entry(frame,width=50)

    frame.pack()
    buttonX.grid(row=0,column=1,columnspan=2)
    buttonY.grid(row=0,column=0)
    button1.grid(row=1,column=0)
    button2.grid(row=1,column=1)
    button5.grid(row=1,column=2)
    button3.grid(row=2,column=0)
    button4.grid(row=2,column=1)
    button6.grid(row=2,column=2)
    radio2.grid(row=3,column=0)
    radio1.grid(row=3,column=2)
    entry.grid(row=4,column=0,columnspan=3)

def IP():
    global frame,buttonY
    ip = socket.gethostbyname(socket.gethostname())
    url = 'http://checkip.dyndns.com/'
    r = requests.get(url)
    G_ip= re.findall(r'\d+', r.text)

    main_frame_delete()
    frame = ttk.Frame(root, padding=16)

    label1 = ttk.Label(frame, text='ローカルIP：')
    label2 = ttk.Label(frame, text='グローバルIP：')
    label3 = ttk.Label(frame, text=ip)
    label4 = ttk.Label(frame, text='.'.join(G_ip))
    button1 = ttk.Button(
        frame,
        text="コピー",
        command=lambda: clip.copy(ip)
    )
    button2 = ttk.Button(
        frame,
        text="コピー",
        command=lambda: clip.copy('.'.join(G_ip))
    )
    buttonX=ttk.Checkbutton(
    frame,
    text=main_checkbox_name,
    onvalue=1,
    offvalue=0,
    variable=window_front,
    command=execute
    )
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    label1.grid(row=1,column=0)
    label3.grid(row=1,column=1)
    button1.grid(row=1,column=2)
    label2.grid(row=2,column=0)
    label4.grid(row=2,column=1)
    button2.grid(row=2,column=2)

def timer():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)
    output_sec=IntVar()
    output_min=IntVar()
    main=StringVar()
    main.set("スタート")
    y=0 # 動作管理
    last_click_time = 0.0

    def start():
        nonlocal y
        nonlocal output_sec
        nonlocal output_min
        nonlocal main
        nonlocal last_click_time
        buttonY.configure(state='disabled')

        if time.time() - last_click_time < 0.5:
            return
        last_click_time = time.time()

        if y==0: # 最初
            main.set("一時停止")
            output_sec.set(entry_sec.get())
            output_min.set(entry_min.get())
            y=1
            root.after(1000,active)
        elif y==1: # 一時停止
            main.set("再開")
            y=2
            button_reset.configure(state='normal')
        elif y==2: # 再開
            main.set("一時停止")
            y=1
            root.after(1000,active)
            button_reset.configure(state='disabled')
        button.configure(text=main.get())

    def active():
        nonlocal y
        nonlocal output_sec
        nonlocal output_min
        nonlocal main
        if y!=1:
            return
        if output_sec.get() > 0:
            output_sec.set(output_sec.get()-1)
            root.after(1000, active)
        elif output_sec.get()==0:
            if output_min.get() > 0:
                output_min.set(output_min.get()-1)
                output_sec.set(59)
                root.after(1000, active)
            elif output_min.get()==0:
                messagebox.showinfo("終了","終了しました。")
                reset()
        button.configure(text=main.get())

    def reset():
        nonlocal y
        nonlocal output_sec
        nonlocal output_min
        nonlocal main
        y=0
        output_min.set(0)
        output_sec.set(0)
        main.set("スタート")
        buttonY.configure(state='normal')
        button.configure(text=main.get())
        button_reset.configure(state='disabled')

    label1 = ttk.Label(frame, text='：',font=("Helvetica", 16),width=2)
    label2 = ttk.Label(frame, text='：',font=("Helvetica", 16),width=2)
    label_sec=ttk.Label(frame, textvariable=output_sec,font=("Helvetica", 16))
    label_min=ttk.Label(frame, textvariable=output_min,font=("Helvetica", 16))
    entry_sec=ttk.Entry(frame,width=5,font=("Helvetica", 12))
    entry_sec.insert(0, "30")
    entry_min=ttk.Entry(frame,width=5,font=("Helvetica", 12))
    entry_min.insert(0, "0")
    button=ttk.Button(frame, text=main.get(), command=start)
    button_reset=ttk.Button(frame, text="リセット", command=reset, state='disabled')
    buttonX=ttk.Checkbutton(
        frame,
        text=main_checkbox_name,
        onvalue=1,
        offvalue=0,
        variable=window_front,
        command=execute
        )
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white",state="normal")

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=2)
    label_min.grid(row=2,column=0)
    label1.grid(row=2,column=1,sticky='E')
    label_sec.grid(row=2,column=2,sticky='W')
    entry_min.grid(row=1,column=0,sticky='E')
    label2.grid(row=1,column=1,sticky='E')
    entry_sec.grid(row=1,column=2,sticky='W')
    button.grid(row=3,column=0)
    button_reset.grid(row=3,column=2)

def mini_win():
    global frame,buttonY
    if buttonY["state"] =="normal":
        buttonY.invoke()
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)
    buttonY=Button(frame,text="メインメニューに戻る",command=lambda:main_frame(0),font=("Helvetica", 14),bg="purple",fg="white")
    frame.pack()
    buttonY.pack(fill=BOTH,expand=True)

def folder():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)
    def gr():
        try:
            x=filedialog.askdirectory()
            s = box.get( 0., END )
            text=s.split(sep="\n")
            for i in text:
                os.makedirs(x+"/"+i,exist_ok=True)
            messagebox.showinfo("終了","終了しました")
        except:
            messagebox.showerror("エラー","うまくフォルダを作成できませんでした。")


    box=ScrolledText(
        frame,
        height=20,
        width=30
    )
    button=ttk.Button(
        frame,
        command=gr,
        text="フォルダ作成"
    )
    buttonX=ttk.Checkbutton(
    frame,
    text=main_checkbox_name,
    onvalue=1,
    offvalue=0,
    variable=window_front,
    command=execute
    )
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    label2=ttk.Label(frame,text="作成フォルダ名を入力してください")

    frame.pack()
    buttonX.pack(side=TOP)
    buttonY.pack(side=TOP)
    label2.pack(side=TOP)
    box.pack(side=TOP)
    button.pack(side=TOP)

def local_pass():
    global frame,buttonY
    main_frame_delete()
    frame=ttk.Frame(root, padding=4)
    def dnd(drop):
        clip.copy(drop.data.replace("{","").replace("}","").replace("\\","/"))
        messagebox.showinfo("終了","終了しました")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',dnd)
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(
        frame,
        text=main_checkbox_name,
        onvalue=1,
        offvalue=0,
        variable=window_front,
        command=execute
        )
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    frame.pack()
    buttonX.pack(side=TOP)
    buttonY.pack(side=TOP)
    label.pack(side=TOP)

def count_down():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def reset_count():
        try:
            count_up.set(int(text.get()))
        except:
            messagebox.showerror("エラー","数字を入力してください")

    count_up=IntVar()
    count_up.set(0)

    label1=ttk.Label(
    frame,
    textvariable=count_up,
    font=("Helvetica", 16)
    )
    button1=ttk.Button(
    frame,
    text="+1",
    command=lambda : count_up.set(count_up.get()+1),
    width=5
    )
    button2=ttk.Button(
        frame,
        text="-1",
        command=lambda : count_up.set(count_up.get()-1),
        width=5
    )
    button3=ttk.Button(
        frame,
        text="リセット",
        command=lambda : count_up.set(0),
        width=8
    )
    text=ttk.Entry(
        frame,
        width=10
    )
    button4=ttk.Button(
        frame,
        text="設定",
        command=reset_count,
        width=8
    )
    buttonX=ttk.Checkbutton(
    frame,
    text=main_checkbox_name,
    onvalue=1,
    offvalue=0,
    variable=window_front,
    command=execute
    )
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=1)
    button1.grid(row=2,column=0)
    button2.grid(row=2,column=1)
    button3.grid(row=2,column=2)
    text.grid(row=3,column=0,columnspan=2)
    button4.grid(row=3,column=2)

def choose():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def choose_rurle():
        try:
            x=box.get( 0., END ).split(sep="\n")
            y=[i for i in x if i != '']
            messagebox.showinfo("抽選結果",random.choice(y))
        except:
            messagebox.showerror("エラー","候補を入力してください")

    def dnd1(drop):
        try:
            x=drop.data.replace("{","").replace("}","").replace("\\","/")
            try:
                with open(x, "r",encoding="utf-8") as file:
                    y=file.read()
            except:
                with open(x, "r",encoding="shift-jis") as file:
                    y=file.read()
            box.delete(0.0,END)
            box.insert(END,y)
        except:
            messagebox.showerror("エラー","うまくファイルを読み込めませんでした")

    def save_file():
        try:
            x=box.get( 0., END )
            y=filedialog.asksaveasfilename(initialdir = os.getcwd(),title = "保存場所を選択",filetypes = [("TXT","*.txt")],initialfile="抽選.txt",defaultextension="txt")
            with open(y, "w",encoding="utf-8") as file:
                file.write(x)
            messagebox.showinfo("終了","保存しました")
        except:
            messagebox.showerror("エラー","うまく保存できませんでした")

    frame_t=ttk.Frame(frame)
    box=ScrolledText(
        frame,
        height=20,
        width=30
    )
    button=ttk.Button(
        frame_t,
        command=choose_rurle,
        text="抽選"
    )
    button1=ttk.Button(
        frame_t,
        command=save_file,
        text="保存"
    )
    buttonX=ttk.Checkbutton(
        frame,
        text=main_checkbox_name,
        onvalue=1,
        offvalue=0,
        variable=window_front,
        command=execute
        )
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")


    frame.pack()
    buttonX.pack(side=TOP)
    buttonY.pack(side=TOP)
    box.pack(side=TOP)
    frame_t.pack(side=TOP)
    button.grid(row=0,column=0)
    button1.grid(row=0,column=1)
    box.drop_target_register(DND_FILES)
    box.dnd_bind('<<Drop>>',dnd1)

def video_dwonload():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def folder():
        x=filedialog.askdirectory()
        return x

    def url():
        x=box.get( 0., END ).split(sep="\n")
        y=[i for i in x if i != '']
        return y

    def fast_video():
        try:
            button_v1.configure(state='disabled')
            button_v3.configure(state='disabled')
            buttonY.configure(state='disabled')
            root.update()
            foldera=folder()
            URL=url()
            ydl_video_opts = {
                'outtmpl':foldera+"/%(title)s"+'_.mp4',
                'format':'best',
                'quiet': True,
                'no-stdout': True,
                'noprogress': True,
                'no_warnings': True
            }
            with YoutubeDL(ydl_video_opts) as ydl:
                ydl.download(URL)
            messagebox.showinfo("終了","ダウンロード終了しました")
            button_v1.configure(state='normal')
            button_v3.configure(state='normal')
            buttonY.configure(state='normal')
            root.update()
        except:
            messagebox.showerror("エラー","うまくダウンロードできませんでした")

    def fast_voice():
        try:
            button_v1.configure(state='disabled')
            button_v3.configure(state='disabled')
            buttonY.configure(state='disabled')
            root.update()
            foldera=folder()
            URL=url()
            folder_path = os.getcwd()+"/temp1/video"
            make_folder(folder_path)
            ydl_video_opts = {
            'outtmpl': folder_path+"/"+"%(title)s.%(ext)s",
            'format': 'bestaudio',
            'quiet': True,
            'no-stdout': True,
            'noprogress': True,
            'extractaudio': True,
            'audioformat': 'webm',
            'no_warnings': True
            }
            with YoutubeDL(ydl_video_opts) as ydl:
                ydl.download(URL)

            urla = []
            for file_name in os.listdir(folder_path):
                if file_name.endswith(".webm"):
                    urla.append(os.path.join(folder_path, file_name))

            for x in urla:
                x2=x.replace(folder_path,"")
                x1= re.sub(r'[\\/*?:"<>|]', '', x2)

                audio_clip = AudioFileClip(x)
                audio_clip.write_audiofile(foldera+"/"+x1.replace("webm","mp3"),logger=None)
                os.remove(x)

            messagebox.showinfo("終了","ダウンロード終了しました")
            button_v1.configure(state='normal')
            button_v3.configure(state='normal')
            buttonY.configure(state='normal')
            root.update()
        except:
            messagebox.showerror("エラー","うまくダウンロードできませんでした")


    root.update()
    box=ScrolledText(frame,
             height=20,
             width=50
    )
    button_v1=ttk.Button(
        frame,
        command=fast_video,
        text="動画ダウンロード"
        )
    button_v3=ttk.Button(
        frame,
        text="音声ダウンロード",
        command=fast_voice
    )
    label_box=ttk.Label(frame,text="URLを改行区切りで入力してください")
    buttonX=ttk.Checkbutton(
        frame,
        text=main_checkbox_name,
        onvalue=1,
        offvalue=0,
        variable=window_front,
        command=execute
    )

    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label_box.grid(row=3,column=0,columnspan=2)
    box.grid(row=4,column=0,columnspan=2)
    button_v1.grid(row=5,column=0)
    button_v3.grid(row=5,column=1)

def image_resize():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def resize(drop):
        try:
            file=drop.data.replace("{","").replace("}","").replace("\\","/")
            x_list=file_mult(file)
            for x in x_list:
                img=Image.open(x)
                img_resize=img.resize((int(entry_w.get()),int(entry_h.get())),Image.LANCZOS)
                y= os.path.basename(x)
                z=x.replace(y,"resize_"+y)
                img_resize.save(z)
            messagebox.showinfo("終了","リサイズ終了しました")
        except:
            messagebox.showerror("エラー","うまくリサイズできませんでした")


    label_w=ttk.Label(frame,text="横幅(px)：")
    label_h=ttk.Label(frame,text="縦幅(px)：")
    entry_w=ttk.Entry(frame)
    entry_h=ttk.Entry(frame)
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',resize)
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(
        frame,
        text=main_checkbox_name,
        onvalue=1,
        offvalue=0,
        variable=window_front,
        command=execute
        )
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label_w.grid(row=1,column=0)
    entry_w.grid(row=1,column=1)
    label_h.grid(row=2,column=0)
    entry_h.grid(row=2,column=1)
    label.grid(row=3,column=0,columnspan=2)

def image_change():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def change(drop):
        file=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(file)
        for x in x_list:
            try:
                img=Image.open(x)
                name=os.path.basename(x)
                youso=os.path.splitext(name)[1]

                if v1.get()==1:
                    x_gif=x.replace(youso,".gif")
                    img.save(x_gif,"gif")
                if v2.get()==1:
                    x_png=x.replace(youso,".png")
                    img.save(x_png,"png",optimize=True)
                if v3.get()==1:
                    x_ico=x.replace(youso,".ico")
                    img.save(x_ico,"ico")
                if v4.get()==1:
                    x_jpg=x.replace(youso,".jpg")
                    img=img.convert("RGB")
                    img.save(x_jpg,"jpeg",quality=95, optimize=True)
                if v5.get()==1:
                    x_pdf=x.replace(youso,".pdf")
                    img.save(x_pdf,"pdf")
                if v6.get()==1:
                    x_webp=x.replace(youso,".webp")
                    img.save(x_webp,"webp")
            except:
                messagebox.showerror("エラー",f"{x}\nを変換できませんでした")
        messagebox.showinfo("終了","変換終了しました")


    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',change)
    v1=IntVar()
    cb1=ttk.Checkbutton(frame,text="GIF",variable=v1, onvalue=1, offvalue=0)
    v2=IntVar()
    cb2=ttk.Checkbutton(frame,text="PNG",variable=v2, onvalue=1, offvalue=0)
    v3=IntVar()
    cb3=ttk.Checkbutton(frame,text="ICO",variable=v3, onvalue=1, offvalue=0)
    v4=IntVar()
    cb4=ttk.Checkbutton(frame,text="JPEG",variable=v4, onvalue=1, offvalue=0)
    v5=IntVar()
    cb5=ttk.Checkbutton(frame,text="PDF",variable=v5, onvalue=1, offvalue=0)
    v6=IntVar()
    cb6=ttk.Checkbutton(frame,text="WEBP",variable=v6, onvalue=1, offvalue=0)
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(
    frame,
    text=main_checkbox_name,
    onvalue=1,
    offvalue=0,
    variable=window_front,
    command=execute
    )
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1,columnspan=5)
    buttonY.grid(row=0,column=0)
    cb1.grid(row=1,column=0)
    cb2.grid(row=1,column=1)
    cb3.grid(row=1,column=2)
    cb4.grid(row=1,column=3)
    cb5.grid(row=1,column=4)
    cb6.grid(row=1,column=5)
    label.grid(row=2,column=0,columnspan=6)

def qr_generate():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def qr_show():
        #try:
            x=clip.paste()
            img = qrcode.make(x)
            folder_path = os.getcwd()+"/temp1/qr_temp"
            make_folder(folder_path)
            z=folder_path+"/temp.png"
            img.save(z)
            root2=Free_window()
            root2.title("QRコード")
            canvas=Canvas(root2,width=img.size[0],height=img.size[1])
            canvas.pack(expand = True, fill = BOTH)
            pil_img1=Image.open(z)
            canvas_img= ImageTk.PhotoImage(pil_img1)
            canvas.canvas_img = canvas_img
            canvas.create_image(pil_img1.width*0.5,pil_img1.height*0.5,image=canvas_img)
        #except:
        #    messagebox.showerror("エラー","うまくQRコードを作成できませんでした")

    def qr_save():
        try:
            x=clip.paste()
            img = qrcode.make(x)
            y=filedialog.asksaveasfilename(
            title = "名前を付けて保存",
            filetypes = [("PNG", ".png")],
            defaultextension = "png",
            initialfile="QR"
            )
            img.save(y)
        except:
            messagebox.showerror("エラー","うまくQRコードを保存できませんでした")

    button1=ttk.Button(frame,text="QRを表示",command=qr_show)
    button2=ttk.Button(frame,text="QRを保存",command=qr_save)
    buttonX=ttk.Checkbutton(
        frame,
        text=main_checkbox_name,
        onvalue=1,
        offvalue=0,
        variable=window_front,
        command=execute
    )
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    button1.grid(row=1,column=0)
    button2.grid(row=1,column=1)

def qr_reader():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def qr_read(drop):
        try:
            x=drop.data.replace("{","").replace("}","").replace("\\","/")
            img = cv2.imdecode(
                np.fromfile(x, dtype=np.uint8),
                cv2.IMREAD_UNCHANGED
                )
            qrDetector = cv2.QRCodeDetector()
            data,_,_ = qrDetector.detectAndDecode(img)
            y=data
            if y=="":
                messagebox.showerror("エラー","うまくQRコードを読み取れませんでした")
                return
            clip.copy(y)
            messagebox.showinfo("終了","認識終了しました")
        except:
            messagebox.showerror("エラー","うまくQRコードを読み取れませんでした")

    def qr_read1():
        try:
            root.iconify()
            make_folder(os.getcwd()+"/temp1/qr_img")
            n=pyautogui.screenshot(os.getcwd()+'/temp1/qr_img/temp.png')
            root1=Free_window()

            def start():
                nonlocal pil_img,pil_img1,canvas_img
                pil_img=Image.open(os.getcwd()+'/temp1/qr_img/temp.png')
                pil_img1=pil_img
                canvas.config(width=pil_img1.width, height=pil_img1.height)
                canvas_img= ImageTk.PhotoImage(pil_img1)
                canvas.canvas_img = canvas_img
                canvas.create_image(pil_img1.width*0.5,pil_img1.height*0.5,image=canvas_img)
                root1.attributes('-fullscreen', True)


            def on_press(event):
                canvas.delete("rect")
                nonlocal start_x
                nonlocal start_y
                start_x = canvas.canvasx(event.x)
                start_y = canvas.canvasy(event.y)

            def on_move_press(event):
                nonlocal end_x, end_y
                end_x, end_y = event.x, event.y
                canvas.delete("rect")
                canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)

            def on_release(event):
                nonlocal end_x, end_y
                end_x, end_y = event.x, event.y
                canvas.delete("rect")
                canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)
                root1.destroy()
                triming()

            def triming():
                nonlocal start_x,start_y,end_x,end_y
                if start_x>end_x:
                    start_x,end_x=end_x,start_x
                if start_y>end_y:
                    start_y,end_y=end_y,start_y
                crop=pil_img.crop((start_x,start_y,end_x,end_y))
                crop.save(os.getcwd()+"/temp1/qr_img/temp1.png")
                img = cv2.imdecode(
                    np.fromfile(os.getcwd()+"/temp1/qr_img/temp1.png", dtype=np.uint8),
                    cv2.IMREAD_UNCHANGED
                    )
                qrDetector = cv2.QRCodeDetector()
                data,_,_ = qrDetector.detectAndDecode(img)
                y=data
                if y=="":
                    messagebox.showerror("エラー","うまくQRコードを読み取れませんでした")
                clip.copy(y)
                os.remove(os.getcwd()+"/temp1/qr_img/temp.png")
                os.remove(os.getcwd()+"/temp1/qr_img/temp1.png")
                messagebox.showinfo("終了","認識終了しました")
                root.deiconify()

            start_x=None
            start_y=None
            end_x=None
            end_y=None

            pil_img=None
            pil_img1=None
            canvas_img=None
            root.update()
            canvas=Canvas(root1,width=300,height=300,bg="gray")
            canvas.bind("<ButtonPress-1>", on_press)
            canvas.bind("<ButtonRelease-1>", on_release)
            canvas.bind("<B1-Motion>", on_move_press)
            canvas.pack(expand=True,side=TOP)
            root.after(10,start)
        except:
            messagebox.showerror("エラー","うまくQRコードを読み取れませんでした")

    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',qr_read)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    button=ttk.Button(frame,text="画面上のQR読み取り",command=qr_read1)

    frame.pack()
    buttonX.pack(side=TOP)
    buttonY.pack(side=TOP)
    label.pack(side=TOP)
    button.pack(side=TOP)

def face_mosaic():
        global frame,buttonY
        main_frame_delete()
        frame = ttk.Frame(root, padding=12)

        def dl_xml(save_path):
            make_folder(os.getcwd()+"/temp1")
            if var.get()==0:
                response = requests.get("https://raw.githubusercontent.com/opencv/opencv/master/data/haarcascades/haarcascade_frontalface_default.xml")
            elif var.get()==2:
                response = requests.get("https://raw.githubusercontent.com/opencv/opencv/master/data/haarcascades/haarcascade_fullbody.xml")
            if response.status_code == 200:
                with open(save_path, 'wb') as file:
                    file.write(response.content)
            else:
                messagebox.showerror("エラー","ファイルのダウンロードに失敗しました。")

        def mosaic(src, ratio=0.1):
            small = cv2.resize(src, None, fx=ratio, fy=ratio, interpolation=cv2.INTER_NEAREST)
            return cv2.resize(small, src.shape[:2][::-1], interpolation=cv2.INTER_NEAREST)

        def mosaic_area(src, x, y, width, height, ratio=0.05):
            dst = src.copy()
            dst[y:y + height, x:x + width] = mosaic(dst[y:y + height, x:x + width], ratio)
            return dst

        def face(drop):
            file=drop.data.replace("{","").replace("}","").replace("\\","/")
            x_list=file_mult(file)
            for x in x_list:
                src = cv2.imdecode(
                    np.fromfile(x, dtype=np.uint8),
                    cv2.IMREAD_UNCHANGED
                    )
                y= os.path.basename(x)
                z=x.replace(y,"mosaic_"+y)
                youso=os.path.splitext(z)[1]
                dl_xml(os.getcwd()+"/temp1/face.xml")
                face_cascade_path =os.getcwd()+"/temp1/face.xml"
                face_cascade = cv2.CascadeClassifier(face_cascade_path)

                src_gray = cv2.cvtColor(src, cv2.COLOR_BGR2GRAY)

                faces = face_cascade.detectMultiScale(src_gray)

                for x, y, w, h in faces:
                    dst_face = mosaic_area(src, x, y, w, h)

                _, buf = cv2.imencode(youso, dst_face)
                buf.tofile(z)
            messagebox.showinfo("終了","モザイク完了しました")


        frame.drop_target_register(DND_FILES)
        frame.dnd_bind('<<Drop>>',face)
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
        var=IntVar()
        var.set(0)
        combo1=ttk.Radiobutton(frame,text="顔",variable=var,value=0)
        combo3=ttk.Radiobutton(frame,text="全身",variable=var,value=2)
        combo4=ttk.Radiobutton(frame,text="任意指定",variable=var,value=3)

        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        combo1.grid(row=1,column=0)
        combo3.grid(row=1,column=1)
        label.grid(row=2,column=0,columnspan=2)

def img_rotate():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def img_rotate1(drop):
            file=drop.data.replace("{","").replace("}","").replace("\\","/")
            x_list=file_mult(file)
            for x in x_list:
                try:
                    img=Image.open(x)
                    img_rotate=img.rotate(int(entry1.get()),resample=Image.BICUBIC,expand=True)
                    name=os.path.basename(x)
                    y=x.replace(name,"rotate_"+name)
                    img_rotate.save(y)
                except:
                    messagebox.showerror("エラー",f"{x}\nをうまく回転できませんでした")
            messagebox.showinfo("終了","回転が完了しました")


    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',img_rotate1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    label1=ttk.Label(frame,text="反時計回り(度)：")
    entry1=ttk.Entry(frame)
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    entry1.insert(0,"180")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label.grid(row=2,column=0,columnspan=2)

def img_gray():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def img_gray1(drop):
        file=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(file)
        for x in x_list:
            try:
                im = cv2.imdecode(
                    np.fromfile(x, dtype=np.uint8),
                    cv2.IMREAD_UNCHANGED
                )
                im_gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
                y= os.path.basename(x)
                z=x.replace(y,"gray_"+y)
                youso=os.path.splitext(z)[1]

                _, buf = cv2.imencode(youso, im_gray)
                buf.tofile(z)
            except:
                messagebox.showerror("エラー",f"{x}\nをうまくグレイスケールできませんでした")
        messagebox.showinfo("終了","グレイスケール完了しました")

    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',img_gray1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.pack(side=TOP)
    buttonY.pack(side=TOP)
    label.pack(side=TOP)

def video_cut():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def video_cut1(drop):
        file=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(file)
        for x in x_list:
            try:
                y= os.path.basename(x)
                z=x.replace(y,"cut_"+y)
                start=60*int(entry_start1.get())+float(entry_start2.get())
                end=60*int(entry_end1.get())+float(entry_end2.get())
                video1 = VideoFileClip(x)
                video=video1.subclip(start, end)
                video.write_videofile(z,logger=None)
                video.close()
            except:
                messagebox.showerror("エラー",f"{x}\nをうまくカットできませんでした")
        messagebox.showinfo("終了","カットが完了しました")

    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',video_cut1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    label1=ttk.Label(frame,text="開始時刻")
    label2=ttk.Label(frame,text="終了時刻")
    label3=ttk.Label(frame,text="：")
    label4=ttk.Label(frame,text="：")
    entry_start1=ttk.Entry(frame,width=5)
    entry_start2=ttk.Entry(frame,width=5)
    entry_end1=ttk.Entry(frame,width=5)
    entry_end2=ttk.Entry(frame,width=5)
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=2)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry_start1.grid(row=1,column=1)
    label3.grid(row=1,column=2)
    entry_start2.grid(row=1,column=3)
    label2.grid(row=2,column=0)
    entry_end1.grid(row=2,column=1)
    label4.grid(row=2,column=2)
    entry_end2.grid(row=2,column=3)
    label.grid(row=3,column=0,columnspan=4)

def img_triming():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    origin=None

    canvas_img=None
    def img_top(drop):
        nonlocal canvas_img,pil_img,z,origin,pil_img_1
        x=drop.data.replace("{","").replace("}","").replace("\\","/")
        y= os.path.basename(x)
        z=x.replace(y,"crop_"+y)
        try:
            pil_img=pil_img_1=Image.open(x)
        except:
            messagebox.showerror("エラー",f"{x}\nをうまく開けませんでした")
            return
        if pil_img.width>pil_img.height:
            origin=0
            pil_img=pil_img.resize((int(ws/2),int(pil_img.height*((ws/2)/pil_img.width))),Image.LANCZOS)
        elif pil_img.width<pil_img.height:
            origin=1
            pil_img=pil_img.resize((int(pil_img.width*((hs/2)/pil_img.height)),int(hs/2)),Image.LANCZOS)
        elif pil_img.width==pil_img.height:
            origin=2
            pil_img=pil_img.resize((int(hs/2),int(hs/2)),Image.LANCZOS)
        canvas.delete("all")
        canvas_img= ImageTk.PhotoImage(pil_img)
        canvas.create_image(pil_img.width/2,pil_img.height/2,image=canvas_img)
        canvas.config(width=pil_img.width, height=pil_img.height)

    def on_press(event):
        canvas.delete("rect")
        nonlocal start_x
        nonlocal start_y
        start_x = canvas.canvasx(event.x)
        start_y = canvas.canvasy(event.y)

    def on_move_press(event):
        nonlocal end_x, end_y
        end_x, end_y = event.x, event.y
        canvas.delete("rect")
        canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)

    def on_release(event):
        nonlocal end_x, end_y
        end_x, end_y = event.x, event.y
        canvas.delete("rect")
        canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)

    def triming():
        nonlocal origin
        try:
            if origin==2:
                start_x1=int(start_x*pil_img_1.height/int(hs/2))
                start_y1=int(start_y*pil_img_1.height/int(hs/2))
                end_x1=int(end_x*pil_img_1.height/int(hs/2))
                end_y1=int(end_y*pil_img_1.height/int(hs/2))
            elif origin==1:
                start_x1=int(start_x*pil_img_1.width/int(pil_img_1.width*((hs/2)/pil_img_1.height)))
                start_y1=int(start_y*pil_img_1.height/int(hs/2))
                end_x1=int(end_x*pil_img_1.width/int(pil_img_1.width*((hs/2)/pil_img_1.height)))
                end_y1=int(end_y*pil_img_1.height/int(hs/2))
            elif origin==0:
                start_x1=int(start_x*pil_img_1.width/int(ws/2))
                start_y1=int(start_y*pil_img_1.height/int(pil_img_1.height*((ws/2)/pil_img_1.width)))
                end_x1=int(end_x*pil_img_1.width/int(ws/2))
                end_y1=int(end_y*pil_img_1.height/int(pil_img_1.height*((ws/2)/pil_img_1.width)))
            if start_x1>end_x1:
                start_x1,end_x1=end_x1,start_x1
            if start_y1>end_y1:
                start_y1,end_y1=end_y1,start_y1
            crop=pil_img_1.crop((start_x1,start_y1,end_x1,end_y1))
            crop.save(z)
            messagebox.showinfo("終了","トリミング完了しました")
        except:
            messagebox.showerror("エラー","うまくトリミングできませんでした")

    start_x=None
    start_y=None
    end_x=None
    end_y=None
    pil_img=None
    z=None
    pil_img_1=None
    root.update()
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',img_top)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    canvas=Canvas(frame,width=300,height=300,bg="gray")
    button1=ttk.Button(frame,text="トリミング",command=triming)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")

    canvas.bind("<ButtonPress-1>", on_press)
    canvas.bind("<ButtonRelease-1>", on_release)
    canvas.bind("<B1-Motion>", on_move_press)

    frame.pack()
    buttonX.pack(side=TOP)
    buttonY.pack(side=TOP)
    canvas.pack(expand=True,side=TOP,fill=BOTH)
    button1.pack(side=TOP)

def img_press():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def change_ratio():
        if var.get()==0:
            label1.config(text="%")
            entry.delete(0, END)
            entry.insert(END,"50")
        elif var.get()==1:
            label1.config(text="色")
            entry.delete(0, END)
            entry.insert(END,"256")

    def img_press1(drop):
        file=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(file)
        for x in x_list:
            try:
                if var.get()==0:
                    img=Image.open(x)
                    name=os.path.basename(x)
                    youso=os.path.splitext(name)[1]
                    make_folder(os.getcwd()+"/temp1/img_compress")
                    name_press=x.replace(name,"press_"+name)
                    youso1=youso.replace(".","")
                    img.save(name_press,"jpeg",quality=int(entry.get()),optimize=True)
                elif var.get()==1:
                    img_press=Image.open(x)
                    name=os.path.basename(x)
                    youso=os.path.splitext(name)[1]
                    img_press=img_press.quantize(int(entry.get()))
                    #img_press=img_press.convert("P",palette=Image.ADAPTIVE,colors=int(entry.get()))
                    name_press=x.replace(name,"press_"+name)
                    youso1=youso.replace(".","")
                    img_press.save(name_press,format=youso1,optimize=True)

            except:
                messagebox.showerror("エラー",f"{x}\nをうまく圧縮できませんでした")
        messagebox.showinfo("終了","圧縮完了しました")

    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',img_press1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    entry=ttk.Entry(frame,width=5)
    label1=ttk.Label(frame,text="%")
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    entry.insert(END,"50")
    var=IntVar()
    var.set(0)
    radio1=ttk.Radiobutton(frame,text="画質圧縮(JPEG)",variable=var,value=0,command=change_ratio)
    radio2=ttk.Radiobutton(frame,text="色圧縮(PNG)",variable=var,value=1,command=change_ratio)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    label1.grid(row=2,column=1)
    entry.grid(row=2,column=0)
    label.grid(row=3,column=0,columnspan=2)

def mouse_click():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def main_frame1():
        global frame,buttonY
        nonlocal t,y
        y=1
        if t.is_alive():
            t.join()
        frame.destroy()
        set_frame1(3)


    def stopping():
        nonlocal t,y
        y=1
        if t.is_alive():
            t.join()
        frame_delete()

    def mouse_click2():
        nonlocal t
        button.configure(state="disabled")
        buttonY.configure(state="disabled")
        t.start()
        root.after(3000,mouse_click1)

    def click_stop():
        nonlocal y
        while True:
            if y==0:
                if key.is_pressed('esc'):
                    y=1
                    break
            else:
                break

    def mouse_click1():
        nonlocal y,x,t

        if y==0:
            if x<int(entry_count.get()):
                if variable.get()=="右クリック":
                    pyautogui.click(button="right")
                else:
                    pyautogui.click(button="left")
                x+=1
                root.after(int(float(entry_between.get())*1000),mouse_click1)
            else:
                y=1
                root.after(1,mouse_click1)
        else:
            x=0
            if t.is_alive():
                t.join()
            button.configure(state="normal")
            buttonY.configure(state="normal")
            t=threading.Thread(target=click_stop)
            messagebox.showinfo("終了","完了しました")
            y=0

    x=0 # 回数
    y=0 # 終了判定
    t=threading.Thread(target=click_stop)
    root.protocol("WM_DELETE_WINDOW", stopping)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    option = ["右クリック","左クリック"]
    variable = StringVar ( )
    combo = ttk.Combobox ( frame , values = option , textvariable = variable,width=8 )
    entry_count=ttk.Entry(frame,width=3)
    entry_between=ttk.Entry(frame,width=3)
    label_count=ttk.Label(frame,text="クリックする回数")
    label_between=ttk.Label(frame,text="クリック間隔")
    label_ex=ttk.Label(frame,text="3秒後に開始し、escキーで終了")
    button=ttk.Button(frame,text="開始",command=mouse_click2,state="normal")
    label_combo=ttk.Label(frame,text="クリックするボタンを選択してください")
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(),font=("Helvetica", 7),bg="gray",fg="white",state="normal")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label_combo.grid(row=1,column=0)
    combo.grid(row=1,column=1)
    label_count.grid(row=2,column=0)
    entry_count.grid(row=2,column=1)
    label_between.grid(row=3,column=0)
    entry_between.grid(row=3,column=1)
    label_ex.grid(row=4,column=0,columnspan=2)
    button.grid(row=5,column=0,columnspan=2)

def video_to_voice():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    def video_to_voice1(drop):
        file=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(file)
        for x in x_list:
            try:
                videoclip=VideoFileClip(x)
                name=os.path.basename(x)
                youso=os.path.splitext(name)[1]
                voice=x.replace(youso,".mp3")
                videoclip.audio.write_audiofile(voice,logger=None)
                videoclip.close()
            except:
                messagebox.showerror("エラー",f"{x}\nをうまく変換できませんでした")
        messagebox.showinfo("終了","完了しました")

    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',video_to_voice1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    label.grid(row=1,column=0,columnspan=2)

def pdf_split():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)
    frame_1=ttk.Frame(frame)
    frame_2=ttk.Frame(frame)

    def pdf_split2(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                name=os.path.basename(x)
                y=x.replace(name,"split_"+name)
                merger = pypdf.PdfMerger()
                start=int(entry1.get())-1
                end=int(entry2.get())
                merger.append(x, pages=(start,end))
                merger.write(y)
                merger.close()
            except:
                messagebox.showerror("エラー",f"{x}\nをうまく分割できませんでした")
        messagebox.showinfo("終了","終了しました")

    def change_frame():
        if var.get()==0:
            entry1.delete(0, END)
            entry2.delete(0, END)
            frame_2.grid_forget()
            frame_1.grid(row=2,column=0,columnspan=2)
            label1.pack()
        elif var.get()==1:
            frame_1.grid_forget()
            frame_2.grid(row=2,column=0,columnspan=2)
            label2.grid(row=0,column=0)
            entry1.grid(row=0,column=1)
            label2_1.grid(row=1,column=0)
            entry2.grid(row=1,column=1)
            label3.grid(row=2,column=0,columnspan=2)
            entry1.insert(END,"1")
            entry2.insert(END,"1")

    def pdf_split1(drop):
        file=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(file)
        for x in x_list:
            try:
                pdf=pypdf.PdfReader(x)
                name=os.path.basename(x)
                page=0
                for i, page in enumerate(pdf.pages):
                    dst_pdf = pypdf.PdfWriter()
                    dst_pdf.add_page(page)
                    output=x.replace(".pdf","")+"_"+str(i+1)+".pdf"
                    dst_pdf.write(output)
            except:
                messagebox.showerror("エラー",f"{x}\nをうまく分割できませんでした")
        messagebox.showinfo("終了","終了しました")

    frame_1.drop_target_register(DND_FILES)
    frame_1.dnd_bind('<<Drop>>',pdf_split1)
    frame_2.drop_target_register(DND_FILES)
    frame_2.dnd_bind('<<Drop>>',pdf_split2)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame_1,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    var=IntVar()
    var.set(0)
    combo1=ttk.Radiobutton(frame,text="全ページ分解",variable=var,value=0,command=change_frame)
    combo2=ttk.Radiobutton(frame,text="ページ抽出",variable=var,value=1,command=change_frame)
    label2=ttk.Label(frame_2,text="開始ページ：")
    label2_1=ttk.Label(frame_2,text="終了ページ：")
    entry1=ttk.Entry(frame_2,width=5)
    entry2=ttk.Entry(frame_2,width=5)
    label3=ttk.Label(frame_2,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    combo1.grid(row=1,column=0)
    combo2.grid(row=1,column=1)
    frame_1.grid(row=2,column=0,columnspan=2)
    label1.pack()

def pdf_merge():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def pdf_merge1(drop):
        file=drop.data.replace("{","").replace("}","").replace("\\","/")
        text.insert(END,"\n"+file)

    def pdf_merge2():
        try:
            path=filedialog.asksaveasfilename(
            title = "名前を付けて保存",
            filetypes = [("PDF", ".pdf")],
            defaultextension = "pdf",
            initialfile="merge.pdf"
                )
            s = text.get( 0., END )
            list=s.split(sep="\n")
            merger = pypdf.PdfMerger()
            list1=list[1:-1]
            for i in list1:
                merger.append(i)
            merger.write(path)
            merger.close()
            messagebox.showinfo("終了","完了しました")
        except:
            messagebox.showerror("エラー","失敗しました")


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    text=ScrolledText(frame,width=30,height=20,wrap=NONE)
    button=ttk.Button(frame,text="開始",command=pdf_merge2)
    text.insert(0.,'ファイルをドロップしてください')
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',pdf_merge1)


    frame.pack()
    buttonX.grid(row=0,column=1,columnspan=2)
    buttonY.grid(row=0,column=0)
    text.grid(row=1,column=0,columnspan=3)
    button.grid(row=3,column=0,columnspan=3)

def red_sheet():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def rgb_to_hex():
        r=int(entry_r.get())
        g=int(entry_g.get())
        b=int(entry_b.get())
        return '#{:02x}{:02x}{:02x}'.format(r, g, b)

    def red_sheet1():
        try:
            root1 = Free_window()
            root1.title('赤シート')
            root1.attributes("-topmost", True)
            frame_1 = ttk.Frame(root1, style="Transparent.TFrame", width="400", height="300")
            root1.wm_attributes("-transparentcolor", "snow")
            root1.attributes("-alpha", float(entry_grass.get())/100)
            style = ttk.Style()
            style.configure("Transparent.TFrame", background=rgb_to_hex())
            style.configure("Transparent.TFrame", borderwidth=0)
            style.configure("Transparent.TFrame", highlightthickness=0)
            style.configure("Transparent.TFrame", relief="flat")
            frame_1.pack(fill="both", expand=True)
        except:
            messagebox.showerror("エラー","RGBを設定してください")

    entry_r=ttk.Entry(frame,width=5)
    entry_g=ttk.Entry(frame,width=5)
    entry_b=ttk.Entry(frame,width=5)
    entry_grass=ttk.Entry(frame,width=5)
    label_grass=ttk.Label(frame,width=18,text="不透明度(0~100%)")
    label_r=ttk.Label(frame,width=10,text="赤(0~255)")
    label_g=ttk.Label(frame,width=10,text="緑(0~255)")
    label_b=ttk.Label(frame,width=10,text="青(0~255)")
    button1=ttk.Button(frame,text="作成",command=red_sheet1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    entry_r.insert(0,"255")
    entry_g.insert(0,"0")
    entry_b.insert(0,"0")
    entry_grass.insert(0,"75")

    frame.pack()
    buttonY.grid(row=0,column=0,columnspan=2)
    buttonX.grid(row=0,column=4,columnspan=2)
    label_r.grid(row=1,column=0)
    entry_r.grid(row=1,column=1)
    label_g.grid(row=1,column=2)
    entry_g.grid(row=1,column=3)
    label_b.grid(row=1,column=4)
    entry_b.grid(row=1,column=5)
    label_grass.grid(row=2,column=1,columnspan=2)
    entry_grass.grid(row=2,column=3)
    button1.grid(row=3,column=0,columnspan=6)

def speed_check():
    global frame,buttonY

    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def speed_check1():
        nonlocal img
        button.config(text="計測中",state="disabled")
        root.update()
        s = speedtest.Speedtest()
        threads = None
        s.get_best_server()
        s.download(threads=threads)
        s.upload(threads=threads)
        url=s.results.share()
        pil=Image.open(io.BytesIO(requests.get(url).content))
        img=ImageTk.PhotoImage(pil)
        canvas.grid(row=1,column=0,columnspan=2)
        canvas.create_image(375,200,image=img)
        messagebox.showinfo("終了","計測が終了しました")
        button.config(text="計測開始",state="normal")


    img=None
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    canvas=Canvas(frame,width=750,height=400)
    button=ttk.Button(frame,text="計測開始",command=speed_check1)

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    button.grid(row=2,column=0,columnspan=2)

def exif():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def get_exif(fname):
        img = Image.open(fname)
        exif = {}
        if img._getexif():
            for k, v in img._getexif().items():
                if k in ExifTags.TAGS:
                    exif[ExifTags.TAGS[k]] = v
            exif.pop('MakerNote')
        return exif

    def print_exif(exif):
        if exif:
            root1 = Free_window()
            root1.title("EXIF")
            frame_1=ttk.Frame(root1,padding=12)
            box=ScrolledText(root1,width=50,height=30,font=(u'メイリオ',8))
            frame_1.pack()
            box.pack(expand=True,fill="both")
            for k, v in exif.items():
                box.insert(END, str(k) + " : " + str(v) + "\n")
        else:
            messagebox.showerror("EXIFなし","EXIF情報がありません")

    def exif1():
        x=entry1.get()
        y=get_exif(x)
        print_exif(y)

    def conv_deg(v):
    # 分数を度に変換
        d = float(v[0][0]) / float(v[0][1])
        m = float(v[1][0]) / float(v[1][1])
        s = float(v[2][0]) / float(v[2][1])
        return d + (m / 60.0) + (s / 3600.0)

    def get_gps(fname):
        img = Image.open(fname)
        exif_data = img._getexif()
        gps = {}
        if exif_data:
            for tag, value in exif_data.items():
                decoded = ExifTags.GPSTAGS.get(tag, tag)
                if decoded == "GPSInfo":
                    gps_data = {}
                    for t in value:
                        sub_decoded = ExifTags.GPSTAGS.get(t, t)
                        gps_data[sub_decoded] = value[t]
                    gps_latitude = conv_deg(gps_data["GPSLatitude"])
                    gps_latitude_ref = gps_data["GPSLatitudeRef"]
                    if gps_latitude_ref != "N":
                        gps_latitude = 0 - gps_latitude
                    gps_longitude = conv_deg(gps_data["GPSLongitude"])
                    gps_longitude_ref = gps_data["GPSLongitudeRef"]
                    if gps_longitude_ref != "E":
                        gps_longitude = 0 - gps_longitude
                    gps['latitude'] = gps_latitude
                    gps['longitude'] = gps_longitude
        return gps if gps else "経度・緯度は記録されていません。"


    def exif2():
        x=entry1.get()
        gps_data = get_gps(x)
        messagebox.showinfo("GPS",gps_data)

    def exif3():
        x=entry1.get()
        img = Image.open(x)
        data = img.getdata()
        mode = img.mode
        size = img.size
        name=os.path.basename(x)
        y=x.replace(name,"exif_"+name)

        new_img = Image.new(mode, size)
        new_img.putdata(data)
        new_img.save(y)
        messagebox.showinfo("完了","EXIFを削除しました")

    button1=ttk.Button(frame,text="EXIFを確認",command=exif1)
    button3=ttk.Button(frame,text="GPSを確認",command=exif2)
    button2=ttk.Button(frame,text="EXIFを削除",command=exif3)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    entry1=ttk.Entry(frame,width=20,font=(u'メイリオ',8))
    label=ttk.Label(frame,text="ファイルをドロップしてください")
    entry1.drop_target_register(DND_FILES)
    entry1.dnd_bind('<<Drop>>',lambda drop:entry1.insert(END,drop.data.replace("{","").replace("}","").replace("\\","/")))

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    button1.grid(row=2,column=0)
    button3.grid(row=2,column=2)
    button2.grid(row=2,column=1)
    entry1.grid(row=1,column=1,columnspan=2)
    label.grid(row=1,column=0)

def memo():
    def on_key_press(event):
        index = text.index("insert").split('.')
        line_number = int(index[0])
        current_line = text.get(f"{line_number}.0", f"{line_number}.end")
        cursor_pos = text.index(INSERT)
        line_end = cursor_pos.split('.')[0] + '.end'
        text_after_cursor = text.get(cursor_pos, line_end)

        if event.keysym == 'Return':
            stripped_line = current_line.lstrip()
            list_match = re.match(r'(\s*)(\d+)\.\s*(.*)', current_line)
            hyphen_list_match = re.match(r'(\s*)-\s+.*', current_line)  # ハイフンとスペースの組み合わせのマッチング
            asterisk_list_match = re.match(r'(\s*)\*\s+.*', current_line)  # アスタリスクとスペースの組み合わせのマッチング

            if hyphen_list_match or asterisk_list_match:  # ハイフンまたはアスタリスクとスペースで始まる場合
                process_list_item(line_number, stripped_line, current_line,text_after_cursor)
                text.delete(cursor_pos, line_end)
                text.mark_set(INSERT, f"{int(cursor_pos.split('.')[0])+1}.end")
                update_numbered_list(line_number + 1)  # 番号付きリストの更新
                return "break"
            elif list_match:
                process_numbered_list_item(line_number, list_match,text_after_cursor)
                text.delete(cursor_pos, line_end)
                text.mark_set(INSERT, f"{int(cursor_pos.split('.')[0])+1}.end")
                update_numbered_list(line_number + 1)  # 番号付きリストの更新
                return "break"

        elif event.keysym == 'Tab':
            adjust_indent(line_number, current_line, event.state & 0x0001)
            return "break"

    def process_list_item(line_number, stripped_line, current_line,text_after_cursor):
        if stripped_line in ('-', '- ', '*', '* '):  # '*' を追加
            text.delete(f"{line_number}.0", f"{line_number}.end")
        else:
            indent = len(current_line) - len(stripped_line)
            new_bullet = '*' if stripped_line.startswith('*') else '-'  # 新しいバレットを決定
            text.insert(f"{line_number}.end", '\n' + ' ' * indent + new_bullet + ' '+text_after_cursor)

    def process_numbered_list_item(line_number, match,text_after_cursor):
        indent, number, content = match.groups()
        if not content:
            text.delete(f"{line_number}.0", f"{line_number}.end")
        else:
            next_number = int(number) + 1
            text.insert(f"{line_number}.end", '\n' + indent + f"{next_number}. "+text_after_cursor)

    def adjust_indent(line_number, current_line, is_shift_pressed):
        indent_match = re.match(r'(\s*)(\-|\*|\d+\.)', current_line)  # '*' を追加
        if indent_match:
            indent, list_type = indent_match.groups()
            if list_type in ('-', '*'):  # '*' を追加
                new_indent = "  " + indent if not is_shift_pressed else indent[2:]
                text.delete(f"{line_number}.0", f"{line_number}.{len(indent)}")
                text.insert(f"{line_number}.0", new_indent)
            else:
                # 番号付きリストの場合
                new_indent = "  " + indent if not is_shift_pressed else indent[2:]
                reset_list_number(line_number, new_indent, is_decreasing_indent=is_shift_pressed)
        update_numbered_list(line_number)


    def reset_list_number(line_number, new_indent, is_decreasing_indent):
        current_line = text.get(f"{line_number}.0", f"{line_number}.end")
        match = re.match(r'(\s*)(\d+)\.', current_line)
        if match:
            if is_decreasing_indent:
                new_number = str(get_previous_number_same_level(line_number, new_indent))
            else:
                new_number = str(get_next_list_number_same_level(line_number, new_indent))
            replacement = f"{new_indent}{new_number}."
            text.replace(f"{line_number}.0", f"{line_number}.{len(match.group(0))}", replacement)


    def get_next_list_number_same_level(line_number, indent):
        next_line_number = line_number - 1
        while next_line_number > 0:
            line = text.get(f"{next_line_number}.0", f"{next_line_number}.end")
            match = re.match(r'(\s*)(\d+)\.', line)
            if match and match.group(1) == indent:
                return int(match.group(2)) + 1
            elif match and len(match.group(1)) < len(indent):
                # 異なるネストが存在する
                break
            next_line_number -= 1
        return 1

    def get_previous_number_same_level(line_number, indent):
        prev_line_number = line_number - 1
        while prev_line_number > 0:
            line = text.get(f"{prev_line_number}.0", f"{prev_line_number}.end")
            match = re.match(r'(\s*)(\d+)\.', line)
            if match:
                if match.group(1) == indent:
                    return int(match.group(2)) + 1
                elif len(match.group(1)) < len(indent):
                    break  # 間に異なるネストが存在する
            prev_line_number -= 1
        return 1

    def update_numbered_list(changed_line_number):
        start_line = find_list_start(changed_line_number)
        line_number = start_line
        last_line_number = int(text.index('end-1c').split('.')[0])

        # 各レベルの番号を保持する辞書
        level_numbers = {}

        while line_number <= last_line_number:
            line = text.get(f"{line_number}.0", f"{line_number}.end")
            indent_level = get_indent_level(line)

            if is_numbered_list_item(line):
                # 同じレベルの以前の番号を取得し、それを1増やす
                number = level_numbers.get(indent_level, 0) + 1
                level_numbers[indent_level] = number

                # より深いレベルの番号をリセット
                for level in list(level_numbers.keys()):
                    if level > indent_level:
                        del level_numbers[level]

                new_line = re.sub(r'\d+\.', f"{number}.", line)
                text.replace(f"{line_number}.0", f"{line_number}.end", new_line)

            elif line.strip() == '':
                break  # 空の行でリスト終了とみなす

            line_number += 1

    def is_numbered_list_item(line):
        # 番号付きリストアイテムかどうかを判断
        return bool(re.match(r'\s*\d+\.', line))

    def find_list_start(line_number):
        # リストの開始行を見つける
        while line_number > 0:
            line = text.get(f"{line_number}.0", f"{line_number}.end")
            if not is_numbered_list_item(line) or line.strip() == '':
                break
            line_number -= 1
        return line_number + 1

    def get_indent_level(line):
        indent_match = re.match(r'(\s*)', line)
        return len(indent_match.group(1)) if indent_match else 0


    def save_text():
        filename = filedialog.asksaveasfilename(
        title = "名前を付けて保存",
        filetypes = [("Text", ".txt"),("Markdwon",".md")],
        defaultextension = "Text"
            )
        if filename != "":
            with open(filename, "w") as f:
                f.write(text.get("1.0", "end-1c"))
            messagebox.showinfo("完了","保存しました")

    def drop_text(drop):
        path=drop.data.replace("{","").replace("}","").replace("\\","/")
        with open(path, "r") as f:
            text.insert(END, f.read())

    root1 = Free_window()
    root1.title('付箋')
    menu=Menu(root1)
    root1.config(menu=menu)
    men1=Menu(menu,tearoff = False)
    menu.add_cascade(label="ファイル",menu=men1)
    men1.add_command(label="中身をコピー",command=lambda:clip.copy(text.get("0.0",END).rstrip('\n')))
    men1.add_command(label="名前をつけて保存",command=save_text)

    frame_1=ttk.Frame(root1,padding=1)
    text=ScrolledText(root1,width=40,height=12,undo=True)
    # ドロップ
    text.drop_target_register(DND_FILES)
    text.dnd_bind('<<Drop>>',drop_text)
    #text.bind('<Return>', insert_bullet_point)

    frame_1.pack()
    text.pack(expand=True,fill="both",side=TOP)
    root1.attributes('-topmost', True)
    text.bind('<KeyPress-Return>', on_key_press)
    text.bind('<KeyPress-Tab>', on_key_press)

def video_to_gif():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def video_gif(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                name=os.path.basename(x)
                youso=os.path.splitext(name)[1]
                y=x.replace(youso,".gif")
                clip = VideoFileClip(x)
                clip.write_gif(y, fps=int(entry1.get()),logger=None)
                clip.close()
            except:
                messagebox.showerror("エラー",f"{x}\nをうまく変換できませんでした")
        messagebox.showinfo("完了","変換を終了します")


    label=ttk.Label(frame,text="FPS：")
    entry1=ttk.Entry(frame,width=20)
    frame.drop_target_register(DND_FILES)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    frame.dnd_bind('<<Drop>>',video_gif)
    entry1.insert(END,"15")
    label1=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label1.grid(row=2,column=0,columnspan=2)

def svg_con():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def svg(drop):
        x = drop.data.replace("{","").replace("}","").replace("\\","/")
        input_list=file_mult(x)
        for input_file in input_list:
            try:
                im = Image.open(input_file)
                w, h = im.size
                output_file = input_file.split(".")[0] + '.svg'

                with open(input_file, "rb") as f:
                    img = base64.b64encode(f.read())
                dwg = svgwrite.Drawing(output_file)
                dwg.add(dwg.image('data:image/jpg;base64,' + img.decode("ascii"),
                                size=(w, h)
                                )
                        )
                dwg.save()
            except:
                messagebox.showerror("エラー",f"{input_file}\nをうまく変換できませんでした")
        messagebox.showinfo("完了","変換が終了しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',svg)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def img_inv():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def inv(drop):
        file = drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(file)
        for x in x_list:
            try:
                name=os.path.basename(x)
                youso=os.path.splitext(name)[1]
                y=x.replace(name,"inv_"+name)
                z=y.replace(youso,".png")
                img = cv2.imdecode(
                    np.fromfile(x, dtype=np.uint8),
                    cv2.IMREAD_UNCHANGED
                    )
                if img.shape[2] == 4:
                    # アルファチャンネルを含む
                    pass
                elif img.shape[2] == 3:
                    # アルファチャンネル追加
                    alpha = np.ones((img.shape[0], img.shape[1], 1), dtype=np.uint8) * 255
                    img = np.concatenate((img, alpha), axis=2)

                img[:, :, 3] = np.where(np.all(img == 255, axis=-1), 0, 255)  # 白でAlphaを0
                _, buf = cv2.imencode(youso, img)
                buf.tofile(z)
            except:
                messagebox.showerror("エラー",f"{x}\nを失敗しました")
        messagebox.showinfo("完了","変換が完了しました")

    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',inv)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def front_window():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def on_release(key):
        if key == keyboard.Key.pause:
            # アクティブなウィンドウ取得
            active_window = win32gui.GetForegroundWindow()
            class_name = win32gui.GetWindowText(active_window)
            # ウィンドウが最前面に固定されているか
            if win32gui.GetWindowLong(active_window, win32con.GWL_EXSTYLE) & win32con.WS_EX_TOPMOST:
                # 最前面表示解除
                win32gui.SetWindowPos(active_window, win32con.HWND_NOTOPMOST, 0, 0, 0, 0, win32con.SWP_NOMOVE | win32con.SWP_NOSIZE)
                toast = winotify.Notification(
                        title="最前面の固定の解除",
                        msg=class_name+"の固定を解除しました",
                        app_id="Yuki's Army knife",
                    )
                toast.show()
            else:
                # 最前面固定
                win32gui.SetWindowPos(active_window, win32con.HWND_TOPMOST, 0, 0, 0, 0, win32con.SWP_NOMOVE | win32con.SWP_NOSIZE)
                toast = winotify.Notification(
                        title="最前面の固定",
                        msg=class_name+"を固定しました",
                        app_id="Yuki's Army knife",
                    )
                toast.show()


    def main_frame1(num):
        listener.stop()
        t.join()
        root.protocol("WM_DELETE_WINDOW", frame_delete)
        main_frame(num)

    def stopping():
        listener.stop()
        t.join()
        frame_delete()

    listener=None
    root.protocol("WM_DELETE_WINDOW", stopping)
    label=ttk.Label(frame,text="pauseキーで\n最前面に表示します",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)
    def listen_thread():
        nonlocal listener
        with keyboard.Listener(on_release=on_release) as listener:
            listener.join()

    t = threading.Thread(target=listen_thread, daemon=True)
    t.start()

def screen_memo():
    global frame,buttonY
    def screen_memo1():
        root.iconify()
        time.sleep(1)
        make_folder(os.getcwd()+"/temp1/img_memo")
        pyautogui.screenshot(os.getcwd()+"/temp1/img_memo/temp.png")
        root.deiconify()
        root1 = Free_window()
        root1.title('スクショメモ')
        frame1=ttk.Frame(root1, padding=0)
        root1.attributes('-fullscreen', True)
        root1.attributes('-topmost', '1')
        root1.resizable(False, False)


        def on_press(event):
            canvas.delete("rect")
            nonlocal start_x
            nonlocal start_y
            start_x = canvas.canvasx(event.x)
            start_y = canvas.canvasy(event.y)

        def on_move_press(event):
            nonlocal end_x, end_y
            end_x, end_y = event.x, event.y
            canvas.delete("rect")
            canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)

        def on_release(event):
            nonlocal end_x, end_y,pmenu,crop
            end_x, end_y = event.x, event.y
            canvas.delete("rect")
            canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)
            pmenu.add_command(label="拡大・縮小", command=size_change)
            pmenu.add_command(label="保存", command=img_get)
            pmenu.add_command(label="コピー", command=img_clip)
            pmenu.add_command(label="終了", command=root1.destroy)
            triming()

        def triming():
            nonlocal crop,start_x,start_y,end_x,end_y
            if start_x > end_x:
                start_x, end_x = end_x, start_x
            if start_y > end_y:
                start_y, end_y = end_y, start_y
            crop=pil_img.crop((start_x,start_y,end_x,end_y))
            crop=crop.resize((int(crop.width*n),int(crop.height*n)))
            root1.attributes('-fullscreen', False)
            canvas_img1= ImageTk.PhotoImage(crop)
            canvas.canvas_img = canvas_img1
            canvas.create_image(crop.width*0.5,crop.height*0.5,image=canvas_img1)
            canvas.config(width=crop.width, height=crop.height)
            canvas.unbind("<ButtonPress-1>")
            canvas.unbind("<ButtonRelease-1>")
            canvas.unbind("<B1-Motion>")
            root1.move_window(canvas)
            root1.bind("<Double-Button-1>",lambda e:root1.iconify())


        def size_change():
            nonlocal n
            def n_set():
                nonlocal n
                n=int(entry.get())/100
                root2.destroy()
                triming()
            root2= Free_window()
            root2.attributes('-topmost', '1')
            frame2=ttk.Frame(root2, padding=3)
            label=ttk.Label(root2,text="拡大・縮小の倍率を入力してください(%)")
            entry=ttk.Entry(root2, width=10)
            button=ttk.Button(root2, text="OK", command= n_set)
            frame2.pack()
            label.pack(side=TOP)
            entry.pack(side=TOP)
            button.pack(side=TOP)

        def showMenu(e):
            pmenu.post(e.x_root, e.y_root)

        def img_get():
            nonlocal crop
            x=filedialog.asksaveasfilename(
            title = "名前を付けて保存",
            filetypes = [("PNG", ".png")],
            defaultextension = "png"
            )
            crop.save(x, format='PNG')
            messagebox.showinfo('保存完了', '保存しました')

        def img_clip():
            nonlocal crop
            try:
                original_image=crop
                output = io.BytesIO()
                make_folder(os.getcwd()+"/temp1/copy_image")
                original_image.save(os.getcwd()+"/temp1/copy_image/temp.png")
                original_image1 = Image.open(os.getcwd()+"/temp1/copy_image/temp.png")
                original_image1.convert('RGB').save(output, 'BMP')
                data = output.getvalue()[14:]
                output.close()
                win32clipboard.OpenClipboard()
                win32clipboard.EmptyClipboard()
                win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
                win32clipboard.CloseClipboard()
                messagebox.showinfo("完了","クリップボードにコピーしました")
            except:
                messagebox.showerror("エラー","クリップボードにコピーできませんでした")



        pmenu = Menu(root1, tearoff=0)
        root1.bind("<Button-3>", showMenu)
        crop=None
        n=1.0
        start_x=None
        start_y=None
        end_x=None
        end_y=None

        root1.update()

        canvas=Canvas(frame1,width=ws,height=hs)
        canvas.bind("<ButtonPress-1>", on_press)
        canvas.bind("<ButtonRelease-1>", on_release)
        canvas.bind("<B1-Motion>", on_move_press)

        pil_img=Image.open(os.getcwd()+"/temp1/img_memo/temp.png")
        canvas.config(width=pil_img.width, height=pil_img.height)
        canvas_img= ImageTk.PhotoImage(pil_img)
        canvas.canvas_img = canvas_img
        canvas.create_image(pil_img.width*0.5,pil_img.height*0.5,image=canvas_img)
        frame1.pack()
        canvas.pack(expand=True,fill=BOTH)

    if intermediate_screen==0:
        main_frame_delete()
        frame = ttk.Frame(root, padding=16)
        button=ttk.Button(frame,text='画像付箋作成',command=screen_memo1,padding=[80,10,80,10])
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        button.grid(row=1,column=0,columnspan=2)
    elif intermediate_screen==1:
        screen_memo1()

def short_url():
    global frame,buttonY
    def short_url1():
        x=clip.paste()
        api_url = "http://tinyurl.com/api-create.php"
        response = requests.get(api_url, params=dict(url=x))
        if response.ok:
            y=response.text.strip()
            clip.copy(y)
            messagebox.showinfo('短縮URL', '短縮URLをクリップボードにコピーしました')
        else:
            messagebox.showerror('エラー', '短縮URLを取得できませんでした')

    if intermediate_screen==0:
        main_frame_delete()
        frame = ttk.Frame(root, padding=16)
        button=ttk.Button(frame,text='短縮URL',command=short_url1,padding=[80,10,80,10])
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        button.grid(row=1,column=0,columnspan=2)
    elif intermediate_screen==1:
        short_url1()

def window_path():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def stopping():
        listener.stop()
        t.join()
        frame_delete()

    def main_frame1(num):
        listener.stop()
        t.join()
        root.protocol("WM_DELETE_WINDOW", frame_delete)
        main_frame(num)

    def key_check():
        nonlocal listener
        with keyboard.Listener(on_release=on_release) as listener:
            listener.join()

    def get_window_process_path(hwnd):
        # ウィンドウのプロセスID取得
        _, pid = win32process.GetWindowThreadProcessId(hwnd)
        # プロセスIDからプロセス取得
        process = psutil.Process(pid)
        # プロセスの実行ファイルのパス取得
        path = process.exe()

        return path

    def on_release(key):
        nonlocal listener
        if key == keyboard.Key.pause:
            # アクティブなウィンドウのハンドル
            hwnd = win32gui.GetForegroundWindow()
            window_title = win32gui.GetWindowText(hwnd)
            if window_title != "":
                path = get_window_process_path(hwnd)
                if os.path.exists(path):
                    # エクスプローラーを開いて、ファイル選択
                    subprocess.Popen(['explorer.exe', '/select,', path]).wait()
                else:
                    messagebox.showerror('エラー', 'ウィンドウのパスを取得できませんでした')
            else:
                messagebox.showerror('エラー', 'ウィンドウのパスを取得できませんでした')


    root.protocol("WM_DELETE_WINDOW", stopping)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(3),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="pauseでウィンドウを実行している\nファイルを直接開きます",font=("Helvetica", 16))

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)
    listener=None
    t=threading.Thread(target=key_check,daemon=True)
    t.start()

def icon_image():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    class SHFILEINFO(ctypes.Structure):
        _fields_ = [("hIcon", ctypes.wintypes.HICON),
                    ("iIcon", ctypes.c_int),
                    ("dwAttributes", ctypes.c_ulong),
                    ("szDisplayName", ctypes.wintypes.WCHAR * 260),
                    ("szTypeName", ctypes.wintypes.WCHAR * 80)]

    SHGFI_ICON = 0x000000100
    SHGFI_USEFILEATTRIBUTES = 0x000000010
    FILE_ATTRIBUTE_NORMAL = 0x00000080

    shell32 = ctypes.WinDLL('shell32')
    shell32.SHGetFileInfoW.argtypes = [ctypes.wintypes.LPCWSTR, ctypes.wintypes.DWORD, ctypes.POINTER(SHFILEINFO), ctypes.wintypes.UINT, ctypes.wintypes.UINT]
    shell32.SHGetFileInfoW.restype = ctypes.wintypes.HANDLE

    def get_folder_icon(folder_path):
        shinfo = SHFILEINFO()
        retval = shell32.SHGetFileInfoW(
            folder_path, 0, ctypes.byref(shinfo),
            ctypes.sizeof(shinfo), SHGFI_ICON
        )
        if retval:
            return shinfo.hIcon
        else:
            return None

    def get_icon(ext):
        shinfo = SHFILEINFO()
        retval = shell32.SHGetFileInfoW(
            ext, FILE_ATTRIBUTE_NORMAL, ctypes.byref(shinfo),
            ctypes.sizeof(shinfo), SHGFI_ICON | SHGFI_USEFILEATTRIBUTES
        )
        if retval:
            return shinfo.hIcon
        else:
            return None

    def icon_to_image(icon, size=(32, 32)):
        hdc = ctypes.windll.user32.GetDC(0)
        hbmp = ctypes.windll.gdi32.CreateCompatibleBitmap(hdc, size[0], size[1])

        hmemdc = ctypes.windll.gdi32.CreateCompatibleDC(hdc)
        ctypes.windll.gdi32.SelectObject(hmemdc, hbmp)
        ctypes.windll.user32.DrawIconEx(hmemdc, 0, 0, icon, size[0], size[1], 0, None, 0x0003)
        ctypes.windll.user32.DestroyIcon(icon)
        ctypes.windll.gdi32.DeleteDC(hmemdc)

        bitmap_bits = ctypes.create_string_buffer(size[0] * size[1] * 4)

        # GetBitmapBitsを呼び出してバッファにビットを格納
        ctypes.windll.gdi32.GetBitmapBits(hbmp, ctypes.sizeof(bitmap_bits), bitmap_bits)

        # PILで画像を作成
        image = Image.frombuffer(
            'RGBA',
            size,
            bitmap_bits,
            'raw',
            'BGRA',
            0,
            1
        )
        ctypes.windll.gdi32.DeleteObject(hbmp)
        ctypes.windll.user32.ReleaseDC(0, hdc)
        return image


    def exe_drop(drop):
        #try:
            y=drop.data.replace("{","").replace("}","").replace("\\","/")
            folder_name=os.path.dirname(y)
            name=os.path.splitext(os.path.basename(y))[0]
            if os.path.isdir(y):
                icon=get_folder_icon("")
            else:
                icon = get_icon(y) #""でドライブアイコン
            if icon:
                img = icon_to_image(icon)
                img.save(f"{folder_name}/{name}.png")
                messagebox.showinfo('完了', 'アイコンを抽出しました')
            else:
                messagebox.showerror('エラー', 'アイコンを抽出できませんでした')

        #except:
        #    messagebox.showerror('エラー', 'エラーが発生しました')


    #def extract_icon_from_exe(exe_path, icon_index=0):
    #    # アイコンの大きさ
    #    icon_size = win32api.GetSystemMetrics(win32con.SM_CXICON), win32api.GetSystemMetrics(win32con.SM_CYICON)
    #    # アイコン抽出
    #    large, small = win32gui.ExtractIconEx(exe_path, icon_index)
    #    # 大きいアイコン
    #    x=filedialog.asksaveasfilename(
    #    title = "名前を付けて保存",
    #    filetypes = [("PNG", ".png")],
    #    defaultextension = "png",
    #    initialfile = "icon"
    #    )
    #    if large:
    #        # Bitmapオブジェクト
    #        hdc = win32ui.CreateDCFromHandle(win32gui.GetDC(0))
    #        hbmp = win32ui.CreateBitmap()
    #        hbmp.CreateCompatibleBitmap(hdc, icon_size[0], icon_size[1])
    #        hdc = hdc.CreateCompatibleDC()
    #        hdc.SelectObject(hbmp)
    #        hdc.DrawIcon((0, 0), large[0])
    #        hdc.DeleteDC()
    #        # PILのImageオブジェクトに変換
    #        bmpinfo = hbmp.GetInfo()
    #        bmpstr = hbmp.GetBitmapBits(True)
    #        icon_image = Image.frombuffer(
    #            'RGBA',
    #            (bmpinfo['bmWidth'], bmpinfo['bmHeight']),
    #            bmpstr, 'raw', 'BGRA', 0, 1
    #        )
    #        icon_image.save(x, "PNG")
    #    # リソース解放
    #    if small:
    #        win32gui.DestroyIcon(small[0])
    #    if large:
    #        win32gui.DestroyIcon(large[0])

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    label2=ttk.Label(frame,text="アイコン抽出するexeをここに\nドラッグ＆ドロップしてください",font=("Helvetica", 16))
    label2.drop_target_register(DND_FILES)
    label2.dnd_bind('<<Drop>>',exe_drop)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label2.grid(row=3,column=0,columnspan=2)

def sleep_stop():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def main_frame1(x):
        root.after_cancel(a)
        main_frame(x)

    def move_mouse(dx, dy):
        flags = win32con.MOUSEEVENTF_MOVE
        win32api.mouse_event(flags, dx, dy, 0, 0)

    def main():
        nonlocal a
        move_mouse_distance = 1
        move_mouse_distance_x = 1
        flag = True

        # マウスの座標取得
        x, y = win32gui.GetCursorPos()
        move_mouse(-move_mouse_distance, 0)
        move_mouse(move_mouse_distance, 0)
        if flag:
            x += move_mouse_distance_x
            flag = False
        else:
            x -= move_mouse_distance_x
            flag = True
        win32api.SetCursorPos((x, y))
        a=root.after(50000, main)

    a=None
    label=ttk.Label(frame,text="スリープを無効にしています。",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(3),font=("Helvetica", 7),bg="gray",fg="white")
    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)
    main()

def txt_get():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def txt_get1(drop):
        try:
            x=drop.data.replace("{","").replace("}","").replace("\\","/")
            f = open(x, 'r')
            data = f.read()
            clip.copy(data)
            f.close()
            messagebox.showinfo('完了', 'クリップボードにコピーしました')
        except:
            messagebox.showerror('エラー', 'テキスト取得を失敗しました')

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',txt_get1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def voice_cut():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def clip(drop):
        file=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(file)
        for x in x_list:
            try:
                y= os.path.basename(x)
                z=x.replace(y,"cut_"+y)
                start=60*int(entry_start1.get())+float(entry_start2.get())
                end=60*int(entry_end1.get())+float(entry_end2.get())
                audio1 = AudioFileClip(x)
                audio=audio1.subclip(start, end)
                audio.write_audiofile(z,logger=None)
                audio1.close()
            except:
                messagebox.showerror("エラー",f"{x}\nをうまく変換できませんでした")
        messagebox.showinfo("終了","カットを終了します")

    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',clip)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    label1=ttk.Label(frame,text="開始時刻")
    label2=ttk.Label(frame,text="終了時刻")
    label3=ttk.Label(frame,text="：")
    label4=ttk.Label(frame,text="：")
    entry_start1=ttk.Entry(frame,width=5)
    entry_start2=ttk.Entry(frame,width=5)
    entry_end1=ttk.Entry(frame,width=5)
    entry_end2=ttk.Entry(frame,width=5)
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=2)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry_start1.grid(row=1,column=1)
    label3.grid(row=1,column=2)
    entry_start2.grid(row=1,column=3)
    label2.grid(row=2,column=0)
    entry_end1.grid(row=2,column=1)
    label4.grid(row=2,column=2)
    entry_end2.grid(row=2,column=3)
    label.grid(row=3,column=0,columnspan=4)

def dumy_human():
    root1 = Free_window()
    root1.title('ダミー作成結果')
    frame1=ttk.Frame(root1, padding=0)
    root1.attributes('-topmost', '1')
    user = RandomUser({'nat': 'US'})
    img = Image.open(io.BytesIO(requests.get(user.get_picture(size='large')).content))

    def save_dumy():
        path=filedialog.asksaveasfilename(
            title = "名前を付けて保存",
            filetypes = [("Text",".txt") ],
            defaultextension = "txt"
            )
        f = open(path, 'w')
        f.writelines([
            f"顔写真：{user.get_picture(size='large')}\n",
            f"First Name：{user.get_first_name()}\n",
            f"Last Name：{user.get_last_name()}\n",
            f"性別：{gender(user.get_gender())}\n",
            f"誕生日：{convert_date(user.get_dob())}\n",
            f"年齢：{user.get_age()}\n",
            f"住所：{user.get_street()}\n",
            f"国籍：{user.get_country()}\n",
            f"ユーザー名：{user.get_username()}\n"
        ])
        f.close()
        messagebox.showinfo('完了', '保存しました')


    def gender(x):
        if x=="male":
            return "男性"
        elif x=="female":
            return "女性"

    def convert_date(date_str):
        date_str=date_str[:-5]
        datetime_obj = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S")
        converted_date = datetime_obj.strftime("%Y-%m-%d")
        return converted_date

    canvas=Canvas(frame1,width=128,height=128)
    num = Menu(canvas, tearoff=0)
    num.add_command(label="コピー", command=lambda: clip.copy(user.get_picture(size='large')))
    canvas.bind("<Button-3>", lambda e:num.post(e.x_root, e.y_root))

    label_r_first=ttk.Label(frame1,text="ファーストネーム：")
    label_r_last=ttk.Label(frame1,text="ラストネーム：")
    label_r_gender=ttk.Label(frame1,text="性別：")
    label_r_birthday=ttk.Label(frame1,text="誕生日：")
    label_r_age=ttk.Label(frame1,text="年齢：")
    label_r_address=ttk.Label(frame1,text="住所：")
    label_r_country=ttk.Label(frame1,text="国籍：")
    label_r_user=ttk.Label(frame1,text="ユーザー名：")
    label_l_first=ttk.Label(frame1,text=user.get_first_name())
    num1 = Menu(label_l_first, tearoff=0)
    num1.add_command(label="コピー", command=lambda: clip.copy(user.get_first_name()))
    label_l_first.bind("<Button-3>", lambda e:num1.post(e.x_root, e.y_root))

    label_l_last=ttk.Label(frame1,text=user.get_last_name())
    num2 = Menu(label_l_first, tearoff=0)
    num2.add_command(label="コピー", command=lambda: clip.copy(user.get_last_name()))
    label_l_last.bind("<Button-3>", lambda e:num2.post(e.x_root, e.y_root))

    label_l_gender=ttk.Label(frame1,text=gender(user.get_gender()))
    num3 = Menu(label_l_gender, tearoff=0)
    num3.add_command(label="コピー", command=lambda: clip.copy(gender(user.get_gender())))
    label_l_gender.bind("<Button-3>", lambda e:num3.post(e.x_root, e.y_root))

    label_l_birthday=ttk.Label(frame1,text=convert_date(user.get_dob()))
    num4 = Menu(label_l_birthday, tearoff=0)
    num4.add_command(label="コピー", command=lambda: clip.copy(convert_date(user.get_dob())))
    label_l_birthday.bind("<Button-3>", lambda e:num4.post(e.x_root, e.y_root))

    label_l_age=ttk.Label(frame1,text=user.get_age())
    num5 = Menu(label_l_age, tearoff=0)
    num5.add_command(label="コピー", command=lambda: clip.copy(user.get_age()))
    label_l_age.bind("<Button-3>", lambda e:num5.post(e.x_root, e.y_root))

    label_l_address=ttk.Label(frame1,text=user.get_street())
    num6 = Menu(label_l_address, tearoff=0)
    num6.add_command(label="コピー", command=lambda: clip.copy(user.get_street()))
    label_l_address.bind("<Button-3>", lambda e:num6.post(e.x_root, e.y_root))

    label_l_country=ttk.Label(frame1,text=user.get_country())
    num7 = Menu(label_l_country, tearoff=0)
    num7.add_command(label="コピー", command=lambda: clip.copy(user.get_country()))
    label_l_country.bind("<Button-3>", lambda e:num7.post(e.x_root, e.y_root))

    label_l_user=ttk.Label(frame1,text=user.get_username())
    num8 = Menu(label_l_user, tearoff=0)
    num8.add_command(label="コピー", command=lambda: clip.copy(user.get_username()))
    label_l_user.bind("<Button-3>", lambda e:num8.post(e.x_root, e.y_root))

    canvas_img= ImageTk.PhotoImage(img)
    canvas.canvas_img = canvas_img
    canvas.create_image(64,64,image=canvas_img)

    Button=ttk.Button(frame1,text="保存",command=save_dumy)

    frame1.pack()
    canvas.grid(row=0,column=0,columnspan=2)
    label_r_first.grid(row=1,column=0)
    label_l_first.grid(row=1,column=1)
    label_r_last.grid(row=2,column=0)
    label_l_last.grid(row=2,column=1)
    label_r_gender.grid(row=3,column=0)
    label_l_gender.grid(row=3,column=1)
    label_r_birthday.grid(row=4,column=0)
    label_l_birthday.grid(row=4,column=1)
    label_r_age.grid(row=5,column=0)
    label_l_age.grid(row=5,column=1)
    label_r_address.grid(row=6,column=0)
    label_l_address.grid(row=6,column=1)
    label_r_country.grid(row=7,column=0)
    label_l_country.grid(row=7,column=1)
    label_r_user.grid(row=8,column=0)
    label_l_user.grid(row=8,column=1)
    Button.grid(row=9,column=0,columnspan=2)

def clock():
    root1 = Free_window()
    root1.title('時計')
    frame_1=ttk.Frame(root1, padding=0)
    root1.attributes('-topmost', '1')
    def day_memo():
        now=datetime.now()
        clip.copy(now.strftime('%Y/%m/%d'))

    def time_memo():
        now=datetime.now()
        clip.copy(now.strftime('%H:%M:%S'))

    def update_label():
        if root1.winfo_exists():
            now=datetime.now()
            if now.weekday()==0:
                week="月"
            elif now.weekday()==1:
                week="火"
            elif now.weekday()==2:
                week="水"
            elif now.weekday()==3:
                week="木"
            elif now.weekday()==4:
                week="金"
            elif now.weekday()==5:
                week="土"
            elif now.weekday()==6:
                week="日"
            label1.config(text=f"{now.strftime('%Y/%m/%d')}\n{now.strftime('%H:%M:%S')}({week})")
            root.after(50, update_label)

    def showMenu(e):
        pmenu.post(e.x_root, e.y_root)

    pmenu = Menu(root1, tearoff=0)
    pmenu.add_command(label="年月日", command=day_memo)
    pmenu.add_command(label="時間", command=time_memo)
    root1.bind("<Button-3>", showMenu)

    label1=ttk.Label(frame_1,font=("Helvetica", 28))
    update_label()
    frame_1.pack()
    label1.pack(side=TOP)

def num_change():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def change1():
        entry10.delete(0, END)
        entry2.delete(0, END)
        entry8.delete(0, END)
        entry16.delete(0, END)
        try:
            if var.get()==10:
                entry10.insert(0,entry_input.get())
                entry2.insert(0,str(bin(int(entry_input.get()))).replace("0b",""))
                entry8.insert(0,str(oct(int(entry_input.get()))).replace("0o",""))
                entry16.insert(0,str(hex(int(entry_input.get()))).replace("0x",""))
            elif var.get()==2:
                entry10.insert(0,str(int(entry_input.get(),2)))
                entry2.insert(0,entry_input.get())
                entry8.insert(0,str(oct(int(entry_input.get(),2))).replace("0o",""))
                entry16.insert(0,str(hex(int(entry_input.get(),2))).replace("0x",""))
            elif var.get()==8:
                entry10.insert(0,str(int(entry_input.get(),8)))
                entry2.insert(0,str(bin(int(entry_input.get(),8))).replace("0b",""))
                entry8.insert(0,entry_input.get())
                entry16.insert(0,str(hex(int(entry_input.get(),8))).replace("0x",""))
            elif var.get()==16:
                entry10.insert(0,str(int(entry_input.get(),16)))
                entry2.insert(0,str(bin(int(entry_input.get(),16))).replace("0b",""))
                entry8.insert(0,str(oct(int(entry_input.get(),16))).replace("0o",""))
                entry16.insert(0,entry_input.get())
        except:
            messagebox.showerror('エラー', '変換できませんでした')

    var=IntVar()
    radio10=ttk.Radiobutton(frame,text="10進数",value=10,variable=var)
    radio16=ttk.Radiobutton(frame,text="16進数",value=16,variable=var)
    radio2=ttk.Radiobutton(frame,text="2進数",value=2,variable=var)
    radio8= ttk.Radiobutton(frame,text="8進数",value=8,variable=var)
    label10=ttk.Label(frame,text="10進数")
    label16=ttk.Label(frame,text="16進数")
    label2=ttk.Label(frame,text="2進数")
    label8=ttk.Label(frame,text="8進数")
    entry10=ttk.Entry(frame,width=30)
    entry16=ttk.Entry(frame,width=30)
    entry2=ttk.Entry(frame,width=30)
    entry8=ttk.Entry(frame,width=30)
    label_input=ttk.Label(frame,text="変換する数値：")
    button=ttk.Button(frame,text="変換",command=change1)
    entry_input=ttk.Entry(frame,width=30)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    var.set(10)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label_input.grid(row=1,column=0)
    entry_input.grid(row=1,column=1)
    radio10.grid(row=2,column=0)
    radio2.grid(row=2,column=1)
    radio8.grid(row=3,column=0)
    radio16.grid(row=3,column=1)
    label10.grid(row=4,column=0)
    label2.grid(row=4,column=1)
    label8.grid(row=6,column=0)
    label16.grid(row=6,column=1)
    entry10.grid(row=5,column=0)
    entry2.grid(row=5,column=1)
    entry8.grid(row=7,column=0)
    entry16.grid(row=7,column=1)
    button.grid(row=8,column=0,columnspan=2)

def sum_paste():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def clipboard_scan():
        nonlocal current,a
        update=clip.paste()
        if update!=current:
            last_line_index = box.index(END)
            next_line_index = f"{last_line_index}+1line"
            box.insert(next_line_index, update.replace("\n","\\n")+'\n')
            current=update
        a=root.after(5, clipboard_scan)

    def sum_paste1():
        nonlocal current
        x=box.get(0.0,END).rstrip("\n").replace("\\n","\n")
        y=x.replace("\n",entry.get())
        root.after_cancel(a)
        clip.copy(y)
        box.delete(0.0,END)
        messagebox.showinfo('完了', 'クリップボードにコピーしました')
        current = clip.paste()
        clipboard_scan()

    def main_frame1():
        root.after_cancel(a)
        main_frame(0)

    def temp_stop():
        nonlocal a,current
        if stop_num.get()==1:
            root.after_cancel(a)
        else:
            current=clip.paste()
            clipboard_scan()

    a=None
    box=ScrolledText(frame,width=50,height=10,wrap=NONE)
    button=ttk.Button(frame,text="合体",command=sum_paste1)
    current = clip.paste()
    entry=ttk.Entry(frame,width=30)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="区切り文字")
    stop_num=IntVar()
    checkbox=ttk.Checkbutton(frame,text="一時停止",onvalue=1,offvalue=0,variable=stop_num,command=temp_stop)
    clipboard_scan()

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    box.grid(row=1,column=0,columnspan=2)
    label.grid(row=2,column=0)
    entry.grid(row=2,column=1)
    button.grid(row=3,column=1)
    checkbox.grid(row=3,column=0)

def copy_paste():
    root1 = Free_window()
    root1.title('連続コピペ')
    frame_1=ttk.Frame(root1, padding=0)
    root1.attributes('-topmost', '1')
    t_end=0
    lock = threading.Lock()

    def frame_delete12():
        nonlocal t_end
        button45.config(state="normal")
        t_end=1
        t1.join()
        root1.after(10, root1.destroy)
    root1.protocol("WM_DELETE_WINDOW", frame_delete12)

    def paste_sub():
        nonlocal current,t_end
        def on_key_event(event):
            nonlocal current
            if event.name == 'v' and key.is_pressed('ctrl'):
                lock.acquire()
                time.sleep(0.001)
                if fifo.get()==0:
                    box.delete(1.0,"1.end")
                    box_sub=box.get('2.0',END).rstrip("\n")
                    box.delete(1.0,END)
                    box.insert(1.0,box_sub)
                elif fifo.get()==1:
                    box.delete('end-1c linestart','end')
                lock.release()
            change1()
        key.on_press(on_key_event)
        while t_end==0:
            update=clip.paste()
            time.sleep(0.03)
            if update!=current and update==clip.paste():
                if box.get(1.0,"1.end")=="":
                    next_line_index="1.0"
                    box.insert(next_line_index, update.replace("\n","\\n"))
                else :
                    last_line_index = box.index(END)
                    next_line_index = f"{last_line_index}+1line"
                    box.insert(next_line_index, "\n"+update.replace("\n","\\n"))
                time.sleep(0.01)
                change1()
            time.sleep(0.03)
        key.unhook_all()


    def change1():
        nonlocal current
        lock.acquire()
        if fifo.get()==0:
            current=box.get(1.0,"1.end").rstrip("\n").replace("\\n","\n")
            clip.copy(current)

        elif fifo.get()==1:
            last_line_start = box.index('end-1c linestart')
            last_line_end = box.index('end-1c lineend')
            current=box.get(last_line_start, last_line_end).rstrip("\n").replace("\\n","\n")
            clip.copy(current)
        lock.release()

    def change2():
        check["state"]="disabled"
        root1.after(1000, lambda:check.config(state="normal"))
        change1()

    button45.config(state="disabled")
    current = clip.paste()
    box=ScrolledText(frame_1,width=50,height=10,wrap=NONE,font=("メイリオ", 10))
    fifo=IntVar()
    check=ttk.Checkbutton(frame_1,text="逆順からペースト",onvalue=1,offvalue=0,variable=fifo,command=change2)

    t1=threading.Thread(target=paste_sub,daemon=True)
    t1.start()

    frame_1.pack()
    check.pack(side=TOP)
    box.pack(side=TOP,fill=BOTH,expand=True)

def length_ruler():
    global frame,buttonY
    def length_ruler1():
        root.iconify()
        make_folder(os.getcwd()+"/temp1/length")
        pyautogui.screenshot(os.getcwd()+"/temp1/length/temp.png")
        root.deiconify()
        root1 = Free_window()
        root1.title('長さ測定')
        frame1=ttk.Frame(root1, padding=0)
        root1.attributes('-fullscreen', True)
        root1.attributes('-topmost', '1')

        def on_press(event):
            canvas.delete("rect")
            nonlocal start_x
            nonlocal start_y
            start_x = canvas.canvasx(event.x)
            start_y = canvas.canvasy(event.y)

        def on_move_press(event):
            nonlocal end_x, end_y
            end_x, end_y = event.x, event.y
            canvas.delete("rect")
            canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)

        def on_release(event):
            nonlocal end_x, end_y
            end_x, end_y = event.x, event.y
            canvas.delete("rect")
            canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)
            triming()

        def triming():
            x=abs(-start_x+end_x)
            y=abs(-start_y+end_y)
            root1.attributes('-fullscreen', False)
            canvas.unbind("<ButtonPress-1>")
            canvas.unbind("<ButtonRelease-1>")
            canvas.unbind("<B1-Motion>")
            root1.destroy()
            messagebox.showinfo('長さ', f'横幅：{x} px\n縦幅：{y} px\n斜め：{round(math.sqrt(x**2+y**2),2)} px')

        start_x=None
        start_y=None
        end_x=None
        end_y=None

        root1.update()
        canvas=Canvas(frame1,width=ws,height=hs)
        canvas.bind("<ButtonPress-1>", on_press)
        canvas.bind("<ButtonRelease-1>", on_release)
        canvas.bind("<B1-Motion>", on_move_press)

        pil_img=Image.open(os.getcwd()+"/temp1/length/temp.png")
        canvas.config(width=pil_img.width, height=pil_img.height)
        canvas_img= ImageTk.PhotoImage(pil_img)
        canvas.canvas_img = canvas_img
        canvas.create_image(pil_img.width*0.5,pil_img.height*0.5,image=canvas_img)
        frame1.pack()
        canvas.pack(expand=True,fill=BOTH)

    main_frame_delete()
    frame = ttk.Frame(root, padding=16)
    button=ttk.Button(frame,text='長さ測定',command=length_ruler1,padding=[80,10,80,10])
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    button.grid(row=1,column=0,columnspan=2)

def voice_read():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)


    def voice_read1():
        button1.config(state="disabled")
        button2.config(state="disabled")
        button2.config(state="disabled")
        delete_folder(os.getcwd()+"/temp1/ai_voice")
        tts = gTTS(box.get(0.0,END), lang=lang.get())
        make_folder(os.getcwd()+"/temp1/ai_voice")
        tts.save(os.getcwd()+"/temp1/ai_voice/temp.mp3")
        button1.config(state="normal")
        button2.config(state="normal")
        button3.config(state="normal")


    def voice_dl():
        path=filedialog.asksaveasfilename(
            title = "名前を付けて保存",
            filetypes = [("mp3",".mp3") ],
            defaultextension = "mp3"
            )
        shutil.copy(os.getcwd()+"\\temp1\\ai_voice\\temp.mp3",path)
        messagebox.showinfo('完了', '保存しました')

    def voice_up():
        playsound(os.path.join(os.getcwd(), "temp1", "ai_voice", "temp.mp3"))


    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    box=ScrolledText(frame,width=50,height=30)
    lang=StringVar()
    lang.set("ja")
    radio1=ttk.Radiobutton(frame,text="日本語",value="ja",variable=lang)
    radio2=ttk.Radiobutton(frame,text="英語",value="en",variable=lang)
    button1=ttk.Button(frame,text="読み取り",command=voice_read1)
    label1=ttk.Label(frame,text="読み上げる文章")
    button2=ttk.Button(frame,text="音声を取得",command=voice_dl,state="disabled")
    button3=ttk.Button(frame,text="再生",command=voice_up,state="disabled")

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=2)
    box.grid(row=2,column=0,columnspan=3)
    label1.grid(row=1,column=1)
    radio1.grid(row=3,column=0)
    radio2.grid(row=3,column=2)
    button1.grid(row=4,column=0)
    button2.grid(row=4,column=2)
    button3.grid(row=4,column=1)

def password():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def get_string(num):
        random_letters = random.choices(string.ascii_letters, k=num)
        random_string = ''.join(random_letters)
        return random_string

    def show_pass():
        def show_pass1():
            entry2.config(show="*")
            entry3.config(show="*")
            button2.config(state="normal")
        entry2.config(show="")
        entry3.config(show="")
        button2.config(state="disabled")
        root.after(5000, show_pass1)


    def url_change():
        if entry1.get()=="":
            domain=get_string(10)
        else:
            parsed_url = urlparse(entry1.get())
            domain = parsed_url.netloc
        return domain

    def hash_num(number, length):
        hash_object = hashlib.sha256(str(number).encode())
        hash_digest = hash_object.hexdigest()
        alpha_num = hash_digest[:length]
        return alpha_num

    def hash_gen(domain,bir):
        i=0
        if entry2.get()=="":
            pass1=get_string(10)
        else:
            pass1=entry2.get()
        hash=domain+pass1+hash_num(spin.get(),8)+bir
        while i<11:
            hash = hashlib.sha512( str( hash ).encode("utf-8") ).hexdigest()
            i+=1
        return hash

    def text_cut(hash):
        decimal_value = num_value=int(hash, 16)
        alphabet_value = ""
        while decimal_value > 0:
            decimal_value, remainder = divmod(decimal_value , 26)
            alphabet_value = chr(ord('a') + remainder) + alphabet_value
        min_length = min(len(str(num_value)), len(alphabet_value))
        num= str(num_value)[:min_length]
        tex= alphabet_value[:min_length]
        return num, tex

    def text_sum(x,y,bir2):
        random.seed(bir2)
        merged=y[0]+x[0]+"".join(random.sample(x+y,len(x+y)))
        random.seed(None)
        return merged

    def generate():
        if birth.get()=="":
            if entry2.get()=="":
                birn=get_string(8)
            else:
                birn=entry2.get()+url_change()
        else:
            birn=birth.get()
        bir=hash_num(birn,8)
        bir2=text_cut(bir)[0]
        domain=url_change()
        hash=hash_gen(domain,bir)
        num,tex=text_cut(hash)
        merged=text_sum(num,tex,bir2)
        if Oomoji1.get()==1:
            random.seed(bir2)
            merged=''.join(random.choice([char.upper(), char]) for char in merged)
            random.seed(None)
            merged=merged[0].upper()+merged[1:]
        if kigou1.get()==1:
            pass_kigou = ["@","_","."]
            if entry2.get()=="":
                pass_seed=get_string(10)
            else:
                pass_seed=entry2.get()+url_change()
            seed_num=int(str(int(hash_num(pass_seed,10),16))[3])
            random.seed(seed_num)
            merged=random.choice(pass_kigou)+merged
            random.seed(None)
        nopass=merged[:var.get()]
        random.seed(bir2)
        password="".join(random.sample(nopass,len(nopass)))
        random.seed(None)
        clip.copy(password)
        messagebox.showinfo('完了', 'クリップボードにコピーしました')


    var=IntVar()
    var.set(8)
    radio8=ttk.Radiobutton(frame,text="8桁",value=8,variable=var)
    radio12=ttk.Radiobutton(frame,text="12桁",value=12,variable=var)
    radio16=ttk.Radiobutton(frame,text="16桁",value=16,variable=var)
    radio24=ttk.Radiobutton(frame,text="24桁",value=24,variable=var)
    Oomoji1=IntVar()
    Oomoji=ttk.Checkbutton(frame,text="大文字を含める",onvalue=1,offvalue=0,variable=Oomoji1)
    kigou1=IntVar()
    kigou=ttk.Checkbutton(frame,text="記号を含める",onvalue=1,offvalue=0,variable=kigou1)
    label1=ttk.Label(frame,text="URL：")
    entry1=ttk.Entry(frame,width=30)
    entry2=ttk.Entry(frame,width=30,show="*")
    label2=ttk.Label(frame,text="マスターパスワード：")
    button=ttk.Button(frame,text="生成",command=generate)
    label3=ttk.Label(frame,text="チェンジ：")
    spin=IntVar()
    spin.set(1)
    Spinbox=ttk.Spinbox(frame,from_=1,to=10,increment=1,textvariable=spin)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    label4=ttk.Label(frame,text="第二パスワード(任意)")
    birth=StringVar()
    entry3=ttk.Entry(frame,width=30,textvariable=birth,show="*")
    label5=ttk.Label(frame,text="注意：パスワードに空白や記号を入れないように")
    button2=ttk.Button(frame,text="パスワードを一時表示",command=show_pass)

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    radio8.grid(row=3,column=0)
    radio12.grid(row=3,column=1)
    radio16.grid(row=4,column=0)
    radio24.grid(row=4,column=1)
    Oomoji.grid(row=5,column=0)
    kigou.grid(row=5,column=1)
    label3.grid(row=6,column=0)
    Spinbox.grid(row=6,column=1)
    label4.grid(row=7,column=0)
    entry3.grid(row=7,column=1)
    label5.grid(row=8,column=0,columnspan=2)
    button.grid(row=9,column=1)
    button2.grid(row=9,column=0)

def color_bullet():
    a=colorchooser.askcolor(title="色を選択")
    root1=Free_window()
    x=(ws/2)-450
    y=(hs/2)-300
    root1.geometry('+%d+%d'%(x,y))
    root1.title("カラーピッカー")
    root1.attributes('-topmost', '1')
    frame_1=ttk.Frame(root1, padding=0)
    def color_bullet1():
        root1.destroy()
        color_bullet()

    button1=ttk.Button(frame_1,text="コピー",command=lambda:clip.copy(f"{a[0]}".replace("(","").replace(")","")))
    button2=ttk.Button(frame_1,text="コピー",command=lambda:clip.copy(a[1]))
    label1=ttk.Label(frame_1,text=f"RGB：{a[0]}",width=25)
    label2=ttk.Label(frame_1,text=f"RRGBGB：{a[1]}",width=25)
    button3=ttk.Button(frame_1,text="再選択",command=color_bullet1)

    frame_1.pack()
    button1.grid(row=0,column=1)
    label1.grid(row=0,column=0)
    button2.grid(row=1,column=1)
    label2.grid(row=1,column=0)
    button3.grid(row=2,column=0,columnspan=2)

def file_day():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def dnd(drop):
        d_list2=entry2.get().split(",")
        d_list3=entry3.get().split(",")
        file=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(file)
        for i in x_list:
            if entry2.get()!="":
                set_time2 = datetime(int(d_list2[0]), int(d_list2[1]), int(d_list2[2]),int(d_list2[3]), int(d_list2[4]), int(d_list2[5])).timestamp()
                sr = os.stat(path=i)
                os.utime(i, (sr.st_atime, set_time2))
            if entry3.get()!="":
                set_time3 = datetime(int(d_list3[0]), int(d_list3[1]), int(d_list3[2]),int(d_list3[3]), int(d_list3[4]), int(d_list3[5])).timestamp()
                win32_setctime.setctime(i, set_time3)
        messagebox.showinfo('完了', '変更しました')

    label4=ttk.Label(frame,text="更新日時：")
    entry2=ttk.Entry(frame,width=30)
    entry3=ttk.Entry(frame,width=30)
    label5=ttk.Label(frame,text="作成日時：")
    label1=ttk.Label(frame,text="yyyy,mm,dd,hh,mm,ssで入力してください")
    label2=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',dnd)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label4.grid(row=1,column=0)
    entry2.grid(row=1,column=1)
    label5.grid(row=2,column=0)
    entry3.grid(row=2,column=1)
    label1.grid(row=3,column=0,columnspan=2)
    label2.grid(row=4,column=0,columnspan=2)

def mp3_wav():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def mp3_wav1(drop):
        file=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(file)
        for x in file_list:
            try:
                clip=AudioFileClip(x)
                name=os.path.basename(x)
                youso=os.path.splitext(name)[1]
                if youso==".mp3":
                    clip.write_audiofile(x.replace(youso,".wav"),logger=None)
                elif youso==".wav":
                    clip.write_audiofile(x.replace(youso,".mp3"),logger=None)
            except:
                messagebox.showerror('エラー', f'{x}\nは変換できませんでした')
        messagebox.showinfo('完了', '終了しました')

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="ここにファイルを\nドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',mp3_wav1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def pc_info():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def main_frame1(n):
        root.after_cancel(a)
        main_frame(n)

    def pc_info1():
        nonlocal a
        root.after_cancel(a)
        label3.config(text=f"CPU使用率：{psutil.cpu_percent(interval=1)}%")
        label4.config(text=f"メモリ使用率：{psutil.virtual_memory().percent}%")
        label5.config(text=f"ディスク使用率：{psutil.disk_usage(path='/').percent}%")
        root.update()
        a=root.after(30000,pc_info1)

    label1=ttk.Label(frame,text=f"物理コア数：{psutil.cpu_count(logical=False)}")
    label2=ttk.Label(frame,text=f"論理コア数：{psutil.cpu_count(logical=True)}")
    label3=ttk.Label(frame,text=f"CPU使用率：{psutil.cpu_percent(interval=1)}%")
    label4=ttk.Label(frame,text=f"メモリ使用率：{psutil.virtual_memory().percent}%")
    label5=ttk.Label(frame,text=f"ディスク使用率：{psutil.disk_usage(path='/').percent}%")
    button=ttk.Button(frame,text="更新",command=pc_info1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0,columnspan=2)
    label2.grid(row=2,column=0,columnspan=2)
    label3.grid(row=3,column=0,columnspan=2)
    label4.grid(row=4,column=0,columnspan=2)
    label5.grid(row=5,column=0,columnspan=2)
    button.grid(row=6,column=0,columnspan=2)
    a=root.after(30000,pc_info1)

def gif_create():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)
    def add_img(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            if box.get("1.0","end-1c")!="":
                box.insert("end","\n")
            box.insert("end",x)

    def gif_create1():
        try:
            f_path=filedialog.asksaveasfilename(
                title = "名前を付けて保存",
                filetypes = [("gif",".gif") ],
                defaultextension = "gif"
                )
            file_list=box.get("1.0","end-1c").split("\n")
            x_list=[]
            for path in file_list:
                if path!="":
                    x_list.append(Image.open(path))
            x_list[0].save(f_path, save_all=True, append_images=x_list[1:], duration=1000/int(entry.get()), loop=0,optimize=False)
            messagebox.showinfo('完了', '作成しました')
        except:
            messagebox.showerror('エラー', '作成できませんでした')

    box=ScrolledText(frame,width=30,height=20,wrap=NONE)
    label1=ttk.Label(frame,text="ここに画像ファイルを\nドロップしてください",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    button=ttk.Button(frame,text="作成",command=gif_create1)
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',add_img)
    entry=ttk.Entry(frame,width=30)
    label_1=ttk.Label(frame,text="FPS：")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label_1.grid(row=1,column=0)
    entry.grid(row=1,column=1)
    label1.grid(row=2,column=0,columnspan=2)
    box.grid(row=3,column=0,columnspan=2)
    button.grid(row=4,column=0,columnspan=2)

def img_blank():
    global frame, buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def add_margin(drop):
        try:
            if var.get() == 1:
                color = (255, 255, 255)  # 白
            elif var.get() == 2:
                color = (0, 0, 0)  # 黒
            elif var.get() == 3:
                color = None  # 透明
            width = int(entry_w.get())
            height = int(entry_h.get())
            files = drop.data.replace("{", "").replace("}", "").replace("\\", "/")
            file_list = file_mult(files)
            for x in file_list:
                name = os.path.basename(x)
                pil_img = Image.open(x)
                name_1=os.path.splitext(name)[0]
                old_w, old_h = pil_img.size
                if old_w <= width and old_h <= height:
                    if color is None:
                        result = Image.new("RGBA", (width, height), (0, 0, 0, 0))
                    else:
                        result = Image.new(pil_img.mode, (width, height), color)
                    result.paste(pil_img, (int((width - old_w) / 2), int((height - old_h) / 2)))
                    result.save(os.path.dirname(x) + "/blank_" + name_1+".png")
                else:
                    messagebox.showerror('エラー', '元画像より大きくしてください')
        except:
            messagebox.showerror('エラー', '作成できませんでした')
        messagebox.showinfo('完了', '終了しました')

    label_w = ttk.Label(frame, text="枠追加後横幅(px)：")
    label_h = ttk.Label(frame, text="枠追加後縦幅(px)：")
    entry_w = ttk.Entry(frame, width=30)
    entry_h = ttk.Entry(frame, width=30)
    label = ttk.Label(frame, text="ここに画像ファイルを\nドロップしてください", font=("Helvetica", 16))
    buttonX = ttk.Checkbutton(frame, text=main_checkbox_name, onvalue=1, offvalue=0, variable=window_front, command=execute)
    buttonY = Button(frame, text="戻る", command=lambda: main_frame(2), font=("Helvetica", 7), bg="gray", fg="white")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>', add_margin)
    var = IntVar()
    var.set(1)
    color_button1 = ttk.Radiobutton(frame, text="白", variable=var, value=1)
    color_button2 = ttk.Radiobutton(frame, text="黒", variable=var, value=2)
    color_button3 = ttk.Radiobutton(frame, text="透明", variable=var, value=3)

    frame.pack()
    buttonX.grid(row=0, column=2)
    buttonY.grid(row=0, column=0)
    label_w.grid(row=1, column=0)
    label_h.grid(row=2, column=0)
    entry_w.grid(row=1, column=2)
    entry_h.grid(row=2, column=2)
    color_button1.grid(row=3, column=0)
    color_button2.grid(row=3, column=1)
    color_button3.grid(row=3, column=2)
    label.grid(row=5, column=0, columnspan=3)

def image_set():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def img_set(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                img=Image.open(x)
                name=os.path.basename(x)
                con = ImageEnhance.Brightness(img)
                result1 = con.enhance(float(entry_con.get()))
                kido = ImageEnhance.Brightness(result1)
                result2 = kido.enhance(float(entry_kido.get()))
                sat=ImageEnhance.Color(result2)
                result3=sat.enhance(float(entry_sat.get()))
                shr=ImageEnhance.Sharpness(result3)
                result4=shr.enhance(float(entry_shr.get()))
                result4.save(os.path.dirname(x)+"/"+"set_"+name)
            except:
                messagebox.showerror('エラー', f'{x}\nは作成できませんでした')
        messagebox.showinfo('完了', '終了しました')

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    label_con=ttk.Label(frame,text="コントラスト：")
    entry_con=ttk.Entry(frame,width=5)
    label_kido=ttk.Label(frame,text="輝度：")
    entry_kido=ttk.Entry(frame,width=5)
    label_sat=ttk.Label(frame,text="彩度：")
    entry_sat=ttk.Entry(frame,width=5)
    label_shr=ttk.Label(frame,text="シャープネス：")
    entry_shr=ttk.Entry(frame,width=5)
    entry_con.insert(END,1.0)
    entry_kido.insert(END,1.0)
    entry_sat.insert(END,1.0)
    entry_shr.insert(END,1.0)
    label=ttk.Label(frame,text="ここに画像ファイルを\nドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',img_set)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label_con.grid(row=1,column=0)
    entry_con.grid(row=1,column=1)
    label_kido.grid(row=2,column=0)
    entry_kido.grid(row=2,column=1)
    label_sat.grid(row=3,column=0)
    entry_sat.grid(row=3,column=1)
    label_shr.grid(row=4,column=0)
    entry_shr.grid(row=4,column=1)
    label.grid(row=5,column=0,columnspan=2)

def img_gamma():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def has_alpha_channel(img):
        return img.mode == 'RGBA' or (img.mode == 'P' and 'transparency' in img.info)

    def gamma_table_alpha(gamma1):
        a_tbl = [x for x in range(256)]
        r_tbl = [min(255, int((x / 255.) ** (1. / gamma1) * 255.)) for x in range(256)]
        return r_tbl*3+a_tbl

    def gamma_table(gamma1):
        r_tbl = [min(255, int((x / 255.) ** (1. / gamma1) * 255.)) for x in range(256)]
        return r_tbl*3

    def gamma(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                img=Image.open(x)
                name=os.path.basename(x)
                if has_alpha_channel(img):
                    result=img.point(gamma_table_alpha(float(entry_gamma.get())))
                else:
                    result=img.point(gamma_table(float(entry_gamma.get())))
                result.save(os.path.dirname(x)+"/"+"gamma_"+name)
            except:
                messagebox.showerror('エラー', f'{x}\nは作成できませんでした')
        messagebox.showinfo('完了', '終了しました')

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    label_gamma=ttk.Label(frame,text="ガンマ値：")
    entry_gamma=ttk.Entry(frame,width=5)
    entry_gamma.insert(END,1.0)
    label=ttk.Label(frame,text="ここに画像ファイルを\nドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',gamma)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label_gamma.grid(row=1,column=0)
    entry_gamma.grid(row=1,column=1)
    label.grid(row=2,column=0,columnspan=2)

def area_mozaic():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    origin=None
    x=None
    frame2=ttk.Frame(frame)

    canvas_img=None
    def img_top(drop):
        nonlocal canvas_img,pil_img,z,origin,pil_img_1,x
        try:
            x=drop.data.replace("{","").replace("}","").replace("\\","/")
            y= os.path.basename(x)
            z=x.replace(y,"mosaic_"+y)
            pil_img=pil_img_1=Image.open(x)
            if pil_img.width>pil_img.height:
                origin=0
                pil_img=pil_img.resize((int(ws/2),int(pil_img.height*((ws/2)/pil_img.width))),Image.LANCZOS)
            elif pil_img.width<pil_img.height:
                origin=1
                pil_img=pil_img.resize((int(pil_img.width*((hs/2)/pil_img.height)),int(hs/2)),Image.LANCZOS)
            elif pil_img.width==pil_img.height:
                origin=2
                pil_img=pil_img.resize((int(hs/2),int(hs/2)),Image.LANCZOS)
            canvas.delete("all")
            canvas_img= ImageTk.PhotoImage(pil_img)
            canvas.create_image(pil_img.width/2,pil_img.height/2,image=canvas_img)
            canvas.config(width=pil_img.width, height=pil_img.height)
        except:
            messagebox.showerror('エラー', '失敗しました')

    def on_press(event):
        canvas.delete("rect")
        nonlocal start_x
        nonlocal start_y
        start_x = canvas.canvasx(event.x)
        start_y = canvas.canvasy(event.y)

    def on_move_press(event):
        nonlocal end_x, end_y
        end_x, end_y = event.x, event.y
        canvas.delete("rect")
        canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)

    def on_release(event):
        nonlocal end_x, end_y
        end_x, end_y = event.x, event.y
        canvas.delete("rect")
        canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)

    def mozaic():
        nonlocal origin
        try:
            if origin==2:
                start_x1=int(start_x*pil_img_1.height/int(hs/2))
                start_y1=int(start_y*pil_img_1.height/int(hs/2))
                end_x1=int(end_x*pil_img_1.height/int(hs/2))
                end_y1=int(end_y*pil_img_1.height/int(hs/2))
            elif origin==1:
                start_x1=int(start_x*pil_img_1.width/int(pil_img_1.width*((hs/2)/pil_img_1.height)))
                start_y1=int(start_y*pil_img_1.height/int(hs/2))
                end_x1=int(end_x*pil_img_1.width/int(pil_img_1.width*((hs/2)/pil_img_1.height)))
                end_y1=int(end_y*pil_img_1.height/int(hs/2))
            elif origin==0:
                start_x1=int(start_x*pil_img_1.width/int(ws/2))
                start_y1=int(start_y*pil_img_1.height/int(pil_img_1.height*((ws/2)/pil_img_1.width)))
                end_x1=int(end_x*pil_img_1.width/int(ws/2))
                end_y1=int(end_y*pil_img_1.height/int(pil_img_1.height*((ws/2)/pil_img_1.width)))
            if start_x1>end_x1:
                start_x1,end_x1=end_x1,start_x1
            if start_y1>end_y1:
                start_y1,end_y1=end_y1,start_y1
            crop=pil_img_1.crop((start_x1,start_y1,end_x1,end_y1))
            small_size = (
                round(crop.width / var.get()),
                round(crop.height / var.get())
            )
            small_image = crop.resize(small_size)
            reverse_image = small_image.resize(
                (crop.width, crop.height),
                Image.NEAREST
                )
            fin_img=Image.open(x)
            fin_img.paste(reverse_image, (start_x1, start_y1))
            fin_img.save(z)
            messagebox.showinfo("終了","モザイク完了しました")
        except:
            messagebox.showerror('エラー', '失敗しました')

    start_x=None
    start_y=None
    end_x=None
    end_y=None
    pil_img=None
    z=None
    pil_img_1=None
    root.update()
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',img_top)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    canvas=Canvas(frame,width=300,height=300,bg="gray")
    button1=ttk.Button(frame,text="加工",command=mozaic)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    var=IntVar()
    var.set(30)
    radio1=ttk.Radiobutton(frame2,text="モザイク",value=30,variable=var)
    radio2=ttk.Radiobutton(frame2,text="ぼかし",value=10,variable=var)

    canvas.bind("<ButtonPress-1>", on_press)
    canvas.bind("<ButtonRelease-1>", on_release)
    canvas.bind("<B1-Motion>", on_move_press)

    frame.pack()
    buttonX.pack(side=TOP)
    buttonY.pack(side=TOP)
    frame2.pack(side=TOP)
    canvas.pack(expand=True,side=TOP,fill=BOTH)
    button1.pack(side=TOP)
    radio1.grid(row=0,column=0)
    radio2.grid(row=0,column=1)

def translate():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def check(x):
        if x==1:
            if var1.get()==1:
                radio1.config(state="normal")
                radio2.config(state="normal")
                df1.set(1)
            else:
                radio1.config(state="disabled")
                radio2.config(state="disabled")
        elif x==2:
            if var2.get()==1:
                radio3.config(state="normal")
                radio4.config(state="normal")
                df2.set(1)
            else:
                radio3.config(state="disabled")
                radio4.config(state="disabled")
        elif x==3:
            if var3.get()==1:
                radio5.config(state="normal")
                radio6.config(state="normal")
                df3.set(1)
            else:
                radio5.config(state="disabled")
                radio6.config(state="disabled")
        elif x==4:
            if var4.get()==1:
                radio7.config(state="normal")
                radio8.config(state="normal")
                df4.set(1)
            else:
                radio7.config(state="disabled")
                radio8.config(state="disabled")

    def trans():
        try:
            x=clip.paste()
            if var1.get()==1:
                if df1.get()==1:
                    x=jaconv.hira2kata(x)
                else:
                    x=jaconv.kata2hira(x)
            if var2.get()==1:
                if df2.get()==2:
                    x=x.upper()
                else:
                    x=x.lower()
            if var3.get()==1:
                if df3.get()==2:
                    x=jaconv.z2h(x,kana=True,digit=True,ascii=True)
                else:
                    x=jaconv.h2z(x,kana=True,digit=True,ascii=True)
            if var4.get()==1:
                if df4.get()==2:
                    x=jaconv.kana2alphabet(x)
                else:
                    x=jaconv.alphabet2kana(x)
            clip.copy(x)
            messagebox.showinfo("終了","変換完了しました")
        except:
            messagebox.showerror('エラー', '失敗しました')

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    var1=IntVar()
    var2=IntVar()
    var3=IntVar()
    var4=IntVar()
    df1=IntVar()
    df2=IntVar()
    df3=IntVar()
    df4=IntVar()
    check1=ttk.Checkbutton(frame,text="ひらがな・カタカナ：",onvalue=1,offvalue=0,variable=var1,command=lambda:check(1))
    check2=ttk.Checkbutton(frame,text="大文字・小文字：",onvalue=1,offvalue=0,variable=var2,command=lambda:check(2))
    check3=ttk.Checkbutton(frame,text="半角・全角：",onvalue=1,offvalue=0,variable=var3,command=lambda:check(3))
    radio1=ttk.Radiobutton(frame,text="ひらがな→カタカナ",value=1,variable=df1,state="disable")
    radio2=ttk.Radiobutton(frame,text="カタカナ→ひらがな",value=2,variable=df1,state="disable")
    radio3=ttk.Radiobutton(frame,text="大文字→小文字",value=1,variable=df2,state="disable")
    radio4=ttk.Radiobutton(frame,text="小文字→大文字",value=2,variable=df2,state="disable")
    radio5=ttk.Radiobutton(frame,text="半角→全角",value=1,variable=df3,state="disable")
    radio6=ttk.Radiobutton(frame,text="全角→半角",value=2,variable=df3,state="disable")
    check4=ttk.Checkbutton(frame,text="ローマ字・ひらがな",onvalue=1,offvalue=0,variable=var4,command=lambda:check(4))
    radio7=ttk.Radiobutton(frame,text="ローマ字→ひらがな",value=1,variable=df4,state="disable")
    radio8=ttk.Radiobutton(frame,text="ひらがな→ローマ字",value=2,variable=df4,state="disable")
    button=ttk.Button(frame,text="変換",command=trans)

    frame.pack()
    buttonX.grid(row=0,column=2)
    buttonY.grid(row=0,column=0)
    check1.grid(row=1,column=0)
    radio1.grid(row=1,column=1)
    radio2.grid(row=1,column=2)
    check2.grid(row=2,column=0)
    radio3.grid(row=2,column=1)
    radio4.grid(row=2,column=2)
    check3.grid(row=3,column=0)
    radio5.grid(row=3,column=1)
    radio6.grid(row=3,column=2)
    check4.grid(row=4,column=0)
    radio7.grid(row=4,column=1)
    radio8.grid(row=4,column=2)
    button.grid(row=5,column=0,columnspan=3)

def video_volume():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def video_top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                if os.path.exists(x)==False:
                    messagebox.showerror("エラー","ファイルが存在しません")
                else:
                    if var.get()==0:
                        clip = VideoFileClip(x)
                        name=os.path.basename(x)
                        newaudioclip = clip.audio.volumex(float(entry1.get()))
                        newclip = clip.set_audio(newaudioclip)
                        z=x.replace(name,"volume_"+name)
                        newclip.write_videofile(z,logger=None)
                    else:
                        audio = AudioFileClip(x)
                        name=os.path.basename(x)
                        z=x.replace(name,"volume_"+name)
                        new_audio = audio.volumex(float(entry1.get()))
                        new_audio.write_audiofile(z,logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("終了","終了しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame,text="音量倍率：")
    entry1=ttk.Entry(frame,width=5)
    label=ttk.Label(frame,text="ここに動画ファイルを\nドラッグ＆ドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',video_top)
    entry1.insert(0,"1.0")
    var=IntVar()
    var.set(0)
    radio1=ttk.Radiobutton(frame,text="動画",value=0,variable=var)
    radio2=ttk.Radiobutton(frame,text="音声",value=1,variable=var)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    radio1.grid(row=2,column=0)
    radio2.grid(row=2,column=1)
    label.grid(row=3,column=0,columnspan=2)

def video_speed():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def video_top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                if os.path.exists(x)==False:
                    messagebox.showerror("エラー","ファイルが存在しません")
                else:
                    if var.get()==0:
                        clip_orig = VideoFileClip(x)
                        name=os.path.basename(x)
                        clip = speedx(clip_orig, factor=float(entry1.get()))
                        z=x.replace(name,"speed_"+name)
                        clip.write_videofile(z,logger=None)
                    else:
                        name=os.path.basename(x)
                        z=x.replace(name,"speed_"+name)
                        audio = AudioFileClip(x)
                        new_audio = audio.fx(speedx, factor=float(entry1.get()))
                        new_audio.write_audiofile(z,logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("終了","終了しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame,text="速度倍率：")
    entry1=ttk.Entry(frame,width=5)
    label=ttk.Label(frame,text="ここに動画ファイルを\nドラッグ＆ドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',video_top)
    entry1.insert(0,"1.0")
    var=IntVar()
    var.set(0)
    radio1=ttk.Radiobutton(frame,text="動画",value=0,variable=var)
    radio2=ttk.Radiobutton(frame,text="音声",value=1,variable=var)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    radio1.grid(row=2,column=0)
    radio2.grid(row=2,column=1)
    label.grid(row=3,column=0,columnspan=2)

def video_shot():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def video_top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                clip = VideoFileClip(x)
                name=os.path.basename(x)
                youso=os.path.splitext(name)[1]
                z=x.replace(youso,".png")
                clip.save_frame(filename=z, t=int(entry1.get())*60+int(entry2.get()))
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("終了","終了しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame,text="スクショ時刻：")
    entry1=ttk.Entry(frame,width=5)
    label2=ttk.Label(frame,text="：")
    entry2=ttk.Entry(frame,width=5)
    label=ttk.Label(frame,text="ここに動画ファイルを\nドラッグ＆ドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',video_top)

    frame.pack()
    buttonX.grid(row=0,column=3)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=1,column=2)
    entry2.grid(row=1,column=3)
    label.grid(row=2,column=0,columnspan=4)

def video_info():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def video_info_top(drop):
        try:
            x=drop.data.replace("{","").replace("}","").replace("\\","/")
            clip = VideoFileClip(x)
            info=f"解像度：{clip.size}\nFPS：{clip.fps}\n長さ：{int(clip.duration//60)}分{int(clip.duration%60)}秒"
            messagebox.showinfo("情報",info)
            clip.close()
        except:
            messagebox.showerror("エラー","失敗しました")

    label=ttk.Label(frame,text="ここに動画ファイルを\nドロップしてください",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',video_info_top)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def md2html():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    def md2html_top(drop):
        def convert_markdown_to_html(markdown_file, html_file):
            with open(markdown_file, 'r', encoding='utf-8') as file:
                markdown_text = file.read()
                html = mistune.markdown(markdown_text)
            with open(html_file, 'w', encoding='utf-8') as file:
                file.write(html)
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                z=x.replace(".md",".html")
                convert_markdown_to_html(x, z)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("終了","終了しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="ここにmdファイルを\nドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',md2html_top)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def video_concat():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def video_concat_top():
        try:
            output=filedialog.asksaveasfilename(
                title = "名前を付けて保存",
                filetypes = [("MP4", ".mp4")],
                defaultextension = "mp4"
                )
            clip1 = VideoFileClip(entry1.get())
            clip2 = VideoFileClip(entry2.get())
            clip = concatenate_videoclips([clip1, clip2])
            clip.write_videofile(output,logger=None)
            messagebox.showinfo("終了","結合完了しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    def drag_video1(drop):
        x=drop.data.replace("{","").replace("}","").replace("\\","/")
        entry1.delete(0,END)
        entry1.insert(0,x)

    def drag_video2(drop):
        x=drop.data.replace("{","").replace("}","").replace("\\","/")
        entry2.delete(0,END)
        entry2.insert(0,x)


    label1=ttk.Label(frame,text="前半の動画をここにドロップ：")
    label2=ttk.Label(frame,text="後半の動画をここにドロップ：")
    entry1=ttk.Entry(frame,width=20)
    entry2=ttk.Entry(frame,width=20)
    button=ttk.Button(frame,text="結合",command=video_concat_top)
    entry1.drop_target_register(DND_FILES)
    entry1.dnd_bind('<<Drop>>',drag_video1)
    entry2.drop_target_register(DND_FILES)
    entry2.dnd_bind('<<Drop>>',drag_video2)
    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',drag_video1)
    label2.drop_target_register(DND_FILES)
    label2.dnd_bind('<<Drop>>',drag_video2)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    button.grid(row=3,column=0,columnspan=2)

def voice2video():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def create_movie_with_sound(img_path, audio_path, movie_out_path):
        img_clip = ImageClip(img_path)
        audio_clip = AudioFileClip(audio_path)
        video_clip = img_clip.set_duration(audio_clip.duration)
        video_clip = video_clip.set_audio(audio_clip)
        video_clip.write_videofile(movie_out_path,
                                fps=30,
                                temp_audiofile='temp-audio.flac',
                                remove_temp=True,
                                logger=None,
                                codec='libx264',
                                audio_codec='flac',
                                ffmpeg_params=['-strict', '-2'])
        return

    def img_top(drop):
        x=drop.data.replace("{","").replace("}","").replace("\\","/")
        entry1.delete(0,END)
        entry1.insert(0,x)

    def voice_top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                name=os.path.basename(x)
                youso=os.path.splitext(name)[1]
                z=x.replace(youso,".mp4")
                create_movie_with_sound(entry1.get(),x,z)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("終了","終了しました")


    label1=ttk.Label(frame,text="音声ファイルをここに\nドロップしてください",font=("Helvetica", 16))
    label2=ttk.Label(frame,text="画像ファイルをドロップしてください：")
    entry1=ttk.Entry(frame,width=25)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    entry1.drop_target_register(DND_FILES)
    entry1.dnd_bind('<<Drop>>',img_top)
    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',voice_top)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label2.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label1.grid(row=2,column=0,columnspan=2)

def compress():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def text_top(drop):
        try:
            x=drop.data.replace("{","").replace("}","").replace("\\","/")
            if os.path.isfile(x):
                messagebox.showinfo("終了","ファイルが設定されています。フォルダを設定してください")
            else:
                def zip_folder(folder_path, zip_name):
                    with zipfile.ZipFile(zip_name, 'w',compression=zipfile.ZIP_DEFLATED,compresslevel=var.get()) as zipf:
                        for root, _, files in os.walk(folder_path):
                            for file in files:
                                file_path = os.path.join(root, file)
                                zipf.write(file_path, os.path.relpath(file_path, folder_path))

                zip_name=x+".zip"
                folder_path = x
                zip_folder(folder_path, zip_name)
                messagebox.showinfo("終了","圧縮完了しました")
        except:
            messagebox.showerror("エラー","失敗しました")


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    var=IntVar()
    var.set(6)
    radio0=ttk.Radiobutton(frame,text="無圧縮",value=0,variable=var)
    radio1=ttk.Radiobutton(frame,text="速度優先",value=1,variable=var)
    radio6=ttk.Radiobutton(frame,text="バランス",value=6,variable=var)
    radio9=ttk.Radiobutton(frame,text="圧縮率優先",value=9,variable=var)
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',text_top)
    label=ttk.Label(frame,text="圧縮レベルを選択してください")
    label1=ttk.Label(frame,text="圧縮するフォルダを\nドロップしてください",font=("Helvetica", 16))

    frame.pack()
    buttonX.grid(row=0,column=3)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)
    radio0.grid(row=2,column=0)
    radio1.grid(row=2,column=1)
    radio6.grid(row=2,column=2)
    radio9.grid(row=2,column=3)
    label1.grid(row=3,column=0,columnspan=4)

def decompress():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        pw=entry.get()
        for f in file_list:
            try:
                output=f.replace(".zip","")
                with open(f, 'rb') as file:
                    # ヘッダ情報
                    header = file.read(30)
                    # フラグ取得
                    general_flags = int.from_bytes(header[6:8], byteorder='little')

                # UTF-8フラグ検知
                if general_flags & 0x0800:
                    with zipfile.ZipFile(f) as zf:
                        zf.extractall(path=output,pwd=pw.encode('utf-8'))
                else:
                    try:
                        with zipfile.ZipFile(f) as z:
                            for info in z.infolist():
                                info.filename = info.orig_filename.encode('cp437').decode('utf-8')
                                if os.sep != "/" and os.sep in info.filename:
                                    info.filename = info.filename.replace(os.sep, "/")
                                z.extract(info,path=output,pwd=pw.encode('utf-8'))
                    except:
                        with zipfile.ZipFile(f) as z:
                            for info in z.infolist():
                                info.filename = info.orig_filename.encode('cp437').decode('cp932')
                                if os.sep != "/" and os.sep in info.filename:
                                    info.filename = info.filename.replace(os.sep, "/")
                                z.extract(info,path=output,pwd=pw.encode('utf-8'))

            except:
                messagebox.showerror("エラー","失敗しました")
        messagebox.showinfo("終了","終了しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)
    label=ttk.Label(frame,text="解凍するファイルを\nドロップしてください",font=("Helvetica", 16))
    entry=ttk.Entry(frame,width=30)
    label_w=ttk.Label(frame,text="パスワードを入力してください")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label_w.grid(row=1,column=0,columnspan=2)
    entry.grid(row=2,column=0,columnspan=2)
    label.grid(row=3,column=0,columnspan=2)

def image_info():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def top(drop):
        try:
            x=drop.data.replace("{","").replace("}","").replace("\\","/")
            img=Image.open(x)
            name=os.path.basename(x)
            size=img.size
            format1=img.format
            format2=img.format_description
            root1=Free_window()
            root1.attributes("-topmost", True)
            root1.title("情報")
            frame1=ttk.Frame(root1)
            labelX1=ttk.Label(frame1,text="サイズ："+str(size))
            labelX2=ttk.Label(frame1,text="フォーマット："+format1)
            labelX3=ttk.Label(frame1,text="フォーマット詳細："+format2)
            labelX4=ttk.Label(frame1,text="ファイル名："+name)
            frame1.pack()
            labelX4.pack(side=TOP)
            labelX1.pack(side=TOP)
            labelX2.pack(side=TOP)
            labelX3.pack(side=TOP)
        except:
            messagebox.showerror("エラー","失敗しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)
    label=ttk.Label(frame,text="情報を表示する画像を\nドロップしてください",font=("Helvetica", 16))

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def window_toumei():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def stopping():
        listener.stop()
        t.join()
        frame_delete()

    def main_frame1(num):
        listener.stop()
        t.join()
        root.protocol("WM_DELETE_WINDOW", frame_delete)
        main_frame(num)

    def key_check():
        nonlocal listener
        with keyboard.Listener(on_release=on_release) as listener:
            listener.join()

    def set_window_opacity(hwnd, opacity):
        # ウィンドウのスタイル
        style = win32gui.GetWindowLong(hwnd, win32con.GWL_EXSTYLE)

        style = style | win32con.WS_EX_LAYERED
        win32gui.SetWindowLong(hwnd, win32con.GWL_EXSTYLE, style)

        # 透明度設定
        win32gui.SetLayeredWindowAttributes(hwnd, 0, int(opacity * 255), win32con.LWA_ALPHA)
        toast = winotify.Notification(
            title="透明度設定",
            msg="透明度を"+str(float(entry.get())*100)+"%に設定しました",
            app_id="Yuki's Army knife",
        )
        toast.show()


    def on_release(key):
        if key == keyboard.Key.pause:
            if entry.get()=="":
                messagebox.showerror("エラー","透明度を入力してください")
            else:
                hwnd = win32gui.GetForegroundWindow()
                set_window_opacity(hwnd, float(entry.get()))


    root.protocol("WM_DELETE_WINDOW", stopping)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(3),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="pauseでウィンドウを実行している\nファイルを直接開きます",font=("Helvetica", 16))
    label1=ttk.Label(frame,text="透明度を入力してください(0~1)")
    entry=ttk.Entry(frame,width=5)
    entry.insert(0,"0.8")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry.grid(row=1,column=1)
    label.grid(row=2,column=0,columnspan=2)
    listener=None
    t=threading.Thread(target=key_check,daemon=True)
    t.start()

def url_change():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def url_change2():
        try:
            if var.get()==0:
                result=urllib.parse.quote(clip.paste())
            else:
                result=urllib.parse.unquote(clip.paste())
            clip.copy(result)
            messagebox.showinfo("結果",result+"\nをクリップボードにコピーしました")
        except:
            messagebox.showerror("エラー","失敗しました")


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    button1=ttk.Button(frame,text="変換",command=url_change2)
    var=IntVar()
    var.set(0)
    radio1=ttk.Radiobutton(frame,text="エンコード(日本語→文字列)",variable=var,value=0)
    radio2=ttk.Radiobutton(frame,text="デコード(文字列→日本語)",variable=var,value=1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    button1.grid(row=2,column=0,columnspan=2)


def voice2read():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def top(drop):
        try:
            x=drop.data.replace("{","").replace("}","").replace("\\","/")
            r = sr.Recognizer()
            with sr.AudioFile(x) as source:
                audio = r.record(source)
            text = r.recognize_google(audio, language=var.get())
            box.delete("1.0",END)
            box.insert(END,text)
        except:
            messagebox.showerror("エラー","失敗しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="wavファイルを\nドロップしてください",font=("Helvetica", 16))
    box=ScrolledText(frame,width=30,height=20)
    button=ttk.Button(frame,text="文字コピー",command=lambda:clip.copy(box.get("1.0",END).strip()))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)
    var=StringVar()
    radio1=ttk.Radiobutton(frame,text="日本語",variable=var,value="ja-JP")
    radio2=ttk.Radiobutton(frame,text="英語",variable=var,value="en-US")
    var.set("ja-JP")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    label.grid(row=2,column=0,columnspan=2)
    box.grid(row=3,column=0,columnspan=2)
    button.grid(row=4,column=0,columnspan=2)

def video_size():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                clip_orig = VideoFileClip(x)
                clip = resize(clip_orig, (int(entry_w.get()), int(entry_h.get())))
                name=os.path.basename(x)
                name1=x.replace(name,"resized_"+name)
                clip.write_videofile(name1,logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("結果","完了しました")

    label_w=ttk.Label(frame,text="横幅：")
    label_h=ttk.Label(frame,text="縦幅：")
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="動画ファイルをここに\nドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)
    entry_w=ttk.Entry(frame,width=10)
    entry_h=ttk.Entry(frame,width=10)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label_w.grid(row=1,column=0)
    entry_w.grid(row=1,column=1)
    label_h.grid(row=2,column=0)
    entry_h.grid(row=2,column=1)
    label.grid(row=3,column=0,columnspan=2)

def video_BW():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                clip_orig = VideoFileClip(x)
                clip = blackwhite(clip_orig)
                name=os.path.basename(x)
                name1=x.replace(name,"BW_"+name)
                clip.write_videofile(name1,logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("結果","終了しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="動画ファイルをここに\nドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def video_light():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                clip_orig = VideoFileClip(x)
                clip = lum_contrast(clip_orig, lum=float(entry1.get()), contrast=float(entry2.get()))
                name=os.path.basename(x)
                name1=x.replace(name,"light_"+name)
                clip.write_videofile(name1,logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("結果","完了しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame,text="明るさ：")
    label2=ttk.Label(frame,text="コントラスト：")
    entry1=ttk.Entry(frame,width=5)
    entry2=ttk.Entry(frame,width=5)
    label3=ttk.Label(frame,text="動画ファイルをここに\nドロップしてください",font=("Helvetica", 16))
    entry1.insert(0,"1.0")
    entry2.insert(0,"1.0")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    label3.grid(row=3,column=0,columnspan=2)

def video_frame():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def convert_color_code_to_rgb(color_code):
        r = int(color_code[1:3], 16)
        g = int(color_code[3:5], 16)
        b = int(color_code[5:7], 16)
        rgb = (r, g, b)
        return rgb

    def top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                clip_orig = VideoFileClip(x)
                clip = margin(clip_orig, mar=int(entry1.get()), color=convert_color_code_to_rgb(entry2.get()))
                name=os.path.basename(x)
                name1=x.replace(name,"frame_"+name)
                clip.write_videofile(name1,logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("結果","完了しました")

    label1=ttk.Label(frame,text="枠組みの幅：")
    label2=ttk.Label(frame,text="枠組みの色(hex)：")
    entry1=ttk.Entry(frame,width=15)
    entry2=ttk.Entry(frame,width=15)
    label3=ttk.Label(frame,text="動画ファイルをここに\nドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    entry1.insert(0,"50")
    entry2.insert(0,"#c0c0c0")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    label3.grid(row=3,column=0,columnspan=2)

def color_code():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def color_change():
        try:
            if entry2_r.get()=="" and entry1.get()!= "":
                color_code=entry1.get()
                r = int(color_code[1:3], 16)
                g = int(color_code[3:5], 16)
                b = int(color_code[5:7], 16)
                root2=Free_window()
                root2.title("変換結果")
                root2.geometry('+%d+%d'%(x,y))
                root2.attributes("-topmost", True)
                result=ttk.Label(root2,text=f"(R,G,B)=({r},{g},{b})",font=("Helvetica", 16))
                button=ttk.Button(root2,text="コピー",command=lambda:clip.copy(f"({r},{g},{b})"))
                result.pack(side=TOP)
                button.pack(side=TOP)

            elif entry1.get()=="" and entry2_r.get()!="" and entry2_g.get()!="" and entry2_b.get()!="":
                rgb=[int(entry2_r.get()),int(entry2_g.get()),int(entry2_b.get())]
                hex_code = '#{:02x}{:02x}{:02x}'.format(rgb[0], rgb[1], rgb[2])
                root2=Free_window()
                root2.title("変換結果")
                root2.geometry('+%d+%d'%(x,y))
                root2.attributes("-topmost", True)
                result=ttk.Label(root2,text="カラーコード："+hex_code,font=("Helvetica", 16))
                button=ttk.Button(root2,text="コピー",command=lambda:clip.copy(hex_code))
                result.pack(side=TOP)
                button.pack(side=TOP)

            else :
                messagebox.showerror("エラー","正しく入力してください")
        except:
            messagebox.showerror("エラー","失敗しました")

    def del_input():
        entry1.delete(0,END)
        entry2_r.delete(0,END)
        entry2_g.delete(0,END)
        entry2_b.delete(0,END)

    entry1=ttk.Entry(frame,width=20)
    entry2_r=ttk.Entry(frame,width=5)
    entry2_g=ttk.Entry(frame,width=5)
    entry2_b=ttk.Entry(frame,width=5)
    button=ttk.Button(frame,text="変換",command=color_change)
    label1=ttk.Label(frame,text="16進数(hex)：")
    label2=ttk.Label(frame,text="RGB：")
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    button1=ttk.Button(frame,text="入力削除",command=del_input)

    frame.pack()
    buttonX.grid(row=0,column=1,columnspan=3)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1,columnspan=3)
    label2.grid(row=2,column=0)
    entry2_r.grid(row=2,column=1)
    entry2_g.grid(row=2,column=2)
    entry2_b.grid(row=2,column=3)
    button.grid(row=3,column=2,columnspan=2)
    button1.grid(row=3,column=0,columnspan=2)

def translate_text():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def translate_text2():

        clip.copy(text1.get("1.0",END))
        translate_lang()

    def translate_lang():
        try:
            translator = Translator()
            text=clip.paste()
            input_num=lang_list.index(box1.get())
            output_num=lang1.index(box2.get())
            input_text=lang2[input_num]
            output_text=lang2[output_num+1]

            if input_text=="auto":
                lang=translator.detect(text).lang
            else:
                lang=input_text
            text1=translator.translate(text, src=lang, dest=output_text).text
            lang_name = LANGUAGES.get(lang)
            root1=Free_window()
            root1.title("翻訳結果")
            root1.geometry('+%d+%d'%(x,y))
            root1.attributes("-topmost", True)
            box=ScrolledText(root1,width=40,height=20)
            label=ttk.Label(root1,text=f"{lang_name}→{LANGUAGES.get(output_text)}",font=("Helvetica", 12))
            box.insert(END,text1)
            button_n=ttk.Button(root1,text="コピー",command=lambda:[clip.copy(text1),
                                                                messagebox.showinfo("完了","コピーしました")])
            label.pack(side=TOP)
            box.pack(side=TOP,expand=True,fill=BOTH)
            button_n.pack(side=TOP)

        except:
            messagebox.showerror("エラー","失敗しました")


    #input_lang=["auto","ja","en"]
    #output_lang=["ja","en"]
    lang1=["英語","日本語","スペイン語","ポルトガル語","ドイツ語","フランス語","イタリア語","ロシア語","アラビア語","トルコ語","韓国語","中国語（簡体字）","中国語（繁体字）","ヒンディー語","ベンガル語","ウクライナ語","ギリシャ語","オランダ語","チェコ語","ポーランド語","タイ語","スウェーデン語","ルーマニア語","フィンランド語","ノルウェー語","デンマーク語","クロアチア語","インドネシア語","フィリピン語","ベトナム語"]
    lang2=["auto","en","ja","es","pt","de","fr","it","ru","ar","tr","ko","zh-CN","zh-TW","hi","bn","uk","el","nl","cs","pl","th","sv","ro","fi","no","da","hr","id","tl","vi"]
    lang_list=["自動","英語","日本語"]
    input_lang=lang_list
    output_lang=lang1


    box1=ttk.Combobox(frame,values=input_lang,width=8,state="readonly")
    box2=ttk.Combobox(frame,values=output_lang,width=8,state="readonly")
    label1=ttk.Label(frame,text="入力言語：")
    label2=ttk.Label(frame,text="出力言語：")
    button=ttk.Button(frame,text="クリップボード翻訳",command=translate_lang)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    box1.current(0)
    box2.current(0)
    box1.configure(state="readonly")
    button1=ttk.Button(frame,text="手動入力翻訳",command=translate_text2)
    text1=ScrolledText(frame,width=40,height=20)

    frame.pack()
    buttonX.grid(row=0,column=1,columnspan=2)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    box1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    box2.grid(row=2,column=1)
    button.grid(row=3,column=0)
    button1.grid(row=3,column=1)
    text1.grid(row=4,column=0,columnspan=2)

def pass_check():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def check_password(password):
        sha1_password = hashlib.sha1(password.encode('utf-8')).hexdigest().upper()
        prefix, suffix = sha1_password[:5], sha1_password[5:]
        url = f"https://api.pwnedpasswords.com/range/{prefix}"
        response = requests.get(url)
        hashes = (line.split(':') for line in response.text.splitlines())
        for h, count in hashes:
            if h == suffix:
                return int(count)
        return 0

    def check_pass1():
        try:
            password = entry.get()
            count = check_password(password)
            if count:
                messagebox.showinfo("結果",f"パスワードは{count}回流出しています")
            else:
                messagebox.showinfo("結果","パスワードは流出していません")
        except:
            messagebox.showerror("エラー","失敗しました")

    label=ttk.Label(frame,text="パスワードを入力してください")
    entry=ttk.Entry(frame,width=30)
    button=ttk.Button(frame,text="チェック",command=check_pass1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0)
    entry.grid(row=1,column=1)
    button.grid(row=2,column=0,columnspan=2)

def voice_delete():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                name=os.path.basename(x)
                name1=x.replace(name,"del_voice_"+name)
                video_path = x
                video = VideoFileClip(video_path)
                video_without_audio = video.without_audio()
                video_without_audio.write_videofile(name1,logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","音声を削除しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="ここに動画ファイルを\nドラッグ&ドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def site_title():
    global frame,buttonY

    def get_url_title(url):
        try:
            response = requests.get(url)
            soup = BeautifulSoup(response.content, 'html.parser')
            title = soup.title.string
            return title
        except:
            return None

    def url_title():
        try:
            url=clip.paste()
            title=get_url_title(url)
            if title==None:
                messagebox.showinfo("結果","タイトルを取得できませんでした")
            else:
                root1=Free_window()
                root1.title("結果")
                root1.configure(bg="gray")
                root1.geometry('+%d+%d'%(x,y))
                root1.attributes("-topmost", True)
                label1=ttk.Label(root1,text=title,font=("Helvetica", 12),wraplength=400,anchor="center")
                frame_1=Frame(root1)
                button1=ttk.Button(frame_1,text="タイトルのみ",width=10,command=lambda:[clip.copy(title),messagebox.showinfo("取得","コピーしました"),root1.focus_set()])
                button2=ttk.Button(frame_1,text="マークダウン",width=10,command=lambda:[clip.copy(f"[{title}]({url})"),messagebox.showinfo("取得","コピーしました"),root1.focus_set()])
                button3=ttk.Button(frame_1,text="HTML",width=10,command=lambda:[clip.copy(f"<a href='{url}'>{title}</a>"),messagebox.showinfo("取得","コピーしました"),root1.focus_set()])
                button4=ttk.Button(frame_1,text="参考文献",width=10,command=lambda:[book(url,title),root1.focus_set()])

                label1.pack(side=TOP)
                frame_1.pack(side=TOP)
                button1.grid(row=0,column=0)
                button2.grid(row=0,column=1)
                button3.grid(row=0,column=2)
                button4.grid(row=0,column=3)
        except:
            messagebox.showerror("エラー","失敗しました")

    def book(url,title):
        dt_now = datetime.now()
        now=dt_now.strftime('%Y-%m-%d')
        x=f"{{著者名}}．“{title}”．{{Webサイトの名称}}．{{更新日付}}．{url}，（参照 {now}）．"
        clip.copy(x)
        messagebox.showinfo("取得","コピーしました")

    if intermediate_screen==0:
        main_frame_delete()
        frame = ttk.Frame(root, padding=16)
        button=ttk.Button(frame,text='サイトタイトル',command=url_title,padding=[80,10,80,10])
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        button.grid(row=1,column=0,columnspan=2)
    elif intermediate_screen==1:
        url_title()

def network_monitor():
    root1=Free_window()
    frame_1=ttk.Frame(root1)
    root1.attributes("-topmost", True)
    root1.title("ネットワークモニター")
    check_ip=0
    stop_a=None

    def network_monitor2():
        nonlocal stop_a
        if check_ip==0:
            stop_a=None
            network_monitor1()
        stop_a=root1.after(60000,network_monitor2)

    def network_monitor1():
        nonlocal check_ip,id_num,destination_ip_list, connection_info_dict, formatted_connection_info
        def get_iid_from_ip(treeview, ip):
            for iid in treeview.get_children():
                if treeview.item(iid)["values"][1] == ip:
                    return iid
        #root1.destroy()
        #network_monitor()
        if check_ip==1:
            return
        check_ip=1
        button.configure(text="更新中",state="disabled")
        button.update()
        # 通信先IPアドレスと実行ファイル
        destination_ip_list1, connection_info_dict1 = get_all_destination_ips_and_executables()
        # 辞書型
        formatted_connection_info1 = {ip: executables[0] for ip, executables in connection_info_dict1.items()}
        ip_list = list(set(destination_ip_list) & set(destination_ip_list1))
        del_ip_list = list(set(destination_ip_list) - set(ip_list))
        create_ip_list = list(set(destination_ip_list1) - set(ip_list))
        for ip in del_ip_list:
            iid1=get_iid_from_ip(tree, ip)
            if iid1 != None:
                tree.delete(iid1)
        dtimes=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for ip in create_ip_list:
            country,org=get_ocation(ip)
            exe_file=formatted_connection_info1[ip]
            tree.insert(parent='', index='end', iid=id_num ,values=(country,ip,org,os.path.basename(exe_file),dtimes))
            id_num=id_num+1
        check_ip=0
        button.config(state="normal",text="更新")
        destination_ip_list=destination_ip_list1
        connection_info_dict=connection_info_dict1
        button.update()

    def detail_record(event):
        def showMenu(e):
            pmenu.post(e.x_root, e.y_root)
        def select_file(file_path):
            subprocess.Popen(['explorer.exe', '/select,', file_path]).wait()
        record_id = tree.focus()
        if tree.item(record_id, 'values')[0] == "None":
            messagebox.showinfo("エラー", "データがありません")
        else:
            record_ip = tree.item(record_id, 'values')[1]
            url1 = f"https://ipapi.co/{record_ip}/json/"
            response1 = requests.get(url1)
            data1 = response1.json()
            root2=Free_window()
            root2.attributes("-topmost", True)
            pmenu = Menu(root2, tearoff=0)
            root2.bind("<Button-3>", showMenu)
            frame_2=ttk.Frame(root2)
            label_ip=ttk.Label(frame_2,text="IPアドレス：")
            label1_ip=ttk.Label(frame_2,text=record_ip)
            label_country=ttk.Label(frame_2,text="国：")
            label1_country=ttk.Label(frame_2,text=data1['country_name'])
            label_connection=ttk.Label(frame_2,text="接続先：")
            label1_connection1=ttk.Label(frame_2,text=data1['org'])
            label_exe=ttk.Label(frame_2,text="実行ファイル：")
            label1_exe=ttk.Label(frame_2,text=connection_info_dict[record_ip][0])
            label_place_num=ttk.Label(frame_2,text="経度 , 緯度：")
            label1_place_num=ttk.Label(frame_2,text=f"{data1['latitude']} , {data1['longitude']}")
            label_send=ttk.Label(frame_2,text="郵便：")
            label1_send=ttk.Label(frame_2,text=data1['postal'])
            label_region=ttk.Label(frame_2,text="地域：")
            label1_region=ttk.Label(frame_2,text=data1['region'])
            label_city=ttk.Label(frame_2,text="都市：")
            label1_city=ttk.Label(frame_2,text=data1['city'])

            pmenu.add_command(label="IPアドレスコピー", command=lambda:[clip.copy(record_ip),messagebox.showinfo("取得","コピーしました")])
            pmenu.add_command(label="接続先コピー", command=lambda:[clip.copy(data1['org']),messagebox.showinfo("取得","コピーしました")])
            pmenu.add_command(label="実行ファイルパスコピー", command=lambda:[clip.copy(connection_info_dict[record_ip][0]),messagebox.showinfo("取得","コピーしました")])
            pmenu.add_command(label="実行ファイルのフォルダを開く", command=lambda:select_file(connection_info_dict[record_ip][0]))
            pmenu.add_command(label="経度・緯度コピー", command=lambda:[clip.copy(f"({data1['latitude']} , {data1['longitude']})"),messagebox.showinfo("取得","コピーしました")])
            pmenu.add_command(label="郵便コピー", command=lambda:[clip.copy(data1['postal']),messagebox.showinfo("取得","コピーしました")])
            pmenu.add_command(label="地域コピー", command=lambda:[clip.copy(data1['region']),messagebox.showinfo("取得","コピーしました")])
            pmenu.add_command(label="都市コピー", command=lambda:[clip.copy(data1['city']),messagebox.showinfo("取得","コピーしました")])

            frame_2.pack()
            label_ip.grid(row=0,column=0)
            label1_ip.grid(row=0,column=1)
            label_country.grid(row=1,column=0)
            label1_country.grid(row=1,column=1)
            label_connection.grid(row=2,column=0)
            label1_connection1.grid(row=2,column=1)
            label_exe.grid(row=3,column=0)
            label1_exe.grid(row=3,column=1)
            label_place_num.grid(row=4,column=0)
            label1_place_num.grid(row=4,column=1)
            label_send.grid(row=5,column=0)
            label1_send.grid(row=5,column=1)
            label_region.grid(row=6,column=0)
            label1_region.grid(row=6,column=1)
            label_city.grid(row=7,column=0)
            label1_city.grid(row=7,column=1)


    def get_all_destination_ips_and_executables(kind='inet'):
        destination_ips = set()
        connection_info = {}
        for conn in psutil.net_connections(kind=kind):
            if conn.status == psutil.CONN_ESTABLISHED:
                destination_ip = conn.raddr.ip
                destination_ips.add(destination_ip)
                if conn.pid is not None:
                    try:
                        proc = psutil.Process(conn.pid)
                        executable = proc.exe()
                        connection_info.setdefault(destination_ip, []).append(executable)
                    except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                        pass
        return list(destination_ips), connection_info

    # 通信先IPアドレスと実行ファイル
    destination_ip_list, connection_info_dict = get_all_destination_ips_and_executables()
    # 辞書型
    formatted_connection_info = {ip: executables[0] for ip, executables in connection_info_dict.items()}

    def get_ocation(ip_address):
        url = f"http://ip-api.com/json/{ip_address}?fields=countryCode,org"
        response = requests.get(url)
        data = response.json()
        if len(data) == 0:
            country_code="None"
            org="None"
        else:
            country_code = data['countryCode']
            org = data['org']
        return country_code, org

    button=ttk.Button(frame_1,text="更新",command=network_monitor1)
    column = ('country', 'ip', 'connection','exe',"time")
    frame_2=ttk.Frame(root1)
    tree = ttk.Treeview(frame_2, columns=column,height=20)
    tree.bind("<<TreeviewSelect>>", detail_record)
    tree.column('#0',width=0, stretch='no')
    tree.column('country', anchor='center', width=60)
    tree.column('ip',anchor='center', width=300)
    tree.column('connection', anchor='center', width=240)
    tree.column('exe', anchor='center', width=240)
    tree.column('time', anchor='center', width=240)

    tree.heading('#0',text='')
    tree.heading('country', text='国',anchor='center')
    tree.heading('ip', text='IPアドレス', anchor='center')
    tree.heading('connection',text='接続先', anchor='center')
    tree.heading('exe',text='実行ファイル', anchor='center')
    tree.heading('time',text='開始時間', anchor='center')

    scrollbar = ttk.Scrollbar(frame_2, orient=VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set)

    id_num=0
    times=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for i in destination_ip_list:
        country,org=get_ocation(i)
        exe_file=formatted_connection_info[i]
        tree.insert(parent='', index='end', iid=id_num ,values=(country,i,org,os.path.basename(exe_file),times))
        id_num=id_num+1

    frame_1.pack(side=TOP)
    button.pack(side=TOP)
    frame_2.pack(side=TOP)
    tree.pack(side=LEFT)
    scrollbar.pack(side=RIGHT, fill=Y)
    srop_a=root1.after(60000,network_monitor2)

    def deswin():
        if srop_a is not None:
            root1.after_cancel(srop_a)
        root1.destroy()
    root1.protocol("WM_DELETE_WINDOW", deswin)

def n_del():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def n_del1():
        try:
            x=clip.paste()
            if var1.get()==1:
                x=x.replace("\n","").replace("\r","").replace("\r\n","")
            if var2.get()==1:
                x=x.replace(".",".\n").replace("。","。\n")
            clip.copy(x)
            messagebox.showinfo("取得","コピーしました")
        except:
            messagebox.showerror("エラー","失敗しました")


    var1=IntVar()
    check1=ttk.Checkbutton(frame,text="改行を削除",variable=var1,onvalue=1,offvalue=0)
    var2=IntVar()
    check2=ttk.Checkbutton(frame,text="区切りで改行",variable=var2,onvalue=1,offvalue=0)
    button=ttk.Button(frame,text="実行",command=n_del1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    var1.set(1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    check1.grid(row=1,column=0)
    check2.grid(row=1,column=1)
    button.grid(row=2,column=0,columnspan=2)

def image_mirror():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                name=os.path.basename(x)
                y=x.replace(name,"mirror_"+name)
                img=Image.open(x)
                im_mirror = ImageOps.mirror(img)
                im_mirror.save(y)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","終了しました")

    label=ttk.Label(frame,text="画像ファイルをここに\nドロップしてください",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def video_mirror():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                name=os.path.basename(x)
                y=x.replace(name,"mirror_"+name)
                mov=VideoFileClip(x)
                if var1.get()==1:
                    mov=mirror_x(mov)
                if var2.get()==1:
                    mov=mirror_y(mov)
                mov.write_videofile(y,logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","ミラー動画を保存しました")


    var1=IntVar()
    var2=IntVar()
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    check1=ttk.Checkbutton(frame,text="左右反転",onvalue=1,offvalue=0,variable=var1)
    check2=ttk.Checkbutton(frame,text="上下反転",onvalue=1,offvalue=0,variable=var2)
    label=ttk.Label(frame,text="動画ファイルをここに\nドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    check1.grid(row=1,column=0)
    check2.grid(row=1,column=1)
    label.grid(row=2,column=0,columnspan=2)

def plain_paste():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)
    thread_stop=0

    def copy_plane():
        nonlocal thread_stop
        x=clip.paste()
        while thread_stop==0:
            x=clip.paste()
            clip.copy(x)
            time.sleep(0.5)

    def main_frame1(n):
        nonlocal thread_stop
        thread_stop=1
        t.join()
        frame.destroy()
        main_frame(n)

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(0),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="クリップボードの内容を\nそのままコピーします",font=("Helvetica", 16))
    t=threading.Thread(target=copy_plane,daemon=True)
    t.start()

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)



def power_control():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def turn_off_display():
        HWND_BROADCAST = 0xFFFF
        WM_SYSCOMMAND = 0x0112
        SC_MONITORPOWER = 0xF170
        MONITOR_OFF = 2

        # ウィンドウハンドル取得
        hwnd = ctypes.windll.user32.FindWindowW(None, None)
        ctypes.windll.user32.SendMessageW(hwnd, WM_SYSCOMMAND, SC_MONITORPOWER, MONITOR_OFF)


    button1=ttk.Button(frame,text="画面を消す",command=turn_off_display,width=12)
    button2=ttk.Button(frame,text="スリープ",width=12,command=lambda:ctypes.windll.PowrProf.SetSuspendState(0, 1, 0))
    button3=ttk.Button(frame,text="休止状態",width=12,command=lambda:ctypes.windll.PowrProf.SetSuspendState(1, 1, 0))
    button4=ttk.Button(frame,text="シャットダウン",width=12,command=lambda:os.system("shutdown /s /t 0"))
    button5=ttk.Button(frame,text="再起動",width=12,command=lambda:os.system("shutdown /r /t 0"))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    button1.grid(row=1,column=0,columnspan=2)
    button2.grid(row=2,column=0)
    button3.grid(row=2,column=1)
    button4.grid(row=3,column=0)
    button5.grid(row=3,column=1)

def sort_list():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def sort_list1():
        try:
            x_list=box.get("1.0","end-1c").rstrip().split("\n")
            if var1.get()==1:
                x_list.sort()
            else:
                x_list.sort(reverse=True)
            box.delete("1.0","end")
            box.insert("1.0","\n".join(x_list))
        except:
            messagebox.showerror("エラー","失敗しました")

    var1=IntVar()
    radio1=ttk.Radiobutton(frame,text="昇順",value=1,variable=var1)
    radio2=ttk.Radiobutton(frame,text="降順",value=2,variable=var1)
    box=ScrolledText(frame,width=30,height=20)
    button=ttk.Button(frame,text="ソート",command=sort_list1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    var1.set(1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    box.grid(row=2,column=0,columnspan=2)
    button.grid(row=3,column=0,columnspan=2)


def pdf_pass():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def change_password(src_path, dst_path, src_password):
        src_pdf = pypdf.PdfReader(src_path)
        src_pdf.decrypt(src_password)

        dst_pdf = pypdf.PdfWriter()
        dst_pdf.clone_reader_document_root(src_pdf)

        d = {key: src_pdf.metadata[key] for key in src_pdf.metadata.keys()}
        dst_pdf.add_metadata(d)

        dst_pdf.write(dst_path)

    def pdf_pass1(drop):
        try:
            x=drop.data.replace("{","").replace("}","").replace("\\","/")
            name=os.path.basename(x)
            y=x.replace(name,"unlock_"+name)
            change_password(x,y,entry.get())
            messagebox.showinfo("完了","パスワードを解除しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    label=ttk.Label(frame,text="パスワードをかけるPDFファイルを\nドロップしてください",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    entry=ttk.Entry(frame,width=30)
    label1=ttk.Label(frame,text="パスワード：")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',pdf_pass1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry.grid(row=1,column=1)
    label.grid(row=2,column=0,columnspan=2)

def bmi():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def bmi1():
        try:
            x1=float(entry1.get())
            x2=float(entry2.get())
            x3=x2/(x1/100)**2
            x4=(x1/100)**2*22
            entry3.delete(0,END)
            entry4.delete(0,END)
            entry3.insert(END,str(round(x3,2)))
            entry4.insert(END,str(round(x4,2))+"kg")
            x6=x2-x4
            if x3<18.5:
                x5="低体重"
            elif x3<25.0:
                x5="普通体重"
            elif x3<30.0:
                x5="肥満（1度）"
            elif x3<35.0:
                x5="肥満（2度）"
            elif x3<40.0:
                x5="肥満（3度）"
            else:
                x5="肥満（4度）"
            entry5.delete(0,END)
            entry5.insert(END,x5)
            entry6.delete(0,END)
            entry6.insert(END,str(round(x6,2))+"kg")
        except:
            messagebox.showerror("エラー","失敗しました")

    label1=ttk.Label(frame,text="身長(cm)：")
    label2=ttk.Label(frame,text="体重(kg)：")
    entry1=ttk.Entry(frame,width=30)
    entry2=ttk.Entry(frame,width=30)
    label3=ttk.Label(frame,text="BMI：")
    label4=ttk.Label(frame,text="適正体重：")
    label5=ttk.Label(frame,text="肥満度判定")
    label6=ttk.Label(frame,text="適正体重との差")
    entry3=ttk.Entry(frame,width=30)
    entry4=ttk.Entry(frame,width=30)
    entry5=ttk.Entry(frame,width=30)
    entry6=ttk.Entry(frame,width=30)
    button=ttk.Button(frame,text="計算",command=bmi1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    button.grid(row=3,column=0,columnspan=2)
    label3.grid(row=4,column=0)
    entry3.grid(row=4,column=1)
    label4.grid(row=5,column=0)
    entry4.grid(row=5,column=1)
    label5.grid(row=6,column=0)
    entry5.grid(row=6,column=1)
    label6.grid(row=7,column=0)
    entry6.grid(row=7,column=1)

def net_info():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def is_url(string):
        url_regex = re.compile(
            r'^(?:http|ftp)s?://'  # http:// or https://
            r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|'  # domain...
            r'localhost|'  # localhost
            r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'  # IP
            r'(?::\d+)?'  # optional port
            r'(?:/?|[/?]\S+)$', re.IGNORECASE)
        return re.match(url_regex, string) is not None

    def net_info1():
        try:
            code=entrya.get()
            if is_url(code):
                parsed_url = urlparse(code)
                domain = parsed_url.netloc
                ip = socket.gethostbyname(domain)
            else:
                ip=code
            url=f"http://ip-api.com/json/{str(ip)}?fields=status,country,regionName,city,district,zip,org"
            response = requests.get(url)
            data = response.json()
            if data['status'] == "fail":
                messagebox.showerror("エラー", "取得に失敗しました")
                return
            else:
                country= data['country']
                regionName = data['regionName']
                city= data['city']
                district = data['district']
                zip_1= data['zip']
                org= data['org']
                root2 = Free_window()
                root2.title("IP情報")
                root2.attributes("-topmost", True)
                root2.geometry('+%d+%d'%(x,y))
                label1=ttk.Label(root2,text="国：")
                label2=ttk.Label(root2,text="地域：")
                label3=ttk.Label(root2,text="都市：")
                label4=ttk.Label(root2,text="区：")
                label5=ttk.Label(root2,text="郵便番号：")
                label6=ttk.Label(root2,text="組織：")
                label7=ttk.Label(root2,text="IPアドレス：")
                entry1=ttk.Entry(root2,width=30)
                entry2=ttk.Entry(root2,width=30)
                entry3=ttk.Entry(root2,width=30)
                entry4=ttk.Entry(root2,width=30)
                entry5=ttk.Entry(root2,width=30)
                entry6=ttk.Entry(root2,width=30)
                entry7=ttk.Entry(root2,width=30)
                entry1.insert(END,country)
                entry2.insert(END,regionName)
                entry3.insert(END,city)
                entry4.insert(END,district)
                entry5.insert(END,zip_1)
                entry6.insert(END,org)
                entry7.insert(END,ip)

                label7.grid(row=0,column=0)
                entry7.grid(row=0,column=1)
                label1.grid(row=0+1,column=0)
                entry1.grid(row=0+1,column=1)
                label2.grid(row=1+1,column=0)
                entry2.grid(row=1+1,column=1)
                label3.grid(row=2+1,column=0)
                entry3.grid(row=2+1,column=1)
                label4.grid(row=3+1,column=0)
                entry4.grid(row=3+1,column=1)
                label5.grid(row=4+1,column=0)
                entry5.grid(row=4+1,column=1)
                label6.grid(row=5+1,column=0)
                entry6.grid(row=5+1,column=1)
        except:
            messagebox.showerror("エラー","失敗しました")

    label1=ttk.Label(frame,text="IPアドレスかURLを入力してください：")
    entrya=ttk.Entry(frame,width=30)
    button1=ttk.Button(frame,text="実行",command=net_info1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0,columnspan=2)
    entrya.grid(row=2,column=0,columnspan=2)
    button1.grid(row=3,column=0,columnspan=2)

def gif_split():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def gif_to_images(gif_path, output_path):
        gif = Image.open(gif_path)
        name=os.path.basename(gif_path)
        for frame in range(gif.n_frames):
            gif.seek(frame)
            frame_image = gif.copy()
            frame_image.save(f"{output_path}/{name}{frame}.png")

    def top(drop):
        try:
            x=drop.data.replace("{","").replace("}","").replace("\\","/")
            out=filedialog.askdirectory(title="保存先を選択してください")
            gif_to_images(x,out)
            messagebox.showinfo("完了","完了しました")
        except:
            messagebox.showerror("エラー","失敗しました")


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="GIFファイルを\n選択してください",font=("Helvetica", 15),padding=10)
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def trapezoid():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    origin=None
    x=None
    point=0
    canvas_img=None

    def scansnap(n_list):
        nonlocal x
        w_ratio = 1.0

        # p1:左上　p2:右上 p3:左下 p4:右下
        p1 = np.array(n_list[0])
        p2 = np.array(n_list[3])
        p3 = np.array(n_list[1])
        p4 = np.array(n_list[2])

        img = cv2.imdecode(
            np.fromfile(x, dtype=np.uint8),
            cv2.IMREAD_UNCHANGED)

        o_width = np.linalg.norm(p2 - p1)
        o_width = math.floor(o_width * w_ratio)
        o_height = np.linalg.norm(p3 - p1)
        o_height = math.floor(o_height)
        src = np.float32([p1, p2, p3, p4])
        dst = np.float32([[0, 0],[o_width, 0],[0, o_height],[o_width, o_height]])

        # 変換行列
        M = cv2.getPerspectiveTransform(src, dst)

        # 射影変換
        output = cv2.warpPerspective(img, M,(o_width, o_height))
        _, buf = cv2.imencode('*.png', output)
        buf.tofile(z)
    def img_top(drop):
        nonlocal canvas_img,pil_img,z,origin,pil_img_1,x
        try:
            x=drop.data.replace("{","").replace("}","").replace("\\","/")
            y= os.path.basename(x)
            z=x.replace(y,"trans_"+y)
            pil_img=pil_img_1=Image.open(x)
            if pil_img.width>pil_img.height:
                origin=0
                pil_img=pil_img.resize((int(ws/2),int(pil_img.height*((ws/2)/pil_img.width))),Image.LANCZOS)
            elif pil_img.width<pil_img.height:
                origin=1
                pil_img=pil_img.resize((int(pil_img.width*((hs/2)/pil_img.height)),int(hs/2)),Image.LANCZOS)
            elif pil_img.width==pil_img.height:
                origin=2
                pil_img=pil_img.resize((int(hs/2),int(hs/2)),Image.LANCZOS)
            canvas.delete("all")
            canvas_img= ImageTk.PhotoImage(pil_img)
            canvas.create_image(pil_img.width/2,pil_img.height/2,image=canvas_img)
            canvas.config(width=pil_img.width, height=pil_img.height)
        except:
            messagebox.showerror("エラー","失敗しました")

    def on_press(event):
        nonlocal point
        nonlocal start_x
        nonlocal start_y,startx_1,starty_1
        nonlocal pointX0,pointY0,pointX1,pointY1,pointX2,pointY2,pointX3,pointY3
        if point==3:
            start_x = canvas.canvasx(event.x)
            start_y = canvas.canvasy(event.y)
            pointX3=start_x
            pointY3=start_y
            canvas.delete("rect")
            canvas.unbind("<Motion>")
            canvas.create_line(start_x, start_y, startx_1, starty_1, tags=f"rect{point}", width=1.5,fill="#06ff06")
            point=0
        elif point==0:
            canvas.delete("rect0")
            canvas.delete("rect1")
            canvas.delete("rect2")
            canvas.delete("rect3")
            start_x = canvas.canvasx(event.x)
            startx_1=start_x
            start_y = canvas.canvasy(event.y)
            starty_1=start_y
            pointX0=start_x
            pointY0=start_y
            point+=1
            canvas.bind("<Motion>", on_move_press)
        elif point==1:
            start_x = canvas.canvasx(event.x)
            start_y = canvas.canvasy(event.y)
            pointX1=start_x
            pointY1=start_y
            point+=1
        elif point==2:
            start_x = canvas.canvasx(event.x)
            start_y = canvas.canvasy(event.y)
            pointX2=start_x
            pointY2=start_y
            point+=1

    def on_move_press(event):
        nonlocal end_x, end_y,point
        end_x, end_y = event.x, event.y
        canvas.delete(f"rect{point}")
        canvas.create_line(start_x, start_y, end_x, end_y, tags=f"rect{point}", width=1.5,fill="#06ff06")


    def mozaic():
        nonlocal origin,x
        try:
            point_list=[[pointX0,pointY0],[pointX1,pointY1],[pointX2,pointY2],[pointX3,pointY3]]
            #left_top = min(point_list, key=lambda coord: coord[0] + coord[1])
            #left_bottom = min(point_list, key=lambda coord: coord[0] - coord[1])
            #right_bottom = max(point_list, key=lambda coord: coord[0] + coord[1])
            #right_top = max(point_list, key=lambda coord: coord[0] - coord[1])
            #point_list=[left_top,left_bottom,right_bottom,right_top]

            pil_img_1=Image.open(x)
            if origin==2:
                for i in range(4):
                    point_list[i][0]=int(point_list[i][0]*pil_img_1.width/int(hs/2))
                    point_list[i][1]=int(point_list[i][1]*pil_img_1.height/int(hs/2))
            elif origin==1:
                for i in range(4):
                    point_list[i][0]=int(point_list[i][0]*pil_img_1.width/int(pil_img_1.width*((hs/2)/pil_img_1.height)))
                    point_list[i][1]=int(point_list[i][1]*pil_img_1.height/int(hs/2))
            elif origin==0:
                for i in range(4):
                    point_list[i][0]=int(point_list[i][0]*pil_img_1.width/int(ws/2))
                    point_list[i][1]=int(point_list[i][1]*pil_img_1.height/int(pil_img_1.height*((ws/2)/pil_img_1.width)))
            scansnap(point_list)
            messagebox.showinfo("完了","完了しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    start_x=None
    start_y=None
    end_x=None
    end_y=None
    pil_img=None
    z=None
    pil_img_1=None
    startx_1=None
    starty_1=None
    pointX1=None
    pointY1=None
    pointX2=None
    pointY2=None
    pointX3=None
    pointY3=None
    pointX0=None
    pointY0=None

    root.update()
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',img_top)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    canvas=Canvas(frame,width=300,height=300,bg="gray")
    button1=ttk.Button(frame,text="台形補正",command=mozaic)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")

    canvas.bind("<ButtonPress-1>", on_press)


    frame.pack()
    buttonX.pack(side=TOP)
    buttonY.pack(side=TOP)
    canvas.pack(expand=True,side=TOP,fill=BOTH)
    button1.pack(side=TOP)

def shadow_delete():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                y= os.path.basename(x)
                z=x.replace(y,"shadow_"+y)
                img = cv2.imdecode(
                    np.fromfile(x, dtype=np.uint8),
                    cv2.IMREAD_UNCHANGED)
                ksize = 51
                blur = cv2.blur(img, (ksize, ksize))
                rij = img/blur
                index_1 = np.where(rij >= 1.00) # 1以上の値があると邪魔なため
                rij[index_1] = 1
                rij_int = np.array(rij*255, np.uint8) # 除算結果が実数値になるため整数に変換
                rij_HSV = cv2.cvtColor(rij_int, cv2.COLOR_BGR2HSV)
                ret, thresh = cv2.threshold(rij_HSV[:,:,2], 0, 255, cv2.THRESH_OTSU)
                rij_HSV[:,:,2] = thresh
                rij_ret = cv2.cvtColor(rij_HSV, cv2.COLOR_HSV2BGR)
                _, buf = cv2.imencode('*.png', rij_ret)
                buf.tofile(z)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","終了しました")


    label=ttk.Label(frame,text="画像フォルダをここに\nドロップしてください",font=("Helvetica", 20))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def video_convert():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                clip=VideoFileClip(x)
                _,youso=os.path.splitext(x)
                if var1.get()==1 and youso!=".mp4":
                    clip.write_videofile(x.replace(youso,".mp4"),codec="libx264",logger=None)
                if var2.get()==1 and youso!=".avi":
                    clip.write_videofile(x.replace(youso,".avi"),codec="libx264",logger=None)
                if var3.get()==1 and youso!=".mov":
                    clip.write_videofile(x.replace(youso,".mov"),codec="libx264",logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","終了しました")

    var1=IntVar()
    check1=ttk.Checkbutton(frame,text="mp4",variable=var1,onvalue=1,offvalue=0)
    var2=IntVar()
    check2=ttk.Checkbutton(frame,text="avi",variable=var2,onvalue=1,offvalue=0)
    var3=IntVar()
    check3=ttk.Checkbutton(frame,text="mov",variable=var3,onvalue=1,offvalue=0)
    label=ttk.Label(frame,text="動画フォルダをここに\nドロップしてください",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)

    frame.pack()
    buttonX.grid(row=0,column=1,columnspan=2)
    buttonY.grid(row=0,column=0)
    check1.grid(row=1,column=0)
    check2.grid(row=1,column=1)
    check3.grid(row=1,column=2)
    label.grid(row=2,column=0,columnspan=3)

def video_fps():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                name=os.path.basename(x)
                y=x.replace(name,"fps"+entry.get()+"_"+name)
                clip = VideoFileClip(x)
                n=clip.set_fps(int(entry.get()))
                n.write_videofile(y,logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","完了しました")


    label1=ttk.Label(frame,text="動画フォルダをここに\nドロップしてください",font=("Helvetica", 16))
    label2=ttk.Label(frame,text="fpsを入力してください：")
    entry=ttk.Entry(frame,width=10)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label2.grid(row=1,column=0)
    entry.grid(row=1,column=1)
    label1.grid(row=2,column=0,columnspan=2)

def video_bgm():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def top1(drop):
        x=drop.data.replace("{","").replace("}","").replace("\\","/")
        entry.delete(0,END)
        entry.insert(0,x)

    def top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                name=os.path.basename(x)
                y=x.replace(name,"bgm_"+name)
                clip = VideoFileClip(x)
                bgm_clip = AudioFileClip(entry.get())
                bgm_loop = bgm_clip.audio_loop(duration=clip.duration)
                n=clip.set_audio(bgm_loop)
                n.write_videofile(y,logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","完了しました")

    label1=ttk.Label(frame,text="動画フォルダをここに\nドロップしてください",font=("Helvetica", 16))
    label2=ttk.Label(frame,text="音楽ファイルを隣にドロップ：")
    entry=ttk.Entry(frame,width=20)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',top)
    label2.drop_target_register(DND_FILES)
    label2.dnd_bind('<<Drop>>',top1)
    entry.drop_target_register(DND_FILES)
    entry.dnd_bind('<<Drop>>',top1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label2.grid(row=1,column=0)
    entry.grid(row=1,column=1)
    label1.grid(row=2,column=0,columnspan=2)

def duplication():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def duplication1():
        try:
            x=box.get("1.0",END)
            text=x.split("\n")
            text1=list(dict.fromkeys(text))
            result = '\n'.join(text1)
            box.delete("1.0",END)
            box.insert("1.0",result)
        except:
            messagebox.showerror("エラー","失敗しました")

    box=ScrolledText(frame,width=40,height=30)
    button=ttk.Button(frame,text="実行",command=duplication1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="テキストを改行区切りで入力してください")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)
    box.grid(row=2,column=0,columnspan=2)
    button.grid(row=3,column=0,columnspan=2)

def video_thumbnail():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def download_image(url, file_name):
        response = requests.get(url)
        with open(file_name, 'wb') as file:
            file.write(response.content)

    def get_thumbnail(url):
        ydl_opts = {
            'format': 'best',
            'quiet': True,
            'no_warnings': True,
            'skip_download': True,
            'writesubtitles': False,
            'writeautomaticsub': False,
            'outtmpl': '%(id)s.%(ext)s',
            'forcethumbnail': True
        }

        with YoutubeDL(ydl_opts) as ydl:
            info_dict = ydl.extract_info(url, download=False)
            thumbnail_url = info_dict.get('thumbnail')
            return thumbnail_url

    def video_thumbnail1():

        video_urls =box.get("1.0",END).strip().split("\n")
        n=0
        y=filedialog.askdirectory()
        for url in video_urls:
            thumbnail_url = get_thumbnail(url)
            download_image(thumbnail_url, y+"/thumbnail"+str(n)+".jpg")
            n+=1
        messagebox.showinfo("完了","完了しました")


    box=ScrolledText(frame,width=40,height=30,wrap=NONE)
    label=ttk.Label(frame,text="動画のURLを改行区切りで入力してください")
    button=ttk.Button(frame,text="サムネ取得",command=video_thumbnail1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)
    box.grid(row=2,column=0,columnspan=2)
    button.grid(row=3,column=0,columnspan=2)

def address_english():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def address_english1():
        nonlocal data,value_list
        adoress_list=entry1.get().split("-")
        url = f"https://madefor.github.io/postal-code-api/api/v1/{adoress_list[0]}/{adoress_list[1]}.json"
        response = requests.get(url)
        data = response.json()
        if data["data"][0]["en"]["prefecture"]=="":
            messagebox.showerror("エラー","英語郵便番号を取得できませんでした")
            return
        else:
            value_list=[]
            for i in range(len(data["data"])):
                x=data["data"][i]["ja"]["prefecture"]+data["data"][i]["ja"]["address1"]+data["data"][i]["ja"]["address2"]+data["data"][i]["ja"]["address3"]
                value_list.append(x)
            select["values"]=value_list
            select.current(0)

            kakasi = pykakasi.kakasi()
            if entry3.get()=="":
                built=""
            else:
                data_l=kakasi.convert(entry3.get())
                hepburn_list = [d['hepburn'] for d in data_l]  # 'hepburn'の値
                built = ''.join(hepburn_list)+","  # 連結
            if entry4.get()=="":
                heya=""
            else:
                heya=f"#{entry4.get()}, "
            line1=heya+built
            if data["data"][0]["en"]["address2"]=="":
                adoress1=data["data"][0]["en"]["address1"].split(",")[0]
                adoress2=data["data"][0]["en"]["address1"].split(",")[1].lstrip()
            else:
                adoress1=data["data"][0]["en"]["address1"]
                adoress2=data["data"][0]["en"]["address2"]
            line2=entry2.get()+", "+adoress1+","
            if data["data"][0]["en"]["prefecture"]=="Tokyo":
                todouhuken="Tokyo"
            elif data["data"][0]["en"]["prefecture"]=="Hokkaido":
                todouhuken="Hokkaido"
            elif data["data"][0]["en"]["prefecture"]=="Kyoto":
                todouhuken="Kyoto-fu"
            elif data["data"][0]["en"]["prefecture"]=="Osaka":
                todouhuken="Osaka-fu"
            else:
                todouhuken=data["data"][0]["en"]["prefecture"]+"-ken"
            line3=adoress2+", "+todouhuken+", "+entry1.get()+","
            line4="JAPAN"
            text=line1+"\n"+line2+"\n"+line3+"\n"+line4
            box.delete("1.0",END)
            box.insert("1.0",text)
        return

    def change_adoress(event):
        num=value_list.index(select.get())
        kakasi = pykakasi.kakasi()
        if entry3.get()=="":
            built=""
        else:
            data_l=kakasi.convert(entry3.get())
            hepburn_list = [d['hepburn'] for d in data_l]  # 'hepburn'の値
            built = ''.join(hepburn_list)+","
        if entry4.get()=="":
            heya=""
        else:
            heya=f"#{entry4.get()}, "
        line1=heya+built
        if data["data"][num]["en"]["address2"]=="":
            adoress1=data["data"][num]["en"]["address1"].split(",")[0]
            adoress2=data["data"][num]["en"]["address1"].split(",")[1]
        else:
            adoress1=data["data"][num]["en"]["address1"]
            adoress2=data["data"][num]["en"]["address2"]
        line2=entry2.get()+", "+adoress1+","
        if data["data"][num]["en"]["prefecture"]=="Tokyo":
            todouhuken="Tokyo"
        elif data["data"][num]["en"]["prefecture"]=="Hokkaido":
            todouhuken="Hokkaido"
        elif data["data"][num]["en"]["prefecture"]=="Kyoto":
            todouhuken="Kyoto-fu"
        elif data["data"][num]["en"]["prefecture"]=="Osaka":
            todouhuken="Osaka-fu"
        else:
            todouhuken=data["data"][num]["en"]["prefecture"]+"-ken"
        line3=adoress2.lstrip()+", "+todouhuken+", "+entry1.get()+","
        line4="JAPAN"
        text=line1+"\n"+line2+"\n"+line3+"\n"+line4
        box.delete("1.0",END)
        box.insert("1.0",text)

    value_list=[]
    data=None
    label1=ttk.Label(frame,text="郵便番号(-区切り)：")
    entry1=ttk.Entry(frame,width=20)
    label2=ttk.Label(frame,text="番地・号(-区切り)：")
    entry2=ttk.Entry(frame,width=20)
    label3=ttk.Label(frame,text="アパート/建物名：")
    entry3=ttk.Entry(frame,width=20)
    label4=ttk.Label(frame,text="部屋名：")
    entry4=ttk.Entry(frame,width=20)
    button1=ttk.Button(frame,text="実行",command=address_english1)
    select=ttk.Combobox(frame,state="readonly",width=60,values=value_list)
    box=ScrolledText(frame,width=60,height=10)
    select.bind('<<ComboboxSelected>>',change_adoress )
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    label3.grid(row=3,column=0)
    entry3.grid(row=3,column=1)
    label4.grid(row=4,column=0)
    entry4.grid(row=4,column=1)
    button1.grid(row=5,column=0,columnspan=2)
    select.grid(row=6,column=0,columnspan=2)
    box.grid(row=7,column=0,columnspan=2)

def input_stop():
    global frame,buttonY
    def input_stop1():
        root1 = Free_window()
        root.iconify()

        AllKeySet = {
        'a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z',
        'A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
        '!','"','#','$','%','&',"'",'(',')','=','~','|','-','^','\\','@','[',';',':',']',',','`','{','}','<','>','?','_',
        '0','1','2','3','4','5','6','7','8','9','.','+','*','/',"num lock",
        "esc","f1","f2","f3","f4","f5","f6","f7","f8","f9","f10","f11","f12","print screen","scroll lock","pause","sys req",
        "insert","delete","home","end","page up","page down",
        "right","down","left","up",
        "半角/全角","tab","caps lock","shift","ctrl","left windows","alt","無変換","space",
        "変換","ひらがな","right alt","right windows","menu","right ctrl","right shift","enter","backspace"
        }

        # キー入力無効化
        for keys in AllKeySet:
            key.block_key(keys)

        def click_close():
            key.unhook_all()
            root.deiconify()
            root1.destroy()

        def middleClick(event):
            nonlocal num
            num+=1
            if num==3:
                click_close()

        num=0
        root1.title('入力停止')
        root1.attributes("-topmost", True)
        frame_1 = ttk.Frame(root1, style="Transparent.TFrame")
        root1.wm_attributes("-transparentcolor", "snow")
        root1.attributes("-alpha", 0.8)
        style = ttk.Style()
        style.configure("Transparent.TFrame", background="gray")
        style.configure("Transparent.TFrame", borderwidth=0)
        style.configure("Transparent.TFrame", highlightthickness=0)
        style.configure("Transparent.TFrame", relief="flat")
        root1.attributes("-fullscreen", True)
        root1.focus_set()
        label=ttk.Label(frame_1,text="入力を停止しました\nミドルクリック3回で解除されます",font=("Helvetica", 20))

        frame_1.pack(fill="both", expand=True)
        label.place(relx=0.5, rely=0.5, anchor=CENTER)
        root1.protocol("WM_DELETE_WINDOW", click_close)
        root1.bind('<Button-2>', middleClick)

    if intermediate_screen==0:
        main_frame_delete()
        frame = ttk.Frame(root, padding=16)
        button=ttk.Button(frame,text='入力停止開始',command=input_stop1,padding=[80,10,80,10])
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        button.grid(row=1,column=0,columnspan=2)
    elif intermediate_screen==1:
        input_stop1()

def renban_list():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=4)

    def renban_list1():
        try:
            text=box.get("1.0","end").strip("\n").split("\n")
            new_text=[]
            for i,x in enumerate(text):
                start=""
                if var2.get()==2:
                    start="["+start
                elif var2.get()==3:
                    start="「"+start
                elif var2.get()==4:
                    start="("+start
                if var1.get()==1:
                    start=start+f"{i+1}"
                elif var1.get()==2:
                    start=start+en1[i]
                elif var1.get()==3:
                    start=start+ja1[i]
                elif var1.get()==4:
                    start=start+ja2[i]
                elif var1.get()==5:
                    start=start+roma[i]
                elif var1.get()==6:
                    start=start+en2[i]
                elif var1.get()==7:
                    start=start;kanzi[i]
                if var2.get()==2:
                    start=start+"]"
                elif var2.get()==3:
                    start=start+"」"
                elif var2.get()==4:
                    start=start+")"
                if var3.get()==1:
                    start=start+"."
                start=entry1.get()+start
                #if var3.get()==1:
                #    end=start+x
                #else:
                end=start+" "+x
                new_text.append(end)
            root2=Free_window()
            root2.title("連番リスト")
            root2.attributes("-topmost", True)
            box1=ScrolledText(root2,width=60,height=30)
            if var4.get()==1:
                box1.insert("1.0","  ".join(new_text))
            else:
                box1.insert("1.0","\n".join(new_text))
            box1.pack(side=TOP)
            button_1=ttk.Button(root2,text="コピー",command=lambda:clip.copy(box1.get("1.0","end").strip("\n")))
            button_1.pack(side=TOP)
        except:
            messagebox.showerror("エラー","失敗しました")


    box=ScrolledText(frame,width=60,height=30)
    label=ttk.Label(frame,text="改行区切りでリストを入力してください")
    button=ttk.Button(frame,text="実行",command=renban_list1)
    label1=ttk.Label(frame,text="連番開始：")
    var1=IntVar()
    var1.set(1)
    radio1_1=ttk.Radiobutton(frame,text="数字",value=1,variable=var1)
    radio1_2=ttk.Radiobutton(frame,text="ローマ字小文字",value=2,variable=var1)
    radio1_3=ttk.Radiobutton(frame,text="ひらがな",value=3,variable=var1)
    radio1_4=ttk.Radiobutton(frame,text="カタカナ",value=4,variable=var1)
    radio1_5=ttk.Radiobutton(frame,text="ローマ数字",value=5,variable=var1)
    radio1_6=ttk.Radiobutton(frame,text="ローマ字大文字",value=6,variable=var1)
    radio1_7=ttk.Radiobutton(frame,text="漢数字",value=7,variable=var1)
    radio1_8=ttk.Radiobutton(frame,text="無し",value=8,variable=var1)
    var2=IntVar()
    label2=ttk.Label(frame,text="記号：")
    radio2_1=ttk.Radiobutton(frame,text="無し",variable=var2,value=1)
    radio2_2=ttk.Radiobutton(frame,text="[]",value=2,variable=var2)
    radio2_3=ttk.Radiobutton(frame,text="「」",value=3,variable=var2)
    radio2_4=ttk.Radiobutton(frame,text="()",value=4,variable=var2)
    var2.set(1)
    label3=ttk.Label(frame,text="その他：")
    var3=IntVar()
    var3.set(0)
    check12=ttk.Checkbutton(frame,text="コンマ",onvalue=1,offvalue=0,variable=var3)
    var4=IntVar()
    var4.set(0)
    check13=ttk.Checkbutton(frame,text="改行なし",onvalue=1,offvalue=0,variable=var4)
    label4=ttk.Label(frame,text="前入れ文字：")
    entry1=ttk.Entry(frame,width=40)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    en1="abcdefghijklmnopqrstuvwxyz"
    ja1="あいうえおかきくけこさしすせそたちつてとなにぬねのはひふへほまみむめもやゆよらりるれろわをん"
    ja2="アイウエオカキクケコサシスセソタチツテトナニヌネノハヒフヘホマミムメモヤユヨラリルレロワヲン"
    en2="ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    roma="ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩⅪⅫ"
    kanzi=["一","二","三","四","五","六","七","八","九","十","十一","十二","十三","十四","十五","十六","十七","十八","十九","二十"]

    frame.pack(fill="both", expand=True)
    buttonX.grid(row=0,column=1,columnspan=2)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    radio1_1.grid(row=1,column=1,sticky=W)
    radio1_2.grid(row=1,column=2,sticky=W)
    radio1_3.grid(row=2,column=1,sticky=W)
    radio1_4.grid(row=2,column=2,sticky=W)
    radio1_5.grid(row=3,column=1,sticky=W)
    radio1_6.grid(row=3,column=2,sticky=W)
    radio1_7.grid(row=4,column=1,sticky=W)
    radio1_8.grid(row=4,column=2,sticky=W)
    label2.grid(row=3+2,column=0)
    radio2_1.grid(row=3+2,column=1,sticky=W)
    radio2_2.grid(row=3+2,column=2,sticky=W)
    radio2_3.grid(row=4+2,column=1,sticky=W)
    radio2_4.grid(row=4+2,column=2,sticky=W)
    label3.grid(row=5+2,column=0)
    check12.grid(row=5+2,column=1,sticky=W)
    check13.grid(row=5+2,column=2,sticky=W)
    label4.grid(row=6+2,column=0)
    entry1.grid(row=6+2,column=1,columnspan=2)
    label.grid(row=7+2,column=0,columnspan=3)
    box.grid(row=8+2,column=0,columnspan=3)
    button.grid(row=9+2,column=0,columnspan=3)

def web_camera():
   global frame,buttonY
   main_frame_delete()
   frame = ttk.Frame(root)

   def main_frame1(num):
       nonlocal stop
       stop=1
       vcap.release()
       time.sleep(0.1)
       main_frame(num)

   def save_path():
       nonlocal path_num,path
       path = filedialog.askdirectory()
       path_label.configure(text=f"保存先：{path}")
       path_num=1

   def update():
       nonlocal canvas1,photo,stop
       if stop==0:
           _, frameS = vcap.read()
           frameD = cv2.cvtColor(frameS, cv2.COLOR_BGR2RGB)
           photo = ImageTk.PhotoImage(image=Image.fromarray(frameD))
           canvas1.create_image(0, 0, image=photo, anchor=NW)
           canvas1.update()
           root.after(30, update)
       else:
           return

   def press_snapshot_button():
       nonlocal path,path_num
       _, frameS = vcap.read()
       frameN = cv2.cvtColor(frameS, cv2.COLOR_BGR2RGB)
       if path_num==1:
           _,buf = cv2.imencode('*.png', cv2.cvtColor( frameN, cv2.COLOR_BGR2RGB ))
           buf.tofile(path+"/frame-" + time.strftime("%Y-%m-%d-%H-%M-%S") + ".jpg")
       else:
           path=filedialog.askdirectory()
           path_label.configure(text=f"保存先：{path}")
           path_num=1
           _,buf = cv2.imencode('*.png', cv2.cvtColor( frameN, cv2.COLOR_BGR2RGB ))
           buf.tofile(path+"/frame-" + time.strftime("%Y-%m-%d-%H-%M-%S") + ".jpg")
   video_source = 0
   vcap = cv2.VideoCapture(video_source)
   width = int(vcap.get(cv2.CAP_PROP_FRAME_WIDTH))
   height = int(vcap.get(cv2.CAP_PROP_FRAME_HEIGHT))
   path_num=0
   photo=None
   path=None
   stop=0

   canvas1 = Canvas(frame)
   canvas1.configure(width=width, height=height)
   snapshot = ttk.Button(frame, text='スクショ',command=press_snapshot_button)
   path_label=ttk.Label(frame,text="保存先：")
   path_button=ttk.Button(frame,text="参照",command=save_path)
   buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
   buttonY=Button(frame,text="戻る",command=lambda:main_frame1(1),font=("Helvetica", 7),bg="gray",fg="white")

   update()

   frame.pack()
   buttonX.grid(row=0,column=1,columnspan=2)
   buttonY.grid(row=0,column=0)
   canvas1.grid(row=1,column=0,columnspan=2)
   snapshot.grid(row=3,column=0)
   path_label.grid(row=2,column=0,columnspan=2)
   path_button.grid(row=3,column=1)

def vidwo_fade():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def top1(drop):
        file=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(file)
        for x in file_list:
            try:
                name=os.path.basename(x)
                clip = VideoFileClip(x)
                if fade_in.get()==1:
                    clip = fadein(clip, int(entry1.get()))
                if fade_out.get()==1:
                    clip = fadeout(clip, int(entry2.get()))
                name_res=x.replace(name,f"fade_{name}")
                clip.write_videofile(name_res,logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","処理が完了しました")

    def select_fade():
        if fade_in.get()==1:
            entry1.configure(state="normal")
        else:
            entry1.configure(state="disabled")
        if fade_out.get()==1:
            entry2.configure(state="normal")
        else:
            entry2.configure(state="disabled")

    label1=ttk.Label(frame,text="動画ファイルをここに\nドロップしてください",font=("Helvetica", 18))
    fade_in=IntVar()
    fade_in.set(0)
    check1=ttk.Checkbutton(frame,text="フェードイン(秒)",onvalue=1,offvalue=0,variable=fade_in,command=select_fade)
    fade_out=IntVar()
    fade_out.set(0)
    check2=ttk.Checkbutton(frame,text="フェードアウト(秒)",onvalue=1,offvalue=0,variable=fade_out,command=select_fade)
    entry1=ttk.Entry(frame,state="disabled",width=10)
    entry2=ttk.Entry(frame,state="disabled",width=10)
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    check1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    check2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    label1.grid(row=3,column=0,columnspan=2)

def aes_file():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    # AES-256オブジェクト作成
    def aes_new(password, iv):
        sha = Crypto.Hash.SHA256.new()
        sha.update(password.encode())
        return AES.new(sha.digest(), AES.MODE_CFB, iv)
    # 暗号化
    def encrypt(data, password):
        iv = Crypto.Random.new().read(AES.block_size)
        return iv + aes_new(password, iv).encrypt(data)
    # 復号化
    def decrypt(data, password):
        iv, cipher = data[:AES.block_size], data[AES.block_size:]
        return aes_new(password, iv).decrypt(cipher)

    def top(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for file in file_list:
            with open(file, 'rb') as f:
                data = f.read()
            if var.get()==1:
                savedata = encrypt(data, entry1.get())
                youso=str(os.path.splitext(file)[1])
                name=os.path.splitext(os.path.basename(file))[0]
                folder=os.path.dirname(file)
                new_name=folder+"/"+name+youso.replace(".","_")
            else:
                savedata = decrypt(data, entry1.get())
                file_name_without_extension, extension = os.path.splitext(file)
                new_name=file_name_without_extension.replace("_", ".") + extension
            with open(new_name, 'wb') as f:
                f.write(savedata)
        messagebox.showinfo("完了","保存しました")

    label1=ttk.Label(frame,text="パスワード：")
    var=IntVar()
    var.set(1)
    radio1=ttk.Radiobutton(frame,text="暗号化",value=1,variable=var)
    radio2=ttk.Radiobutton(frame,text="復号化",value=2,variable=var)
    entry1=ttk.Entry(frame,width=30)
    label=ttk.Label(frame,text="ファイルをドロップ\nしてください",font=("Helvetica", 18))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    radio1.grid(row=2,column=0)
    radio2.grid(row=2,column=1)
    label.grid(row=3,column=0,columnspan=2)

def video_str():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def download_srt_file(video_id):
        try:
            id1=video_id.replace("https://www.youtube.com/watch?v=","")
            transcript_list = YouTubeTranscriptApi.list_transcripts(id1)
            if var3.get()==2:
                transcript = transcript_list.find_generated_transcript([var1.get()])
            else:
                transcript = transcript_list.find_manually_created_transcript([var1.get()])
            srt_data = transcript.fetch()
            srt_content = dict_to_srt(srt_data)
            if var2.get()==1:
                file_path=filedialog.asksaveasfilename(
                    title = "名前を付けて保存",
                    filetypes = [("SRT", ".srt")],
                    defaultextension = "srt",
                    initialfile="text"
                    )
                with open(file_path, "w", encoding="utf-8") as file:
                    file.write(srt_content)
            else:
                num=0
                new_list=[]
                text_list=srt_content.split("\n")
                for x in text_list:
                    num+=1
                    if num%4==3:
                        new_list.append(x)
                    elif num%4==4:
                        new_list.append(x)
                new_text="\n".join(new_list)
                file_path=filedialog.asksaveasfilename(
                    title = "名前を付けて保存",
                    filetypes = [("TXT", ".txt")],
                    defaultextension = "txt",
                    initialfile="text"
                    )
                with open(file_path, "w", encoding="utf-8") as file:
                    file.write(new_text)
            messagebox.showinfo("完了","保存しました")
        except:
                messagebox.showerror("エラー","字幕取得失敗")

    def dict_to_srt(data):
        srt = ""
        for i, entry in enumerate(data, start=1):
            srt += str(i) + "\n"
            if i < len(data):
                srt += format_time(entry['start']) + " --> " + format_time(data[i]["start"]) + "\n"
            else:
                srt += format_time(entry['start']) + " --> " + format_time(entry['start'] + entry['duration']) + "\n"
            srt += entry['text'] + "\n\n"
        return srt

    def format_time(seconds):
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        seconds1 = int(seconds % 60)
        milliseconds = int((seconds % 1) * 1000)
        return f"{hours:02d}:{minutes:02d}:{seconds1:02d},{milliseconds:03d}"


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    box=ttk.Entry(frame,width=30)
    var1=StringVar()
    var1.set("ja")
    radio1_1=ttk.Radiobutton(frame,text="日本語",value="ja",variable=var1)
    radio1_2=ttk.Radiobutton(frame,text="英語",value="en",variable=var1)
    var2=IntVar()
    var2.set(1)
    radio2_1=ttk.Radiobutton(frame,text="SRT",value=1,variable=var2)
    radio2_2=ttk.Radiobutton(frame,text="TXT",value=2,variable=var2)
    var3=IntVar()
    var3.set(1)
    radio3_1=ttk.Radiobutton(frame,text="手動字幕",value=1,variable=var3)
    radio3_2=ttk.Radiobutton(frame,text="自動字幕",value=2,variable=var3)
    button=ttk.Button(frame,text="変換",command=lambda:download_srt_file(box.get().strip().strip("\n")))
    label=ttk.Label(frame,text="URL：")
    label1=ttk.Label(frame,text="言語：")
    label2=ttk.Label(frame,text="書式：")
    label3=ttk.Label(frame,text="種類：")

    frame.pack()
    buttonX.grid(row=0,column=1,columnspan=2)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    radio1_1.grid(row=1,column=1)
    radio1_2.grid(row=1,column=2)
    label2.grid(row=2,column=0)
    radio2_1.grid(row=2,column=1)
    radio2_2.grid(row=2,column=2)
    label3.grid(row=3,column=0)
    radio3_1.grid(row=3,column=1)
    radio3_2.grid(row=3,column=2)
    label.grid(row=4,column=0)
    box.grid(row=4,column=1,columnspan=2)
    button.grid(row=5,column=0,columnspan=3)

def delete2file():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def dnd(drop):
        file=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(file)
        for x in x_list:
            try:
                dod_wipe_file(x)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","終了しました")

    def dod_wipe_file(file_path):
        if os.path.isfile(file_path):
            with open(file_path, 'rb+') as file:
                file_size = os.path.getsize(file_path)
                # 上書き
                file.seek(0)
                file.write(b'0' * file_size)
                # 書き換え
                file.seek(0)
                random_data = bytearray(os.urandom(file_size))
                file.write(random_data)
                # 上書き
                file.seek(0)
                file.write(b'0' * file_size)
                file.close()
            os.remove(file_path)

        elif os.path.isdir(file_path):
            for roots, dirs, files in os.walk(file_path, topdown=False):
                for name in files:
                    file = os.path.join(roots, name)
                    if os.path.isfile(file):
                        dod_wipe_file(file)
                for name in dirs:
                    dir_path = os.path.join(roots, name)
                    if os.path.isdir(dir_path):
                        shutil.rmtree(dir_path)
            shutil.rmtree(file_path)

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="ファイル・フォルダを\nドロップしてください",font=("Helvetica", 18))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',dnd)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)


def window_center():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def on_combo_click(event):
        combo.config(values=[a for a in get_window_titles() if a != ''])

    def get_window_titles():
        def enum_windows_callback(hwnd, titles):
            if win32gui.IsWindowVisible(hwnd):
                titles.append(win32gui.GetWindowText(hwnd))

        titles = []
        win32gui.EnumWindows(enum_windows_callback, titles)
        return titles

    def center_window(window_handle):
        # ウィンドウの位置とサイズを取得
        rect = win32gui.GetWindowRect(window_handle)
        window_width = rect[2] - rect[0]
        window_height = rect[3] - rect[1]

        screen_width = win32api.GetSystemMetrics(0)
        screen_height = win32api.GetSystemMetrics(1)

        x = int((screen_width - window_width) / 2)
        y = int((screen_height - window_height) / 2)
        left, top, right, bottom = win32gui.GetWindowRect(window_handle)
        width = right - left
        height = bottom - top
        win32gui.SetWindowPos(window_handle, 0, x, y, width, height, 0)

    def window_center1():
        window_handle = win32gui.FindWindow(None, combo.get())
        if window_handle != 0:
            # 中央配置
            center_window(window_handle)
        else:
            messagebox.showerror("エラー","ウィンドウが見つかりませんでした")

    combo=ttk.Combobox(frame,values=[a for a in get_window_titles() if a != ''],state="readonly",width=30)
    label1=ttk.Label(frame,text="ウィンドウを選択してください")
    button=ttk.Button(frame,text="中心に移動",command=window_center1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    combo.bind('<Button-1>', on_combo_click)


    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0,columnspan=2)
    combo.grid(row=2,column=0,columnspan=2)
    button.grid(row=4,column=0,columnspan=2)

def image_join():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def dnd(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            box.insert("end",x+"\n")

    def get_concat_h_multi_resize(im_list, resample=Image.BICUBIC):
        try:
            min_height = min(im.height for im in im_list)
            im_list_resize = [im.resize((int(im.width * min_height / im.height), min_height),resample=resample)
                                for im in im_list]
            total_width = sum(im.width for im in im_list_resize)
            dst = Image.new('RGB', (total_width, min_height))
            pos_x = 0
            for im in im_list_resize:
                dst.paste(im, (pos_x, 0))
                pos_x += im.width
            path=filedialog.asksaveasfilename(
                title = "名前を付けて保存",
                filetypes = [("PNG",".png"),("JPEG",".jpg"),("GIF",".gif"),("BMP",".bmp")],
                defaultextension = "png",
                initialfile="concat"
                )
            dst.save(path)
            messagebox.showinfo("完了","保存しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    def get_concat_v_multi_resize(im_list, resample=Image.BICUBIC):
        try:
            min_width = min(im.width for im in im_list)
            im_list_resize = [im.resize((min_width, int(im.height * min_width / im.width)),resample=resample)
                            for im in im_list]
            total_height = sum(im.height for im in im_list_resize)
            dst = Image.new('RGB', (min_width, total_height))
            pos_y = 0
            for im in im_list_resize:
                dst.paste(im, (0, pos_y))
                pos_y += im.height
            path=filedialog.asksaveasfilename(
                title = "名前を付けて保存",
                filetypes = [("PNG",".png"),("JPEG",".jpg"),("GIF",".gif"),("BMP",".bmp")],
                defaultextension = "png",
                initialfile="concat"
                )
            dst.save(path)
            messagebox.showinfo("完了","保存しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    box=ScrolledText(frame,width=50,height=20,wrap="none")
    button_w=ttk.Button(frame,text="横に結合",command=lambda:get_concat_h_multi_resize([Image.open(a) for a in box.get("1.0","end-1c").strip().split("\n")]))
    button_h=ttk.Button(frame,text="縦に結合",command=lambda:get_concat_v_multi_resize([Image.open(a) for a in box.get("1.0","end-1c").strip().split("\n")]))
    label=ttk.Label(frame,text="画像をドロップしてください")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',dnd)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)
    box.grid(row=2,column=0,columnspan=2)
    button_w.grid(row=3,column=0)
    button_h.grid(row=3,column=1)

def get_copyimage():
    global frame,buttonY
    def get_copyimage1():
        try:
            img = ImageGrab.grabclipboard()
            path=filedialog.asksaveasfilename(
                title = "名前を付けて保存",
                filetypes = [("PNG",".png"),("JPEG",".jpg"),("GIF",".gif"),("BMP",".bmp")],
                defaultextension = "png",
                initialfile="copy_image"
                )
            img.save(path)
        except:
            messagebox.showerror("エラー","失敗しました")

    if intermediate_screen==0:
        main_frame_delete()
        frame = ttk.Frame(root, padding=16)
        button=ttk.Button(frame,text='コピー画像取得',command=get_copyimage1,padding=[80,10,80,10])
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        button.grid(row=1,column=0,columnspan=2)
    elif intermediate_screen==1:
        get_copyimage1()

def text_search():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def search_text():
        keyword = entry.get()
        text.tag_remove("highlight", "1.0", END)  # クリア

        if keyword:
            start = "1.0"
            while True:
                start = text.search(keyword, start, stopindex=END)  # 検索
                if not start:
                    break
                end = f"{start}+{len(keyword)}c"  # 終了位置
                text.tag_add("highlight", start, end)  # タグ追加
                start = end  # 検索開始位置更新

    def clear_highlight():
        text.tag_remove("highlight", "1.0", END)  # タグを削除

    text = ScrolledText(frame,width=60,height=30)
    entry = ttk.Entry(frame, width=40)
    search_button = ttk.Button(frame, text="検索", command=search_text)
    clear_button = ttk.Button(frame, text="クリア", command=clear_highlight)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    text.grid(row=1,column=0,columnspan=2)
    entry.grid(row=2,column=0,columnspan=2)
    search_button.grid(row=3,column=0)
    clear_button.grid(row=3,column=1)

    text.tag_configure("highlight", background="red")

def site_image():
    global frame,buttonY
    def site_image1():
        try:
            url=clip.paste()
            path=filedialog.askdirectory()
            response = requests.get(url)
            html_content = response.text

            soup = BeautifulSoup(html_content, "html.parser")

            image_urls = []
            for img_tag in soup.find_all("img"):
                image_url = img_tag.get("src")
                absolute_image_url = urljoin(url, image_url)
                image_urls.append(absolute_image_url)

            n=0
            for image_url in image_urls:
                if image_url.startswith("data:image/"):
                    image_data = datauri.parse(image_url).data
                else:
                    response = requests.get(image_url)
                    image_data = response.content
                filename = image_url.split("/")[-1]
                parsed_url = urllib.parse.urlparse(filename)
                filename_without_query = urllib.parse.unquote(parsed_url.path)
                youso = os.path.splitext(filename_without_query)[1]
                file_name1 = path + f"/image{n}" + youso
                with open(file_name1, "wb") as f:
                    f.write(image_data)
                if os.path.isfile(file_name1) and "." not in f"image{n}" + youso:
                    os.remove(file_name1)
                else:
                    n += 1

            messagebox.showinfo("完了","画像の保存が完了しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    if intermediate_screen==0:
        main_frame_delete()
        frame = ttk.Frame(root, padding=16)
        button=ttk.Button(frame,text='サイト画像DL',command=site_image1,padding=[80,10,80,10])
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        button.grid(row=1,column=0,columnspan=2)
    elif intermediate_screen==1:
        site_image1()

def dice():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def dice_execute():
        try:
            num1=int(entry1.get())
            num2=int(entry2.get())
            num_list=[random.randint(1, num1) for j in range(num2)]
            num=sum(num_list)
            label2["text"]=f"結果：{num}"
        except:
            messagebox.showerror("エラー","数字を入力してください")


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame,text="D")
    entry1=ttk.Entry(frame,width=5)
    entry2=ttk.Entry(frame,width=5)
    entry1.insert(END,"6")
    entry2.insert(END,"1")
    button=ttk.Button(frame,text="実行",command=dice_execute)
    label2=ttk.Label(frame,font=("Helvetica", 18),text="結果： ")

    frame.pack()
    buttonX.grid(row=0,column=1,columnspan=2)
    buttonY.grid(row=0,column=0)
    entry1.grid(row=1,column=0)
    label1.grid(row=1,column=1)
    entry2.grid(row=1,column=2)
    label2.grid(row=2,column=1)
    button.grid(row=3,column=0,columnspan=3)

def dummy_txt():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def dumy_tet1():
        try:
            text=dumy_text()[:int(entry1.get())]
            result = ""
            if int(entry2.get())==0:
                result=text
            else:
                for i in range(0, len(text), int(entry2.get())):
                    result += text[i:i+10] + "\n"
            clip.copy(result)
            messagebox.showinfo("完了","ダミーテキストをクリップボードにコピーしました")
        except:
            messagebox.showerror("エラー","入力値が不正です")


    label1=ttk.Label(frame,text="文字数")
    entry1=ttk.Entry(frame,width=5)
    label2=ttk.Label(frame,text="区切り基準")
    entry2=ttk.Entry(frame,width=5)
    entry2.insert(END,"0")
    button=ttk.Button(frame,text="実行",command=dumy_tet1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    button.grid(row=3,column=0,columnspan=2)

def dummy_img():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def dumy_img1():
        url=f"https://via.placeholder.com/{int(entry1.get())}x{int(entry2.get())}"
        response = requests.get(url)
        if response.status_code == 200:
            path=filedialog.asksaveasfilename(
                title = "名前を付けて保存",
                filetypes = [("JPG",".jpg")],
                defaultextension = "jpg",
                initialfile="dumy_img"
                )
            with open(path, "wb") as file:
                file.write(response.content)
            messagebox.showinfo("完了","ダミー画像を保存しました")
        else:
            messagebox.showerror("エラー","ダミー画像の取得に失敗しました")

    label1=ttk.Label(frame,text="横幅：")
    label2=ttk.Label(frame,text="縦幅：")
    entry1=ttk.Entry(frame,width=8)
    entry2=ttk.Entry(frame,width=8)
    button=ttk.Button(frame,text="実行",command=dumy_img1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")


    frame.pack()
    buttonX.grid(row=0,column=2,columnspan=2)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=1,column=2)
    entry2.grid(row=1,column=3)
    button.grid(row=3,column=0,columnspan=4)

def unix_time():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def unix_time1():
        try:
            if var.get()==0:
                dt = datetime.strptime(entry1.get(), "%Y-%m-%d %H:%M:%S")
                timestamp = int(dt.timestamp())
                entry2.delete(0,END)
                entry2.insert(END,timestamp)
            else:
                dt = datetime.fromtimestamp(int(entry1.get()))
                formatted_dt = dt.strftime("%Y-%m-%d %H:%M:%S")
                entry2.delete(0,END)
                entry2.insert(END,formatted_dt)
        except:
            messagebox.showerror("エラー","入力値が不正です")

    def unix_time2():
        if var.get()==0:
            label1.configure(text="YYYY-MM-DD HH:MM:SS：")
            label2.configure(text="UNIX時間：")
        else:
            label1.configure(text="UNIX時間：")
            label2.configure(text="YYYY-MM-DD HH:MM:SS：")


    label1=ttk.Label(frame,text="通常時間(YYYY-MM-DD HH:MM:SS)")
    label2=ttk.Label(frame,text="UNIX時間")
    label_brank=ttk.Label(frame)
    label_brank1=ttk.Label(frame)
    entry1=ttk.Entry(frame,width=20)
    entry2=ttk.Entry(frame,width=20)
    var=IntVar()
    var.set(0)
    radio1=ttk.Radiobutton(frame,text="通常時間→UNIX時間",variable=var,value=0,command=unix_time2)
    radio2=ttk.Radiobutton(frame,text="UNIX時間→通常時間",variable=var,value=1,command=unix_time2)
    button=ttk.Button(frame,text="実行",command=unix_time1)
    unix_time2()
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")


    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    label_brank.grid(row=2,column=0)
    label1.grid(row=3,column=0)
    entry1.grid(row=3,column=1)
    button.grid(row=4,column=0,columnspan=2)
    label_brank1.grid(row=5,column=0)
    label2.grid(row=6,column=0)
    entry2.grid(row=6,column=1)

def web_video_info():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def web_video_info1():
        ydl_opts = {
            'skip_download': True,
            'format': 'best',
            'noprogress': True,
            'no_warnings': True,
            'no-stdout': True,
        }
        with YoutubeDL(ydl_opts) as ydl:
            info_dict = ydl.extract_info(entry1.get(), download=False)
        box.delete("1.0","end")
        box.insert("end",f"タイトル：{info_dict['title']}\n")
        box.insert("end",f"チャンネル名：{info_dict['uploader']}\n")
        box.insert("end",f"再生回数：{info_dict['view_count']}\n")
        box.insert("end",f"投稿日時：{info_dict['upload_date'][:4]}-{info_dict['upload_date'][4:6]}-{info_dict['upload_date'][6:8]}\n")
        box.insert("end",f"動画時間：{int(info_dict['duration'])//60}分{int(info_dict['duration'])%60}秒\n")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    entry1=ttk.Entry(frame,width=40)
    label1=ttk.Label(frame,text="URL：")
    button=ttk.Button(frame,text="実行",command=web_video_info1)
    box=ScrolledText(frame,width=50,height=20)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    button.grid(row=2,column=0,columnspan=2)
    box.grid(row=3,column=0,columnspan=2)

def empty_line():
    global frame,buttonY
    def empty_line1():
        try:
            text=clip.paste()
            lines = text.split("\n")
            non_empty_lines = [line for line in lines if line.strip()]
            result = "\n".join(non_empty_lines)
            clip.copy(result)
            messagebox.showinfo("完了","空行を削除しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    if intermediate_screen==0:
        main_frame_delete()
        frame = ttk.Frame(root, padding=16)
        button=ttk.Button(frame,text='空行削除',command=empty_line1,padding=[80,10,80,10])
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        button.grid(row=1,column=0,columnspan=2)
    elif intermediate_screen==1:
        empty_line1()

def mp3_tag():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def check():
        if var1.get()==1:
            entry1.configure(state="normal")
        else:
            entry1.configure(state="disabled")
        if var2.get()==1:
            entry2.configure(state="normal")
        else:
            entry2.configure(state="disabled")
        if var3.get()==1:
            entry3.configure(state="normal")
        else:
            entry3.configure(state="disabled")
        if var4.get()==1:
            entry4.configure(state="normal")
        else:
            entry4.configure(state="disabled")

    def dnd(drop):
        nonlocal file1
        try:
            file1=drop.data.replace("{","").replace("}","").replace("\\","/")
            try:
                tags = EasyID3(file1)
            except:
                audio = ID3()
                audio.save(file1)
                tags = EasyID3(file1)
            if tags.get("title")!=None:
                label2_1.configure(text=f"タイトル：{tags['title'][0]}")
            if tags.get("artist")!=None:
                label2_2.configure(text=f"アーティスト：{tags['artist'][0]}")
            if tags.get("album")!=None:
                label2_3.configure(text=f"アルバム：{tags['album'][0]}")
            if tags.get("tracknumber")!=None:
                label2_4.configure(text=f"トラック番号：{tags['tracknumber'][0]}")
        except:
            messagebox.showerror("エラー","ファイル読み込みに失敗しました")

    def mp3_tag1():
        try:
            tags = EasyID3(file1)
            if var1.get()==1:
                tags['title'] = entry1.get()
            if var2.get()==1:
                tags['artist'] = entry2.get()
            if var3.get()==1:
                tags['album'] = entry3.get()
            if var4.get()==1:
                tags['tracknumber'] = entry4.get()
            tags.save()
            messagebox.showinfo("完了","タグを変更しました")
        except:
            messagebox.showerror("エラー","タグの変更に失敗しました")


    var1=IntVar()
    check1=ttk.Checkbutton(frame,text="タイトル",onvalue=1,offvalue=0,variable=var1,command=check)
    var2=IntVar()
    check2=ttk.Checkbutton(frame,text="アーティスト",onvalue=1,offvalue=0,variable=var2,command=check)
    var3=IntVar()
    check3=ttk.Checkbutton(frame,text="アルバム",onvalue=1,offvalue=0,variable=var3,command=check)
    entry1=ttk.Entry(frame,width=20,state="disabled")
    entry2=ttk.Entry(frame,width=20,state="disabled")
    entry3=ttk.Entry(frame,width=20,state="disabled")
    var4=IntVar()
    check4=ttk.Checkbutton(frame,text="トラック番号",onvalue=1,offvalue=0,variable=var4,command=check)
    entry4=ttk.Entry(frame,width=20,state="disabled")
    label=ttk.Label(frame,text="mp3ファイルを\nドロップしてください",font=("Helvetica", 18))
    label2_1=ttk.Label(frame,text="タイトル：")
    label2_2=ttk.Label(frame,text="アーティスト：")
    label2_3=ttk.Label(frame,text="アルバム：")
    label2_4=ttk.Label(frame,text="トラック番号：")
    button=ttk.Button(frame,text="タグ変更",command=mp3_tag1)
    file1=None
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',dnd)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label2=ttk.Label(frame,text="タグ変更（チェックを入れてから入力）")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)
    label2_1.grid(row=2,column=0,columnspan=2)
    label2_2.grid(row=3,column=0,columnspan=2)
    label2_3.grid(row=4,column=0,columnspan=2)
    label2_4.grid(row=5,column=0,columnspan=2)
    label2.grid(row=6,column=0,columnspan=2)
    check1.grid(row=7,column=0)
    entry1.grid(row=7,column=1)
    check2.grid(row=8,column=0)
    entry2.grid(row=8,column=1)
    check3.grid(row=9,column=0)
    entry3.grid(row=9,column=1)
    check4.grid(row=10,column=0)
    entry4.grid(row=10,column=1)
    button.grid(row=11,column=0,columnspan=2)

def pdf_text():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def dnd(drop):
        #try:
            file1=drop.data.replace("{","").replace("}","").replace("\\","/")
            reader = pypdf.PdfReader(file1)
            number_of_pages = len(reader.pages)
            root1=Free_window()
            root1.title("PDFからテキスト抽出")
            box=ScrolledText(root1,width=60,height=30)
            root1.attributes("-topmost",True)
            box.pack()
            for i in range(number_of_pages):
                page = reader.pages[i]
                text = f"ページ{i+1}\n"+page.extract_text()+"\n"
                box.insert("end",text)
            messagebox.showinfo("完了","テキストを抽出しました")
        #except:
        #    messagebox.showerror("エラー","失敗しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="PDFファイルを\nドロップしてください",font=("Helvetica", 18))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',dnd)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def file_hash():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def dnd(drop):
        try:
            file_1=drop.data.replace("{","").replace("}","").replace("\\","/")
            with open(file_1, 'rb') as file:
                fileData = file.read()
            if var.get()==0:
                hash = hashlib.md5(fileData).hexdigest()
            elif var.get()==1:
                hash = hashlib.sha1(fileData).hexdigest()
            elif var.get()==2:
                hash = hashlib.sha256(fileData).hexdigest()
            elif var.get()==3:
                hash = hashlib.sha512(fileData).hexdigest()
            box.delete("1.0","end")
            box.insert("end",hash)
        except:
            messagebox.showerror("エラー","失敗しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    var=IntVar()
    radio1=ttk.Radiobutton(frame,text="MD5",variable=var,value=0)
    radio2=ttk.Radiobutton(frame,text="SHA1",variable=var,value=1)
    radio3=ttk.Radiobutton(frame,text="SHA256",variable=var,value=2)
    radio4=ttk.Radiobutton(frame,text="SHA512",variable=var,value=3)
    label=ttk.Label(frame,text="ファイルを\nドロップしてください",font=("Helvetica", 18))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',dnd)
    var.set(0)
    box=ScrolledText(frame,width=60,height=5)
    button=ttk.Button(frame,text="コピー",command=lambda:clip.copy(box.get("1.0","end-1c")))

    frame.pack()
    buttonX.grid(row=0,column=2,columnspan=2)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    radio3.grid(row=1,column=2)
    radio4.grid(row=1,column=3)
    label.grid(row=2,column=0,columnspan=4)
    box.grid(row=3,column=0,columnspan=4)
    button.grid(row=4,column=0,columnspan=4)

def text_hash():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def text_hash_execute():
        try:
            if var.get()==0:
                hs = hashlib.md5(entry1.get( 0., END ).rstrip().encode()).hexdigest()
            elif var.get()==1:
                hs = hashlib.sha1(entry1.get( 0., END ).rstrip().encode()).hexdigest()
            elif var.get()==2:
                hs = hashlib.sha256(entry1.get( 0., END ).rstrip().encode()).hexdigest()
            elif var.get()==3:
                hs = hashlib.sha512(entry1.get( 0., END ).rstrip().encode()).hexdigest()
            entry2.delete("1.0","end")
            entry2.insert("end",hs)
        except:
            messagebox.showerror("エラー","失敗しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    entry1=ScrolledText(frame,width=64,height=5)
    entry2=ScrolledText(frame,width=64,height=3)
    button=ttk.Button(frame,text="ハッシュ化",command=text_hash_execute)
    var=IntVar()
    var.set(0)
    radio1=ttk.Radiobutton(frame,text="MD5",variable=var,value=0)
    radio2=ttk.Radiobutton(frame,text="SHA1",variable=var,value=1)
    radio3=ttk.Radiobutton(frame,text="SHA256",variable=var,value=2)
    radio4=ttk.Radiobutton(frame,text="SHA512",variable=var,value=3)

    frame.pack()
    buttonX.grid(row=0,column=2,columnspan=2)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    radio3.grid(row=1,column=2)
    radio4.grid(row=1,column=3)
    entry1.grid(row=2,column=0,columnspan=4)
    button.grid(row=3,column=0,columnspan=4)
    entry2.grid(row=4,column=0,columnspan=4)

def image_base64():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def dnd(drop):
        try:
            path=drop.data.replace("{","").replace("}","").replace("\\","/")
            with open(path, "rb") as file:
                file_data = file.read()
                b64_data = base64.b64encode(file_data).decode('utf-8')
                clip.copy(b64_data)
            messagebox.showinfo("完了","クリップボードにコピーしました")
        except:
            messagebox.showerror("エラー","失敗しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="画像を\nドロップしてください",font=("Helvetica", 18))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',dnd)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def url_shotcut():
    global frame,buttonY
    def url_shotcut1():
        try:
            url=clip.paste()
            text="[InternetShortcut]\nURL="+url
            path=filedialog.asksaveasfilename(
                title = "名前を付けて保存",
                filetypes = [("URL", ".url")],
                defaultextension = "url",
                initialfile="shortcut"
                )
            with open(path,"w") as file:
                file.write(text)
            messagebox.showinfo("完了","ショートカットを作成しました")
        except:
            messagebox.showerror("エラー","保存に失敗しました")

    if intermediate_screen==0:
        main_frame_delete()
        frame = ttk.Frame(root, padding=16)
        button=ttk.Button(frame,text='URLショトカ',command=url_shotcut1,padding=[80,10,80,10])
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        button.grid(row=1,column=0,columnspan=2)
    elif intermediate_screen==1:
        url_shotcut1()

def explore_restart():
    global frame,buttonY
    def explorer_restart1():
        os.system("taskkill /f /im explorer.exe")
        os.system("start explorer.exe")

    if intermediate_screen==0:
        main_frame_delete()
        frame = ttk.Frame(root, padding=16)
        button=ttk.Button(frame,text='エクスプローラー再起動',command=explorer_restart1,padding=[80,10,80,10])
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        button.grid(row=1,column=0,columnspan=2)
    elif intermediate_screen==1:
        explorer_restart1()

def godmode():
    path = os.path.join(os.getcwd(), "temp1", "god_mode")
    godmode_folder = os.path.join(path, "神モード.{ED7BA470-8E54-465E-825C-99712043E01C}")
    os.makedirs(godmode_folder, exist_ok=True)
    subprocess.Popen(['explorer', godmode_folder])

def cursive():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def cursive_change(text):
        math_script_mapping = {
            'A': '𝒜', 'B': 'ℬ', 'C': '𝒞', 'D': '𝒟', 'E': 'ℰ', 'F': 'ℱ', 'G': '𝒢',
            'H': 'ℋ', 'I': 'ℐ', 'J': '𝒥', 'K': '𝒦', 'L': 'ℒ', 'M': 'ℳ', 'N': '𝒩',
            'O': '𝒪', 'P': '𝒫', 'Q': '𝒬', 'R': 'ℛ', 'S': '𝒮', 'T': '𝒯', 'U': '𝒰',
            'V': '𝒱', 'W': '𝒲', 'X': '𝒳', 'Y': '𝒴', 'Z': '𝒵',
            'a': '𝒶', 'b': '𝒷', 'c': '𝒸', 'd': '𝒹', 'e': 'ℯ', 'f': '𝒻', 'g': 'ℊ',
            'h': '𝒽', 'i': '𝒾', 'j': '𝒿', 'k': '𝓀', 'l': '𝓁', 'm': '𝓂', 'n': '𝓃',
            'o': 'ℴ', 'p': '𝓅', 'q': '𝓆', 'r': '𝓇', 's': '𝓈', 't': '𝓉', 'u': '𝓊',
            'v': '𝓋', 'w': '𝓌', 'x': '𝓍', 'y': '𝓎', 'z': '𝓏'
        }
        converted_text = ''.join(math_script_mapping.get(char, char) for char in text)
        return converted_text

    def cursive1():
        try:
            text=box.get( 0., END ).rstrip()
            text1=cursive_change(text)
            root1=Free_window()
            root1.title("変換結果")
            root1.attributes("-topmost",True)
            root1.geometry('+%d+%d'%(x,y))
            box1=ScrolledText(root1,width=56,height=20)
            box1.insert(1.0,text1)
            button_c=ttk.Button(root1,text="コピー",command=lambda:clip.copy(text1))
            box1.pack()
            button_c.pack()
        except:
            messagebox.showerror("エラー","失敗しました")

    box=ScrolledText(frame,width=56,height=20)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    button=ttk.Button(frame,text="変換",command=cursive1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    box.grid(row=1,column=0,columnspan=2)
    button.grid(row=2,column=0,columnspan=2)

def warikan():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def warikan1():
        try:
            money=int(entry1.get())
            people=int(entry2.get())
            money1=money//people
            money2=money%people
            label3_after["text"]=str(money1)
            label4_after["text"]=str(money2)
            mannenn1=money1//10000
            mannenn2=money1%10000
            label5_after["text"]=str(mannenn1)+f"({mannenn1*people})"
            gosennenn1=mannenn2//5000
            goseennenn2=mannenn2%5000
            label13_after["text"]=str(gosennenn1)+f"({gosennenn1*people})"
            sennenn1=goseennenn2//1000
            sennenn2=goseennenn2%1000
            label6_after["text"]=str(sennenn1)+f"({sennenn1*people})"
            gohyaku1=sennenn2//500
            gohyaku2=sennenn2%500
            label7_after["text"]=str(gohyaku1)+f"({gohyaku1*people})"
            hyaku1=gohyaku2//100
            hyaku2=gohyaku2%100
            label8_after["text"]=str(hyaku1)+f"({hyaku1*people})"
            goju1=hyaku2//50
            goju2=hyaku2%50
            label9_after["text"]=str(goju1)+f"({goju1*people})"
            ju1=goju2//10
            ju2=goju2%10
            label10_after["text"]=str(ju1)+f"({ju1*people})"
            go1=ju2//5
            go2=ju2%5
            label11_after["text"]=str(go1)+f"({go1*people})"
            ichi1=go2//1
            label12_after["text"]=str(ichi1)+f"({ichi1*people})"

        except:
            messagebox.showerror("エラー","入力が不正です")

    frame_left=ttk.Frame(frame)
    frame_right=ttk.Frame(frame)
    label1=ttk.Label(frame,text="金額(円)：")
    entry1=ttk.Entry(frame,width=20)
    label2=ttk.Label(frame,text="人数(人)：")
    entry2=ttk.Entry(frame,width=20)
    button=ttk.Button(frame,text="計算",command=warikan1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    label3=ttk.Label(frame_left,text="1人あたり(円)：",font=("Helvetica", 15),anchor="w")
    label4=ttk.Label(frame_left,text="端数(円)：",font=("Helvetica", 15),anchor="w")
    label5=ttk.Label(frame_left,text="1万円札枚数：",font=("Helvetica", 15),anchor="w")
    label6=ttk.Label(frame_left,text="1千円札枚数：",font=("Helvetica", 15),anchor="w")
    label7=ttk.Label(frame_left,text="500円玉枚数：",font=("Helvetica", 15),anchor="w")
    label8=ttk.Label(frame_left,text="100円玉枚数：",font=("Helvetica", 15),anchor="w")
    label9=ttk.Label(frame_left,text="50円玉枚数：",font=("Helvetica", 15),anchor="w")
    label10=ttk.Label(frame_left,text="10円玉枚数：",font=("Helvetica", 15),anchor="w")
    label11=ttk.Label(frame_left,text="5円玉枚数：",font=("Helvetica", 15),anchor="w")
    label12=ttk.Label(frame_left,text="1円玉枚数：",font=("Helvetica", 15),anchor="w")
    label13=ttk.Label(frame_left,text="5千円札枚数：",font=("Helvetica", 15),anchor="w")

    label3_after=ttk.Label(frame_right,text="0",font=("Helvetica", 15),anchor="w")
    label4_after=ttk.Label(frame_right,text="0",font=("Helvetica", 15),anchor="w")
    label5_after=ttk.Label(frame_right,text="0",font=("Helvetica", 15),anchor="w")
    label6_after=ttk.Label(frame_right,text="0",font=("Helvetica", 15),anchor="w")
    label7_after=ttk.Label(frame_right,text="0",font=("Helvetica", 15),anchor="w")
    label8_after=ttk.Label(frame_right,text="0",font=("Helvetica", 15),anchor="w")
    label9_after=ttk.Label(frame_right,text="0",font=("Helvetica", 15),anchor="w")
    label10_after=ttk.Label(frame_right,text="0",font=("Helvetica", 15),anchor="w")
    label11_after=ttk.Label(frame_right,text="0",font=("Helvetica", 15),anchor="w")
    label12_after=ttk.Label(frame_right,text="0",font=("Helvetica", 15),anchor="w")
    label13_after=ttk.Label(frame_right,text="0",font=("Helvetica", 15),anchor="w")



    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=2)
    frame_left.grid(row=4,column=0)
    frame_right.grid(row=4,column=1)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1,columnspan=2)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1,columnspan=2)
    button.grid(row=3,column=0,columnspan=3)
    label3.pack(fill="x",expand=True)
    label4.pack(fill="x",expand=True)
    label5.pack(fill="x",expand=True)
    label13.pack(fill="x",expand=True)
    label6.pack(fill="x",expand=True)
    label7.pack(fill="x",expand=True)
    label8.pack(fill="x",expand=True)
    label9.pack(fill="x",expand=True)
    label10.pack(fill="x",expand=True)
    label11.pack(fill="x",expand=True)
    label12.pack(fill="x",expand=True)

    label3_after.pack(fill="x",expand=True)
    label4_after.pack(fill="x",expand=True)
    label5_after.pack(fill="x",expand=True)
    label13_after.pack(fill="x",expand=True)
    label6_after.pack(fill="x",expand=True)
    label7_after.pack(fill="x",expand=True)
    label8_after.pack(fill="x",expand=True)
    label9_after.pack(fill="x",expand=True)
    label10_after.pack(fill="x",expand=True)
    label11_after.pack(fill="x",expand=True)
    label12_after.pack(fill="x",expand=True)




def color_picker():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def showMenu(e):
        pmenu.post(e.x_root, e.y_root)

    def main_frame1(n):
        nonlocal l
        l=1
        listener.stop()
        root1.after_cancel(a)
        root1.destroy()
        main_frame(n)

    def handler():
        if buttonY["state"] =="normal":
            buttonY.invoke()
        root1.destroy()

    def on_release(key):
        nonlocal l
        if key == keyboard.Key.pause:
            l=1
            root1.overrideredirect(False)
            return False

    def update_position():
        nonlocal l,a
        x1, y1 = pyautogui.position()
        #x1 = max(0, min(x, root1.winfo_screenwidth()-1))
        #y1 = max(0, min(y, root1.winfo_screenheight()-1))
        if root1.winfo_screenwidth()/2 < x1:
            if root1.winfo_screenheight()/2<y1:
                root1.geometry(f"+{x1-50-root1.winfo_width()}+{y1-50-root1.winfo_height()}")
            else:
                root1.geometry(f"+{x1-50-root1.winfo_width()}+{y1+50}")
        else:
            if root1.winfo_screenheight()/2<y1:
                root1.geometry(f"+{x1+50}+{y1-50-root1.winfo_height()}")
            else:
                root1.geometry(f"+{x1+50}+{y1+50}")

        def get_pixel_color(x, y):
            try:
                pixel_color = pyautogui.pixel(x, y)
            except OSError:
                pixel_color = "Invalid position"
            return pixel_color



        rgb_color = get_pixel_color(x1, y1)
        if type(rgb_color) == tuple:
            label2.config(text=f"Mouse Position: ({x1}, {y1})")
            label3.config(text=f"RGB Color: {rgb_color}")
            label4.config(text=f"Hex Color: #{rgb_color[0]:02x}{rgb_color[1]:02x}{rgb_color[2]:02x}")
            label5.config(text="           ")
            label5.config(bg=f"#{rgb_color[0]:02x}{rgb_color[1]:02x}{rgb_color[2]:02x}")
        else:
            label2.config(text=f"Mouse Position: ({x1}, {y1})")
            label3.config(text="RGB Color:取得失敗")
            label4.config(text="Hex Color:取得失敗")
            label5.config(bg="white")
            label5.config(text="取得失敗")


        if l==0:
            a=root1.after(10, update_position)

    label1=ttk.Label(frame,text="pauseキーで色取得",font=("Helvetica", 15))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    label1.grid(row=1,column=0,columnspan=2)

    root1 = Free_window()
    label2 = Label(root1, text="", font=("Helvetica", 15))
    label2.pack()
    label3 = Label(root1, text="", font=("Helvetica", 15))
    label3.pack()
    label4 = Label(root1, text="", font=("Helvetica", 15))
    label4.pack()
    label5 = Label(root1, text="           ", font=("Helvetica", 15))
    label5.pack()

    def listen_thread():
        nonlocal listener
        with keyboard.Listener(on_release=on_release) as listener:
            listener.join()

    listener=None
    l=0
    a=None
    t = threading.Thread(target=listen_thread, daemon=True)
    t.start()
    root1.attributes("-topmost", True)
    #root1.overrideredirect(True)
    pmenu = Menu(root1, tearoff=0)
    pmenu.add_command(label="RGBコピー", command=lambda: clip.copy(label3.cget("text").replace("RGB Color: ","")))
    pmenu.add_command(label="16進数コピー", command=lambda: clip.copy(label4.cget("text").replace("Hex Color: ","")))
    root1.bind("<Button-3>", showMenu)
    update_position()
    root1.protocol("WM_DELETE_WINDOW", handler)

def magnifier():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)
    img_size = 500
    img_double_size = img_size * 0.5
    tk_image = None
    long_s=int(img_double_size/2)
    root1 = Free_window()

    def magnifier1():
        nonlocal img_double_size,long_s
        img_double_size = img_size/float(entry1.get())
        long_s=int(img_double_size/2)

    def pil_to_tk(pil_image):
        nonlocal tk_image
        tk_image = ImageTk.PhotoImage(pil_image)
        return tk_image

    def main_frame1(n):
        nonlocal l
        l=1
        listener.stop()
        root1.after_cancel(a)
        root1.destroy()
        main_frame(n)

    def handler():
        if buttonY["state"] =="normal":
            buttonY.invoke()
        root1.destroy()

    def on_release(key):
        nonlocal l
        if key == keyboard.Key.pause:
            l=1
            root1.overrideredirect(False)
            return False

    def update_position():
        nonlocal l,a,tk_image,img_double_size
        x, y = pyautogui.position()
        x1 = max(0, min(x, root1.winfo_screenwidth()-1))
        y1 = max(0, min(y, root1.winfo_screenheight()-1))
        if root1.winfo_screenwidth()/2 < x:
            if root1.winfo_screenheight()/2<y:
                root1.geometry(f"+{x1-long_s-root1.winfo_width()}+{y1-long_s-root1.winfo_height()}")
            else:
                root1.geometry(f"+{x1-long_s-root1.winfo_width()}+{y1+long_s}")
        else:
            if root1.winfo_screenheight()/2<y:
                root1.geometry(f"+{x1+long_s}+{y1-long_s-root1.winfo_height()}")
            else:
                root1.geometry(f"+{x1+long_s}+{y1+long_s}")


        pil_image = ImageGrab.grab(bbox=(x - img_double_size // 2,
                                            y - img_double_size // 2,
                                            x + img_double_size // 2,
                                            y + img_double_size // 2))

        pil_image = pil_image.resize((img_size, img_size), Image.LANCZOS)

        tk_image = pil_to_tk(pil_image)
        canvas.create_image(
            tk_image.width() // 2,
            tk_image.height() // 2,
            image=tk_image)
        if l==0:
            a=root1.after(10, update_position)

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(3),font=("Helvetica", 7),bg="gray",fg="white")
    entry1=ttk.Entry(frame,width=10)
    button1=ttk.Button(frame,text="拡大率変更(倍)",command=magnifier1)

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    entry1.grid(row=1,column=0)
    button1.grid(row=1,column=1)

    canvas = Canvas(root1, width=img_size, height=img_size)
    canvas.pack()

    def listen_thread():
        nonlocal listener
        with keyboard.Listener(on_release=on_release) as listener:
            listener.join()

    listener=None
    l=0
    a=None
    t = threading.Thread(target=listen_thread, daemon=True)
    t.start()
    entry1.insert(0,2)
    root1.attributes("-topmost", True)
    root1.overrideredirect(True)
    root1.after_idle(root1.after, 0, update_position)
    root1.protocol("WM_DELETE_WINDOW", handler)

def excel_csv():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def widget_set():
        if var.get()==0:
            label1.grid(row=2,column=1)
            spinbox.grid(row=2,column=0)
            label["text"]="Excelファイルを\nドロップしてください"
        elif var.get()==1:
            label1.grid_forget()
            spinbox.grid_forget()
            label["text"]="CSVファイルを\nドロップしてください"

    def dnd(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                if var.get()==0:
                    workbook = openpyxl.load_workbook(x)
                    sheet_names = workbook.sheetnames
                    y=x.replace(".xlsx",f"_{spinbox.get()}.csv")
                    if int(spinbox.get()) < len(sheet_names):
                        target_sheet_name = sheet_names[int(spinbox.get())]
                        sheet = workbook[target_sheet_name]
                        with open(y, 'w', newline='') as csv_file:
                            csv_writer = csv.writer(csv_file)
                            for row in sheet.iter_rows():
                                csv_writer.writerow([cell.value for cell in row])
                    else:
                        messagebox.showerror("エラー",f"{x}の{spinbox.get()}ページが存在しません")
                    workbook.close()
                elif var.get()==1:
                    y=x.replace(".csv",".xlsx")
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    with open(x, 'r', encoding='utf-8') as csvfile:
                        csvreader = csv.reader(csvfile)
                        for row in csvreader:
                            sheet.append(row)
                    workbook.save(y)
                    workbook.close()
            except:
                messagebox.showerror("エラー","失敗しました")
        messagebox.showinfo("完了","終了しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="Excelファイルを\nドロップしてください",font=("Helvetica", 20))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',dnd)
    label1=ttk.Label(frame,text="ページ目")
    spinbox=Spinbox(frame,from_=0,to=1000,width=5,increment=1)
    var=IntVar()
    var.set(0)
    radio1=ttk.Radiobutton(frame,text="Excel→CSV",variable=var,value=0,command=widget_set)
    radio2=ttk.Radiobutton(frame,text="CSV→Excel",variable=var,value=1,command=widget_set)


    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    label1.grid(row=2,column=1)
    spinbox.grid(row=2,column=0)
    label.grid(row=3,column=0,columnspan=2)

def unicode_normalize():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def unicode_normalize1():
        try:
            s=clip.paste()
            if var.get()==0:
                x=unicodedata.normalize("NFKC",s)
            elif var.get()==1:
                x=unicodedata.normalize("NFKD",s)
            elif var.get()==2:
                x=unicodedata.normalize("NFC",s)
            elif var.get()==3:
                x=unicodedata.normalize("NFD",s)
            clip.copy(x)
            messagebox.showinfo("完了","変換が完了しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    var=IntVar()
    var.set(3)
    radio4=ttk.Radiobutton(frame,text="NFKC",variable=var,value=0)
    radio3=ttk.Radiobutton(frame,text="NFKD",variable=var,value=1)
    radio2=ttk.Radiobutton(frame,text="NFC",variable=var,value=2)
    radio1=ttk.Radiobutton(frame,text="NFD",variable=var,value=3)
    button=ttk.Button(frame,text="変換",command=unicode_normalize1)

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=2,columnspan=2)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    radio3.grid(row=1,column=2)
    radio4.grid(row=1,column=3)
    button.grid(row=3,column=0,columnspan=4)

def office_image():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def dnd(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        make_folder(os.getcwd()+"/temp1/office_zip")
        put_path=filedialog.askdirectory(title="保存先を選択してください")
        for x in file_list:
            try:
                make_folder(os.getcwd()+"/temp1/office_zip")
                if os.path.splitext(x)[1]==".docx":
                    shutil.copyfile(x, os.getcwd()+"/temp1/office_zip/temp.docx")
                    z=os.getcwd()+"/temp1/office_zip/temp.docx"
                    os.rename(z,z.replace(".docx",".zip"))
                    y=z.replace(".docx",".zip")
                elif os.path.splitext(x)[1]==".pptx":
                    shutil.copyfile(x, os.getcwd()+"/temp1/office_zip/temp.pptx")
                    z=os.getcwd()+"/temp1/office_zip/temp.pptx"
                    os.rename(z,z.replace(".pptx",".zip"))
                    y=z.replace(".pptx",".zip")
                elif os.path.splitext(x)[1]==".xlsx":
                    shutil.copyfile(x, os.getcwd()+"/temp1/office_zip/temp.xlsx")
                    z=os.getcwd()+"/temp1/office_zip/temp.xlsx"
                    os.rename(z,z.replace(".xlsx",".zip"))
                    y=z.replace(".xlsx",".zip")
                with zipfile.ZipFile(y,'r') as inputFile:
                    inputFile.extractall(path=os.getcwd()+"/temp1/office_zip")
                if os.path.splitext(x)[1]==".docx":
                    img_files = os.listdir(os.getcwd()+"/temp1/office_zip/word/media")
                    for img in img_files:
                        if os.path.isfile(os.path.join(os.getcwd()+"/temp1/office_zip/word/media", img)):
                            shutil.move(os.path.join(os.getcwd()+"/temp1/office_zip/word/media", img), put_path)
                elif os.path.splitext(x)[1]==".pptx":
                    img_files = os.listdir(os.getcwd()+"/temp1/office_zip/ppt/media")
                    for img in img_files:
                        if os.path.isfile(os.path.join(os.getcwd()+"/temp1/office_zip/ppt/media", img)):
                            shutil.move(os.path.join(os.getcwd()+"/temp1/office_zip/ppt/media", img), put_path)
                elif os.path.splitext(x)[1]==".xlsx":
                    img_files = os.listdir(os.getcwd()+"/temp1/office_zip/xl/media")
                    for img in img_files:
                        if os.path.isfile(os.path.join(os.getcwd()+"/temp1/office_zip/xl/media", img)):
                            shutil.move(os.path.join(os.getcwd()+"/temp1/office_zip/xl/media", img), put_path)
                delete_folder(os.getcwd()+"/temp1/office_zip")
            except:
                messagebox.showerror("エラー","失敗しました")
        messagebox.showinfo("完了","変換が終了しました")


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="Officeファイルを\nドロップしてください",font=("Helvetica", 20))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',dnd)

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    label.grid(row=1,column=0,columnspan=2)

def gif_video():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def gif_video1():
        try:
            full_img= ImageGrab.grab()
            button1["state"]="disabled"
            button2["state"]="normal"
            root1=Free_window()
            root1.attributes("-topmost", True)
            def start():
                nonlocal pil_img,pil_img1,canvas_img
                pil_img=full_img
                canvas.config(width=pil_img.width, height=pil_img.height)
                canvas_img= ImageTk.PhotoImage(pil_img)
                canvas.canvas_img = canvas_img
                canvas.create_image(pil_img.width*0.5,pil_img.height*0.5,image=canvas_img)
                root1.attributes('-fullscreen', True)

            def on_press(event):
                canvas.delete("rect")
                nonlocal start_x
                nonlocal start_y
                start_x = canvas.canvasx(event.x)
                start_y = canvas.canvasy(event.y)

            def on_move_press(event):
                nonlocal end_x, end_y
                end_x, end_y = event.x, event.y
                canvas.delete("rect")
                canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)

            def on_release(event):
                nonlocal end_x, end_y
                end_x, end_y = event.x, event.y
                canvas.delete("rect")
                canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)
                root1.destroy()
                t=threading.Thread(target=gif_sc,daemon=True)
                t.start()

            def gif_sc():
                nonlocal start_x,start_y,end_x,end_y
                if start_x>end_x:
                    start_x,end_x=end_x,start_x
                if start_y>end_y:
                    start_y,end_y=end_y,start_y
                gif_list=[]
                num=0
                make_folder(os.getcwd()+"/temp1/gif_sc")
                while end_loop==0:
                    a=time.time()
                    img=ImageGrab.grab(bbox=(start_x, start_y, end_x, end_y)).quantize()
                    img.save(os.getcwd()+"/temp1/gif_sc/"+str(num)+".png")
                    gif_list.append(os.getcwd()+"/temp1/gif_sc/"+str(num)+".png")
                    num+=1
                    b=time.time()
                    time_ab=b-a
                    time_sum=max(float(1/int(entry1.get()))-time_ab,0)
                    time.sleep(time_sum)
                image_clip = ImageSequenceClip(gif_list, fps=int(entry1.get()))
                image_clip.write_videofile(os.getcwd()+"/temp1/gif_sc/temp.mp4",codec='libx264', audio=False,logger=None)
                path=filedialog.asksaveasfilename(title="保存先を選択してください",filetypes=[("GIF","*.gif")],defaultextension=".gif")
                video_clip = VideoFileClip(os.getcwd()+"/temp1/gif_sc/temp.mp4")
                video_clip.write_gif(path, fps=int(entry1.get()),logger=None)
                video_clip.close()
                image_clip.close()
                delete_folder(os.getcwd()+"/temp1/gif_sc")
                messagebox.showinfo("終了","Gifを作成しました")

            start_x=None
            start_y=None
            end_x=None
            end_y=None
            pil_img=None
            pil_img1=None
            canvas_img=None
            root.update()
            canvas=Canvas(root1,width=300,height=300,bg="gray")
            canvas.bind("<ButtonPress-1>", on_press)
            canvas.bind("<ButtonRelease-1>", on_release)
            canvas.bind("<B1-Motion>", on_move_press)
            canvas.pack(expand=True,side=TOP)
            root.after(10,start)
        except:
            messagebox.showerror("エラー","失敗しました")

    def end():
        nonlocal end_loop
        end_loop=1
        button1["state"]="normal"
        button2["state"]="disabled"

    end_loop=0
    label1=ttk.Label(frame,text="FPS：")
    entry1=ttk.Entry(frame,width=10)
    button1=ttk.Button(frame,text="撮影開始",command=gif_video1)
    button2=ttk.Button(frame,text="撮影終了",command=end, state="disabled")
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    button1.grid(row=2,column=0)
    button2.grid(row=2,column=1)
    entry1.insert(0,15)

def metrome():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)
    onoff=0

    def main_frame1(n):
        nonlocal onoff
        onoff=1
        if t.is_alive():
            t.join()
        main_frame(n)

    def metrome1():
        nonlocal t,onoff
        button2["state"]="normal"
        cycle=int(entry3.get())
        button1["state"]="disabled"
        root.update()
        tone = int(entry4.get())-9
        freq = int(440*2**(tone/12)) #周波数
        sleep_time = (60/int(entry1.get())) - (int(entry2.get())/1000) #停止時間
        if sleep_time>0:
            for i in range(cycle):
                if onoff==0:
                    winsound.Beep(freq, int(int(entry2.get())))
                    time.sleep(sleep_time)
                else:
                    return
        else:
            messagebox.showerror("エラー","BPMが高すぎます")
        button1["state"]="normal"
        t=threading.Thread(target=metrome1,daemon=True)
        button2["state"]="disabled"
        onoff=0
        root.update()
        return

    def on():
        nonlocal onoff,t
        onoff=1
        t.join()
        button1["state"]="normal"
        t=threading.Thread(target=metrome1,daemon=True)
        onoff=0
        button2["state"]="disabled"
        return

    t=threading.Thread(target=metrome1,daemon=True)
    label1=ttk.Label(frame,text="BPM：")
    entry1=ttk.Entry(frame,width=10)
    label2=ttk.Label(frame,text="音の長さ(ms)：")
    entry2=ttk.Entry(frame,width=10)
    label3=ttk.Label(frame,text="くり返し回数：")
    entry3=ttk.Entry(frame,width=10)
    label4=ttk.Label(frame,text="音の高さ(0~)：")
    entry4=ttk.Entry(frame,width=10)
    button1=ttk.Button(frame,text="開始",command=lambda:t.start())
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(1),font=("Helvetica", 7),bg="gray",fg="white")
    button2=ttk.Button(frame,text="停止",command=on, state="disabled")

    entry1.insert(0,60)
    entry2.insert(0,50)
    entry3.insert(0,5)
    entry4.insert(0,9)

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    label3.grid(row=3,column=0)
    entry3.grid(row=3,column=1)
    label4.grid(row=4,column=0)
    entry4.grid(row=4,column=1)
    button1.grid(row=5,column=0)
    button2.grid(row=5,column=1)

def ogp():
    global frame,buttonY
    def ogp1():
        img=None
        img_url=None
        def handler():
            nonlocal img,img_url
            try:
                clipboard=urllib.parse.quote(clip.paste())
                url=f"https://kotsukotsu-ogp-api.vercel.app/api/ogp?url={clipboard}"
                response = requests.get(url)
                data = response.json()
                if data.get("image") is None:
                    img_url=""
                else:
                    img_url=data["image"]
                if data.get("title") is None:
                    title=""
                else:
                    title=data["title"]
                if data.get("description") is None:
                    description=""
                else:
                    description=data["description"]

                parsed_url = urlparse(img_url)
                path = parsed_url.path
                extension = os.path.splitext(path)[1]
                file_name=f"temp{extension}"

                entry1.insert(0,title)
                box.insert(1.0,description)

                if img_url=="":
                    return
                response = requests.get(img_url)
                make_folder(os.getcwd()+"/temp1/ogp")
                with open(os.getcwd()+"/temp1/ogp"+"/"+file_name, 'wb') as f:
                    f.write(response.content)
                pil=Image.open(os.getcwd()+"/temp1/ogp"+"/"+file_name)
                img=ImageTk.PhotoImage(pil)
                w,h=pil.size

                canvas.config(width=w,height=h)
                canvas.canvas_img = img

                canvas.create_image(w*0.5,h*0.5,image=img)
                canvas.update()
            except:
                messagebox.showerror("エラー","失敗しました")

        root1=Free_window()
        root1.attributes("-topmost", True)
        canvas=Canvas(root1)
        entry1=ttk.Entry(root1,width=75)
        label1=ttk.Label(root1,text="説明：")
        box=Text(root1,width=75,height=5)


        entry1.grid(row=0,column=0,columnspan=2)
        label1.grid(row=1,column=0,columnspan=2)
        box.grid(row=2,column=0,columnspan=2)
        canvas.grid(row=3,column=0,columnspan=2)
        handler()

        def showMenu(e):
            pmenu.post(e.x_root, e.y_root)
        def copy():
            clip.copy(img_url)
        pmenu = Menu(root1, tearoff=0)
        pmenu.add_command(label="画像URLコピー", command=copy)
        root1.bind("<Button-3>", showMenu)

    if intermediate_screen==0:
        main_frame_delete()
        frame = ttk.Frame(root, padding=16)
        button=ttk.Button(frame,text='OGP取得',command=ogp1,padding=[80,10,80,10])
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        button.grid(row=1,column=0,columnspan=2)
    elif intermediate_screen==1:
        ogp1()

def copy_image():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def dnd(drop):
        try:
            path=drop.data.replace("{","").replace("}","").replace("\\","/")
            original_image = Image.open(path)
            output = io.BytesIO()
            make_folder(os.getcwd()+"/temp1/copy_image")
            original_image.save(os.getcwd()+"/temp1/copy_image/temp.png")
            original_image1 = Image.open(os.getcwd()+"/temp1/copy_image/temp.png")
            original_image1.convert('RGB').save(output, 'BMP')
            data = output.getvalue()[14:]
            output.close()
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
            win32clipboard.CloseClipboard()
            messagebox.showinfo("完了","クリップボードにコピーしました")
        except:
            messagebox.showerror("エラー","失敗しました")

    label=ttk.Label(frame,text="画像をここに\nドロップしてください",font=("Helvetica", 20))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',dnd)

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    label.grid(row=1,column=0,columnspan=2)

def cloud_download():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def cloud_download1():
        url=clip.paste()
        try:
            if var.get()==0:
                pattern = r'\/d/(.*?)\/view'
                matches = re.findall(pattern, url)
                if matches==[]:
                    pattern = r'\/d/(.*?)\/edit'
                matches = re.findall(pattern, url)
                url1="https://drive.google.com/u/4/uc?id="+matches[0]+"&export=download"
            elif var.get()==1:
                url1=url.replace("dl=0","dl=1")
            clip.copy(url1)
            messagebox.showinfo("完了","DLリンクをコピーしました")
        except:
            messagebox.showerror("エラー","失敗しました")

    var=IntVar()
    var.set(0)
    radio1=Radiobutton(frame,text="Google Drive",variable=var,value=0)
    radio2=Radiobutton(frame,text="Dropbox",variable=var,value=1)
    button=ttk.Button(frame,text="DLリンク変更コピー",command=cloud_download1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    button.grid(row=2,column=0,columnspan=2)

def voice_concat():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def voice_concat_top():
        try:
            output=filedialog.asksaveasfilename(
                title = "名前を付けて保存",
                filetypes = [("MP3", ".mp3"),("WAV", ".wav")],
                defaultextension = "mp3"
                )
            clip1 = AudioFileClip(entry1.get())
            clip2 = AudioFileClip(entry2.get())
            clip = concatenate_audioclips([clip1, clip2])
            clip.write_audiofile(output,logger=None)
            messagebox.showinfo("終了","結合完了しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    def drag_voice1(drop):
        x=drop.data.replace("{","").replace("}","").replace("\\","/")
        entry1.delete(0,END)
        entry1.insert(0,x)

    def drag_voice2(drop):
        x=drop.data.replace("{","").replace("}","").replace("\\","/")
        entry2.delete(0,END)
        entry2.insert(0,x)


    label1=ttk.Label(frame,text="前半の音声をここにドロップ：")
    label2=ttk.Label(frame,text="後半の音声をここにドロップ：")
    entry1=ttk.Entry(frame,width=20)
    entry2=ttk.Entry(frame,width=20)
    button=ttk.Button(frame,text="結合",command=voice_concat_top)
    entry1.drop_target_register(DND_FILES)
    entry1.dnd_bind('<<Drop>>',drag_voice1)
    entry2.drop_target_register(DND_FILES)
    entry2.dnd_bind('<<Drop>>',drag_voice2)
    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',drag_voice1)
    label2.drop_target_register(DND_FILES)
    label2.dnd_bind('<<Drop>>',drag_voice2)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    button.grid(row=3,column=0,columnspan=2)

def json_format():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def json_formatA(txt):
        root1=Free_window()
        root1.attributes("-topmost", True)
        box=ScrolledText(root1,width=75,height=20)
        box.pack(expand=True,side=TOP,fill=BOTH)
        box.insert(1.0,txt)

    def json_format1():
        s=clip.paste()
        try:
            x=json.loads(s)
            json_formatA(json.dumps(x,indent=2,ensure_ascii=False))

        except:
            messagebox.showerror("エラー","失敗しました")

    def json_format2(drop):
        file1=drop.data.replace("{","").replace("}","").replace("\\","/")
        try:
            with open(file1, "r",encoding="utf-8") as f:
                dic = f.read()
            x=json.loads(dic)
            json_formatA(json.dumps(x,indent=2,ensure_ascii=False))
        except:
            messagebox.showerror("エラー","失敗しました")


    label=ttk.Label(frame,text="JSONをここに\nドロップしてください",font=("Helvetica", 20))
    button=ttk.Button(frame,text="クリップボード入力整形",command=json_format1)
    label.drop_target_register(DND_FILES)
    label.dnd_bind('<<Drop>>',json_format2)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    label.grid(row=1,column=0,columnspan=2)
    button.grid(row=2,column=0,columnspan=2)

def env_word():
    root1=Free_window()
    root1.attributes("-topmost", True)
    root1.title("環境変数")
    root1.geometry("850x500")

    def copy_to_clipboard(text):
        clip.copy(text)

    def on_treeview_click(event):
        x, y = event.x, event.y
        item = tree.identify_row(y)
        if item:
            tree.selection_set(item)
        else:
            # アイテムがない場合は選択を解除する
            tree.selection_remove(tree.selection())
            return
        #item = tree.selection()[0]
        key, value = tree.item(item, 'values')
        menu = Menu(root, tearoff=0)
        menu.add_command(label="keyをコピー", command=lambda: copy_to_clipboard(key))
        menu.add_command(label="Valueをコピー", command=lambda: copy_to_clipboard(value))
        x, y = root1.winfo_pointerx(), root1.winfo_pointery()
        menu.post(x, y)

    def display_data(data):
        for key, value in data.items():
            tree.insert('', 'end', values=(key, value))

    env_dic=os.environ

    tree = ttk.Treeview(root1, columns=("Key", "Value"), show="headings")
    tree.heading("Key", text="Key")
    tree.heading("Value", text="Value")
    display_data(env_dic)
    tree.pack(expand=True, fill=BOTH,side=LEFT)

    scrollbar = ttk.Scrollbar(root1, orient=VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.pack(side=RIGHT, fill=Y)

    tree.bind("<Button-3>", on_treeview_click)

def seiwa():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def seiwa1():
        year=entry1.get()
        if var.get()==0:
            try:
                year1=int(year)
                w2j = jeraconv.W2J()
                result1=w2j.convert(year1, 1, 1, return_type='list')
                if result1[1]<0:
                    messagebox.showerror("エラー","古すぎるため変換できません")
                    return
                result=f"{result1[0]} {result1[1]}年"
                entry2.delete(0,END)
                entry2.insert(0,result)
            except:
                messagebox.showerror("エラー","変換できませんでした")
        elif var.get()==1:
            try:
                j2w = jeraconv.J2W()
                result=j2w.convert(year)
                entry2.delete(0,END)
                entry2.insert(0,result)
            except:
                messagebox.showerror("エラー","変換できませんでした")

    var=IntVar()
    var.set(0)
    radio1=Radiobutton(frame,text="西暦→和暦",variable=var,value=0)
    radio2=Radiobutton(frame,text="和暦→西暦",variable=var,value=1)
    entry1=ttk.Entry(frame,width=20)
    label1=ttk.Label(frame,text="年を入力")
    button=ttk.Button(frame,text="変換",command=seiwa1)
    entry2=ttk.Entry(frame,width=20)
    label2=ttk.Label(frame,text="結果")
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    label1.grid(row=2,column=0)
    entry1.grid(row=2,column=1)
    button.grid(row=3,column=0,columnspan=2)
    label2.grid(row=4,column=0)
    entry2.grid(row=4,column=1)

def wallpaper():
    global button143
    root1=Free_window()
    root1.attributes("-topmost", True)
    root1.title("壁紙スライド")

    def delete_window():
        if a is not None:
            root1.after_cancel(a)
        button143["state"]="normal"
        root1.destroy()

    file_list=[]
    a=None
    SPI_SETDESKWALLPAPER = 20
    SystemParametersInfo = ctypes.windll.user32.SystemParametersInfoW
    count=0
    memo=None

    def set_wallpaper(image_list):
        nonlocal count,a
        result = SystemParametersInfo(SPI_SETDESKWALLPAPER, 0, image_list[count], 3)
        if result==False:
            messagebox.showerror("エラー","失敗しました")
        count=(count+1)%len(image_list)
        if count==0 and wallpaper_random.get()==1:
            image_list1=random.sample(image_list,len(image_list))
        else:
            image_list1=image_list
        a=root1.after(int(entry2.get())*1000,lambda:set_wallpaper(image_list1))

    def wallpaper1():
        nonlocal file_list
        folder_path=filedialog.askdirectory(title="フォルダを選択してください")
        file_list = [os.path.join(folder_path, file_name) for file_name in os.listdir(folder_path)]
        label1["text"]=folder_path

    def wallpaper2():
        nonlocal a
        if a is not  None:
            root1.after_cancel(a)
        if file_list==[]:
            messagebox.showerror("エラー","フォルダを選択してください")
        elif entry2.get()=="":
            messagebox.showerror("エラー","時間間隔を入力してください")
        else:
            if wallpaper_random.get()==1:
                new=random.sample(file_list,len(file_list))
            else:
                new=file_list
            set_wallpaper(new)

    def wallpaper3():
        nonlocal a,memo
        try:
            if a is not  None:
                root1.after_cancel(a)
            if file_list==[]:
                messagebox.showerror("エラー","フォルダを選択してください")
            else:
                if wallpaper_random.get()==1:
                    new=random.sample(file_list,len(file_list))
                else:
                    new=file_list
                set_wallpaper1(new)
        except:
            messagebox.showerror("エラー","失敗しました")

    def set_wallpaper1(image_list):
        nonlocal count,a,hand_list
        result = SystemParametersInfo(SPI_SETDESKWALLPAPER, 0, image_list[count], 3)
        if result==False:
            messagebox.showerror("エラー","失敗しました")
        count=(count+1)%len(image_list)
        if count==0 and wallpaper_random.get()==1:
            hand_list=random.sample(image_list,len(image_list))
        else:
            hand_list=image_list

    frame2=ttk.Frame(root1)
    root1.protocol("WM_DELETE_WINDOW", delete_window)
    hand_list=None
    button1=ttk.Button(frame2,text="壁紙のあるフォルダを指定",command=wallpaper1)
    label1=ttk.Label(frame2,text="壁紙のあるフォルダを指定してください")
    wallpaper_random=IntVar()
    check1=ttk.Checkbutton(frame2,text="ランダム",onvalue=1,offvalue=0,variable=wallpaper_random)
    button2=ttk.Button(frame2,text="壁紙スライドショー開始",command=wallpaper2)
    label2=ttk.Label(frame2,text="壁紙スライドショー感覚(秒)：")
    entry2=ttk.Entry(frame2,width=10)
    button3=ttk.Button(frame2,text="手動切替",command=wallpaper3)

    frame2.pack()
    button1.grid(row=0,column=0)
    label1.grid(row=0,column=1)
    label2.grid(row=1,column=0)
    entry2.grid(row=1,column=1)
    check1.grid(row=2,column=0)
    button2.grid(row=2,column=1)
    button3.grid(row=3,column=0,columnspan=2)
    button143["state"]="disabled"

def pdf_image():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def pdf_image1(drop):
        try:
            file1=drop.data.replace("{","").replace("}","").replace("\\","/")
            reader=pypdf.PdfReader(file1)
            path=filedialog.askdirectory(title="保存先を選択してください")
            for i in range(len(reader.pages)):
                page=reader.pages[i]
                count = 0
                for image_file_object in page.images:
                    with open(path+f"/{i+1}_"+str(count) + image_file_object.name, "wb") as fp:
                        fp.write(image_file_object.data)
                        count += 1
            messagebox.showinfo("完了","画像を保存しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    label1=ttk.Label(frame,text="PDFをここに\nドロップしてください",font=("Helvetica", 20))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',pdf_image1)

    frame.pack()
    buttonY.grid(row=0,column=0)
    buttonX.grid(row=0,column=1)
    label1.grid(row=1,column=0,columnspan=2)

def pdf_press():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def pdf_press1(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        try:
            for file1 in file_list:
                if var.get()==0:
                    reader = pypdf.PdfReader(file1)
                    writer = pypdf.PdfWriter()
                    name=os.path.basename(file1)
                    path=file1.replace(name,"duplication_"+name)
                    for page in reader.pages:
                        writer.add_page(page)
                    writer.add_metadata(reader.metadata)
                    with open(path, "wb") as fp:
                        writer.write(fp)
                elif var.get()==1:
                    reader = pypdf.PdfReader(file1)
                    writer = pypdf.PdfWriter()
                    name=os.path.basename(file1)
                    path=file1.replace(name,"image_"+name)
                    for page in reader.pages:
                        writer.add_page(page)
                    for page in writer.pages:
                        for img in page.images:
                            img.replace(img.image, quality=60)
                    with open(path, "wb") as f:
                        writer.write(f)
                elif var.get()==2:
                    reader = pypdf.PdfReader(file1)
                    writer = pypdf.PdfWriter()
                    name=os.path.basename(file1)
                    path=file1.replace(name,"compress_"+name)
                    for page in reader.pages:
                        writer.add_page(page)
                    for page in writer.pages:
                        page.compress_content_streams(level=9)  # This is CPU intensive!
                    with open(path, "wb") as f:
                        writer.write(f)
            messagebox.showinfo("完了","処理が完了しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    var=IntVar()
    var.set(0)
    radio1=ttk.Radiobutton(frame,text="オブジェクト重複削除",variable=var,value=0)
    radio2=ttk.Radiobutton(frame,text="画像画質圧縮",variable=var,value=1)
    radio3=ttk.Radiobutton(frame,text="PDF可逆圧縮",variable=var,value=2)
    label1=ttk.Label(frame,text="PDFをここに\nドロップしてください",font=("Helvetica", 20))
    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',pdf_press1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1,columnspan=2)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    radio3.grid(row=1,column=2)
    label1.grid(row=2,column=0,columnspan=3)

def pdf_rotate():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def pdf_rotate1(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for path in file_list:
            try:
                reader = pypdf.PdfReader(path)
                writer = pypdf.PdfWriter()
                name=os.path.basename(path)
                path=path.replace(name,f"rotate{int(entry2.get())-1}_"+name)
                for page in reader.pages:
                    writer.add_page(page)
                writer.pages[int(entry2.get())-1].rotate(int(entry1.get()))
                with open(path, "wb") as fp:
                    writer.write(fp)
            except:
                messagebox.showerror("エラー",f"{path}\nで失敗しました")
        messagebox.showinfo("完了","回転しました")

    label1=ttk.Label(frame,text="PDFをここに\nドロップしてください",font=("Helvetica", 20))
    label2=ttk.Label(frame,text="回転角度(90度間隔)")
    entry1=ttk.Entry(frame,width=10)
    label3=ttk.Label(frame,text="ページ")
    entry2=ttk.Entry(frame,width=10)
    entry1.insert(0,90)
    entry2.insert(0,1)
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',pdf_rotate1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label2.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label3.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    label1.grid(row=3,column=0,columnspan=2)

def window_kill():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def get_active_window_info():
        hwnd = win32gui.GetForegroundWindow()
        _, pid = win32process.GetWindowThreadProcessId(hwnd)
        return hwnd, pid

    def kill_process_by_pid(pid):
        try:
            os.system(f"taskkill /F /PID {pid}")
        except Exception as e:
            messagebox.showerror("エラー",f"プロセスを終了できませんでした: {e}")

    def on_release(key):
        if key == keyboard.Key.pause:
                hwnd, pid = get_active_window_info()
                if pid:
                    kill_process_by_pid(pid)
                else:
                    messagebox.showerror("エラー","失敗しました")

    def main_frame1(num):
        listener.stop()
        t.join()
        root.protocol("WM_DELETE_WINDOW", frame_delete)
        main_frame(num)

    def stopping():
        listener.stop()
        t.join()
        frame_delete()

    listener=None
    root.protocol("WM_DELETE_WINDOW", stopping)
    label=ttk.Label(frame,text="pauseキーで\nウィンドウを強制終了させます",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)
    def listen_thread():
        nonlocal listener
        with keyboard.Listener(on_release=on_release) as listener:
            listener.join()

    t = threading.Thread(target=listen_thread, daemon=True)
    t.start()

def time_calc():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def time_calc1():
        try:
            if combo.get()=="基準年月日から指定日数後の年月日を計算":
                dt1=datetime.strptime(entry1.get(),'%Y/%m/%d')
                dt2=timedelta(days=int(entry2.get()))
                dt3=str((dt1+dt2).strftime("%Y/%m/%d"))
            elif combo.get()=="基準年月日から指定年月日までの日数を計算":
                dt1=datetime.strptime(entry1.get(),'%Y/%m/%d')
                dt2=datetime.strptime(entry2.get(),'%Y/%m/%d')
                dt3=str(dt1-dt2).split(",")[0].replace("days","日後").replace("day","日後")
            label3["text"]="結果："+dt3
        except:
            messagebox.showerror("エラー","失敗しました")

    def on_combobox_select(event):
        try:
            if combo.get()=="基準年月日から指定日数後の年月日を計算":
                label2["text"]="指定日数(-で前日)"
            elif combo.get()=="基準年月日から指定年月日までの日数を計算":
                label2["text"]="指定年月日(年/月/日)"
            entry2.delete(0,END)
            label3["text"]="結果："
        except:
            messagebox.showerror("エラー","失敗しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    entry1=ttk.Entry(frame,width=20)
    combo=ttk.Combobox(frame,values=["基準年月日から指定日数後の年月日を計算","基準年月日から指定年月日までの日数を計算"],width=40,state="readonly")
    combo.current(0)
    entry2=ttk.Entry(frame,width=20)
    label1=ttk.Label(frame,text="基準年月日(年/月/日)")
    label2=ttk.Label(frame,text="指定日数(-で前日)")
    button=ttk.Button(frame,text="計算",command=time_calc1)
    label3=ttk.Label(frame,text="結果：")
    combo.bind("<<ComboboxSelected>>", on_combobox_select)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    combo.grid(row=2,column=0,columnspan=2)
    label2.grid(row=3,column=0)
    entry2.grid(row=3,column=1)
    button.grid(row=4,column=0,columnspan=2)
    label3.grid(row=5,column=0,columnspan=2)

def golden_ratio():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def golden_ratio1():
        try:
            if var.get()==0:
                result=round(int(entry1.get())*1.618033988749894848204586834365,2)
                entry2.delete(0,END)
                entry2.insert(0,f"幅：{result} & 高さ：{int(entry1.get())}")
            elif var.get()==1:
                result=round(int(entry1.get())/1.618033988749894848204586834365,2)
                entry2.delete(0,END)
                entry2.insert(0,f"幅：{int(entry1.get())} & 高さ：{result}")
        except:
            messagebox.showerror("エラー","失敗しました")

    var=IntVar()
    var.set(0)
    label1=ttk.Label(frame,text="基準の長さ：")
    radio1=ttk.Radiobutton(frame,text="基準を幅",variable=var,value=0)
    radio2=ttk.Radiobutton(frame,text="基準を高さ",variable=var,value=1)
    button=ttk.Button(frame,text="計算",command=golden_ratio1)
    label2=ttk.Label(frame,text="黄金比：")
    entry1=ttk.Entry(frame,width=20)
    entry2=ttk.Entry(frame,width=30)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    radio1.grid(row=2,column=0)
    radio2.grid(row=2,column=1)
    button.grid(row=3,column=0,columnspan=2)
    label2.grid(row=4,column=0)
    entry2.grid(row=4,column=1)

def piano():
    root1=Free_window()
    root1.attributes("-topmost", True)
    root1.title("電子音ピアノ")

    def piano1(n):
        duration = 50 #1音の長さ：通常 50, 推奨 10~100 長くしすぎるとBPMに影響
        tone_d = n-9
        freq = int(440*2**(tone_d/12)) #周波数
        winsound.Beep(freq, int(duration)) #440hz, 100ミリ秒の音

    button1=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(1))
    button3=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(3))
    button5=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(5))
    button6=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(6))
    button8=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(8))
    button10=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(10))
    button11=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(11))
    button13=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(13))
    button15=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(15))
    button16=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(16))
    button17=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(17))
    button18=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(18))
    button20=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(20))
    button22=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(22))
    button23=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(23))
    button25=Button(root1,width=5,height=10,bg="lightgray",command=lambda:piano1(25))


    button2=Button(root1,width=3,height=5,bg="black",command=lambda:piano1(2))
    button4=Button(root1,width=3,height=5,bg="black",command=lambda:piano1(4))
    button7=Button(root1,width=3,height=5,bg="black",command=lambda:piano1(7))
    button9=Button(root1,width=3,height=5,bg="black",command=lambda:piano1(9))
    button12=Button(root1,width=3,height=5,bg="black",command=lambda:piano1(12))
    button14=Button(root1,width=3,height=5,bg="black",command=lambda:piano1(14))
    button16=Button(root1,width=3,height=5,bg="black",command=lambda:piano1(16))
    button19=Button(root1,width=3,height=5,bg="black",command=lambda:piano1(19))
    button21=Button(root1,width=3,height=5,bg="black",command=lambda:piano1(21))
    button24=Button(root1,width=3,height=5,bg="black",command=lambda:piano1(24))

    button1.grid(row=0,column=0)
    button2.grid(row=0,column=0,columnspan=2,sticky=N)
    button3.grid(row=0,column=1)
    button4.grid(row=0,column=1,columnspan=2,sticky=N)
    button5.grid(row=0,column=2)
    button6.grid(row=0,column=3)
    button7.grid(row=0,column=3,columnspan=2,sticky=N)
    button8.grid(row=0,column=4)
    button9.grid(row=0,column=4,columnspan=2,sticky=N)
    button10.grid(row=0,column=5)
    button11.grid(row=0,column=6)
    button12.grid(row=0,column=6,columnspan=2,sticky=N)
    button13.grid(row=0,column=7)
    button14.grid(row=0,column=0+7,columnspan=2,sticky=N)
    button15.grid(row=0,column=1+7)
    button16.grid(row=0,column=1+7,columnspan=2,sticky=N)
    button17.grid(row=0,column=2+7)
    button18.grid(row=0,column=3+7)
    button19.grid(row=0,column=3+7,columnspan=2,sticky=N)
    button20.grid(row=0,column=4+7)
    button21.grid(row=0,column=4+7,columnspan=2,sticky=N)
    button22.grid(row=0,column=5+7)
    button23.grid(row=0,column=6+7)
    button24.grid(row=0,column=6+7,columnspan=2,sticky=N)
    button25.grid(row=0,column=7+7)

def json_csv():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def json_csv1(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for x in file_list:
            try:
                if var.get()==0:
                    y=x.replace(".json",".csv")
                    with open(x, 'r') as f:
                        json_dict = json.load(f)
                    with open(y, 'w', newline='') as f:
                        writer = csv.DictWriter(f, fieldnames=json_dict[0].keys(),
                                                doublequote=True,
                                                quoting=csv.QUOTE_ALL)
                        writer.writeheader()
                        writer.writerows(json_dict)
                elif var.get()==1:
                    y=x.replace(".csv",".json")
                    with open(x, 'r') as f:
                        d_reader = csv.DictReader(f)
                        d_list = [row for row in d_reader]
                    with open(y, 'w') as f:
                        json.dump(d_list, f)
            except:
                messagebox.showerror("エラー",f"{x}\nでは失敗しました")
        messagebox.showinfo("完了","終了しました")


    var=IntVar()
    var.set(0)
    radio1=ttk.Radiobutton(frame,text="JSON→CSV",variable=var,value=0)
    radio2=ttk.Radiobutton(frame,text="CSV→JSON",variable=var,value=1)
    label1=ttk.Label(frame,text="ファイルをここに\nドロップしてください",font=("Helvetica", 20))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',json_csv1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    label1.grid(row=2,column=0,columnspan=2)

def exchange():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)
    url = f"https://api.coingecko.com/api/v3/exchange_rates"
    response = requests.get(url)
    data = response.json()

    def golden_ratio1():
        try:
            for currency in data["rates"]:
                if data["rates"][currency]["name"] == combox1.get():
                    value1 = data["rates"][currency]["value"]
                if data["rates"][currency]["name"] == combox2.get():
                    value2 = data["rates"][currency]["value"]

            values=float(entry1.get())*value2/value1
            entry2.delete(0,END)
            entry2.insert(0,round(values,4))
        except:
            messagebox.showerror("エラー","失敗しました")

    values1=[data["rates"][currency]["name"] for currency in data["rates"]]
    values2=values1.copy()

    combox1=Searchbox(frame,values=values1,width=30)
    combox2=Searchbox(frame,values=values2,width=30)
    label1=ttk.Label(frame,text="変換元：")
    label2=ttk.Label(frame,text="変換先：")
    label3=ttk.Label(frame,text="変換元の金額：")
    entry1=ttk.Entry(frame,width=40)
    button=ttk.Button(frame,text="計算",command=golden_ratio1)
    label4=ttk.Label(frame,text="変換後の金額：")
    entry2=ttk.Entry(frame,width=40)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")

    entry1.insert(0,1)
    combox1.set("US Dollar")
    combox2.set("Japanese Yen")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    combox1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    combox2.grid(row=2,column=1)
    label3.grid(row=3,column=0)
    entry1.grid(row=3,column=1)
    button.grid(row=4,column=0,columnspan=2)
    label4.grid(row=5,column=0)
    entry2.grid(row=5,column=1)

def score_calc():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)
    num1=0
    num2=0

    def result_reset():
        nonlocal num1,num2
        num1=0
        num2=0
        result_math()

    def result_math():
        if num1==0 and num2==0:
            sum_q=0
            suc_rate=0
        else:
            sum_q=num1+num2
            suc_rate=round(num1/sum_q*100,2)
        label1["text"]=f"正解数：{num1}"
        label2["text"]=f"不正回数：{num2}"
        label3["text"]=f"総問題数：{sum_q}"
        label4["text"]=f"正答率：{suc_rate}%"

    def score_calc1(num):
        nonlocal num1,num2
        if num==1:
            num1+=1
            result_math()
        elif num==2:
            if num1!=0:
                num1-=1
                result_math()
        elif num==3:
            num2+=1
            result_math()
        elif num==4:
            if num2!=0:
                num2-=1
                result_math()


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame,text=f"正解数：{num1}")
    label2=ttk.Label(frame,text=f"不正回数：{num2}")
    button1_1=ttk.Button(frame,text="+1",command=lambda:score_calc1(1),width=8)
    button1_2=ttk.Button(frame,text="-1",command=lambda:score_calc1(2),width=8)
    button2_1=ttk.Button(frame,text="+1",command=lambda:score_calc1(3),width=8)
    button2_2=ttk.Button(frame,text="-1",command=lambda:score_calc1(4),width=8)
    button3=ttk.Button(frame,text="リセット",command=result_reset)
    label3=ttk.Label(frame,text="総問題数：",font=("Helvetica", 10))
    label4=ttk.Label(frame,text="正答率：",font=("Helvetica", 10))

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    button1_1.grid(row=1,column=1)
    button1_2.grid(row=1,column=2)
    label2.grid(row=2,column=0)
    button2_1.grid(row=2,column=1)
    button2_2.grid(row=2,column=2)
    button3.grid(row=3,column=0,columnspan=3)
    label3.grid(row=4,column=0,columnspan=3)
    label4.grid(row=5,column=0,columnspan=3)

def video_time_image():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def video_time_image1(drop):
        try:
            path=drop.data.replace("{","").replace("}","").replace("\\","/")
            name=os.path.basename(path)

            def drawTextRightBottom(image, text) :
                hmargin = 5
                color = (0,0,0)
                thickness   = 1
                fontFace    = cv2.FONT_HERSHEY_SIMPLEX
                fontScale   = 0.5
                image_size = image.shape[:2]

                # テキストの描画サイズ取得
                sz  = cv2.getTextSize(text, fontFace, fontScale, thickness)

                # テキスト描画位置
                x = image_size[1] - sz[0][0] - hmargin
                y = image_size[0] - sz[1]

                # 矩形描画位置
                rx = x
                ry = image_size[0] - sz[0][1] - sz[1]

                cv2.rectangle(image, (rx, ry), (image_size[1]-1, image_size[0]-1), (255, 255, 255), -1, cv2.LINE_AA)
                cv2.putText(image, text, (x, y), fontFace, fontScale, color, thickness, cv2.LINE_AA)

            #cap = cv2.VideoCapture(path)
            cap = cv2.VideoCapture(u"{path}".format(path=path))
            if not cap.isOpened():
                sys.exit()

            #動画プロパティ取得
            v_w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))    # 縦の大きさ
            v_h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))   # 横の大きさ
            fps = cap.get(cv2.CAP_PROP_FPS)                 # フレームレート
            fcnt =  cap.get(cv2.CAP_PROP_FRAME_COUNT)       # フレーム数

            div_horizontal = 5  # タイリング横個数
            div_vertical = 5    # タイリング縦個数

            # 1タイル画像の縮小率
            sc = min(1 / div_horizontal, 1 / div_vertical)
            loop_cnt = div_horizontal * div_vertical
            resimage = np.zeros((v_h, v_w, 3),np.uint8)
            resize_width = int(v_w * sc)
            resize_height = int(v_h * sc)

            for i in range(0, loop_cnt):
                # フレーム位置算出
                if i == 0:
                    # 初回
                    fpos = 0
                    fpos2 = int((i+1 * fcnt) / loop_cnt)
                    # 先頭だけ5秒を読み込み
                    fpos1 = int(5.1 * fps)
                    if fpos2 < fpos1 :
                        fpos = int(fpos2 / 2)
                    else:
                        fpos = fpos1
                else:
                    fpos = int(i * fcnt / loop_cnt)

                # フレームシーク 読み出し位置指定
                cap.set(cv2.CAP_PROP_POS_FRAMES, fpos)
                ret, frame = cap.read()

                # 画像縮小
                resize_img = cv2.resize(frame, None, fx=sc, fy=sc, interpolation=cv2.INTER_CUBIC)

                # 画像の描画位置計算
                iy = i // div_horizontal
                ix = i % div_horizontal
                y = iy * resize_height
                x = ix * resize_width

                # 画像の時間を右下に描画
                totalsec = fpos // fps
                mi = totalsec // 60
                se = totalsec % 60
                text = '%3d:%02d' % (mi, se)
                drawTextRightBottom(resize_img, text)

                # 画像転送
                resimage[y:y+resize_height, x:x + resize_width, :] = resize_img

                # 画像ファイル出力
                #cv2.imwrite(path.replace(name,"digest.png"), resimage)
                _, buf = cv2.imencode('*.png', resimage)
                buf.tofile(path.replace(name,"digest.png"))
            messagebox.showinfo("完了","終了しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame,text="動画をここに\nドロップしてください",font=("Helvetica", 20))
    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',video_time_image1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0,columnspan=2)

def pdf_insert():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def pdf_insert1_1(drop):
        path=drop.data.replace("{","").replace("}","").replace("\\","/")
        entry1.delete(0,END)
        entry1.insert(0,path)

    def pdf_insert1_2(drop):
        path=drop.data.replace("{","").replace("}","").replace("\\","/")
        entry2.delete(0,END)
        entry2.insert(0,path)

    def pdf_insert1():
        try:
            merger = pypdf.PdfMerger()
            merger.append(entry1.get())
            merger.merge(int(entry3.get())-1,entry2.get())
            pdf_path=filedialog.asksaveasfilename(filetypes=[("PDF","*.pdf")],initialfile="new.pdf")
            merger.write(pdf_path)
            merger.close()
            messagebox.showinfo("完了","終了しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame,text="挿入されるPDF：")
    entry1=ttk.Entry(frame,width=40)
    label2=ttk.Label(frame,text="挿入するPDF：")
    entry2=ttk.Entry(frame,width=40)
    label3=ttk.Label(frame,text="挿入するページ：")
    entry3=ttk.Entry(frame,width=10)
    button=ttk.Button(frame,text="実行",command=pdf_insert1)
    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',pdf_insert1_1)
    label2.drop_target_register(DND_FILES)
    label2.dnd_bind('<<Drop>>',pdf_insert1_2)
    entry1.drop_target_register(DND_FILES)
    entry1.dnd_bind('<<Drop>>',pdf_insert1_1)
    entry2.drop_target_register(DND_FILES)
    entry2.dnd_bind('<<Drop>>',pdf_insert1_2)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    label3.grid(row=3,column=0)
    entry3.grid(row=3,column=1)
    button.grid(row=4,column=0,columnspan=2)

def word_book():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def word_book1(drop):
        try:
            question_num=0
            answer_list=[]

            def create_csv_file(incorrect_questions):
                try:
                    path=filedialog.asksaveasfilename(filetypes=[("CSV","*.csv")],initialfile="incorrect_questions.csv",defaultextension=".csv")
                    with open(path, mode="w", newline="",encoding="shift_jis",) as file:
                        writer = csv.writer(file)
                        for question in incorrect_questions:
                            if question[0] and question[1]:  # Skip empty lines
                                writer.writerow([question[0].encode('cp932').decode('shift_jis'), question[1].encode('cp932').decode('shift_jis')])

                    messagebox.showinfo("完了", "誤答問題CSVを作成しました: incorrect_questions.csv")
                except Exception as e:
                    messagebox.showerror("エラー", f"CSV作成エラー: {str(e)}")

            def create_book():
                nonlocal question_num,answer_list

                def create_result():
                    nonlocal question_num,answer_list
                    frame_question.destroy()
                    result_list=[]
                    incorrect_questions = []
                    sum_question=0
                    sum_correct=0
                    for i in range(len(word_text)):
                        if answer_list[i]==0:
                            result_session=[f"{i+1}","◯",f"{word_text[i][0]}",f"{word_text[i][1]}"]
                            sum_question+=1
                            sum_correct+=1
                        elif answer_list[i]==1:
                            result_session=[f"{i+1}","✕",f"{word_text[i][0]}",f"{word_text[i][1]}"]
                            sum_question+=1
                            incorrect_questions.append(word_text[i])
                        result_list.append(result_session)
                    frame_tree=ttk.Frame(root1)
                    label_sum_question=ttk.Label(root1,text=f"総問題数：{sum_question}\n正解数：{sum_correct}\n正解率：{round(sum_correct/sum_question*100,2)}%",font=("Helvetica", 15),anchor=CENTER)
                    tree=ScrolledTree(frame_tree,arrRows=result_list,columns=["番号","正誤","問題","答え"],arrColWidth=[50,50,500,500],arrSortType=["num","name","name","name"],height=20)

                    label_sum_question.pack(side=TOP,fill="both")
                    frame_tree.pack(side=TOP,fill="both", expand=True)
                    tree.pack(side=LEFT,fill="both", expand=True)

                        # Add a button to create CSV with incorrect questions
                    button_create_csv = ttk.Button(frame_tree, text="誤答問題CSV作成", command=lambda: create_csv_file(incorrect_questions))
                    button_create_csv.pack(side=TOP)

                    frame_tree.pack(side=TOP, fill="both", expand=True)
                    tree.pack(side=LEFT, fill="both", expand=True)


                def next_question(num):
                    nonlocal question_num,answer_list
                    if num==0:
                        answer_list.append(0)
                        question_num+=1
                        if question_num==len(word_text):
                            create_result()
                            messagebox.showinfo("完了","終了しました")
                            return
                    elif num==1:
                        answer_list.append(1)
                        question_num+=1
                        if question_num==len(word_text):
                            create_result()
                            messagebox.showinfo("完了","終了しました")
                            return
                    label_question["text"]=word_text[question_num][0]
                    label_question_num["text"]=f"番号：{question_num+1}"

                def answer_open():
                    if label_question["text"]==word_text[question_num][0]:
                        label_question["text"]=word_text[question_num][1]
                        button_answer["text"]="問題を表示"
                    else:
                        label_question["text"]=word_text[question_num][0]
                        button_answer["text"]="答えを表示"


                root1=Free_window()
                root1.attributes("-topmost", True)
                root1.title("単語帳")
                frame_question=ttk.Frame(root1)
                label_question=ttk.Label(frame_question,text="",font=("Helvetica", 30),width=30,anchor=CENTER,wraplength=300)
                button_answer=ttk.Button(frame_question,text="答えを表示",command=answer_open)
                button_correct_answer=Button(frame_question,text="◯",command=lambda:next_question(0),font=("Helvetica", 15))
                button_wrong_answer=Button(frame_question,text="✕",command=lambda:next_question(1),font=("Helvetica", 15))
                label_question_num=ttk.Label(frame_question,text=f"番号：1",anchor=CENTER,font=("Helvetica", 15))

                frame_question.pack()
                label_question.grid(row=3,column=0,columnspan=2)
                button_answer.grid(row=2,column=0,columnspan=2)
                button_correct_answer.grid(row=1,column=0)
                button_wrong_answer.grid(row=1,column=1)
                label_question_num.grid(row=0,column=0,columnspan=2)

                next_question(2)


            path=drop.data.replace("{","").replace("}","").replace("\\","/")
            word_text=[]
            try:
                with open(path, 'r') as f:
                    reader = csv.reader(f)
                    for line in reader:
                        if line!=["",""]:
                            word_text.append(line)
            except:
                with open(path, 'r',encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for line in reader:
                        if line!=["",""]:
                            word_text.append(line)
            if var2.get()==1:
                random.shuffle(word_text)
            if var1.get()==1:
                for i in range(len(word_text)):
                    word_text[i][0],word_text[i][1]=word_text[i][1],word_text[i][0]
            create_book()
        except:
            messagebox.showerror("エラー","失敗しました")


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame,text="単語帳CSVをここに\nドロップしてください\n(1列目に問題・2列目に答え)",font=("Helvetica", 20))
    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',word_book1)
    var1=IntVar()
    var1.set(0)
    var2=IntVar()
    var2.set(0)
    radio1=ttk.Checkbutton(frame,text="問答入れ替え",variable=var1,onvalue=1,offvalue=0)
    radio2=ttk.Checkbutton(frame,text="ランダム",variable=var2,onvalue=1,offvalue=0)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    label1.grid(row=2,column=0,columnspan=2)

def color_name():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def color_name1():

        try:
            try:
                url="https://www.thecolorapi.com/id?"
                if var1.get()==0:
                    url=url+f"rgb={entry1.get()}"
                elif var1.get()==1:
                    url=url+f"hex={entry1.get()[1:]}"
                response = requests.get(url)
                data = response.json()
                if data["cmyk"]["fraction"]["c"]==None:
                    messagebox.showerror("エラー","入力が不正です")
                    return
                name=data["name"]["value"]
                hex=data["hex"]["value"]
                rgb=data["rgb"]["value"]
                rgb=rgb.replace("rgb(","").replace(")","")
            except:
                if var1.get()==0:
                    rgb1=entry1.get().split(",")
                    rgb=f"{rgb1[0]},{rgb1[1]},{rgb1[2]}"
                    rgb2=[int(rgb1[0]),int(rgb1[1]),int(rgb1[2])]
                    hex ="#{:02x}{:02x}{:02x}".format(*rgb2)
                elif var1.get()==1:
                    hex=entry1.get()
                    rgb_veiw = tuple(int(hex[i:i+2], 16) for i in (1, 3, 5))
                    rgb=f"{rgb_veiw[0]},{rgb_veiw[1]},{rgb_veiw[2]}"
                name="不明"
        except:
            messagebox.showerror("エラー","失敗しました")
            return

        root1=Free_window()
        root1.attributes("-topmost", True)
        root1.title("色情報")
        menu=[["名前コピー",lambda:clip.copy(name)],
              ["HEXコピー",lambda:clip.copy(hex)],
              ["RGBコピー",lambda:clip.copy(rgb)],
              ["標本画像URLコピー",lambda:clip.copy(data["image"]["named"])],
              ["名前検索",lambda:webbrowser.open(f"https://www.google.com/search?q={name}")],
              ]

        RightClickMenu(root1,menu)

        label1=ttk.Label(root1,text=f"名前：{name}",font=("Helvetica", 20))
        label2=ttk.Label(root1,text=f"HEX：{hex}",font=("Helvetica", 20))
        label3=ttk.Label(root1,text=f"RGB：{rgb}",font=("Helvetica", 20))
        label4=Label(root1,bg=hex.lower(),height=3,relief="groove")

        label1.pack(expand=True,fill="both")
        label2.pack(expand=True,fill="both")
        label3.pack(expand=True,fill="both")
        label4.pack(expand=True,fill="both")



    var1=IntVar()
    var1.set(0)
    radio1=ttk.Radiobutton(frame,text="RGB(,区切り)",variable=var1,value=0)
    radio2=ttk.Radiobutton(frame,text="HEX(#~)",variable=var1,value=1)
    entry1=ttk.Entry(frame,width=40)
    button1=ttk.Button(frame,text="検索",command=color_name1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    entry1.grid(row=2,column=0,columnspan=2)
    button1.grid(row=3,column=0,columnspan=2)

def desc_mascot():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def desc_mascot1(drop):
        path=drop.data.replace("{","").replace("}","").replace("\\","/")
        if os.path.splitext(path)[1]!=".png":
            messagebox.showerror("エラー","PNGファイルをドロップしてください")
            return
        root1=Free_window()
        root1.move_window(root1)
        root1.attributes("-topmost", True)
        root1.title("マスコット")
        root1.resizable(False, False)

        canvas=Canvas(root1,bd=0, highlightthickness=0)
        image=ImageTk.PhotoImage(file=path)
        width=image.width()
        height=image.height()
        root1.geometry(f"{width}x{height}")
        canvas.config(width=width,height=height)
        canvas.canvas_img = image
        canvas.create_image(0,0,image=image,anchor=NW)
        canvas.pack(fill="both")
        canvas.update()
        root1.wm_attributes("-transparentcolor", canvas["bg"])
        root1.overrideredirect(True)
        RightClickMenu(root1,[["終了",root1.destroy]])


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame,text="マスコットのPNG画像を\nドロップしてください",font=("Helvetica", 20))
    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',desc_mascot1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0,columnspan=2)

def auto_oparation():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)


    def operation_add():
        nonlocal action_list,kurikaesu
        action_name=search.get()
        if kurikaesu==1:
            messagebox.showerror("エラー","繰り返すは最後の一回のみ使用できます")
            return

        def on_press(event):
            nonlocal start_x,canvas,action_list,start_y
            start_x = canvas.canvasx(event.x)
            start_y = canvas.canvasy(event.y)
            root1.destroy()
            action_list.append([len(action_list)+1,action_name,f"{int(start_x)},{int(start_y)}"])
            tree.configure_arr(action_list)
            tree.update()
            return

        if search.get()=="右クリック":
            start_x=None
            start_y=None
            root.iconify()
            time.sleep(2)
            make_folder(os.getcwd()+"/temp1/img_auto")
            pyautogui.screenshot(os.getcwd()+"/temp1/img_auto/temp.png")
            root.deiconify()
            root1 = Free_window()
            root1.title(search.get())
            pil_img=Image.open(os.getcwd()+"/temp1/img_auto/temp.png")
            canvas=Canvas(root1,width=pil_img.width,height=pil_img.height)
            canvas_img= ImageTk.PhotoImage(file=os.getcwd()+"/temp1/img_auto/temp.png")
            canvas.canvas_img = canvas_img
            canvas.create_image(pil_img.width*0.5,pil_img.height*0.5,image=canvas_img)
            canvas.pack(expand=True,fill=BOTH)
            canvas.bind("<ButtonPress-1>", on_press)
            root1.update()
            root1.attributes('-fullscreen', True)
            root1.attributes('-topmost', '1')
            root1.resizable(False, False)

        elif search.get()=="左クリック":
            start_x=None
            start_y=None
            root.iconify()
            time.sleep(2)
            make_folder(os.getcwd()+"/temp1/img_auto")
            pyautogui.screenshot(os.getcwd()+"/temp1/img_auto/temp.png")
            root.deiconify()
            root1 = Free_window()
            root1.title(search.get())
            pil_img=Image.open(os.getcwd()+"/temp1/img_auto/temp.png")
            canvas=Canvas(root1,width=pil_img.width,height=pil_img.height)
            canvas_img= ImageTk.PhotoImage(file=os.getcwd()+"/temp1/img_auto/temp.png")
            canvas.canvas_img = canvas_img
            canvas.create_image(pil_img.width*0.5,pil_img.height*0.5,image=canvas_img)
            canvas.pack(expand=True,fill=BOTH)
            canvas.bind("<ButtonPress-1>", on_press)
            root1.update()
            root1.attributes('-fullscreen', True)
            root1.attributes('-topmost', '1')
            root1.resizable(False, False)

        elif search.get()=="左ダブルクリック":
            start_x=None
            start_y=None
            root.iconify()
            time.sleep(2)
            make_folder(os.getcwd()+"/temp1/img_auto")
            pyautogui.screenshot(os.getcwd()+"/temp1/img_auto/temp.png")
            root.deiconify()
            root1 = Free_window()
            root1.title(search.get())
            pil_img=Image.open(os.getcwd()+"/temp1/img_auto/temp.png")
            canvas=Canvas(root1,width=pil_img.width,height=pil_img.height)
            canvas_img= ImageTk.PhotoImage(file=os.getcwd()+"/temp1/img_auto/temp.png")
            canvas.canvas_img = canvas_img
            canvas.create_image(pil_img.width*0.5,pil_img.height*0.5,image=canvas_img)
            canvas.pack(expand=True,fill=BOTH)
            canvas.bind("<ButtonPress-1>", on_press)
            root1.update()
            root1.attributes('-fullscreen', True)
            root1.attributes('-topmost', '1')
            root1.resizable(False, False)

        elif search.get()=="待機":
            action_name=search.get()
            inputdata = simpledialog.askstring("待機時間", "待機時間を入力してください(秒)",)
            if inputdata==None:
                return
            else:
                if int(inputdata)>0:
                    action_list.append([len(action_list)+1,action_name,f"{inputdata}"])
                    tree.configure_arr(action_list)

        elif search.get()=="繰り返す":
            action_name=search.get()
            inputdata = simpledialog.askstring("繰り返し回数", "繰り返す回数を入力してください(回)",)
            if inputdata==None:
                return
            else:
                if int(inputdata)>0:
                    action_list.append([len(action_list)+1,action_name,f"{inputdata}"])
                    tree.configure_arr(action_list)
                    kurikaesu=1

        elif search.get()=="キーボード入力":
            key_con=None
            input_num_str = simpledialog.askstring("同時入力数", "同時入力数を入力してください(回)",)
            input_num=int(input_num_str)
            def key_contain():
                nonlocal key_con,action_list
                key_con1=[]
                while True:
                    key_con = key.read_key()
                    if key_con in key_con1:
                        continue
                    key_con1.append(key_con)
                    if len(key_con1)==input_num:
                        key_con_str="+".join(key_con1)
                        root2.destroy()
                        action_list.append([len(action_list)+1,action_name,f"{key_con_str}"])
                        tree.configure_arr(action_list)
                        break
            action_name=search.get()
            root2=Free_window()
            ws=root2.winfo_screenwidth()
            hs=root2.winfo_screenheight()
            x2=(ws/2)-400
            y2=(hs/2)-200
            root2.geometry('+%d+%d'%(x2,y2))
            root2.title("キーボード入力")
            root2.attributes("-topmost", True)
            label1=ttk.Label(root2,text="キーボード入力を\nしてください",font=("Helvetica", 20))
            label1.pack(expand=True,fill="both")
            t1=threading.Thread(target=key_contain,daemon=True)
            t1.start()

        elif search.get()=="テキスト入力":
            def save_text():
                nonlocal action_list
                action_list.append([len(action_list)+1,action_name,f"{text.get('1.0', 'end -1c')}"])
                tree.configure_arr(action_list)
                root2.destroy()

            action_name=search.get()
            root2=Free_window()
            ws=root2.winfo_screenwidth()
            hs=root2.winfo_screenheight()
            x2=(ws/2)-400
            y2=(hs/2)-200
            root2.geometry('+%d+%d'%(x2,y2))
            root2.title("キーボード入力")
            root2.attributes("-topmost", True)
            label1=ttk.Label(root2,text="テキスト入力してください",font=("Helvetica", 20),anchor=CENTER)
            text=ScrolledText(root2,width=50,height=10)
            button1=ttk.Button(root2,text="保存",command=save_text)
            label1.pack(expand=True,fill="both")
            text.pack(expand=True,fill="both")
            button1.pack(expand=True,fill="both")

        elif search.get()=="マウス移動":
            def on_press1(event):
                nonlocal start_x,canvas,action_list,start_y
                start_x = canvas.canvasx(event.x)
                start_y = canvas.canvasy(event.y)
                root1.destroy()
                action_list.append([len(action_list)+1,action_name,f"{duration},{int(start_x)},{int(start_y)}"])
                tree.configure_arr(action_list)
                tree.update()
                return


            start_x=None
            start_y=None
            duration_str=simpledialog.askstring("移動時間", "移動時間を入力してください(秒)",)
            duration=float(duration_str)
            root.iconify()
            time.sleep(2)
            make_folder(os.getcwd()+"/temp1/img_auto")
            pyautogui.screenshot(os.getcwd()+"/temp1/img_auto/temp.png")
            root.deiconify()
            root1 = Free_window()
            root1.title(search.get())
            pil_img=Image.open(os.getcwd()+"/temp1/img_auto/temp.png")
            canvas=Canvas(root1,width=pil_img.width,height=pil_img.height)
            canvas_img= ImageTk.PhotoImage(file=os.getcwd()+"/temp1/img_auto/temp.png")
            canvas.canvas_img = canvas_img
            canvas.create_image(pil_img.width*0.5,pil_img.height*0.5,image=canvas_img)
            canvas.pack(expand=True,fill=BOTH)
            canvas.bind("<ButtonPress-1>", on_press1)
            root1.update()
            root1.attributes('-fullscreen', True)
            root1.attributes('-topmost', '1')
            root1.resizable(False, False)

        elif search.get()=="ドラッグ":
            def on_press1(event):
                nonlocal start_x,canvas,action_list,start_y
                start_x = canvas.canvasx(event.x)
                start_y = canvas.canvasy(event.y)
                root1.destroy()
                action_list.append([len(action_list)+1,action_name,f"{duration},{int(start_x)},{int(start_y)}"])
                tree.configure_arr(action_list)
                tree.update()
                return

            start_x=None
            start_y=None
            duration_str=simpledialog.askstring("移動時間", "移動時間を入力してください(秒)",)
            duration=float(duration_str)
            root.iconify()
            time.sleep(2)
            make_folder(os.getcwd()+"/temp1/img_auto")
            pyautogui.screenshot(os.getcwd()+"/temp1/img_auto/temp.png")
            root.deiconify()
            root1 = Free_window()
            root1.title(search.get())
            pil_img=Image.open(os.getcwd()+"/temp1/img_auto/temp.png")
            canvas=Canvas(root1,width=pil_img.width,height=pil_img.height)
            canvas_img= ImageTk.PhotoImage(file=os.getcwd()+"/temp1/img_auto/temp.png")
            canvas.canvas_img = canvas_img
            canvas.create_image(pil_img.width*0.5,pil_img.height*0.5,image=canvas_img)
            canvas.pack(expand=True,fill=BOTH)
            canvas.bind("<ButtonPress-1>", on_press1)
            root1.update()
            root1.attributes('-fullscreen', True)
            root1.attributes('-topmost', '1')
            root1.resizable(False, False)

        elif search.get()=="サイトを開く":
            url=simpledialog.askstring("URL", "URLを入力してください",)
            if url==None:
                return
            else:
                action_list.append([len(action_list)+1,action_name,f"{url}"])
                tree.configure_arr(action_list)

        elif search.get()=="実行ファイルを実行":
            path=filedialog.askopenfilename(filetypes=[("実行ファイル","*.exe")])
            if path=="":
                return
            else:
                action_list.append([len(action_list)+1,action_name,f"{path}"])
                tree.configure_arr(action_list)

        elif search.get()=="ファイルを開く":
            path1=filedialog.askopenfilename(filetypes=[("実行ファイル","*.exe")],title="ソフトを指定する")
            path2=filedialog.askopenfilename(title="開くファイルを指定する")
            action_list.append([len(action_list)+1,action_name,f"{path1},{path2}"])
            tree.configure_arr(action_list)


        return


    def tree_click():
        nonlocal action_list,kurikaesu
        record_values = tree.item(tree.focus(), 'values')
        if record_values[1]=="繰り返す":
            kurikaesu=0
        action_list.pop(int(record_values[0])-1)
        for i in range(len(action_list)):
            action_list[i][0]=i+1
        tree.configure_arr(action_list)


    def operation_execute():
        t1=threading.Thread(target=operation_execute_t,daemon=True)
        t2=threading.Thread(target=operation_execute_t2,daemon=True)
        t1.start()
        t2.start()

    def operation_execute_t2():
        nonlocal stop_key
        while stop_key==0:
            if key.read_key() == "esc":
                stop_key=1
                break

    def operation_execute_t():
        nonlocal stop_key
        try:
            opp_num=0
            while opp_num!=-1:
                for x_list in action_list:
                    if stop_key==1:
                        return
                    if x_list[1]=="右クリック":
                        time.sleep(float(entry1.get()))
                        x,y=x_list[2].split(",")
                        x,y=int(x),int(y)
                        pyautogui.click(x=x,y=y,button="right")
                    elif x_list[1]=="左クリック":
                        time.sleep(float(entry1.get()))
                        x,y=x_list[2].split(",")
                        x,y=int(x),int(y)
                        pyautogui.click(x=x,y=y,button="left")
                    elif x_list[1]=="左ダブルクリック":
                        time.sleep(float(entry1.get()))
                        x,y=x_list[2].split(",")
                        x,y=int(x),int(y)
                        pyautogui.doubleClick(x=x,y=y,button="left")
                    elif x_list[1]=="待機":
                        time.sleep(int(x_list[2]))
                    elif x_list[1]=="繰り返す":
                        if opp_num==int(x_list[2])-1:
                            opp_num=-1
                        else:
                            opp_num+=1
                    elif x_list[1]=="キーボード入力":
                        time.sleep(float(entry1.get()))
                        key_pressed=x_list[2]
                        key.press_and_release(key_pressed)
                    elif x_list[1]=="テキスト入力":
                        time.sleep(float(entry1.get()))
                        text=x_list[2]
                        key.write(text)
                    elif x_list[1]=="マウス移動":
                        time.sleep(float(entry1.get()))
                        duration_str,x,y=x_list[2].split(",")
                        x,y=int(x),int(y)
                        duration=float(duration_str)
                        pyautogui.moveTo(x=x,y=y,duration=duration)
                    elif x_list[1]=="ドラッグ":
                        time.sleep(float(entry1.get()))
                        duration_str,x,y=x_list[2].split(",")
                        x,y=int(x),int(y)
                        duration=float(duration_str)
                        pyautogui.dragTo(x=x,y=y,duration=duration)
                    elif x_list[1]=="サイトを開く":
                        time.sleep(float(entry1.get()))
                        url=x_list[2]
                        webbrowser.open(url)
                    elif x_list[1]=="実行ファイルを実行":
                        time.sleep(float(entry1.get()))
                        path=x_list[2]
                        subprocess.Popen(path)
                    elif x_list[1]=="ファイルを開く":
                        time.sleep(float(entry1.get()))
                        path1,path2=x_list[2].split(",")
                        subprocess.Popen([path1,path2])
                if opp_num==0:
                    opp_num=-1
            stop_key=1
        except:
            messagebox.showerror("エラー","失敗しました")

    def operation_save():
        path=filedialog.asksaveasfilename(filetypes=[("CSV","*.csv")],initialfile="save_opp.csv")
        with open(path, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerows(action_list)

    def operation_load():
        nonlocal action_list
        path=filedialog.askopenfilename(filetypes=[("CSV","*.csv")])
        with open(path, "r") as csv_file:
            csv_reader = csv.reader(csv_file)
            action_list = list(csv_reader)
        tree.configure_arr(action_list)

    def tree_move_up():
        nonlocal action_list
        try:
            record_values = tree.item(tree.focus(), 'values')
            if action_list[int(record_values[0])-1][1]=="繰り返す":
                messagebox.showerror("エラー","繰り返すは最後の一回のみ使用できます")
            elif action_list[int(record_values[0])-2][1]=="繰り返す":
                messagebox.showerror("エラー","繰り返すは最後の一回のみ使用できます")
            action_list[int(record_values[0])-1],action_list[int(record_values[0])-2]=action_list[int(record_values[0])-2],action_list[int(record_values[0])-1]
            action_list[int(record_values[0])-1][0]=int(record_values[0])
            action_list[int(record_values[0])-2][0]=int(record_values[0])-1
            tree.configure_arr(action_list)
        except:
            messagebox.showerror("エラー","失敗しました")

    def tree_move_down():
        nonlocal action_list
        try:
            record_values = tree.item(tree.focus(), 'values')
            if action_list[int(record_values[0])][1]=="繰り返す":
                messagebox.showerror("エラー","繰り返すは最後の一回のみ使用できます")
            elif action_list[int(record_values[0])-1][1]=="繰り返す":
                messagebox.showerror("エラー","繰り返すは最後の一回のみ使用できます")
            action_list[int(record_values[0])-1],action_list[int(record_values[0])]=action_list[int(record_values[0])],action_list[int(record_values[0])-1]
            action_list[int(record_values[0])-1][0]=int(record_values[0])
            action_list[int(record_values[0])][0]=int(record_values[0])+1
            tree.configure_arr(action_list)
        except:
            messagebox.showerror("エラー","失敗しました")

    def tree_delete():
        nonlocal action_list
        action_list=[]
        tree.configure_arr(action_list)

    stop_key=0
    kurikaesu=0
    action=["左クリック","右クリック","左ダブルクリック","待機","キーボード入力","テキスト入力","マウス移動","ドラッグ","サイトを開く","実行ファイルを実行","ファイルを開く","繰り返す"]
    search=ttk.Combobox(frame,values=action,width=20,state="readonly")
    button1=ttk.Button(frame,text="追加",command=operation_add)
    action_list=[]
    tree=ScrolledTree(frame,arrRows=action_list,columns=["番号","操作","内容"],arrColWidth=[100,100,200],arrSortType=["num","name","name"],height=20)
    button2=ttk.Button(frame,text="セーブ",command=operation_save)
    button3=ttk.Button(frame,text="ロード",command=operation_load)
    button4=ttk.Button(frame,text="実行",command=operation_execute)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    button5=ttk.Button(frame,text="削除",command=tree_click)
    button6=ttk.Button(frame,text="全削除",command=tree_delete)
    button7=ttk.Button(frame,text="上へ",command=tree_move_up)
    button8=ttk.Button(frame,text="下へ",command=tree_move_down)
    label1=ttk.Label(frame,text="操作感覚（秒）：")
    entry1=ttk.Entry(frame,width=10)
    entry1.insert(0,"0.5")


    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    search.grid(row=1,column=0)
    button1.grid(row=1,column=1)
    button5.grid(row=1,column=2)
    button7.grid(row=3,column=0)
    button8.grid(row=3,column=1)
    button6.grid(row=3,column=2)
    tree.grid(row=2,column=0,columnspan=3)
    label1.grid(row=4,column=0)
    entry1.grid(row=4,column=1,columnspan=2)
    button2.grid(row=5,column=1)
    button3.grid(row=5,column=2)
    button4.grid(row=5,column=0)


def compulsion_paste():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def on_release(key1):
        if key1 == keyboard.Key.pause:
            key.write(clip.paste())


    def main_frame1(num):
        listener.stop()
        t.join()
        root.protocol("WM_DELETE_WINDOW", frame_delete)
        main_frame(num)

    def stopping():
        listener.stop()
        t.join()
        frame_delete()

    listener=None
    root.protocol("WM_DELETE_WINDOW", stopping)
    label=ttk.Label(frame,text="pauseキーで\n強制ペーストを行います",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(0),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)
    def listen_thread():
        nonlocal listener
        with keyboard.Listener(on_release=on_release) as listener:
            listener.join()

    t = threading.Thread(target=listen_thread, daemon=True)
    t.start()

def window_tray():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def on_release(key):

        if key == keyboard.Key.pause:
            buttonY["state"] = "disabled"

            def quit_icon():
                icon1.stop()
                label["text"]="pauseキーで\n一つのウィンドウをタスクトレイに収納します"
                buttonY["state"] = "normal"
                root.protocol("WM_DELETE_WINDOW", stopping)

            def stopping1():
                quit_icon()
                stopping()

            def quit_action():
                _, pid = win32process.GetWindowThreadProcessId(active_window)
                # The command to run
                cmd = f'taskkill /F /PID {pid}'

                # Run the command and ignore the output
                subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                #cmd = f'taskkill /F /PID {pid}'
                #os.system(cmd)
                icon1.stop()
                label["text"]="pauseキーで\n一つのウィンドウをタスクトレイに収納します"
                buttonY["state"] = "normal"
                root.protocol("WM_DELETE_WINDOW", stopping)

            def open_action():
                # ウィンドウが非表示かどうかを判定
                is_visible = win32gui.IsWindowVisible(active_window)

                if is_visible==0:
                    if active_window != 0:
                        # ウィンドウを再表示する
                        win32gui.ShowWindow(active_window, win32con.SW_SHOW)
                else:
                    if active_window != 0:
                        # ウィンドウを非表示にする
                        win32gui.ShowWindow(active_window, win32con.SW_HIDE)

            def extract_icon_from_exe(exe_path, icon_index=0):
                # アイコンの大きさ
                icon_size = win32api.GetSystemMetrics(win32con.SM_CXICON), win32api.GetSystemMetrics(win32con.SM_CYICON)
                # アイコン抽出
                large, small = win32gui.ExtractIconEx(exe_path, icon_index)
                # 大きいアイコン
                if large:
                    # Bitmapオブジェクト
                    hdc = win32ui.CreateDCFromHandle(win32gui.GetDC(0))
                    hbmp = win32ui.CreateBitmap()
                    hbmp.CreateCompatibleBitmap(hdc, icon_size[0], icon_size[1])
                    hdc = hdc.CreateCompatibleDC()
                    hdc.SelectObject(hbmp)
                    hdc.DrawIcon((0, 0), large[0])
                    hdc.DeleteDC()
                    # PILのImageオブジェクトに変換
                    bmpinfo = hbmp.GetInfo()
                    bmpstr = hbmp.GetBitmapBits(True)
                    icon_image = Image.frombuffer(
                        'RGBA',
                        (bmpinfo['bmWidth'], bmpinfo['bmHeight']),
                        bmpstr, 'raw', 'BGRA', 0, 1
                    )
                # リソース解放
                if small:
                    win32gui.DestroyIcon(small[0])
                if large:
                    win32gui.DestroyIcon(large[0])
                return icon_image

            def get_window_process_path(hwnd):
                # ウィンドウのプロセスID取得
                _, pid = win32process.GetWindowThreadProcessId(hwnd)
                # プロセスIDからプロセス取得
                process = psutil.Process(pid)
                # プロセスの実行ファイルのパス取得
                path = process.exe()

                return path

            root.protocol("WM_DELETE_WINDOW", stopping1)
            # アクティブなウィンドウ取得
            active_window = win32gui.GetForegroundWindow()
            if active_window != 0:
                # アクティブなウィンドウを非表示にする
                win32gui.ShowWindow(active_window, win32con.SW_HIDE)
                class_name = win32gui.GetWindowText(active_window)
                path = get_window_process_path(active_window)
                image=extract_icon_from_exe(path)
                task_list=[]
                task_list.append(pystray.MenuItem(text="ウィンドウ開閉", action=open_action, default=True))
                task_list.append(pystray.MenuItem(text="終了", action=quit_icon))
                task_list.append(pystray.MenuItem(text="強制終了", action=quit_action))
                task_menu=pystray.Menu(*task_list)

                icon1 = pystray.Icon(
                    name=class_name,
                    icon=image,
                    title=class_name,
                    menu=task_menu
                )
                label["text"]=f"{class_name}\nをタスクトレイに収納しました"
                icon1.run()


    def main_frame1(num):
        listener.stop()
        t.join()
        root.protocol("WM_DELETE_WINDOW", frame_delete)
        main_frame(num)

    def stopping():
        listener.stop()
        t.join()
        frame_delete()

    listener=None
    root.protocol("WM_DELETE_WINDOW", stopping)
    label=ttk.Label(frame,text="pauseキーで\n一つのウィンドウをタスクトレイに収納します",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)
    def listen_thread():
        nonlocal listener
        with keyboard.Listener(on_release=on_release) as listener:
            listener.join()

    t = threading.Thread(target=listen_thread, daemon=True)
    t.start()

def code_change():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def code_change1(drop):
        paths=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(paths)
        try:
            for x in x_list:
                if string_code.get()=="utf8":
                    ec=os.path.splitext(x)[1]
                    y=x.replace(ec,"_utf8"+ec)
                    # Read from the Shift-JIS encoded file
                    with open(x, 'r', encoding='shift_jis') as file:
                        content = file.read()
                    # Write to a new file with UTF-8 encoding
                    with open(y, 'w', encoding='utf-8') as file:
                        file.write(content)
                elif string_code.get()=="sjis":
                    ec=os.path.splitext(x)[1]
                    y=x.replace(ec,"_sjis"+ec)
                    # Read from the UTF-8 encoded file
                    with open(x, 'r', encoding='utf-8') as file:
                        content = file.read()
                    # Write to a new file with Shift-JIS encoding
                    with open(y, 'w', encoding='shift_jis') as file:
                        file.write(content)
            messagebox.showinfo("完了","変換が完了しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    string_code=StringVar()
    radio1=ttk.Radiobutton(frame,text="UTF-8→Shift-JIS",value="sjis",variable=string_code)
    radio2=ttk.Radiobutton(frame,text="Shift-JIS→UTF-8",value="utf8",variable=string_code)
    label1=ttk.Label(frame,text="ファイルを\nドロップしてください",font=("Helvetica", 20))
    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',code_change1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    string_code.set("sjis")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    label1.grid(row=2,column=0,columnspan=2)

def code_check():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def code_check1(drop):
        try:
            path=drop.data.replace("{","").replace("}","").replace("\\","/")
            with open(path, 'rb') as f:
                b = f.read()
            enc = detect(b)
            if enc["encoding"]=="Windows-1252":
                messagebox.showinfo("結果","shift-jis(Windows-1252の可能性もあり)です")
            elif enc["encoding"]==None:
                messagebox.showinfo("結果","判別できませんでした")
            else:
                messagebox.showinfo("結果",f"{enc['encoding']}の可能性が高いです")
        except:
            messagebox.showerror("エラー","失敗しました")


    label1=ttk.Label(frame,text="ファイルを\nドロップしてください",font=("Helvetica", 20))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")

    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',code_check1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0,columnspan=2)

def mouse_highlight():
    global mause_highlight_hwnd
    if mause_highlight_hwnd!=0:
        mause_highlight_hwnd=0
        return

    def mouse_move():
        while True:
            try:
                if mause_highlight_hwnd==0:
                    root1.destroy()
                    break
                x1, y1 = pyautogui.position()
                root1.geometry('+{x}+{y}'.format(x=x1-50, y=y1-50))
                time.sleep(0.03)
            except:
                break

    def remove_title_bar():
        global mause_highlight_hwnd

        hwnd = ctypes.windll.user32.GetParent(root1.winfo_id())
        mause_highlight_hwnd=hwnd
        style = win32gui.GetWindowLong(hwnd, win32con.GWL_EXSTYLE)

        # ウィンドウスタイルの更新：透明度とレイヤードウィンドウの設定
        new_style = style | win32con.WS_EX_LAYERED | win32con.WS_EX_TRANSPARENT
        win32gui.SetWindowLong(hwnd, win32con.GWL_EXSTYLE, new_style)

        # 透明度を50%に設定
        win32gui.SetLayeredWindowAttributes(hwnd, 0, int(0.5 * 255), win32con.LWA_ALPHA)

        # ウィンドウのタイトルバーと枠を削除
        style = win32api.GetWindowLong(hwnd, win32con.GWL_STYLE)
        style = style & ~win32con.WS_CAPTION & ~win32con.WS_THICKFRAME
        win32api.SetWindowLong(hwnd, win32con.GWL_STYLE, style)

        # ウィンドウの位置やサイズを変更せずにスタイルを更新
        win32gui.SetWindowPos(hwnd, None, 0, 0, 0, 0,
                            win32con.SWP_NOMOVE | win32con.SWP_NOSIZE | win32con.SWP_FRAMECHANGED | win32con.SWP_NOREDRAW)

    def delete_window():
        global mause_highlight_hwnd
        mause_highlight_hwnd=0
        root1.destroy()


    root1=Free_window(root)
    root1.geometry("100x100")
    root1.attributes("-topmost", True)
    canvas=Canvas(root1)
    canvas.pack(fill=BOTH, expand=True)
    canvas.create_rectangle(0, 0, 300, 300, fill='yellow')

    # タイトルバーを削除する前にウィンドウを表示
    root1.update_idletasks()
    remove_title_bar()
    t=threading.Thread(target=mouse_move, daemon=True)
    t.start()
    root1.protocol("WM_DELETE_WINDOW",delete_window )

def image_excel():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def image_excel1(drop):
        paths=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(paths)
        for x in x_list:
            try:
                kakutyou=os.path.splitext(x)[1]
                excel_file=x.replace(kakutyou,".xlsx")
                image = cv2.imdecode(
                    np.fromfile(x, dtype=np.uint8),
                    cv2.IMREAD_UNCHANGED
                    )
                wb = Workbook()
                ws = wb.active

                col_name = [chr(i) for i in range(65, 91)]

                i = 0
                for j in range(26, len(image[0])):
                    col_name.append(col_name[i] + col_name[j % 26])
                    if (j + 1) % 26 == 0:
                        i += 1
                    if len(col_name) == len(image[0]):
                        break

                for gyo in range(len(image)):
                    for retu in range(len(image[gyo])):
                        red = hex(image[gyo][retu][2])[2:].zfill(2)
                        green = hex(image[gyo][retu][1])[2:].zfill(2)
                        blue = hex(image[gyo][retu][0])[2:].zfill(2)

                        ws.column_dimensions[col_name[retu]].width = 0.3
                        ws.row_dimensions[gyo + 1].height = 1.5

                        cell_name = col_name[retu] + str(gyo + 1)
                        cell = ws[cell_name]
                        color = red + green + blue
                        cell.fill = PatternFill(patternType='solid', fgColor=color)

                wb.save(excel_file)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","変換が完了しました")


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="画像ファイルを\nドロップしてください",font=("Helvetica", 20))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',image_excel1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def markdown_edit():

    def save_md():
        path=filedialog.asksaveasfilename(filetypes=[("Markdown","*.md")],defaultextension=".md",initialfile="markdown.md")
        sample_text=text.get("1.0",END)
        with open(path, 'w',encoding="utf-8") as f:
            f.write(sample_text)

    def save_html():
        path=filedialog.asksaveasfilename(filetypes=[("HTML","*.html")],defaultextension=".html",initialfile="markdown.html")
        sample_text=text.get("1.0",END)
        md = markdown.Markdown(extensions=["extra","codehilite","sane_lists",
                                    "smarty","wikilinks","nl2br","admonition",
                                    "meta",'mdx_truly_sane_lists',"toc"],)
        html1=md.convert(sample_text).replace("<table>\n","<table border='1'>\n")
        html1=html1.replace("<body>","<body style='font-family:Meiryo;'>")
        html1=convert_to_html_links(html1)
        with open(path, 'w',encoding="utf-8") as f:
            f.write(html1)

    def make_table():
        def generate_markdown_table(rows, columns):
            # Creating the header row
            header_row = "| " + " | ".join(["Header" for i in range(columns)]) + " |"
            # Creating the separator row
            separator_row = "|" + "---|" * columns
            # Creating the data rows
            data_rows = [
                "| " + " | ".join(["Data" for j in range(columns)]) + " |"
                for _ in range(rows - 1)
            ]
            # Combining all parts
            markdown_table = "\n".join([header_row, separator_row] + data_rows)
            return markdown_table

        table_size=simpledialog.askstring("表の大きさ","行数,列数").split(",")
        width,height=int(table_size[0]),int(table_size[1])
        table_text=generate_markdown_table(width,height)
        text.insert(INSERT, table_text+"\n")

    def make_h():
        text.insert(INSERT, "[TOC]\n")

    def make_code():
        try:
            # 選択されたテキストの開始と終了インデックスを取得
            start_index = text.tag_ranges("sel")[0]
            end_index = text.tag_ranges("sel")[1]
            # 選択されたテキストを取得
            selected_text = text.get(start_index, end_index)
            # テキストを**で囲む
            enclosed_text = f"`{selected_text}`"
            # 元のテキストを削除して新しいテキストを挿入
            text.delete(start_index, end_index)
            text.insert(start_index, enclosed_text)
        except:
            text.insert(INSERT, "`コード`")

    def make_bold():
        try:
            # 選択されたテキストの開始と終了インデックスを取得
            start_index = text.tag_ranges("sel")[0]
            end_index = text.tag_ranges("sel")[1]
            # 選択されたテキストを取得
            selected_text = text.get(start_index, end_index)
            # テキストを**で囲む
            enclosed_text = f"**{selected_text}**"
            # 元のテキストを削除して新しいテキストを挿入
            text.delete(start_index, end_index)
            text.insert(start_index, enclosed_text)
        except:
            text.insert(INSERT, "**太字**")

    def make_ital():
        try:
            # 選択されたテキストの開始と終了インデックスを取得
            start_index = text.tag_ranges("sel")[0]
            end_index = text.tag_ranges("sel")[1]
            # 選択されたテキストを取得
            selected_text = text.get(start_index, end_index)
            # テキストを**で囲む
            enclosed_text = f"*{selected_text}*"
            # 元のテキストを削除して新しいテキストを挿入
            text.delete(start_index, end_index)
            text.insert(start_index, enclosed_text)
        except:
            text.insert(INSERT, "*斜体*")

    def make_strong():
        try:
            # 選択されたテキストの開始と終了インデックスを取得
            start_index = text.tag_ranges("sel")[0]
            end_index = text.tag_ranges("sel")[1]
            # 選択されたテキストを取得
            selected_text = text.get(start_index, end_index)
            # テキストを**で囲む
            enclosed_text = f"***{selected_text}***"
            # 元のテキストを削除して新しいテキストを挿入
            text.delete(start_index, end_index)
            text.insert(start_index, enclosed_text)
        except:
            text.insert(INSERT, "***強調***")

    def make_link():
        text.insert(INSERT, "[リンク名](URL)")

    def make_annotation():
        text.insert(INSERT, "[^注釈名]\n\n[^注釈名]:注釈内容")

    def make_text_color():
        rgb1=colorchooser.askcolor(title="文字色")[1]
        rgb2=colorchooser.askcolor(title="背景色")[1]
        try:
            # 選択されたテキストの開始と終了インデックスを取得
            start_index = text.tag_ranges("sel")[0]
            end_index = text.tag_ranges("sel")[1]
            # 選択されたテキストを取得
            selected_text = text.get(start_index, end_index)
            # テキストを**で囲む
            enclosed_text = f"<span style='color: {rgb1}; background-color:{rgb2};'>{selected_text}</span>"
            # 元のテキストを削除して新しいテキストを挿入
            text.delete(start_index, end_index)
            text.insert(start_index, enclosed_text)
        except:
            text.insert(INSERT, f"<span style='color: {rgb1};background-color:{rgb2}; '>文字色</span>")

    def make_del():
        try:
            # 選択されたテキストの開始と終了インデックスを取得
            start_index = text.tag_ranges("sel")[0]
            end_index = text.tag_ranges("sel")[1]
            # 選択されたテキストを取得
            selected_text = text.get(start_index, end_index)
            # テキストを**で囲む
            enclosed_text = f"~~{selected_text}~~"
            # 元のテキストを削除して新しいテキストを挿入
            text.delete(start_index, end_index)
            text.insert(start_index, enclosed_text)
        except:
            text.insert(INSERT, "~~取り消し~~")

    def on_key_press(event):
        index = text.index("insert").split('.')
        line_number = int(index[0])
        current_line = text.get(f"{line_number}.0", f"{line_number}.end")
        cursor_pos = text.index(INSERT)
        line_end = cursor_pos.split('.')[0] + '.end'
        text_after_cursor = text.get(cursor_pos, line_end)

        if event.keysym == 'Return':
            stripped_line = current_line.lstrip()
            list_match = re.match(r'(\s*)(\d+)\.\s*(.*)', current_line)
            hyphen_list_match = re.match(r'(\s*)-\s+.*', current_line)  # ハイフンとスペースの組み合わせのマッチング
            asterisk_list_match = re.match(r'(\s*)\*\s+.*', current_line)  # アスタリスクとスペースの組み合わせのマッチング

            if hyphen_list_match or asterisk_list_match:  # ハイフンまたはアスタリスクとスペースで始まる場合
                process_list_item(line_number, stripped_line, current_line,text_after_cursor)
                text.delete(cursor_pos, line_end)
                text.mark_set(INSERT, f"{int(cursor_pos.split('.')[0])+1}.end")
                update_numbered_list(line_number + 1)  # 番号付きリストの更新
                return "break"
            elif list_match:
                process_numbered_list_item(line_number, list_match,text_after_cursor)
                text.delete(cursor_pos, line_end)
                text.mark_set(INSERT, f"{int(cursor_pos.split('.')[0])+1}.end")
                update_numbered_list(line_number + 1)  # 番号付きリストの更新
                return "break"

        elif event.keysym == 'Tab':
            adjust_indent(line_number, current_line, event.state & 0x0001)
            return "break"

    def process_list_item(line_number, stripped_line, current_line,text_after_cursor):
        if stripped_line in ('-', '- ', '*', '* '):  # '*' を追加
            text.delete(f"{line_number}.0", f"{line_number}.end")
        else:
            indent = len(current_line) - len(stripped_line)
            new_bullet = '*' if stripped_line.startswith('*') else '-'  # 新しいバレットを決定
            text.insert(f"{line_number}.end", '\n' + ' ' * indent + new_bullet + ' '+text_after_cursor)

    def process_numbered_list_item(line_number, match,text_after_cursor):
        indent, number, content = match.groups()
        if not content:
            text.delete(f"{line_number}.0", f"{line_number}.end")
        else:
            next_number = int(number) + 1
            text.insert(f"{line_number}.end", '\n' + indent + f"{next_number}. "+text_after_cursor)

    def adjust_indent(line_number, current_line, is_shift_pressed):
        indent_match = re.match(r'(\s*)(\-|\*|\d+\.)', current_line)  # '*' を追加
        if indent_match:
            indent, list_type = indent_match.groups()
            if list_type in ('-', '*'):  # '*' を追加
                new_indent = "  " + indent if not is_shift_pressed else indent[2:]
                text.delete(f"{line_number}.0", f"{line_number}.{len(indent)}")
                text.insert(f"{line_number}.0", new_indent)
            else:
                # 番号付きリストの場合
                new_indent = "  " + indent if not is_shift_pressed else indent[2:]
                reset_list_number(line_number, new_indent, is_decreasing_indent=is_shift_pressed)
        update_numbered_list(line_number)


    def reset_list_number(line_number, new_indent, is_decreasing_indent):
        current_line = text.get(f"{line_number}.0", f"{line_number}.end")
        match = re.match(r'(\s*)(\d+)\.', current_line)
        if match:
            if is_decreasing_indent:
                new_number = str(get_previous_number_same_level(line_number, new_indent))
            else:
                new_number = str(get_next_list_number_same_level(line_number, new_indent))
            replacement = f"{new_indent}{new_number}."
            text.replace(f"{line_number}.0", f"{line_number}.{len(match.group(0))}", replacement)

    def get_next_list_number_same_level(line_number, indent):
        next_line_number = line_number - 1
        while next_line_number > 0:
            line = text.get(f"{next_line_number}.0", f"{next_line_number}.end")
            match = re.match(r'(\s*)(\d+)\.', line)
            if match and match.group(1) == indent:
                return int(match.group(2)) + 1
            elif match and len(match.group(1)) < len(indent):
                # 異なるネストが存在する
                break
            next_line_number -= 1
        return 1

    def get_previous_number_same_level(line_number, indent):
        prev_line_number = line_number - 1
        while prev_line_number > 0:
            line = text.get(f"{prev_line_number}.0", f"{prev_line_number}.end")
            match = re.match(r'(\s*)(\d+)\.', line)
            if match:
                if match.group(1) == indent:
                    return int(match.group(2)) + 1
                elif len(match.group(1)) < len(indent):
                    break  # 間に異なるネストが存在する
            prev_line_number -= 1
        return 1

    def update_numbered_list(changed_line_number):
        start_line = find_list_start(changed_line_number)
        line_number = start_line
        last_line_number = int(text.index('end-1c').split('.')[0])

        # 各レベルの番号を保持する辞書
        level_numbers = {}

        while line_number <= last_line_number:
            line = text.get(f"{line_number}.0", f"{line_number}.end")
            indent_level = get_indent_level(line)

            if is_numbered_list_item(line):
                # 同じレベルの以前の番号を取得し、それを1増やす
                number = level_numbers.get(indent_level, 0) + 1
                level_numbers[indent_level] = number

                # より深いレベルの番号をリセット
                for level in list(level_numbers.keys()):
                    if level > indent_level:
                        del level_numbers[level]

                new_line = re.sub(r'\d+\.', f"{number}.", line)
                text.replace(f"{line_number}.0", f"{line_number}.end", new_line)

            elif line.strip() == '':
                break  # 空の行でリスト終了とみなす

            line_number += 1

    def is_numbered_list_item(line):
        # 番号付きリストアイテムかどうかを判断
        return bool(re.match(r'\s*\d+\.', line))

    def find_list_start(line_number):
        # リストの開始行を見つける
        while line_number > 0:
            line = text.get(f"{line_number}.0", f"{line_number}.end")
            if not is_numbered_list_item(line) or line.strip() == '':
                break
            line_number -= 1
        return line_number + 1

    def get_indent_level(line):
        indent_match = re.match(r'(\s*)', line)
        return len(indent_match.group(1)) if indent_match else 0

    def image_drop(drop):
        path=drop.data.replace("{","").replace("}","").replace("\\","/")
        # mdファイルの場合のみ分岐
        if os.path.splitext(path)[1]==".md":
            with open(path, 'r',encoding="utf-8") as f:
                sample_text=f.read()
                text.delete("1.0",END)
                text.insert(INSERT, sample_text)
            return

        image_text=f'![{path.split("/")[-1]}](file:///{path})'
        text.insert(INSERT, image_text)

    def find_urls(text):
        # 正規表現を用いてURLを見つける
        url_pattern = r'https?://[^\s<>"]+|www\.[^\s<>"]+'
        return re.findall(url_pattern, text)

    def convert_to_html_links(text):
        urls = find_urls(text)
        urls=list(set(urls))
        for url in urls:
            # HTMLリンクタグの形式に変換
            html_link = f'<a href="{url}">{url}</a>'
            text = text.replace(url, html_link)
        return text

    def load_new_page(url):

        url1=str(url.replace("file:///",""))
        url2=url1.replace(os.getcwd(),"")
        pattern1=r"/#.+"
        if re.match(pattern1,url2):
            url_new="file:///"+os.getcwd()+url2.replace(":","")
            frame_html.load_url(url_new)
            return
        elif os.path.isfile(url1):
            os.startfile(url1)
            return
        webbrowser.open(url)

    def get_html():
        nonlocal num,html
        sample_text=text.get("1.0",END)
        md = markdown.Markdown(extensions=["extra","codehilite","sane_lists",
                                    "smarty","wikilinks","nl2br","admonition",
                                    "meta",'mdx_truly_sane_lists',"toc"],)
        #sample_text=sample_text.replace("\n- [ ] ","\n<input type='checkbox' checked='checked'> ")
        #if sample_text[:6]=="- [ ] ":
        #    sample_text="<input type='checkbox' name='' value=''> "+sample_text[5:]
        html1=md.convert(sample_text).replace("<table>\n","<table border='1'>\n")
        html1=html1.replace("<body>","<body style='font-family:Meiryo;'>")
        html1=convert_to_html_links(html1)
        pattern = r"~~(.*?)~~"
        html1 = re.sub(pattern, r"<s>\1</s>", html1)
        html1=html1.replace("<blockquote>","<blockquote style='font-style: italic; background-color: lightgray'>")
        soup = BeautifulSoup(html1, 'html.parser')
        for tag in soup.find_all(id=True):
            tag['id'] = tag['id'].replace(':', '')
        html1=str(soup)
        num=frame_html.yview()[0]
        if html1!=html:
            html=html1
            frame_html.load_html(html1)
            frame_html.yview_moveto(num)
        root1.after(1000,get_html)

    num=0.0
    html=""
    root1 = Free_window(root)
    root1.geometry("1500x600")
    root1.title("Markdown Editor")
    menubar = Menu(root1)
    paned_window = PanedWindow(root1, showhandle=False)

    paned_window.pack(fill=BOTH, expand=True)
    frame_l=ttk.Frame(paned_window)
    frame_r=ttk.Frame(paned_window)
    paned_window.add(frame_l)
    paned_window.add(frame_r)
    root1.config(menu=menubar)
    # menubarを親として設定メニューを作成と表示
    setting_menu = Menu(menubar, tearoff=0)
    menubar.add_cascade(label='ファイル', menu=setting_menu)
    setting_menu.add_command(label='マークダウンとして保存', command=save_md)
    setting_menu.add_command(label='HTMLとして保存', command=save_html)

    frame_html = HtmlFrame(frame_r,messages_enabled = False,horizontal_scrollbar=True) #create the HTML browser
    frame_html.enable_stylesheets(enabled=True)
    frame_html.on_link_click(load_new_page)
    frame_html.set_maximum_thread_count(19)
    frame_html.on_done_loading(lambda: frame_html.yview_moveto(num))

    root1.after(100,get_html)

    text=scrolledtext.ScrolledText(frame_l,wrap=WORD,undo=True)

    #frame_l.pack(side=LEFT,fill="both", expand=True)
    #frame_r.pack(side=RIGHT,fill="both", expand=True)
    text.pack(fill="both", expand=True)
    frame_html.pack(fill="both", expand=True)
    #text.bind('<Return>', insert_bullet_point)
    text.drop_target_register(DND_FILES)
    text.dnd_bind('<<Drop>>', image_drop)
    text.bind('<KeyPress-Return>', on_key_press)
    text.bind('<KeyPress-Tab>', on_key_press)

    RightClickMenu(text,[["コード",make_code],
                        ["太字",make_bold],["斜体",make_ital],
                        ["強調",make_strong],["取り消し線",make_del],
                        ["リンク",make_link],["表",make_table],["目次",make_h],
                        ["注釈",make_annotation],["文字色",make_text_color]])


def heic_convert():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def heic_convert1(drop):
        paths=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(paths)
        for x in x_list:
            try:
                y=x.replace(".heic",".png")
                y=y.replace(".HEIC",".png")
                heif_file = pillow_heif.read_heif(x)
                image = Image.frombytes(
                    heif_file.mode,
                    heif_file.size,
                    heif_file.data,
                    "raw",
                    heif_file.mode,
                    heif_file.stride,
                )
                image.save(y, 'PNG')
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","変換が終了しました")


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="HEICファイルを\nドロップしてください",font=("Helvetica", 20))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',heic_convert1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def totp_generator():
    global frame,buttonY

    def add_data():
        nonlocal savejson
        try:
            root.iconify()
            make_folder(os.getcwd()+"/temp1/qr_img")
            n=pyautogui.screenshot(os.getcwd()+'/temp1/qr_img/temp.png')
            root1=Free_window()

            def start():
                nonlocal pil_img,pil_img1,canvas_img
                pil_img=Image.open(os.getcwd()+'/temp1/qr_img/temp.png')
                pil_img1=pil_img
                canvas.config(width=pil_img1.width, height=pil_img1.height)
                canvas_img= ImageTk.PhotoImage(pil_img1)
                canvas.canvas_img = canvas_img
                canvas.create_image(pil_img1.width*0.5,pil_img1.height*0.5,image=canvas_img)
                root1.attributes('-fullscreen', True)


            def on_press(event):
                canvas.delete("rect")
                nonlocal start_x
                nonlocal start_y
                start_x = canvas.canvasx(event.x)
                start_y = canvas.canvasy(event.y)

            def on_move_press(event):
                nonlocal end_x, end_y
                end_x, end_y = event.x, event.y
                canvas.delete("rect")
                canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)

            def on_release(event):
                nonlocal end_x, end_y
                end_x, end_y = event.x, event.y
                canvas.delete("rect")
                canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)
                root1.destroy()
                triming()

            def triming():
                try:
                    nonlocal start_x,start_y,end_x,end_y,savejson
                    if start_x>end_x:
                        start_x,end_x=end_x,start_x
                    if start_y>end_y:
                        start_y,end_y=end_y,start_y
                    crop=pil_img.crop((start_x,start_y,end_x,end_y))
                    crop.save(os.getcwd()+"/temp1/qr_img/temp1.png")
                    img = cv2.imdecode(
                        np.fromfile(os.getcwd()+"/temp1/qr_img/temp1.png", dtype=np.uint8),
                        cv2.IMREAD_UNCHANGED
                        )
                    qrDetector = cv2.QRCodeDetector()
                    data,_,_ = qrDetector.detectAndDecode(img)
                    uri=data
                    if y=="":
                        messagebox.showerror("エラー","うまくQRコードを読み取れませんでした")
                        return
                    os.remove(os.getcwd()+"/temp1/qr_img/temp.png")
                    os.remove(os.getcwd()+"/temp1/qr_img/temp1.png")
                    # 正規表現を定義
                    pattern = re.compile(r'otpauth://totp/([^%]+)%3A([^?]+)\?secret=([^&]+).*issuer=([^&]+)')
                    # URIから情報を抽出
                    match = pattern.match(uri)
                    if match:
                        site = unquote(match.group(1))
                        username = unquote(match.group(2))
                        secret_id = unquote(match.group(3))
                    else:
                        pattern = re.compile(r'otpauth://totp/([^%]+):([^?]+)\?secret=([^&]+).*issuer=([^&]+)')
                        match = pattern.match(uri)
                        site = unquote(match.group(1))
                        username = unquote(match.group(2))
                        secret_id = unquote(match.group(3))
                    savejson+= [{"site_name": site, "my_name": username, "code": secret_id, "uri": uri}]
                    action_list=[[item["site_name"],item["my_name"] ,pyotp.TOTP(item["code"]).now()] for item in savejson]
                    tree.configure_arr(action_list)
                    json_save()
                    root.deiconify()
                    messagebox.showinfo("完了","追加が完了しました")
                except:
                    messagebox.showerror("エラー","うまくQRコードを読み取れませんでした")
                    root.deiconify()
                    return

            start_x=None
            start_y=None
            end_x=None
            end_y=None

            pil_img=None
            pil_img1=None
            canvas_img=None
            root.update()
            canvas=Canvas(root1,width=300,height=300,bg="gray")
            canvas.bind("<ButtonPress-1>", on_press)
            canvas.bind("<ButtonRelease-1>", on_release)
            canvas.bind("<B1-Motion>", on_move_press)
            canvas.pack(expand=True,side=TOP)
            root.after(10,start)
        except:
            messagebox.showerror("エラー","うまくQRコードを読み取れませんでした")
            root.deiconify()
            return

    def json_save():
        json_data = json.dumps(savejson, indent=2).encode('utf-8')
        savedata = encrypt(json_data, pass_hash)
        make_folder(os.getcwd()+"/config/totp_data")
        with open(os.getcwd()+"/config/totp_data/totp_list", 'wb') as f:
            f.write(savedata)


    def update_time():
        nonlocal action_list
        try:
            if  len(savejson)!=0:
                action_list1=action_list.copy()
                action_list=[[item["site_name"],item["my_name"] ,pyotp.TOTP(item["code"]).now()] for item in savejson]
                if action_list1!=action_list:
                    tree.configure_arr(action_list)
            label_time.config(text="残り時間："+str(30-int(time.time())%30))
            root.after(200,update_time)
        except:
            return

    def num_copy():
        try:
            record_id = tree.focus()
            record_values = tree.item(record_id, 'values')
            clip.copy(record_values[2])
        except:
            messagebox.showerror("エラー","行を指定してください")

    def delete_data():
        ok_no=messagebox.askyesno("確認","削除してもよろしいですか？")
        if ok_no!=True:
            return
        try:
            record_id = tree.focus()
            record_values = tree.item(record_id, 'values')
            for i in range(len(savejson)):
                if savejson[i]["site_name"]==record_values[0] and savejson[i]["my_name"]==record_values[1] and pyotp.TOTP(savejson[i]["code"]).now()==record_values[2]:
                    del savejson[i]
                    break
            action_list=[[item["site_name"],item["my_name"] ,pyotp.TOTP(item["code"]).now()] for item in savejson]
            tree.configure_arr(action_list)
            json_save()
            messagebox.showinfo("完了","削除しました")
        except:
            messagebox.showerror("エラー","行を指定してください")

    def show_qr():
        try:
            record_id = tree.focus()
            record_values = tree.item(record_id, 'values')
            for i in range(len(savejson)):
                if savejson[i]["site_name"]==record_values[0] and savejson[i]["my_name"]==record_values[1] and pyotp.TOTP(savejson[i]["code"]).now()==record_values[2]:
                    uri1=savejson[i]["uri"]
                    break
            make_folder(os.getcwd()+"/temp1/qr_img")
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(uri1)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            img.save(os.getcwd()+"/temp1/qr_img/temp.png")
            root1=Free_window()
            root1.title("QRコード")
            # 最前面
            root1.attributes('-topmost', True)
            canvas=Canvas(root1,width=300,height=300,bg="gray")
            canvas.pack(expand=True,side=TOP)
            pil_img=Image.open(os.getcwd()+"/temp1/qr_img/temp.png")
            pil_img1=pil_img
            canvas.config(width=pil_img1.width, height=pil_img1.height)
            canvas_img= ImageTk.PhotoImage(pil_img1)
            canvas.canvas_img = canvas_img
            canvas.create_image(pil_img1.width*0.5,pil_img1.height*0.5,image=canvas_img)
        except:
            messagebox.showerror("エラー","行を指定してください")


    def change_pass():
        nonlocal savejson,pass_hash
        new_pass=simpledialog.askstring("パスワード","新しいパスワードを入力してください")
        if new_pass==None:
            return
        pass_hash=hashlib.sha512(f"{str(new_pass)}YukisArmyknife".rstrip().encode()).hexdigest()
        for i in range(10):
            pass_hash=hashlib.sha512(pass_hash.rstrip().encode()).hexdigest()
        json_save()
        messagebox.showinfo("完了","パスワードを変更しました")


    def aes_new(password, iv):
        sha = Crypto.Hash.SHA256.new()
        sha.update(password.encode())
        return AES.new(sha.digest(), AES.MODE_CFB, iv)

    def decrypt(data, password):
        iv, cipher = data[:AES.block_size], data[AES.block_size:]
        return aes_new(password, iv).decrypt(cipher)

        # 暗号化
    def encrypt(data, password):
        iv = Crypto.Random.new().read(AES.block_size)
        return iv + aes_new(password, iv).encrypt(data)

    try:
        if os.path.exists(os.getcwd()+"/config/totp_data/totp_list"):
            with open(os.getcwd()+"/config/totp_data/totp_list", 'rb') as f:
                totp_data=f.read()
            password=simpledialog.askstring("パスワード","パスワードを入力してください")
            if password==None:
                return
            pass_hash=hashlib.sha512(f"{str(password)}YukisArmyknife".rstrip().encode()).hexdigest()
            for i in range(10):
                pass_hash=hashlib.sha512(pass_hash.rstrip().encode()).hexdigest()
            savedata = decrypt(totp_data, pass_hash)
            savejson=json.loads(savedata)
        else:
            messagebox.showerror("初期設定","パスワードを設定してください")
            new_pass=simpledialog.askstring("パスワード","パスワードを入力してください")
            pass_hash=hashlib.sha512(f"{str(new_pass)}YukisArmyknife".rstrip().encode()).hexdigest()
            for i in range(10):
                pass_hash=hashlib.sha512(pass_hash.rstrip().encode()).hexdigest()
            savejson=[]
            messagebox.showinfo("初期設定","初期設定が完了しました")
    except:
        messagebox.showerror("エラー","パスワードが間違っています")
        return


    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    if len(savejson)==0:
        action_list=[]
    else:
        action_list=[[item["site_name"],item["my_name"] ,pyotp.TOTP(item["code"]).now()] for item in savejson]

    tree=ScrolledTree(frame,arrRows=action_list,columns=["サイト","ユーザー名","パス"],arrColWidth=[150,150,100],arrSortType=["name","name","num"],height=20)
    label_time=ttk.Label(frame,text="残り時間：")
    button_add=ttk.Button(frame,text="追加",command=add_data)
    RightClickMenu(tree,[["コピー",num_copy],["QR表示",show_qr],["削除",delete_data]],tree_rclick=True)
    button_change=ttk.Button(frame,text="パスワード変更",command=change_pass)

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=2)
    buttonY.grid(row=0,column=0)
    label_time.grid(row=1,column=0)
    button_change.grid(row=1,column=1)
    button_add.grid(row=1,column=2)
    tree.grid(row=2,column=0,columnspan=3)

    root.after(10,update_time)

def video_equally():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def video_cut1(drop):
        paths=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(paths)
        cut_time=int(entry1.get().split(":")[0])*60+int(entry1.get().split(":")[1])
        for x in x_list:
            name,exp=os.path.splitext(x)[0],os.path.splitext(x)[1]
            if os.path.splitext(x)[1]==".mp4" or os.path.splitext(x)[1]==".mov" or os.path.splitext(x)[1]==".avi":
                video_clip = VideoFileClip(x)
                check_video=1
            elif os.path.splitext(x)[1]==".mp3" or os.path.splitext(x)[1]==".wav":
                video_clip = AudioFileClip(x)
                check_video=0
            duration = video_clip.duration
            num_splits = int(duration // cut_time)
            for i in range(num_splits):
                start_time = i * cut_time
                end_time = (i + 1) * cut_time

                # 動画を指定の時間で切り抜く
                subclip = video_clip.subclip(start_time, end_time)
                if check_video==1:
                    # 分割した動画を保存
                    subclip.write_videofile(f"{name}_{i+1}{exp}",logger=None)
                elif check_video==0:
                    # 分割した音声を保存
                    subclip.write_audiofile(f"{name}_{i+1}{exp}",logger=None)

            # 元の動画クリップを閉じる
            video_clip.close()
        messagebox.showinfo("完了","分割が終了しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    label=ttk.Label(frame,text="動画・音声ファイルを\nドロップしてください",font=("Helvetica", 20))
    label1=ttk.Label(frame,text="分割間隔(分:秒)")
    entry1=ttk.Entry(frame,width=10)

    label.drop_target_register(DND_FILES)
    label.dnd_bind('<<Drop>>',video_cut1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label.grid(row=2,column=0,columnspan=2)

def video_audio_split():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def video_cut1(drop):
        paths=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(paths)
        cut_time=int(entry1.get().split(":")[0])*60+int(entry1.get().split(":")[1])
        for x in x_list:
            name,exp=os.path.splitext(x)[0],os.path.splitext(x)[1]
            if os.path.splitext(x)[1]==".mp4" or os.path.splitext(x)[1]==".mov" or os.path.splitext(x)[1]==".avi":
                video_clip = VideoFileClip(x)
                check_video=1
            elif os.path.splitext(x)[1]==".mp3" or os.path.splitext(x)[1]==".wav":
                video_clip = AudioFileClip(x)
                check_video=0
            duration = video_clip.duration
            start_time1 = 0
            end_time1 = cut_time
            end_time2 = duration

            # 動画を指定の時間で切り抜く
            subclip1 = video_clip.subclip(start_time1, end_time1)
            subclip2 = video_clip.subclip(end_time1, end_time2)
            if check_video==1:
                # 分割した動画を保存
                subclip1.write_videofile(f"{name}_1{exp}",logger=None)
                subclip2.write_videofile(f"{name}_2{exp}",logger=None)
            elif check_video==0:
                # 分割した音声を保存
                subclip1.write_audiofile(f"{name}_1{exp}",logger=None)
                subclip2.write_audiofile(f"{name}_2{exp}",logger=None)

            # 元の動画クリップを閉じる
            video_clip.close()
        messagebox.showinfo("完了","分割が終了しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    label=ttk.Label(frame,text="動画・音声ファイルを\nドロップしてください",font=("Helvetica", 20))
    label1=ttk.Label(frame,text="分割時間(分:秒)")
    entry1=ttk.Entry(frame,width=10)

    label.drop_target_register(DND_FILES)
    label.dnd_bind('<<Drop>>',video_cut1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label.grid(row=2,column=0,columnspan=2)

def formula_image():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    pil1=None

    def formula_image1():
        nonlocal pil1
        try:
            text=entry1.get()
            url="https://api.excelapi.org/math/tex2image?math="+text
            pil=Image.open(io.BytesIO(requests.get(url).content))
            pil1=pil.copy()
            width,height=pil.size
            width1=1500
            height1=width1*height//width
            pil=pil.resize((width1,height1))
            canvas.config(width=width1, height=height1)
            canvas_img= ImageTk.PhotoImage(pil)
            canvas.canvas_img = canvas_img
            canvas.create_image(width1*0.5,height1*0.5,image=canvas_img)
            repaired_position()
        except:
            messagebox.showerror("エラー","うまく変換できませんでした")

    def save_image():
        try:
            path=filedialog.asksaveasfilename(filetypes=[("PNG",".png")],
                                            defaultextension=".png",
                                            initialfile="数式.png")
            pil1.save(path)
            messagebox.showinfo("完了","保存が完了しました")
        except:
            messagebox.showerror("エラー","うまく保存できませんでした")

    def save_copy():
        try:
            original_image1=pil1.copy()
            output = io.BytesIO()
            data = original_image1.getdata()
            new_data = []
            for item in data:
                if item[3] == 0:
                    new_data.append((255, 255, 255, 255))
                else:
                    new_data.append(item)
            original_image1.putdata(new_data)
            original_image1.convert('RGB').save(output, 'BMP')
            data = output.getvalue()[14:]
            output.close()
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
            win32clipboard.CloseClipboard()
            messagebox.showinfo("完了","コピーが完了しました")
        except:
            messagebox.showerror("エラー","うまくコピーできませんでした")



    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    entry1=ttk.Entry(frame,width=30)
    label1=ttk.Label(frame,text="数式を入力してください")
    button=ttk.Button(frame,text="変換",command=formula_image1)
    canvas=Canvas(frame,width=300,height=300)

    RightClickMenu(canvas,[["画像を保存",save_image],["画像をコピー",save_copy]])

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    button.grid(row=2,column=0,columnspan=2)
    canvas.grid(row=3,column=0,columnspan=2)

def video_rotate():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def video_rotate1(drop):
        paths=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(paths)
        rotate_angle=int(entry1.get())
        for x in x_list:
            try:
                exp=os.path.splitext(x)[1]
                y=x.replace(exp,"_rotate"+exp)
                clip_orig = VideoFileClip(x)
                # 動画を90度回転
                clip = rotate(clip_orig, angle=rotate_angle, resample='bilinear', expand=True)
                # 回転した動画を保存
                clip.write_videofile(y,logger=None)
                # 元の動画クリップを閉じる
                clip_orig.close()
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","回転が終了しました")


    label=ttk.Label(frame,text="動画ファイルを\nドロップしてください",font=("Helvetica", 20))
    label1=ttk.Label(frame,text="回転角度")
    entry1=ttk.Entry(frame,width=10)
    entry1.insert(END,"90")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',video_rotate1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label.grid(row=2,column=0,columnspan=2)

def video_normalize():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def video_normalize1(drop):
        paths=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(paths)
        for x in x_list:
            try:
                exp=os.path.splitext(x)[1]
                y=x.replace(exp,"_normalize"+exp)
                if var.get()==0:
                    videoclip = VideoFileClip(x)
                    normalized_clip = videoclip.fx(audio_normalize)
                    normalized_clip.write_videofile(y,logger=None)
                elif var.get()==1:
                    audioclip = AudioFileClip(x)
                    normalized_clip = audioclip.fx(audio_normalize)
                    normalized_clip.write_audiofile(y,logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","正規化が終了しました")

    var=IntVar()
    var.set(0)
    radio1=ttk.Radiobutton(frame,text="動画",value=0,variable=var)
    radio2=ttk.Radiobutton(frame,text="音声",value=1,variable=var)
    label=ttk.Label(frame,text="ファイルを\nドロップしてください",font=("Helvetica", 20))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',video_normalize1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    label.grid(row=2,column=0,columnspan=2)

def video_loop():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def video_loop1(drop):
        paths=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(paths)
        for x in x_list:
            try:
                exp=os.path.splitext(x)[1]
                y=x.replace(exp,"_loop"+exp)
                clip_time=int(entry1.get().split(":")[0])*60+int(entry1.get().split(":")[1])
                if var.get()==0:
                    clip = VideoFileClip(x)
                    clip_loop =loop(clip, n=None)
                    clip=clip_loop.subclip(0,clip_time)
                    clip.write_videofile(y,logger=None)
                elif var.get()==1:
                    clip = AudioFileClip(x)
                    clip_loop =loop(clip, n=None)
                    clip=clip_loop.subclip(0,clip_time)
                    clip.write_audiofile(y,logger=None)
                clip.close()
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","ループが終了しました")


    var=IntVar()
    var.set(0)
    radio1=ttk.Radiobutton(frame,text="動画",value=0,variable=var)
    radio2=ttk.Radiobutton(frame,text="音声",value=1,variable=var)
    label=ttk.Label(frame,text="ファイルを\nドロップしてください",font=("Helvetica", 20))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',video_loop1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame,text="合計時間(分:秒)")
    entry1=ttk.Entry(frame,width=10)
    entry1.insert(END,"3:00")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    label1.grid(row=2,column=0)
    entry1.grid(row=2,column=1)
    label.grid(row=3,column=0,columnspan=2)

def video_invert():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def video_invert1(drop):
        paths=drop.data.replace("{","").replace("}","").replace("\\","/")
        x_list=file_mult(paths)
        for x in x_list:
            ext=os.path.splitext(x)[1]
            y=x.replace(ext,"_invert"+ext)
            clip=VideoFileClip(x)
            clip = invert_colors(clip)
            clip.write_videofile(y,logger=None)
            clip.close()
        messagebox.showinfo("完了","反転が終了しました")


    label=ttk.Label(frame,text="動画ファイルを\nドロップしてください",font=("Helvetica", 20))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',video_invert1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)

def audio_fade():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root)

    def top1(drop):
        file=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(file)
        for x in file_list:
            try:
                name=os.path.basename(x)
                clip = AudioFileClip(x)
                if fade_in.get()==1:
                    clip = audio_fadein(clip, int(entry1.get()))
                if fade_out.get()==1:
                    clip = audio_fadeout(clip, int(entry2.get()))
                name_res=x.replace(name,f"fade_{name}")
                clip.write_audiofile(name_res,logger=None)
            except:
                messagebox.showerror("エラー",f"{x}\nは失敗しました")
        messagebox.showinfo("完了","処理が完了しました")

    def select_fade():
        if fade_in.get()==1:
            entry1.configure(state="normal")
        else:
            entry1.configure(state="disabled")
        if fade_out.get()==1:
            entry2.configure(state="normal")
        else:
            entry2.configure(state="disabled")

    label1=ttk.Label(frame,text="音声ファイルをここに\nドロップしてください",font=("Helvetica", 18))
    fade_in=IntVar()
    fade_in.set(0)
    check1=ttk.Checkbutton(frame,text="フェードイン(秒)",onvalue=1,offvalue=0,variable=fade_in,command=select_fade)
    fade_out=IntVar()
    fade_out.set(0)
    check2=ttk.Checkbutton(frame,text="フェードアウト(秒)",onvalue=1,offvalue=0,variable=fade_out,command=select_fade)
    entry1=ttk.Entry(frame,state="disabled",width=10)
    entry2=ttk.Entry(frame,state="disabled",width=10)
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',top1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(1),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    check1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    check2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    label1.grid(row=3,column=0,columnspan=2)

def json_yaml():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def save_file(data):
        try:
            if var.get()==0:
                path=filedialog.asksaveasfilename(filetypes=[("YAML",".yaml")],
                                                initialfile="json_data.yaml",
                                                defaultextension=".yaml")
            elif var.get()==1:
                path=filedialog.asksaveasfilename(filetypes=[("JSON",".json")],
                                                    initialfile="yaml_data.json",
                                                    defaultextension=".json")
            if path=="":
                return
            with open(path, "w", encoding="utf-8") as file:
                file.write(data)
            messagebox.showinfo("完了","保存が完了しました")
        except:
            messagebox.showerror("エラー","うまく保存できませんでした")

    def json_ymal3(data):
        try:
            if var.get()==0:
                dict_data = json.loads(data)
                data1 = yaml.dump(dict_data)
            elif var.get()==1:
                dict_data = yaml.load(data, Loader=yaml.FullLoader)
                data1 = json.dumps(dict_data, indent=2)
            root1=Free_window()
            root1.title("変換結果")
            root1.attributes('-topmost', True)
            text=ScrolledText(root1,width=50,height=20)
            text.insert(END,data1)
            text.pack(expand=True,fill="both")
            button=ttk.Button(root1,text="ファイルに保存",command=lambda:save_file(data1))
            button.pack()
        except:
            messagebox.showerror("エラー","うまく変換できませんでした")

    def json_yaml2():
        try:
            data=clip.paste()
            json_ymal3(data)
        except:
            messagebox.showerror("エラー","データ取得に失敗しました")

    def json_yaml1(drop):
        try:
            path=drop.data.replace("{","").replace("}","").replace("\\","/")
            with open(path, "r", encoding="utf-8") as file:
                data = file.read()
            json_ymal3(data)
        except:
            messagebox.showerror("エラー","データ取得に失敗しました")


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    var=IntVar()
    var.set(0)
    radio1=ttk.Radiobutton(frame,text="JSON→YAML",value=0,variable=var)
    radio2=ttk.Radiobutton(frame,text="YAML→JSON",value=1,variable=var)
    label=ttk.Label(frame,text="ファイルを\nドロップしてください",font=("Helvetica", 20))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',json_yaml1)
    button=ttk.Button(frame,text="クリップボード変換",command=json_yaml2)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    radio1.grid(row=1,column=0)
    radio2.grid(row=1,column=1)
    label.grid(row=2,column=0,columnspan=2)
    button.grid(row=3,column=0,columnspan=2)

def image_superimposed():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    origin=None
    x=None
    point=0
    canvas_img=None

    def scansnap(n_list):
        nonlocal x
        try:
            # p1:左上　p2:右上 p3:左下 p4:右下
            p1 = np.array(n_list[0])
            p2 = np.array(n_list[3])
            p3 = np.array(n_list[2])
            p4 = np.array(n_list[1])

            img_base = cv2.imdecode(
                np.fromfile(x, dtype=np.uint8),
                cv2.IMREAD_UNCHANGED)
            height_img_base, width_img_base, _ = img_base.shape

            img_overlay = cv2.imdecode(
                np.fromfile(overlay_image_path, dtype=np.uint8),
                cv2.IMREAD_UNCHANGED)
            img_overlay= cv2.cvtColor(img_overlay, cv2.COLOR_BGR2RGB)
            height_img_overlay, width_img_overlay, _ = img_overlay.shape

            img_overlay_coord = np.array([[0, 0], [width_img_overlay, 0], [width_img_overlay, height_img_overlay], [0, height_img_overlay]], dtype=np.float32)

            coord_lst = np.array([p1, p2,p3,p4], dtype=np.float32)

            mat = cv2.getPerspectiveTransform(img_overlay_coord, coord_lst)
            perspective_image = cv2.warpPerspective(img_overlay, mat, (width_img_base, height_img_base))

            perspective_image = cv2.cvtColor(perspective_image, cv2.COLOR_BGR2BGRA)
            perspective_image[np.all(perspective_image == [0, 0, 0, 255], axis=2)] = [0, 0, 0, 0]

            test = Image.fromarray(perspective_image).convert("RGBA")
            bg = Image.fromarray(img_base).convert("RGBA")

            img_clear = Image.new("RGBA", bg.size, (255, 255, 255, 0))
            img_clear.paste(test)

            final = Image.alpha_composite(bg, img_clear)
            ext=os.path.splitext(x)[1]
            y=x.replace(ext,"_trapezoid"+ext)
            final.save(y, format="PNG")
        except:
            messagebox.showerror("エラー","失敗しました")


    def img_top(drop):
        nonlocal canvas_img,pil_img,z,origin,pil_img_1,x
        try:
            x=drop.data.replace("{","").replace("}","").replace("\\","/")
            y= os.path.basename(x)
            z=x.replace(y,"trans_"+y)
            pil_img=pil_img_1=Image.open(x)
            if pil_img.width>pil_img.height:
                origin=0
                pil_img=pil_img.resize((int(ws/2),int(pil_img.height*((ws/2)/pil_img.width))),Image.LANCZOS)
            elif pil_img.width<pil_img.height:
                origin=1
                pil_img=pil_img.resize((int(pil_img.width*((hs/2)/pil_img.height)),int(hs/2)),Image.LANCZOS)
            elif pil_img.width==pil_img.height:
                origin=2
                pil_img=pil_img.resize((int(hs/2),int(hs/2)),Image.LANCZOS)
            canvas.delete("all")
            canvas_img= ImageTk.PhotoImage(pil_img)
            canvas.create_image(pil_img.width/2,pil_img.height/2,image=canvas_img)
            canvas.config(width=pil_img.width, height=pil_img.height)
            repaired_position()
        except:
            messagebox.showerror("エラー","失敗しました")

    def on_press(event):
        nonlocal point
        nonlocal start_x
        nonlocal start_y,startx_1,starty_1
        nonlocal pointX0,pointY0,pointX1,pointY1,pointX2,pointY2,pointX3,pointY3
        if point==3:
            start_x = canvas.canvasx(event.x)
            start_y = canvas.canvasy(event.y)
            pointX3=start_x
            pointY3=start_y
            canvas.delete("rect")
            canvas.unbind("<Motion>")
            canvas.create_line(start_x, start_y, startx_1, starty_1, tags=f"rect{point}", width=1.5,fill="#06ff06")
            point=0
        elif point==0:
            canvas.delete("rect0")
            canvas.delete("rect1")
            canvas.delete("rect2")
            canvas.delete("rect3")
            start_x = canvas.canvasx(event.x)
            startx_1=start_x
            start_y = canvas.canvasy(event.y)
            starty_1=start_y
            pointX0=start_x
            pointY0=start_y
            point+=1
            canvas.bind("<Motion>", on_move_press)
        elif point==1:
            start_x = canvas.canvasx(event.x)
            start_y = canvas.canvasy(event.y)
            pointX1=start_x
            pointY1=start_y
            point+=1
        elif point==2:
            start_x = canvas.canvasx(event.x)
            start_y = canvas.canvasy(event.y)
            pointX2=start_x
            pointY2=start_y
            point+=1

    def on_move_press(event):
        nonlocal end_x, end_y,point
        end_x, end_y = event.x, event.y
        canvas.delete(f"rect{point}")
        canvas.create_line(start_x, start_y, end_x, end_y, tags=f"rect{point}", width=1.5,fill="#06ff06")


    def mozaic():
        nonlocal origin,x,overlay_image_path
        try:
            point_list=[[pointX0,pointY0],[pointX1,pointY1],[pointX2,pointY2],[pointX3,pointY3]]
            overlay_image_path=entry1.get()
            pil_img_1=Image.open(x)
            if origin==2:
                for i in range(4):
                    point_list[i][0]=int(point_list[i][0]*pil_img_1.width/int(hs/2))
                    point_list[i][1]=int(point_list[i][1]*pil_img_1.height/int(hs/2))
            elif origin==1:
                for i in range(4):
                    point_list[i][0]=int(point_list[i][0]*pil_img_1.width/int(pil_img_1.width*((hs/2)/pil_img_1.height)))
                    point_list[i][1]=int(point_list[i][1]*pil_img_1.height/int(hs/2))
            elif origin==0:
                for i in range(4):
                    point_list[i][0]=int(point_list[i][0]*pil_img_1.width/int(ws/2))
                    point_list[i][1]=int(point_list[i][1]*pil_img_1.height/int(pil_img_1.height*((ws/2)/pil_img_1.width)))
            scansnap(point_list)
            messagebox.showinfo("完了","完了しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    def img_top1(drop):
        path=drop.data.replace("{","").replace("}","").replace("\\","/")
        entry1.delete(0,END)
        entry1.insert(END,path)

    start_x=None
    start_y=None
    end_x=None
    end_y=None
    pil_img=None
    z=None
    pil_img_1=None
    startx_1=None
    starty_1=None
    pointX1=None
    pointY1=None
    pointX2=None
    pointY2=None
    pointX3=None
    pointY3=None
    pointX0=None
    pointY0=None
    overlay_image_path=None

    root.update()

    frame_t=ttk.Frame(frame)
    label1=ttk.Label(frame_t,text="重ねる画像のパス(ドロップしてください)")
    entry1=ttk.Entry(frame_t,width=30)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    canvas=Canvas(frame,width=300,height=300,bg="gray")
    button1=ttk.Button(frame,text="重畳表示",command=mozaic)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    canvas.drop_target_register(DND_FILES)
    canvas.dnd_bind('<<Drop>>',img_top)
    canvas.bind("<ButtonPress-1>", on_press)
    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',img_top1)
    entry1.drop_target_register(DND_FILES)
    entry1.dnd_bind('<<Drop>>',img_top1)

    frame.pack()
    buttonX.pack(side=TOP)
    buttonY.pack(side=TOP)
    frame_t.pack(side=TOP)
    label1.grid(row=0,column=0)
    entry1.grid(row=0,column=1)
    canvas.pack(expand=True,side=TOP,fill=BOTH)
    button1.pack(side=TOP)

def pdf_search():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def pdf_search1(drop):
        def file_open():
            record_id = tree.focus()
            record_values = tree.item(record_id, 'values')
            os.startfile(record_values[2])

        try:
            if entry.get()=="":
                messagebox.showerror("エラー","検索ワードを入力してください")
                return
            path=drop.data.replace("{","").replace("}","").replace("\\","/")
            tree_list=[]
            if os.path.isfile(path):
                reader = pypdf.PdfReader(path)
                number_of_pages = len(reader.pages)
                search_string = entry.get()
                for i in range(number_of_pages):
                    page = reader.pages[i]
                    text = page.extract_text()
                    name=os.path.basename(path)
                    if search_string in text:
                        path=path.replace("\\","/")
                        tree_list.append([name,i+1,path])
            elif os.path.isdir(path):
                pdf_files = []
                # フォルダ内のすべてのファイルとサブディレクトリを取得
                for root, dirs, files in os.walk(path):
                    for file in files:
                        # ファイルがPDFで終わる場合、リストに追加
                        if file.lower().endswith('.pdf'):
                            pdf_files.append(os.path.join(root, file))

                for x in pdf_files:
                    reader = pypdf.PdfReader(x)
                    number_of_pages = len(reader.pages)
                    search_string = entry.get()
                    for i in range(number_of_pages):
                        page = reader.pages[i]
                        text = page.extract_text()
                        name=os.path.basename(x)
                        if search_string in text:
                            x=x.replace("\\","/")
                            tree_list.append([name,i+1,x])

            root1=Free_window()
            root1.title("検索結果")
            root1.attributes('-topmost', True)
            label_t=ttk.Label(root1,text=f"検索テキスト：{entry.get()}　　検索結果：{len(tree_list)}件")
            tree=ScrolledTree(root1,
                            columns=["ファイル名","ページ数","パス"],
                            arrRows=tree_list,
                            arrColWidth=[300,100,800],
                            arrSortType=["name","num","name"])
            label_t.pack(expand=True,fill="both")
            tree.pack(expand=True,fill="both")
            tree.focus_set()
            RightClickMenu(tree,func=[["ファイルを開く",file_open]],tree_rclick=True)
        except:
            messagebox.showerror("エラー","失敗しました")


    label1=ttk.Label(frame,text="PDFを含むフォルダを\nドロップしてください",font=("Helvetica", 20))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',pdf_search1)
    label2=ttk.Label(frame,text="検索ワード：")
    entry=ttk.Entry(frame,width=30)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label2.grid(row=1,column=0)
    entry.grid(row=1,column=1)
    label1.grid(row=2,column=0,columnspan=2)

def qr_camera():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def detect_qr_code(frame):
        qr_code_detector = cv2.QRCodeDetector()
        return qr_code_detector.detectAndDecode(frame)

    def display_frame(canvas, frame):
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        frame = Image.fromarray(frame)
        photo = ImageTk.PhotoImage(frame)

        canvas.config(width=photo.width(), height=photo.height())
        canvas.create_image(0, 0, anchor=NW, image=photo)
        canvas.photo = photo

    def update():
        try:
            ret, frame = cap.read()

            if ret:
                try:
                    decoded_text, points, _ = detect_qr_code(frame)
                except:
                    decoded_text, points = None, None

                if points is not None and len(points) == 4 and all(len(point) == 2 for point in points):
                    for i in range(len(points)):
                        start_point = tuple(map(int, points[i]))
                        end_point = tuple(map(int, points[(i + 1) % len(points)]))
                        try:
                            cv2.line(frame, start_point, end_point, (0, 255, 0), 3)
                        except:
                            pass

                display_frame(canvas, frame)

                if decoded_text:
                    clip.copy(decoded_text)
                    messagebox.showinfo("QRコード",decoded_text)

            root.after(10, update)
        except:
            cap.release()
            return

    def main_frame1(n):
        cap.release()
        main_frame(n)

    def update1():
        update()
        repaired_position()

    cap = cv2.VideoCapture(0)

    canvas = Canvas(frame)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(1),font=("Helvetica", 7),bg="gray",fg="white")
    frame.pack()
    buttonX.pack(side=TOP)
    buttonY.pack(side=TOP)
    canvas.pack(padx=10, pady=10)
    root.after(10, update1)

def text_image():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def text_image1():
        def save_file():
            path=filedialog.asksaveasfilename(filetypes=[("PNG",".png")],
                                              initialfile="text_image.png",
                                              defaultextension=".png")
            if path=="":
                return
            image.save(path)
            messagebox.showinfo("完了","保存が完了しました")

        # 画像サイズと背景色の指定
        width, height = entry.get().split(",")
        width=int(width)
        height=int(height)
        background_color = (255, 255, 255)  # 白色
        max_font_size = width // 5

        # 画像の作成
        image = Image.new("RGB", (width, height), background_color)
        draw = ImageDraw.Draw(image)

        # フォントとテキストの設定（適切な日本語フォントのパスを指定）
        font_size = width // 5
        font_path = 'C:\Windows\Fonts\meiryo.ttc'
        font = ImageFont.truetype(font_path, font_size)

        text = text1.get("1.0", "end -1c")

        # テキストの描画位置の計算
        text_bbox = draw.textbbox((0, 0), text, font=font)
        text_width, text_height = text_bbox[2] - text_bbox[0], text_bbox[3] - text_bbox[1]

        while text_width > width or text_height > height:
            max_font_size -= 1
            font = ImageFont.truetype(font_path, max_font_size)
            text_bbox = draw.textbbox((0, 0), text, font=font)
            text_width, text_height = text_bbox[2] - text_bbox[0], text_bbox[3] - text_bbox[1]

        x = (width - text_width) // 2
        y = (height - text_height) // 2
        # テキストの描画
        draw.text((x, y), text, font=font, fill=(0, 0, 0))  # テキストの色を指定

        root1=Free_window()
        root1.title("画像")
        root1.attributes('-topmost', True)
        canvas=Canvas(root1)
        canvas.pack()
        canvas_img= ImageTk.PhotoImage(image)
        canvas.create_image(width/2,height/2,image=canvas_img)
        canvas.config(width=width, height=height)
        canvas.image=canvas_img
        RightClickMenu(canvas,func=[["画像を保存",save_file]])


    label1=ttk.Label(frame,text="テキストを入力してください")
    text1=ScrolledText(frame,width=50,height=10)
    label2=ttk.Label(frame,text="画像大きさ(横,縦)：")
    entry=ttk.Entry(frame,width=10)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    button=ttk.Button(frame,text="画像作成",command=text_image1)
    entry.insert(END,"1500,800")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0,columnspan=2)
    text1.grid(row=2,column=0,columnspan=2)
    label2.grid(row=3,column=0)
    entry.grid(row=3,column=1)
    button.grid(row=4,column=0,columnspan=2)

def thumbnail_md():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def thumbnail_md1():
        try:
            url=entry1.get()
            pattern = re.compile(r'(?:https?://)?(?:www\.)?(?:youtube\.com/.*\?(?:.*&)?v=|youtu\.be/)([^"&?/\s]{11,})')
            match = pattern.search(url)
            if match:
                # マッチした部分（動画ID）を返す
                id=match.group(1)
            else:
                messagebox.showerror("エラー","URLが正しくありません")
                return
            new_url=f"https://img.youtube.com/vi/{id}/maxresdefault.jpg"
            new_md=f"[![{entry2.get()}]({new_url})]({url})"
            clip.copy(new_md)
            messagebox.showinfo("完了","クリップボードにコピーしました")
        except:
            messagebox.showerror("エラー","失敗しました")


    label1=ttk.Label(frame,text="youtube動画のURL：")
    entry1=ttk.Entry(frame,width=20)
    label2=ttk.Label(frame,text="テキスト：")
    entry2=ttk.Entry(frame,width=20)
    button=ttk.Button(frame,text="作成",command=thumbnail_md1)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    button.grid(row=3,column=0,columnspan=2)

def text_template():
    root1=Free_window()
    root1.title("テキストテンプレート")
    root1.attributes('-topmost', True)

    def create_dict(list_data):
        # リストから辞書作成
        dict_data={}
        for i,x in enumerate(list_data):
            dict_data[f"key{i+1}"]=x
        return dict_data

    def add_text():
        nonlocal data
        messagebox.showinfo("追加","コピーしているテキストを追加します")
        new_text=clip.paste()
        if new_text=="":
            messagebox.showerror("エラー","テキストがありません")
            return
        data.append([new_text])
        listbox.insert_inlist(data)
        new_dict=create_dict(data)
        with open(os.getcwd()+f"/config/text_template/{searchbox.get()}.json", "w",encoding="utf-8") as file:
            file.write(json.dumps(new_dict,indent=4,ensure_ascii=False))

    def add_page():
        nonlocal data
        name=simpledialog.askstring("ページ追加","ページ名を入力してください")
        if name==None:
            return
        if name=="":
            messagebox.showerror("エラー","ページ名がありません")
            return

        with open(os.getcwd()+f"/config/text_template/{name}.json", "w",encoding="utf-8") as file:
            json.dump({}, file)
        file_names = os.listdir(os.getcwd()+"/config/text_template")
        file_list=[]
        for x in file_names:
            if os.path.splitext(x)[1]==".json":
                file_list.append(os.path.splitext(x)[0])
        searchbox["values"]=file_list
        searchbox["state"]="normal"
        searchbox.delete(0,END)
        searchbox.insert(END,name)
        searchbox["state"]="readonly"
        data=[]
        listbox.insert_inlist(data)


    def page_change(e):
        nonlocal data
        with open(os.getcwd()+f"/config/text_template/{searchbox.get()}.json", "r",encoding="utf-8") as file:
            json_data =json.load(file)
        data=list(json_data.values())
        listbox.insert_inlist(data)

    def delete_text():
        nonlocal data
        if messagebox.askyesno("確認","選択したテキストを削除しますか？"):
            select=listbox.curselection()
            if select==():
                messagebox.showerror("エラー","テキストが選択されていません")
                return
            data.pop(select[0])
            listbox.insert_inlist(data)
            new_dict=create_dict(data)
            with open(os.getcwd()+f"/config/text_template/{searchbox.get()}.json", "w",encoding="utf-8") as file:
                file.write(json.dumps(new_dict,indent=4,ensure_ascii=False))

    def up_text():
        nonlocal data
        select=listbox.curselection()
        if select==():
            messagebox.showerror("エラー","テキストが選択されていません")
            return
        if select[0]==0:
            messagebox.showerror("エラー","これ以上上に移動できません")
            return
        data[select[0]-1],data[select[0]]=data[select[0]],data[select[0]-1]
        listbox.insert_inlist(data)
        new_dict=create_dict(data)
        with open(os.getcwd()+f"/config/text_template/{searchbox.get()}.json", "w",encoding="utf-8") as file:
            file.write(json.dumps(new_dict,indent=4,ensure_ascii=False))

    def down_text():
        nonlocal data
        select=listbox.curselection()
        if select==():
            messagebox.showerror("エラー","テキストが選択されていません")
            return
        if select[0]==len(data)-1:
            messagebox.showerror("エラー","これ以上下に移動できません")
            return
        data[select[0]+1],data[select[0]]=data[select[0]],data[select[0]+1]
        listbox.insert_inlist(data)
        new_dict=create_dict(data)
        with open(os.getcwd()+f"/config/text_template/{searchbox.get()}.json", "w",encoding="utf-8") as file:
            file.write(json.dumps(new_dict,indent=4,ensure_ascii=False))

    def page_delete():
        nonlocal data
        if messagebox.askyesno("確認","選択したページを削除しますか？"):
            if searchbox.get()=="デフォルト":
                messagebox.showerror("エラー","デフォルトページは削除できません")
            os.remove(os.getcwd()+f"/config/text_template/{searchbox.get()}.json")
            file_names = os.listdir(os.getcwd()+"/config/text_template")
            file_list=[]
            for x in file_names:
                if os.path.splitext(x)[1]==".json":
                    file_list.append(os.path.splitext(x)[0])
            searchbox["values"]=file_list
            searchbox["state"]="normal"
            searchbox.delete(0,END)
            searchbox.insert(END,"デフォルト")
            searchbox["state"]="readonly"
            with open(os.getcwd()+f"/config/text_template/{searchbox.get()}.json", "r",encoding="utf-8") as file:
                json_data =json.load(file)
            data=list(json_data.values())
            listbox.insert_inlist(data)

    def text_get(e):
        text1=listbox.curselection()
        if text1==():
            return
        clip.copy(data[text1[0]][0])
        messagebox.showinfo("完了","クリップボードにコピーしました")

    if not os.path.exists(os.getcwd()+"/config/text_template/デフォルト.json"):
        make_folder(os.getcwd()+"/config/text_template")
        with open(os.getcwd()+"/config/text_template/デフォルト.json", "w",encoding="utf-8") as file:
            json.dump({}, file)
    with open(os.getcwd()+"/config/text_template/デフォルト.json", "r",encoding="utf-8") as file:
        json_data =json.load(file)
    data=list(json_data.values())

    file_names = os.listdir(os.getcwd()+"/config/text_template")
    file_list=[]
    for x in file_names:
        if os.path.splitext(x)[1]==".json":
            file_list.append(os.path.splitext(x)[0])

    frame_t=ttk.Frame(root1)
    searchbox=Searchbox(frame_t,values=file_list,width=50)
    searchbox.insert(0,"デフォルト")
    searchbox["state"]="readonly"
    listbox=ScrolledList(frame_t,inlist=data,width=50,height=20,selectmode="single")
    button1=ttk.Button(frame_t,text="定型文追加",command=add_text)
    button2=ttk.Button(frame_t,text="ページ追加",command=add_page)
    # serachboxを変更したときにページ切り替え
    searchbox.bind("<<ComboboxSelected>>",page_change)
    RightClickMenu(listbox,
                   func=[["削除",delete_text],
                                 ["上に移動",up_text],
                                 ["下に移動",down_text]],
                   list_rclick=True)
    #listbox.bind('<<ListboxSelect>>', text_get)
    listbox.bind('<Double-Button-1>', text_get)
    button3=ttk.Button(frame_t,text="ページを削除",command=page_delete)

    frame_t.pack()
    searchbox.grid(row=0,column=0,columnspan=3)
    button1.grid(row=1,column=0)
    button2.grid(row=1,column=1)
    button3.grid(row=1,column=2)
    listbox.grid(row=2,column=0,columnspan=3)
    root1.update_idletasks()
    ws=root.winfo_screenwidth()
    hs=root.winfo_screenheight()
    ws1=root1.winfo_width()
    hs1=root1.winfo_height()
    x=(ws/2)-(ws1/2)
    y=(hs/2)-(hs1/2)
    root1.geometry('+%d+%d'%(x,y))

def naming_rule():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    translator = Translator()

    def naming_rule1():
        code_text=entry1.get()
        if code_text=="":
            messagebox.showerror("エラー","テキストがありません")
            return
        if var.get()==1:
            code_text = translator.translate(code_text, src='ja', dest='en').text
        text_list=code_text.split(" ")

        # camelCase
        camelCase =""
        for i in range(len(text_list)):
            if i==0:
                camelCase+=text_list[i].lower()
            else:
                camelCase+=text_list[i].capitalize()
        entry2.delete(0,END)
        entry2.insert(END,camelCase)

        # PascalCase
        PascalCase =""
        for i in range(len(text_list)):
            PascalCase+=text_list[i].capitalize()
        entry3.delete(0,END)
        entry3.insert(END,PascalCase)

        # snake_case
        snake_case =""
        for i in range(len(text_list)):
            if i==0:
                snake_case+=text_list[i].lower()
            else:
                snake_case+="_"+text_list[i].lower()
        entry4.delete(0,END)
        entry4.insert(END,snake_case)

        # CONSTANT_CASE
        CONSTANT_CASE =""
        for i in range(len(text_list)):
            if i==0:
                CONSTANT_CASE+=text_list[i].upper()
            else:
                CONSTANT_CASE+="_"+text_list[i].upper()
        entry5.delete(0,END)
        entry5.insert(END,CONSTANT_CASE)

        # kebab-case
        kebab_case =""
        for i in range(len(text_list)):
            if i==0:
                kebab_case+=text_list[i].lower()
            else:
                kebab_case+="-"+text_list[i].lower()
        entry6.delete(0,END)
        entry6.insert(END,kebab_case)

        # Train-Case
        Train_Case =""
        for i in range(len(text_list)):
            if i==0:
                Train_Case+=text_list[i].capitalize()
            else:
                Train_Case+="-"+text_list[i].capitalize()
        entry7.delete(0,END)
        entry7.insert(END,Train_Case)

    label1=ttk.Label(frame,text="変換元：")
    entry1=ttk.Entry(frame,width=75)
    var=IntVar()
    var.set(0)
    check1=ttk.Checkbutton(frame,text="翻訳モード",variable=var,onvalue=1,offvalue=0)
    button=ttk.Button(frame,text="変換",command=naming_rule1)
    label2=ttk.Label(frame,text="camelCase：")
    entry2=ttk.Entry(frame,width=75)
    label3=ttk.Label(frame,text="PascalCase：")
    entry3=ttk.Entry(frame,width=75)
    label4=ttk.Label(frame,text="snake_case：")
    entry4=ttk.Entry(frame,width=75)
    label5=ttk.Label(frame,text="CONSTANT_CASE：")
    entry5=ttk.Entry(frame,width=75)
    label6=ttk.Label(frame,text="kebab-case：")
    entry6=ttk.Entry(frame,width=75)
    label7=ttk.Label(frame,text="Train-Case：")
    entry7=ttk.Entry(frame,width=75)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    check1.grid(row=2,column=0)
    button.grid(row=2,column=1)
    label2.grid(row=3,column=0)
    entry2.grid(row=3,column=1)
    label3.grid(row=4,column=0)
    entry3.grid(row=4,column=1)
    label4.grid(row=5,column=0)
    entry4.grid(row=5,column=1)
    label5.grid(row=6,column=0)
    entry5.grid(row=6,column=1)
    label6.grid(row=7,column=0)
    entry6.grid(row=7,column=1)
    label7.grid(row=8,column=0)
    entry7.grid(row=8,column=1)

def csv_md():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def csv_to_markdown(csv_file_path):
        # CSVファイルを読み込む
        with open(csv_file_path, 'r') as file:
            lines = file.readlines()
        # ヘッダーとデータを分割
        header = lines[0].strip().split(',')
        data = [line.strip().split(',') for line in lines[1:]]
        # Markdown形式の表に変換
        markdown_table = '| ' + ' | '.join(header) + ' |\n'
        markdown_table += '| ' + ' | '.join(['---'] * len(header)) + ' |\n'
        markdown_table += '\n'.join(['| ' + ' | '.join(row) + ' |' for row in data])
        return markdown_table

    def markdown_to_csv(markdown_table):
        # Markdown形式の表を行ごとに分割
        lines = markdown_table.strip().split('\n')
        # ヘッダーとデータを抽出
        header = [cell.strip() for cell in lines[0].strip('|').split('|')]
        data = [[cell.strip() for cell in line.strip('|').split('|')] for line in lines[2:]]
        # CSV形式に変換
        csv_data = [','.join(header)]
        csv_data.extend([','.join(row) for row in data])
        return '\n'.join(csv_data)

    def csv_md1():
        try:
            text=clip.paste()
            csv_text=markdown_to_csv(text)
            path=filedialog.asksaveasfilename(filetypes=[("CSV",".csv")],
                                            initialfile="table.csv",
                                            defaultextension=".csv")
            with open(path, "w") as file:
                file.write(csv_text)
            messagebox.showinfo("完了","完了しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    def csv_md2(drop):
        try:
            path=drop.data.replace("{","").replace("}","").replace("\\","/")
            md_text=csv_to_markdown(path)
            clip.copy(md_text)
            messagebox.showinfo("完了","クリップボードにコピーしました")
        except:
            messagebox.showerror("エラー","失敗しました")

    label=ttk.Label(frame,text="CSVファイルをドロップしてください\n(CSVファイル→Markdown(コピー))",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    button=ttk.Button(frame,text="Markdown表（コピー）→CSVファイル",command=csv_md1)
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',csv_md2)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)
    button.grid(row=2,column=0,columnspan=2)

def copy_replace():
    global frame,buttonY

    def copy_replace1():

        def copy_replace2():
            clip.copy(box.get("1.0", "end -1c"))
            messagebox.showinfo("完了","クリップボードにコピーしました")
            root1.destroy()

        text=clip.paste()
        root1=Free_window()
        root1.title("コピー書き換え")
        root1.attributes('-topmost', True)

        box=ScrolledText(root1,width=50,height=10)
        box.insert(END,text)
        box.pack(expand=True,fill="both")
        button=ttk.Button(root1,text="書き換え",command=copy_replace2)
        button.pack()


    if intermediate_screen==0:
        main_frame_delete()
        frame = ttk.Frame(root, padding=16)
        button=ttk.Button(frame,text='コピー書き換え',command=copy_replace1,padding=[80,10,80,10])
        buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
        buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
        frame.pack()
        buttonX.grid(row=0,column=1)
        buttonY.grid(row=0,column=0)
        button.grid(row=1,column=0,columnspan=2)
    elif intermediate_screen==1:
        copy_replace1()


def rsa_encrypt():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def create_rsa_key():
        # 鍵の生成
        rsa_key = RSA.generate(2048)
        path=filedialog.askdirectory()

        # 公開鍵をファイルに保存
        with open(path+"/public_key.pem", "wb") as f:
            f.write(rsa_key.publickey().export_key())

        # 秘密鍵をファイルに保存
        with open(path+"/private_key.pem", "wb") as f:
            f.write(rsa_key.export_key())

    def drop_file(drop):
        entry1.delete(0,END)
        entry1.insert(END,drop.data.replace("{","").replace("}","").replace("\\","/"))


    def change_frame():
        if radio_var.get()==1:
            label1.config(text="公開鍵ドロップ：")
            label2.config(text="暗号化したいファイルを\nドロップしてください")
            entry1.delete(0,END)

        elif radio_var.get()==2:
            label1.config(text="秘密鍵ドロップ：")
            label2.config(text="復号化したいファイルを\nドロップしてください")
            entry1.delete(0,END)

    def rsa_encrypt1(drop):
        path=drop.data.replace("{","").replace("}","").replace("\\","/")
        if radio_var.get()==1:
            try:
                ext=os.path.splitext(path)[1]
                name=os.path.splitext(path)[0]
                file_name=name+ext
                new_path=path.replace(file_name,name+"_"+ext[1:]+".bin")
                aes_key = os.urandom(16)  # 16バイトのAES鍵を生成
                public_key_path=entry1.get()
                    # 対称鍵でファイルを暗号化
                with open(path, 'rb') as f:
                    data = f.read()
                cipher_aes = AES.new(aes_key, AES.MODE_GCM)
                cipher_aes.update(b'yukisarmyknife')  # 追加のデータを使用して認証タグを生成
                encrypted_data, tag = cipher_aes.encrypt_and_digest(data)

                # 暗号化された対称鍵をRSAで暗号化
                with open(public_key_path, 'rb') as f:
                    public_key = RSA.import_key(f.read())

                cipher_rsa = PKCS1_OAEP.new(public_key)
                encrypted_aes_key = cipher_rsa.encrypt(aes_key)

                # 暗号化された対称鍵と暗号化されたデータを出力ファイルに書き込む
                with open(new_path, 'wb') as f:
                    f.write(encrypted_aes_key)
                    f.write(cipher_aes.nonce)
                    f.write(tag)
                    f.write(encrypted_data)
                messagebox.showinfo("完了","完了しました")
            except:
                messagebox.showerror("エラー","失敗しました")

        elif radio_var.get()==2:
            try:
                ext=os.path.splitext(path)[1]
                name=os.path.splitext(path)[0]
                ext1=name.split("_")[-1]
                origin_name=name.replace("_"+ext1,"")+"."+ext1
                new_path=path.replace(name+ext,origin_name)
                with open(path, 'rb') as f:
                    # 暗号化された対称鍵と暗号化されたデータを読み込む
                    encrypted_aes_key = f.read(256)  # RSA 2048ビット暗号化時の最大サイズ
                    nonce = f.read(16)
                    tag = f.read(16)
                    encrypted_data = f.read()

                # RSAで暗号化された対称鍵を復号化
                with open(entry1.get(), 'rb') as f:
                    private_key = RSA.import_key(f.read())

                cipher_rsa = PKCS1_OAEP.new(private_key)
                aes_key = cipher_rsa.decrypt(encrypted_aes_key)

                # 対称鍵でデータを復号化
                cipher_aes = AES.new(aes_key, AES.MODE_GCM, nonce=nonce)
                cipher_aes.update(b'yukisarmyknife')  # 暗号化時と同じ追加データを使用して認証タグを生成

                decrypted_data = cipher_aes.decrypt_and_verify(encrypted_data, tag)
                with open(new_path, 'wb') as f:
                    f.write(decrypted_data)
                    messagebox.showinfo("完了","完了しました")
            except:
                messagebox.showerror("エラー","失敗しました")



    radio_var=IntVar()
    radio_var.set(1)
    radio1=ttk.Radiobutton(frame,text="暗号化",value=1,variable=radio_var,command=change_frame)
    radio2=ttk.Radiobutton(frame,text="復号化",value=2,variable=radio_var,command=change_frame)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    button1=ttk.Button(frame,text="鍵生成",command=create_rsa_key)
    label1=ttk.Label(frame,text="公開鍵ドロップ：")
    entry1=ttk.Entry(frame,width=20)
    label2=ttk.Label(frame,text="暗号化したいファイルを\nドロップしてください",font=("Helvetica", 16))
    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',drop_file)
    entry1.drop_target_register(DND_FILES)
    entry1.dnd_bind('<<Drop>>',drop_file)
    label2.drop_target_register(DND_FILES)
    label2.dnd_bind('<<Drop>>',rsa_encrypt1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    button1.grid(row=1,column=0,columnspan=2)
    radio1.grid(row=2,column=0)
    radio2.grid(row=2,column=1)
    label1.grid(row=3,column=0)
    entry1.grid(row=3,column=1)
    label2.grid(row=4,column=0,columnspan=2)

def text_diff():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    dmp = diff_match_patch.diff_match_patch()

    def text_diff1():
        text_1=text1.get("1.0", "end -1c")
        text_2=text2.get("1.0", "end -1c")
        diff = dmp.diff_main(text_1, text_2)
        text3.config(state="normal")
        text4.config(state="normal")
        text3.delete("1.0", "end")
        text4.delete("1.0", "end")
        text3.tag_configure("r", foreground="#FF0000")
        text4.tag_configure("g", foreground="#00FF00")

        for i in diff:
            if i[0] == 0:
                text3.insert("end", i[1])
                text4.insert("end", i[1])
            elif i[0] == -1:
                text3.insert("end", i[1], "r")
            elif i[0] == 1:
                text4.insert("end", i[1], "g")

        text3.config(state="disabled")
        text4.config(state="disabled")

    text1=ScrolledText(frame,width=30,height=10)
    text2=ScrolledText(frame,width=30,height=10)
    button=ttk.Button(frame,text="比較",command=text_diff1,width=10)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(0),font=("Helvetica", 7),bg="gray",fg="white")
    text3=ScrolledText(frame,width=30,height=10,state="disabled")
    text4=ScrolledText(frame,width=30,height=10,state="disabled")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    text1.grid(row=1,column=0)
    text2.grid(row=1,column=1)
    button.grid(row=2,column=0,columnspan=2)
    text3.grid(row=3,column=0)
    text4.grid(row=3,column=1)

def image_round():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    origin=None
    canvas_img=None

    def crop_circle(image, output_path):
        # 画像の高さ, 幅を取得
        height, width = image.shape[:2]
        # 円の中心座標, 半径を計算
        center = (width // 2, height // 2)
        radius = min(center)
        # 円形のマスクを作成
        mask = np.zeros((height, width), dtype=np.uint8)
        cv2.circle(mask, center, radius, (255, 255, 255), -1, lineType=cv2.LINE_AA)
        # 透明チャンネルを作成
        alpha = np.zeros((height, width), dtype=np.uint8)
        alpha[:] = 255
        # マスクと透明チャンネルを結合
        result = cv2.merge([image, alpha])
        # 外側のみを透明にする
        result = cv2.bitwise_and(result, result, mask=mask)
        # 結果を保存
        retval, buf = cv2.imencode('*.png', result)
        buf.tofile(output_path)
        cv2.imwrite(output_path, result)

    def pil2cv(image):
        new_image = np.array(image, dtype=np.uint8)
        if new_image.ndim == 2:  # モノクロ
            pass
        elif new_image.shape[2] == 3:  # カラー
            new_image = cv2.cvtColor(new_image, cv2.COLOR_RGB2BGR)
        elif new_image.shape[2] == 4:  # 透過
            new_image = cv2.cvtColor(new_image, cv2.COLOR_RGBA2BGRA)
        return new_image


    def img_top(drop):
        nonlocal canvas_img,pil_img,z,origin,pil_img_1
        x=drop.data.replace("{","").replace("}","").replace("\\","/")
        y= os.path.basename(x)
        ext=os.path.splitext(y)[1]
        z=x.replace(y,"crop_"+y)
        z=z.replace(ext,".png")
        try:
            pil_img=pil_img_1=Image.open(x)
        except:
            messagebox.showerror("エラー",f"{x}\nをうまく開けませんでした")
            return
        if pil_img.width>pil_img.height:
            origin=0
            pil_img=pil_img.resize((int(ws/2),int(pil_img.height*((ws/2)/pil_img.width))),Image.LANCZOS)
        elif pil_img.width<pil_img.height:
            origin=1
            pil_img=pil_img.resize((int(pil_img.width*((hs/2)/pil_img.height)),int(hs/2)),Image.LANCZOS)
        elif pil_img.width==pil_img.height:
            origin=2
            pil_img=pil_img.resize((int(hs/2),int(hs/2)),Image.LANCZOS)
        canvas.delete("all")
        canvas_img= ImageTk.PhotoImage(pil_img)
        canvas.create_image(pil_img.width/2,pil_img.height/2,image=canvas_img)
        canvas.config(width=pil_img.width, height=pil_img.height)

    def on_press(event):
        canvas.delete("rect")
        nonlocal start_x
        nonlocal start_y
        start_x = canvas.canvasx(event.x)
        start_y = canvas.canvasy(event.y)

    def on_move_press(event):
        nonlocal end_x, end_y
        end_x, end_y = event.x, event.y
        canvas.delete("rect")
        canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)

    def on_release(event):
        nonlocal end_x, end_y
        end_x, end_y = event.x, event.y
        canvas.delete("rect")
        canvas.create_rectangle(start_x, start_y, end_x, end_y, outline="lightgreen", tags="rect", width=1.5)

    def triming():
        nonlocal origin
        try:
            if origin==2:
                start_x1=int(start_x*pil_img_1.height/int(hs/2))
                start_y1=int(start_y*pil_img_1.height/int(hs/2))
                end_x1=int(end_x*pil_img_1.height/int(hs/2))
                end_y1=int(end_y*pil_img_1.height/int(hs/2))
            elif origin==1:
                start_x1=int(start_x*pil_img_1.width/int(pil_img_1.width*((hs/2)/pil_img_1.height)))
                start_y1=int(start_y*pil_img_1.height/int(hs/2))
                end_x1=int(end_x*pil_img_1.width/int(pil_img_1.width*((hs/2)/pil_img_1.height)))
                end_y1=int(end_y*pil_img_1.height/int(hs/2))
            elif origin==0:
                start_x1=int(start_x*pil_img_1.width/int(ws/2))
                start_y1=int(start_y*pil_img_1.height/int(pil_img_1.height*((ws/2)/pil_img_1.width)))
                end_x1=int(end_x*pil_img_1.width/int(ws/2))
                end_y1=int(end_y*pil_img_1.height/int(pil_img_1.height*((ws/2)/pil_img_1.width)))
            if start_x1>end_x1:
                start_x1,end_x1=end_x1,start_x1
            if start_y1>end_y1:
                start_y1,end_y1=end_y1,start_y1
            crop=pil_img_1.crop((start_x1,start_y1,end_x1,end_y1))
            #crop.save(z)
            cv2_image=pil2cv(crop)
            crop_circle(cv2_image,z)

            messagebox.showinfo("終了","トリミング完了しました")
        except:
            messagebox.showerror("エラー","うまくトリミングできませんでした")

    start_x=None
    start_y=None
    end_x=None
    end_y=None
    pil_img=None
    z=None
    pil_img_1=None
    root.update()
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',img_top)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    canvas=Canvas(frame,width=300,height=300,bg="gray")
    button1=ttk.Button(frame,text="丸トリミング",command=triming)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")

    canvas.bind("<ButtonPress-1>", on_press)
    canvas.bind("<ButtonRelease-1>", on_release)
    canvas.bind("<B1-Motion>", on_move_press)

    frame.pack()
    buttonX.pack(side=TOP)
    buttonY.pack(side=TOP)
    canvas.pack(expand=True,side=TOP,fill=BOTH)
    button1.pack(side=TOP)

def file_type():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def file_type1(drop):
        try:
            path=drop.data.replace("{","").replace("}","").replace("\\","/")
            with open(path, "rb") as f:
                res = m.identify_bytes(f.read())
            messagebox.showinfo("ファイルの種類",res.output.ct_label)
        except:
            messagebox.showerror("エラー","失敗しました")

    m = Magika()
    label=ttk.Label(frame,text="ファイルの種類を調べる\n(ドロップしてください)",font=("Helvetica", 16))
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label.grid(row=1,column=0,columnspan=2)
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',file_type1)

def webp_compress():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)

    def check_button():
        if lossless.get()==1:
            label1["text"]="圧縮度："
        elif lossless.get()==0:
            label1["text"]="品質："

    def webp_compress1(drop):
        files=drop.data.replace("{","").replace("}","").replace("\\","/")
        file_list=file_mult(files)
        for file in file_list:
            name=os.path.basename(file)
            ext=os.path.splitext(name)[1]
            img=Image.open(file)
            name=name.replace(ext,".webp")
            name1=file.replace(name,"compress_"+name)
            img.save(name1, "WEBP", quality=quality.get(),lossless=lossless.get(),method=6)
        messagebox.showinfo("完了","完了しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    quality=IntVar()
    quality.set(75)
    lossless=IntVar()
    lossless.set(0)
    label1=ttk.Label(frame,text="品質：")
    entry1=ttk.Entry(frame,textvariable=quality,width=5)
    label2=ttk.Label(frame,text="ロスレス：")
    check1=ttk.Checkbutton(frame,variable=lossless,onvalue=1,offvalue=0,command=check_button)
    label=ttk.Label(frame,text="WEBPファイルを\nドロップしてください",font=("Helvetica", 16))
    frame.drop_target_register(DND_FILES)
    frame.dnd_bind('<<Drop>>',webp_compress1)

    frame.pack()
    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    check1.grid(row=2,column=1)
    label.grid(row=3,column=0,columnspan=2)

def file_delete():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    frame.pack()

    def drop_file(drop):
        path=drop.data.replace("{","").replace("}","").replace("\\","/")
        entry1.delete(0,END)
        entry1.insert(END,path)

    def file_delete1():
        directory=entry1.get()
        pattern=entry2.get()
        for filename in os.listdir(directory):
            # ファイル名が正規表現に一致するか確認
            if re.match(pattern, filename):
                # 一致するファイルを削除
                if var1.get()==1:
                    if os.path.isfile(os.path.join(directory, filename)):
                        os.remove(os.path.join(directory, filename))
                # ディレクトリかどうかをチェック
                if var2.get()==1:
                    if os.path.isdir(os.path.join(directory, filename)):
                        shutil.rmtree(os.path.join(directory, filename))
        messagebox.showinfo("完了","完了しました")


    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(3),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame,text="フォルダをドロップ：")
    entry1=ttk.Entry(frame,width=40)
    label2=ttk.Label(frame,text="削除するファイル名（正規表現）：")
    entry2=ttk.Entry(frame,width=40)
    button=ttk.Button(frame,text="削除",command=file_delete1)
    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',drop_file)
    entry1.drop_target_register(DND_FILES)
    entry1.dnd_bind('<<Drop>>',drop_file)
    var1=IntVar()
    var1.set(1)
    var2=IntVar()
    var2.set(1)
    check1=ttk.Checkbutton(frame,text="ファイル削除",onvalue=1,offvalue=0,variable=var1)
    check2=ttk.Checkbutton(frame,text="フォルダ削除",onvalue=1,offvalue=0,variable=var2)

    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    check1.grid(row=3,column=0)
    check2.grid(row=3,column=1)
    button.grid(row=4,column=0,columnspan=2)

def process_priority():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    frame.pack()

    def process_priority1(path):
        check_num=0
        for proc in psutil.process_iter(['pid', 'name']):
        # 指定したプロセス名と一致する場合
            if proc.info['name'] == path:
                if var.get()=="通常":
                    priority=psutil.NORMAL_PRIORITY_CLASS
                elif var.get()=="やや高":
                    priority=psutil.ABOVE_NORMAL_PRIORITY_CLASS
                elif var.get()=="高":
                    priority=psutil.HIGH_PRIORITY_CLASS
                elif var.get()=="やや低":
                    priority=psutil.BELOW_NORMAL_PRIORITY_CLASS
                elif var.get()=="低":
                    priority=psutil.IDLE_PRIORITY_CLASS
                try:
                    p = psutil.Process(proc.info['pid'])
                    # プロセスの優先度を変更
                    p.nice(priority)
                    check_num=1
                except psutil.AccessDenied:
                    check_num=2
                except psutil.NoSuchProcess:
                    check_num=3
        if check_num==1:
            messagebox.showinfo("完了",f"プロセス {path} の優先度を {var.get()} に設定しました。")
        elif check_num==2:
            messagebox.showerror("エラー",f"プロセス {path} の優先度を変更する権限がありません。")
        elif check_num==3:
            messagebox.showerror("エラー",f"プロセス {path} が見つかりませんでした。")

    def stopping():
        listener.stop()
        t.join()
        frame_delete()

    def main_frame1(num):
        listener.stop()
        t.join()
        root.protocol("WM_DELETE_WINDOW", frame_delete)
        main_frame(num)

    def key_check():
        nonlocal listener
        with keyboard.Listener(on_release=on_release) as listener:
            listener.join()

    def get_window_process_path(hwnd):
        # ウィンドウのプロセスID取得
        _, pid = win32process.GetWindowThreadProcessId(hwnd)
        # プロセスIDからプロセス取得
        process = psutil.Process(pid)
        # プロセスの実行ファイルのパス取得
        path = process.exe()

        return path

    def on_release(key):
        nonlocal listener
        try:
            if key == keyboard.Key.pause:
                # アクティブなウィンドウのハンドル
                hwnd = win32gui.GetForegroundWindow()
                window_title = win32gui.GetWindowText(hwnd)
                if window_title != "":
                    path = get_window_process_path(hwnd)
                    if os.path.exists(path):
                        # エクスプローラーを開いて、ファイル選択
                        name=os.path.basename(path)
                        process_priority1(name)
                    else:
                        messagebox.showerror('エラー', 'ウィンドウのパスを取得できませんでした')
                else:
                    messagebox.showerror('エラー', 'ウィンドウのパスを取得できませんでした')
        except:
            listener.join()
            return

    t=threading.Thread(target=key_check,daemon=True)
    listener=None
    label1=ttk.Label(frame,text="プロセス優先度")
    var=StringVar()
    var="通常"
    radio1=ttk.Radiobutton(frame,text="通常",value="通常",variable=var)
    radio2=ttk.Radiobutton(frame,text="やや高",value="やや高",variable=var)
    radio3=ttk.Radiobutton(frame,text="高",value="高",variable=var)
    radio4=ttk.Radiobutton(frame,text="やや低",value="やや低",variable=var)
    radio5=ttk.Radiobutton(frame,text="低",value="低",variable=var)
    #button1=ttk.Button(frame,text="ウィンドウから指定",command=window_set)
    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame1(3),font=("Helvetica", 7),bg="gray",fg="white")
    label=ttk.Label(frame,text="pauseでウィンドウを実行している\nファイルを直接開きます",font=("Helvetica", 16))

    buttonX.grid(row=0,column=2)
    buttonY.grid(row=0,column=0,columnspan=2)
    label1.grid(row=1,column=0,columnspan=3)
    radio1.grid(row=2,column=0)
    radio2.grid(row=2,column=1)
    radio3.grid(row=2,column=2)
    radio4.grid(row=3,column=0)
    radio5.grid(row=3,column=1)
    #button1.grid(row=4,column=0,columnspan=3)
    label.grid(row=4,column=0,columnspan=3)
    root.protocol("WM_DELETE_WINDOW", stopping)
    t.start()

def code_image():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    frame.pack()

    def code_img():
        try:
            index_num=alias_list.index(box1.get())
            lexer = get_lexer_by_name(alias_list1[index_num][0])
            if var1.get()==1:
                num_show=True
            else:
                num_show=False
            formatter = ImageFormatter(style=box2.get(),
                                    encoding='utf-8',
                                    font_name='BIZ UDGothic & BIZ UDPGothic (TrueType);',
                                    line_number_separator=False,
                                    line_numbers=num_show,
                                    )
            path=filedialog.asksaveasfilename(defaultextension=".png",filetypes=[("PNG","*.png")],initialfile="code.png")
            with open(path, 'wb') as f:
                highlight(box3.get("1.0", "end-1c"), lexer, formatter, outfile=f)
            messagebox.showinfo("完了","完了しました")
        except:
            messagebox.showerror("エラー","失敗しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")

    alias_list = [i[0] for i in get_all_lexers()]
    alias_list1 = [i[1] for i in get_all_lexers()]
    label1=ttk.Label(frame,text="言語を指定：")
    box1=Searchbox(frame,values=alias_list,width=20)
    label2=ttk.Label(frame,text="スタイルを指定：")
    box2=ttk.Combobox(frame,values=list(STYLE_MAP.keys()),width=20,state="readonly")
    box2.set("default")
    label3=ttk.Label(frame,text="コードを入力：")
    box3=ScrolledText(frame,width=30,height=10)
    var1=IntVar()
    var1.set(1)
    check=ttk.Checkbutton(frame,text="行番号表示",onvalue=1,offvalue=0,variable=var1)


    button=ttk.Button(frame,text="画像化",command=code_img)

    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    box1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    box2.grid(row=2,column=1)
    label3.grid(row=3,column=0,columnspan=2)
    check.grid(row=4,column=0,columnspan=2)
    box3.grid(row=5,column=0,columnspan=2)
    button.grid(row=6,column=0,columnspan=2)

def image_compare():
    global frame,buttonY
    main_frame_delete()
    frame = ttk.Frame(root, padding=12)
    frame.pack()

    def drop_file1(drop):
        path=drop.data.replace("{","").replace("}","").replace("\\","/")
        entry1.delete(0,END)
        entry1.insert(END,path)

    def drop_file2(drop):
        path=drop.data.replace("{","").replace("}","").replace("\\","/")
        entry2.delete(0,END)
        entry2.insert(END,path)

    def image_compare1():

        def save_img():
            try:
                path=filedialog.asksaveasfilename(defaultextension=".png",filetypes=[("PNG","*.png")],initialfile="compare.png")
                image_pil.save(path,"png",optimize=True)
                messagebox.showinfo("完了","保存しました")
            except:
                messagebox.showerror("エラー","失敗しました")

        try:
            img1=cv2_save(entry1.get())
            img2=cv2_save(entry2.get())
            img1_gray = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
            img2_gray = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)
            diff = cv2.absdiff(img1_gray, img2_gray)
            thresh = 30
            diff[diff < thresh] = 0
            diff[diff >= thresh] = 255
            contours, hierarchy = cv2.findContours(diff, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            for contour in contours:
                (x, y, w, h) = cv2.boundingRect(contour)
                cv2.rectangle(img1, (x, y), (x + w, y + h), (0, 0, 255), 2)

            img1 = cv2.cvtColor(img1, cv2.COLOR_BGR2RGB)
            image_pil = Image.fromarray(img1)
            image_tk  = ImageTk.PhotoImage(image_pil)
            root1=Free_window()
            root1.title("比較結果")
            canvas = Canvas(root1, width=image_pil.width, height=image_pil.height)
            canvas.pack()
            canvas.create_image(0, 0, image=image_tk, anchor='nw')
            canvas.image = image_tk
            RightClickMenu(canvas,[("保存",save_img)])
        except:
            messagebox.showerror("エラー","失敗しました")

    buttonX=ttk.Checkbutton(frame,text=main_checkbox_name,onvalue=1,offvalue=0,variable=window_front,command=execute)
    buttonY=Button(frame,text="戻る",command=lambda:main_frame(2),font=("Helvetica", 7),bg="gray",fg="white")
    label1=ttk.Label(frame,text="比較元画像をドロップ：")
    label2=ttk.Label(frame,text="比較先画像をドロップ：")
    entry1=ttk.Entry(frame,width=20)
    entry2=ttk.Entry(frame,width=20)
    button=ttk.Button(frame,text="比較",command=image_compare1)

    label1.drop_target_register(DND_FILES)
    label1.dnd_bind('<<Drop>>',drop_file1)
    entry1.drop_target_register(DND_FILES)
    entry1.dnd_bind('<<Drop>>',drop_file1)
    label2.drop_target_register(DND_FILES)
    label2.dnd_bind('<<Drop>>',drop_file2)
    entry2.drop_target_register(DND_FILES)
    entry2.dnd_bind('<<Drop>>',drop_file2)

    buttonX.grid(row=0,column=1)
    buttonY.grid(row=0,column=0)
    label1.grid(row=1,column=0)
    entry1.grid(row=1,column=1)
    label2.grid(row=2,column=0)
    entry2.grid(row=2,column=1)
    button.grid(row=3,column=0,columnspan=2)



# GUI
time.sleep(0.1)

try:
   ctypes.windll.user32.SetProcessDPIAware()
   ctypes.windll.shcore.SetProcessDpiAwareness(True)
except AttributeError:
    pass

if os.path.exists(os.getcwd()+"/temp1"):
    shutil.rmtree(os.getcwd()+"/temp1", ignore_errors=True)
make_folder(os.getcwd()+"/config")
if not os.path.exists(os.getcwd()+"/config/style.txt"):
    with open(os.getcwd()+"/config/style.txt", "w") as file:
        file.write("0")
with open(os.getcwd()+"/config/style.txt", "r") as file:
    style_setting = file.read()

if not os.path.exists(os.getcwd()+"/config/adv_setting.json"):
    data={"width_num":4,
          "front_screen":1,
          "multi_window":0,
          "auto_update_chack":0,
          "show_center":0,
          "intermediate_screen":1,
          "hide_mainwindow":0,}
    with open(os.getcwd()+"/config/adv_setting.json", "w") as file:
        json.dump(data, file)
with open(os.getcwd()+"/config/adv_setting.json", "r") as file:
    adv_setting =json.load(file)
if "width_num" in adv_setting:
    main_frame_width=adv_setting["width_num"]
else:
    main_frame_width=4
if "front_screen" in adv_setting:
    front_screen=adv_setting["front_screen"]
else:
    front_screen=1
if "multi_window" in adv_setting:
    multi_window=adv_setting["multi_window"]
else:
    multi_window=0
if "auto_update_chack" in adv_setting:
    auto_update_chack=adv_setting["auto_update_chack"]
else:
    auto_update_chack=0
if "show_center" in adv_setting:
    show_center=adv_setting["show_center"]
else:
    show_center=0
if "intermediate_screen" in adv_setting:
    intermediate_screen=adv_setting["intermediate_screen"]
else:
    intermediate_screen=1
if "hide_mainwindow" in adv_setting:
    hide_mainwindow=adv_setting["hide_mainwindow"]
else:
    hide_mainwindow=0

if multi_window==0:
    WindowName = "Yuki's army knife"
    WindowHandle = win32gui.FindWindow(None, WindowName)
    if 0 != WindowHandle:
        root=Tk()
        root.withdraw()
        messagebox.showerror("エラー","すでに起動しています")
        sys.exit()

main_checkbox_name="表示ショトカ無効"
main_show_shrotcut=0


root = Tk()
s = ttk.Style()
if style_setting.strip()=="modern":
    s.theme_use("adapta")
elif style_setting.strip()=="dark":
    s.theme_use("black")
elif style_setting.strip()=="digital":
    s.theme_use("aquativo")
elif style_setting.strip()=="natural":
    s.theme_use("clearlooks")
elif style_setting.strip()=="systematic":
    s.theme_use("elegance")
elif style_setting.strip()=="keramik":
    s.theme_use("keramik")
elif style_setting.strip()=="wood-like":
    s.theme_use("kroc")
elif style_setting.strip()=="plastik":
    s.theme_use("plastik")
elif style_setting.strip()=="radiance":
    s.theme_use("radiance")
elif style_setting.strip()=="XPblue":
    s.theme_use("winxpblue")
elif style_setting.strip()=="sand":
    s.theme_use("scidsand")
elif style_setting.strip()=="classic":
    s.theme_use("clam")
elif style_setting.strip()=="modern_dark":
    sv_ttk.set_theme("dark")
elif style_setting.strip()=="modern_light":
    sv_ttk.set_theme("light")
s.configure("TNotebook", tabposition='n')

root.resizable(False, False)
root.title("Yuki's army knife")
photo = my_icon.get_photo_image4icon()
root.iconphoto(True, photo)
root.bind("<Escape>", esc_key_pressed)
ws=root.winfo_screenwidth()
hs=root.winfo_screenheight()
x=(ws/2)-250
y=(hs/2)-300
root.geometry('+%d+%d'%(x,y))
window_front=IntVar()
set_frame1(0)
root.withdraw()
# ウィンドウ最小化
#root.bind("<Unmap>", on_minimize)
# 起動
task_area=threading.Thread(target=taskarea,daemon=True)
task_area.start()
listener_def=None
listener_window_t=threading.Thread(target=listener_window,daemon=True)
listener_window_t.start()

if front_screen==1:
    root.attributes("-topmost", True)

if auto_update_chack==1:
    root.after(1,check_new_ver)

# app_global_setting
mause_highlight_hwnd=0

root.mainloop()
